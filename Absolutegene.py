import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import plotly.io as pio
import scipy.stats as stats
from io import BytesIO
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfgen import canvas
from reportlab.platypus import (Table, TableStyle, SimpleDocTemplate, Paragraph,
                                 Spacer, PageBreak, Image as RLImage, HRFlowable, Flowable)
from reportlab.lib.units import inch
from reportlab.pdfbase.ttfonts import TTFont
from reportlab import pdfbase
from reportlab.pdfbase import pdfmetrics
import matplotlib.pyplot as plt
import matplotlib.font_manager as _mpl_fm
import streamlit.components.v1 as components
import os
import json
import urllib.request
import glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

# ── Version tracking (independent per analysis engine) ────────────────────────
GENEQUANTIFY_VERSION = "2.0.0"  # bumped: this release adds the dPCR analysis mode
ABSOLUTEGENE_VERSION = "1.4.0"
PLATFORM_VERSION = "2.0.0"      # combined GeneQuantify platform version

try:
    plt.rcParams['font.family'] = 'DejaVu Sans'
    plt.rcParams['axes.unicode_minus'] = False
except Exception:
    pass

# ─── RDML PARSER ──────────────────────────────────────────────────────────────
def parse_rdml(file_bytes):
    """
    Parse an RDML file (.rdml is a ZIP containing rdml_data.xml).
    Returns a dict: {target_name: {'unkn': [cq,...], 'ref': [cq,...]}}
    Also returns a flat DataFrame for inspection.
    """
    import zipfile, io
    try:
        import xml.etree.ElementTree as ET
    except ImportError:
        return None, "xml.etree.ElementTree not available"

    rows = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as zf:
            xml_name = next((n for n in zf.namelist() if n.endswith(".xml")), None)
            if xml_name is None:
                return None, "No XML found inside RDML file."
            with zf.open(xml_name) as xf:
                tree = ET.parse(xf)
        root = tree.getroot()

        # RDML uses namespaces like {http://www.rdml.org/rdml_v1_2.rng}
        ns_raw = root.tag.split("}")[0].lstrip("{") if "}" in root.tag else ""
        ns = f"{{{ns_raw}}}" if ns_raw else ""

        # Build lookup dicts for targets and samples
        target_types = {}   # id -> type ('toi' or 'ref')
        target_names = {}   # id -> name
        for t in root.findall(f"{ns}target"):
            tid  = t.get("id", "")
            tname = t.findtext(f"{ns}commercialAssay") or tid
            ttype = t.findtext(f"{ns}type") or "toi"
            target_types[tid] = ttype
            target_names[tid] = tname

        sample_types = {}  # id -> type ('unkn', 'ntc', 'std', etc.)
        for s in root.findall(f"{ns}sample"):
            sid   = s.get("id", "")
            stype = s.findtext(f"{ns}type") or "unkn"
            sample_types[sid] = stype

        for exp in root.findall(f"{ns}experiment"):
            for run in exp.findall(f"{ns}run"):
                for react in run.findall(f"{ns}react"):
                    sample_id = react.findtext(f"{ns}sample") or react.get("id", "")
                    stype = sample_types.get(sample_id, "unkn")
                    for data_el in react.findall(f"{ns}data"):
                        target_id = data_el.findtext(f"{ns}tar") or ""
                        cq_text   = data_el.findtext(f"{ns}cq")
                        try:
                            cq = float(cq_text) if cq_text else None
                        except ValueError:
                            cq = None
                        rows.append({
                            "Sample":      sample_id,
                            "SampleType":  stype,
                            "Target":      target_names.get(target_id, target_id),
                            "TargetType":  target_types.get(target_id, "toi"),
                            "Cq":          cq,
                        })

    except zipfile.BadZipFile:
        # Some RDML files are plain XML, not ZIP
        try:
            tree = ET.parse(io.BytesIO(file_bytes))
            root = tree.getroot()
            ns_raw = root.tag.split("}")[0].lstrip("{") if "}" in root.tag else ""
            ns = f"{{{ns_raw}}}" if ns_raw else ""
            target_types = {}
            target_names = {}
            for t in root.findall(f"{ns}target"):
                tid   = t.get("id", "")
                ttype = t.findtext(f"{ns}type") or "toi"
                target_types[tid] = ttype
                target_names[tid] = tid
            sample_types = {}
            for s in root.findall(f"{ns}sample"):
                sid   = s.get("id", "")
                stype = s.findtext(f"{ns}type") or "unkn"
                sample_types[sid] = stype
            for exp in root.findall(f"{ns}experiment"):
                for run in exp.findall(f"{ns}run"):
                    for react in run.findall(f"{ns}react"):
                        sample_id = react.findtext(f"{ns}sample") or react.get("id", "")
                        stype = sample_types.get(sample_id, "unkn")
                        for data_el in react.findall(f"{ns}data"):
                            target_id = data_el.findtext(f"{ns}tar") or ""
                            cq_text   = data_el.findtext(f"{ns}cq")
                            try:
                                cq = float(cq_text) if cq_text else None
                            except ValueError:
                                cq = None
                            rows.append({
                                "Sample":     sample_id,
                                "SampleType": stype,
                                "Target":     target_names.get(target_id, target_id),
                                "TargetType": target_types.get(target_id, "toi"),
                                "Cq":         cq,
                            })
        except Exception as e:
            return None, f"RDML parse error: {e}"
    except Exception as e:
        return None, f"RDML parse error: {e}"

    if not rows:
        return None, "No reaction data found in RDML file."

    df = pd.DataFrame(rows)
    return df, None


# ─── RDES PARSER ──────────────────────────────────────────────────────────────
def parse_rdes(file_bytes):
    """
    Parse an RDES file (tab-separated, .tsv / .csv / .txt).
    Required columns: Well, Sample, Sample Type, Target, Target Type, Dye, Cq
    Returns DataFrame with columns: Sample, SampleType, Target, TargetType, Cq
    """
    try:
        content = file_bytes.decode("utf-8", errors="replace")
        lines   = [l.rstrip("\r") for l in content.split("\n") if l.strip()]
        if not lines:
            return None, "Empty RDES file."

        header = [h.strip() for h in lines[0].split("\t")]
        required = ["Well", "Sample", "Sample Type", "Target", "Target Type", "Dye", "Cq"]
        missing  = [c for c in required if c not in header]
        if missing:
            return None, f"RDES file missing required columns: {missing}"

        rows = []
        for line in lines[1:]:
            if not line.strip():
                continue
            cells = line.split("\t")
            row   = dict(zip(header, cells))
            cq_raw = row.get("Cq", "").strip()
            try:
                cq = float(cq_raw.replace(",", ".")) if cq_raw and cq_raw != "-1.0" else None
            except ValueError:
                cq = None
            rows.append({
                "Sample":     row.get("Sample", "").strip(),
                "SampleType": row.get("Sample Type", "unkn").strip(),
                "Target":     row.get("Target", "").strip(),
                "TargetType": row.get("Target Type", "toi").strip(),
                "Cq":         cq,
            })

        if not rows:
            return None, "No data rows found in RDES file."
        return pd.DataFrame(rows), None

    except Exception as e:
        return None, f"RDES parse error: {e}"


# ─── RDML/RDES → session_state mapper ─────────────────────────────────────────





st.set_page_config(
    page_title="GeneQuantify",
    page_icon="🧬",
    layout="wide"
)

# ═══════════════════════════════════════════════════════════════════════════════
# GLOBAL CSS (shared by both qPCR and dPCR analysis modes)
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<style>
.block-container { padding-top: 3rem !important; padding-bottom: 1rem !important; }
div[data-testid="stAlert"] { padding: 5px 10px !important; font-size: 12px !important; }
div[data-testid="stNumberInput"] { margin-bottom: 0 !important; }
div[data-testid="stRadio"]       { margin-bottom: 0 !important; }
section[data-testid="stSidebar"] > div:first-child { padding-top: 0.8rem !important; }
section[data-testid="stSidebar"] hr { margin: 4px 0 !important; }
</style>
""", unsafe_allow_html=True)

if 'language' not in st.session_state:
    st.session_state.language = "English"

# ═══════════════════════════════════════════════════════════════════════════════
# ANALYSIS MODE SELECTOR — GeneQuantify (qPCR, relative quantification) vs
# AbsoluteGene (dPCR/ddPCR, absolute Poisson-based quantification). This is a
# single platform offering both engines; switching modes does not lose your
# work in the other mode (each mode's inputs live under separate session_state
# keys), but only one mode's tabs are shown at a time.
# ═══════════════════════════════════════════════════════════════════════════════
if 'analysis_mode' not in st.session_state:
    st.session_state.analysis_mode = "qPCR"

_mode_options = {
    "qPCR": "🧬 qPCR — GeneQuantify (Ct-based, relative quantification)",
    "dPCR": "🧪 dPCR / ddPCR — AbsoluteGene (partition-based, absolute quantification)",
}
_mode_display = st.sidebar.radio(
    "🔬 Analysis Mode / Analiz Modu",
    options=list(_mode_options.values()),
    index=0 if st.session_state.analysis_mode == "qPCR" else 1,
    key="analysis_mode_radio"
)
analysis_mode = "qPCR" if _mode_display == _mode_options["qPCR"] else "dPCR"
st.session_state.analysis_mode = analysis_mode
st.sidebar.caption(
    "GeneQuantify: relative gene expression / CNV from Ct values."
    if analysis_mode == "qPCR" else
    "AbsoluteGene: absolute quantification / CNV / VAF from dPCR partition counts."
)
st.sidebar.divider()

if analysis_mode == "qPCR":
    flags = {
        "Türkçe": "🇹🇷",
        "English": "🇬🇧",
        "Deutsch": "🇩🇪",
        "Français": "🇫🇷",
        "Español": "🇪🇸",
        "العربية": "🇸🇦"
    }
    default_index = list(flags.keys()).index(st.session_state.language)

    # Sidebar — logo küçük ve ortalı
    try:
        import base64 as _b64, os as _os
        _logo_path = "geneq.jpg"
        if _os.path.exists(_logo_path):
            with open(_logo_path, "rb") as _f:
                _logo_b64 = _b64.b64encode(_f.read()).decode()
            st.sidebar.markdown(
                f"<div style='text-align:center;padding:4px 0 6px 0;'>"
                f"<img src='data:image/jpeg;base64,{_logo_b64}' width='130' style='border-radius:8px;'/>"
                f"</div>",
                unsafe_allow_html=True
            )
    except Exception:
        pass
    selected_language = st.sidebar.selectbox(
        "🌐 Language",
        options=[f"{flags[lang]} {lang}" for lang in flags],
        index=default_index,
        label_visibility="collapsed"
    )

    try:
        selected_language_name = selected_language.split(' ', 1)[1]
        selected_flag = flags[selected_language_name]
    except KeyError:
        selected_language_name = selected_language
        selected_flag = None

    st.sidebar.markdown(
        f"<div style='font-size:11px;color:#888;margin:-6px 0 4px 0;'>"
        f"Language / Dil / Sprache / Langue / Idioma / لغة</div>",
        unsafe_allow_html=True
    )

    st.markdown("""
    <style>
    [data-testid="stSidebarCollapseButton"] {
        display: block !important;
        visibility: visible !important;
        opacity: 1 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    st.sidebar.divider()
    instruction_clicked = st.sidebar.button("📘 User Guide", use_container_width=True)

    if instruction_clicked or selected_language_name == "Instruction":

        @st.dialog("📘 GeneQuantify — User Guide", width="large")
        def show_guide():
            st.markdown("""
    <style>
    .guide-section { background:#f8f9fa; border-left:4px solid #2196F3; padding:10px 16px; border-radius:4px; margin-bottom:12px; }
    .guide-formula { background:#1e1e2e; color:#cdd6f4; font-family:monospace; padding:10px 14px; border-radius:6px; font-size:14px; margin:8px 0; }
    .guide-warn { background:#fff3cd; border-left:4px solid #ffc107; padding:8px 14px; border-radius:4px; }
    .guide-ok   { background:#d4edda; border-left:4px solid #28a745; padding:8px 14px; border-radius:4px; }
    </style>
    """, unsafe_allow_html=True)

            tab1, tab2, tab3, tab4, tab5 = st.tabs(["📥 Data Input", "🧮 Calculations", "📊 Statistics", "⚙️ Settings", "⚖️ Disclaimer"])

            with tab1:
                st.markdown("### 📥 Data Input Format")
                st.info("GeneQuantify accepts Cq values entered as a column — one value per line. Compatible with direct **Excel/spreadsheet copy–paste**.")
                col1, col2 = st.columns(2)
                with col1:
                    st.markdown("**✅ Correct format**")
                    st.code("23.15\n22.90\n25.20\n24.88\n23.45", language="text")
                with col2:
                    st.markdown("**✅ Also accepted (comma as decimal)**")
                    st.code("23,15\n22,90\n25,20\n24,88\n23,45", language="text")
                st.markdown("""
    **Guidelines:**
    - Minimum **3 replicates** recommended per group (required for outlier detection)
    - All groups for the same gene should have the **same number of replicates** (app auto-trims to shortest)
    - Enter one target gene and one reference gene per section
    - Multiple reference genes (geNorm normalization) can be enabled in the settings
    """)
                st.markdown("### 📋 Example Study Design")
                st.dataframe({
                    "Group": ["Control","Control","Control","Patient 1","Patient 1","Patient 1"],
                    "Target Cq": [23.1, 22.9, 25.2, 27.3, 28.1, 26.8],
                    "Reference Cq": [18.2, 17.9, 18.5, 18.3, 18.0, 18.6],
                }, use_container_width=True)

            with tab2:
                st.markdown("### 🧮 Calculation Methods")
                st.markdown("#### 1. Classic ΔΔCq Method (Livak & Schmittgen, 2001)")
                st.code("ΔCq        = Cq(target) − Cq(reference)\nΔΔCq       = ΔCt(sample) − ΔCt(control)\nFold Change = 2^(−ΔΔCq)", language="text")
                st.markdown("""
    **Assumptions:**
    - Target and reference gene efficiencies are both ~100% (E ≈ 2.0)
    - Efficiency difference between target and reference < 10%
    - If these assumptions are violated, use the **Pfaffl method** instead
    """)
                st.markdown("#### 2. Pfaffl Method (Pfaffl, 2001)")
                st.code("Ratio = (E_target ^ ΔCq_target) / (E_ref ^ ΔCt_ref)\n\nwhere:\n  ΔCt_target = Ct_control(target) − Ct_sample(target)\n  ΔCt_ref    = Ct_control(ref)    − Ct_sample(ref)", language="text")
                st.info("The Pfaffl method accounts for primer-specific efficiencies and is more accurate when E differs between genes.")

                st.markdown("#### 3. Amplification Efficiency (E)")
                st.code("E = 10^(−1 / slope)\n\nPerfect efficiency: E = 2.0 (100%)\nAcceptable range:   E = 1.8 – 2.2  (90–110%)\nSlope range:        −3.10 to −3.58", language="text")
                st.markdown("""
    **How to obtain E:**
    | Method | Description |
    |--------|-------------|
    | Standard Curve | Run 4–5 serial dilutions; qPCR software reports slope → use built-in calculator |
    | LinRegPCR | Free software; calculates E from raw fluorescence |
    | qBase+ / CFX Maestro | Automated E calculation |
    | Primer datasheet | Manufacturer-validated E for commercial kits |
    """)
                st.markdown("#### 4. Multiple Reference Genes (geNorm, Vandesompele 2002)")
                st.code("Normalization Factor (NF) = arithmetic mean of reference gene Cq values\nGeNorm M-value < 0.5  → Excellent stability\nGeNorm M-value 0.5–1.0 → Acceptable\nGeNorm M-value ≥ 1.0  → Unstable — consider excluding", language="text")

            with tab3:
                st.markdown("### 📊 Statistical Decision Pathway")
                st.markdown("""
    The app automatically selects the appropriate statistical test:

    ```
    Input ΔCq values
          │
          ▼
    Shapiro-Wilk normality test (p > 0.05 = normal)
          │
          ├── Both groups NORMAL ──▶ Levene's test (equal variance?)
          │                               │
          │                    ┌──────────┴──────────┐
          │                  YES (p>0.05)           NO (p≤0.05)
          │                    │                     │
          │             Student's t-test       Welch's t-test
          │
          └── Any group NON-NORMAL ──▶ Mann-Whitney U test
    ```

    **Multi-group (≥3 groups):**
    ```
    Normal + Equal variance     → One-way ANOVA → Tukey HSD post-hoc
    Normal + Unequal variance   → Welch ANOVA   → Games-Howell post-hoc
    Any non-normal              → Kruskal-Wallis → Dunn's test post-hoc
    ```
    """)
                st.markdown("### 🔢 Multiple Testing Correction")
                st.markdown("""
    When analyzing **multiple target genes**, the false positive risk increases:

    | Method | Controls | Best for |
    |--------|----------|----------|
    | **Bonferroni** | Family-wise error rate (FWER) | Few genes, conservative |
    | **FDR (Benjamini-Hochberg)** | False discovery rate | Many genes, more power |

    **Rule of thumb:** Report both. Use FDR for exploratory studies (≥5 genes), Bonferroni for confirmatory studies.
    """)

            with tab4:
                st.markdown("### ⚙️ Settings Guide")
                st.markdown("""
    #### Gene & Group Count
    Set the number of target genes and patient groups before entering data.  
    Each target gene gets its own Ct input sections for control and each patient group.

    #### Reference Genes
    - **1 reference gene:** Simpler, but less reliable normalization
    - **2+ reference genes:** Recommended (MIQE guidelines); geNorm stability automatically calculated
    - Reference genes should be stably expressed across all conditions

    #### Outlier Detection
    | Setting | Description |
    |---------|-------------|
    | **Grubbs test** | Best for small samples (n=3–8); detects single extreme outliers |
    | **IQR method** | Better for larger samples; flags values outside Q1−k×IQR / Q3+k×IQR |
    | **Alpha (Grubbs)** | Significance threshold; 0.05 recommended |
    | **k multiplier (IQR)** | 1.5 = standard Tukey fence; 3.0 = extreme outliers only |

    > ⚠️ Outlier exclusion **requires biological or technical justification**. The app flags candidates — the researcher decides.

    #### Efficiency Threshold
    If |E_target − E_ref| exceeds this threshold (default 10%), a warning is shown and Pfaffl method is recommended.
    """)

            with tab5:
                st.markdown("### ⚖️ Disclaimer & Citation")
                st.warning("""
    **For Research & Education Use Only**

    This application is intended for research, education, and preliminary laboratory analysis only.  
    It is **NOT** designed or validated for clinical diagnosis, treatment decisions, or patient management.

    **Users are responsible for:**
    - Verifying the accuracy of entered Cq data
    - Appropriate interpretation of results
    - Confirming findings using validated laboratory methods

    The developers are **not liable** for any decisions, losses, or damages arising from application use.  
    All clinical decisions must be made by qualified professionals.
    """)
                st.markdown("""
    **How to cite GeneQuantify:**
    Yalçınkaya B (2026). GeneQuantify: a web-based tool for qPCR gene expression and copy number variation analysis. *Molecular and Cellular Biochemistry*. https://doi.org/10.1007/s11010-026-05621-y

    **References:**
    - Livak KJ & Schmittgen TD. *Methods* 2001;25:402–408. (ΔΔCq method)
    - Pfaffl MW. *Nucleic Acids Res* 2001;29(9):e45. (Pfaffl method)
    - Vandesompele J et al. *Genome Biol* 2002;3(7). (geNorm)
    - Bustin SA et al. *Clin Chem* 2009;55(4):611–622. (MIQE guidelines)
    - Grubbs FE. *Technometrics* 1969;11(1):1–21. (Outlier detection)
    - Benjamini Y & Hochberg Y. *J R Stat Soc B* 1995;57(1):289–300. (FDR)

    **Contact:** mailtoburhanettin@gmail.com
    """)

        show_guide()
    
    language_map = {
        "Türkçe": "tr",
        "Español": "es",
        "English": "en",
        "Français": "fr",
        "Deutsch": "de",
        "العربية": "ar"
    }

    language_code = language_map.get(selected_language_name, "en")  

    translations = {
        "tr": {
            "title": "🧬 GeneQuantify: Gen Ekspresyonu ve Kopya Sayısı Varyasyonu (CNV) Analizi",
            "tab_data": "Veri Girişi",
            "tab_results": "Sonuçlar",
            "tab_report": "Rapor",
            "subtitle": "B. Yalçınkaya tarafından geliştirildi",
            "patient_data_header": "📊 Hasta ve Kontrol Grubu Verisi Girin",
            "num_target_genes": "🔹 Hedef Gen Sayısını Girin",
            "num_patient_groups": "🔹 Hasta Grubu Sayısını Girin",
            "sample_number": "Örnek Numarası",
            "Grup": "Grup",
            "x_axis_title": "Grup Adı",
            "ct_value": "Cq Değeri",
            "reference_ct": "Referans Cq",
            "delta_ct_control": "ΔCq (Kontrol)",
            "delta_ct_patient": "ΔCq (Hasta)",
            "warning_empty_input": "⚠️ Dikkat: Verileri alt alta yazın veya boşluk içeren hücre olmayacak şekilde excelden kopyalayıp yapıştırın.",
            "download_csv": "📥 CSV İndir",
            "generate_pdf": "📥 PDF Raporu Hazırla",
            "pdf_report": "Gen Ekspresyon Analizi Raporu",
            "statistics": "istatistiksel Sonuçlar",
            "nil_mine": "📊 Sonuçlar",
            "gr_tbl": "📋 Giriş Verileri Tablosu",
            "control_group": "🧬 Kontrol Grubu",
            "ctrl_trgt_ct": "🟦 Kontrol Grubu Hedef Gen {i} Cq Değerleri",
            "ctrl_ref_ct": "🟦 Kontrol Grubu Referans Gen {i} Cq Değerleri",
            "hst_trgt_ct": "🩸 Hasta Grubu Hedef Gen {j} Cq Değerleri",
            "hst_ref_ct": "🩸 Hasta Grubu Referans Gen {j} Cq Değerleri",
            "warning_control_ct": "⚠️ Dikkat: Kontrol Grubu {i} verilerini alt alta yazın veya boşluk içeren hücre olmayacak şekilde Excel'den kopyalayıp yapıştırın.",
            "warning_patient_cq": "⚠️ Dikkat: Hasta grubu Cq verilerini alt alta yazın veya boşluk içeren hücre olmayacak şekilde Excel'den kopyalayıp yapıştırın.",
            "target_gene": "Hedef Gen",
            "reference_gene": "Referans Gen",
            "target_ct": "Hedef Gen Cq",
            "distribution_graph": "Dağılım Grafiği",
            "error_missing_control_data": "⚠️ Hata: Kontrol Grubu için Hedef Gen {i} verileri eksik!",
            "control_group_avg": "Kontrol Grubu Ortalama",
            "avg": "Ortalama",
            "control": "Kontrol",
            "sample": "Örnek",
            "patient": "Hasta",
            "delta_ct_distribution": "ΔCq Dağılımı",
            "delta_ct_value": "ΔCq Değeri",
            "parametric": "Parametrik",
            "non_parametric": "Nonparametrik",
            "t_test": "t-test",
            "mann_whitney_u_test": "Mann-Whitney U testi",
            "welch_t_test": "welch_t_testi",
            "significant": "Anlamlı",
            "insignificant": "Anlamsız",
            "test_type": "Test Türü",
            "test_method": "Kullanılan Test",
            "test_pvalue": "Test P-değeri",
            "significance": "Anlamlılık",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "Gen Ekspresyon Değişimi (2^(-ΔΔCq))",
            "regulation_status": "Regülasyon Durumu",
            "no_change": "Değişim Yok",
            "upregulated": "Yukarı Regüle",
            "downregulated": "Aşağı Regüle",
            "report_title": "Gen Ekspresyon Analizi Raporu",
            "input_data_table": "Giriş Verileri Tablosu",
            "results": "Sonuçlar",
            "statistical_results": "📈 İstatistiksel Sonuçlar",
            "statistical_evaluation": "İstatistiksel Değerlendirme",
            "target_gene": "Hedef Gen",
            "patient_group": "🩸 Hasta Grubu",
            "expression_change": "Gen Ekspresyon Değişimi",
            "generate_pdf": "PDF Oluştur",
            "pdf_report": "Gen Ekspresyon Raporu",
            "error_no_data": "Veri bulunamadı, PDF oluşturulamadı.",
            # Efficiency translations
            "efficiency_header": "🔬 Amplifikasyon Etkinliği (Efficiency) Doğrulaması",
            "efficiency_method": "Efficiency Giriş Yöntemi",
            "efficiency_manual": "Manuel E değeri gir",
            "efficiency_slope": "Slope (eğim) ile hesapla",
            "efficiency_target_label": "Hedef Gen {i} Efficiency (E)",
            "efficiency_ref_label": "Referans Gen {i} Efficiency (E)",
            "efficiency_target_slope_label": "Hedef Gen {i} Slope",
            "efficiency_ref_slope_label": "Referans Gen {i} Slope",
            "efficiency_threshold": "Kabul edilebilir efficiency farkı eşiği (%)",
            "efficiency_ok": "✅ Efficiency farkı kabul edilebilir ({diff:.1f}%)",
            "efficiency_warning": "⚠️ Efficiency farkı eşiği aşıyor ({diff:.1f}%) — ΔΔCq yöntemi güvenilir olmayabilir!",
            "efficiency_target_pct": "Hedef Gen Efficiency",
            "efficiency_ref_pct": "Referans Gen Efficiency",
            "efficiency_diff": "Fark",
            "pfaffl_result": "Pfaffl Oranı",
            "pfaffl_header": "Pfaffl Metodu Sonuçları",
            "classic_ddct": "Klasik ΔΔCq Sonucu (2^(-ΔΔCq))",
            "pfaffl_ratio": "Pfaffl Oranı",
            "method_comparison": "📊 Yöntem Karşılaştırması",
            "efficiency_note": "Not: E=2.0 mükemmel etkinliği (100%) temsil eder. Kabul edilen aralık: 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "İstatistiksel değerlendirme sürecinde veri dağılımı Shapiro-Wilk testi ile analiz edilmiştir. "
                "Normallik sağlanırsa, gruplar arasındaki varyans eşitliği Levene testi ile kontrol edilmiştir. "
                "Varyans eşitliği varsa bağımsız örneklem t-testi, yoksa Welch t-testi uygulanmıştır. "
                "Normal dağılım sağlanmazsa, parametrik olmayan Mann-Whitney U testi kullanılmıştır. "
                "Sonuçların anlamlılığı p < 0.05 kriterine göre belirlenmiştir. "
                "<b>Öneri ve destekleriniz için:</b> Burhanettin Yalçınkaya - mail: mailtoburhanettin@gmail.com"
            ),
            # Outlier section
            "outlier_section_title": "### 🔍 Aykırı Değer Tespiti Ayarları",
            "outlier_enable": "Aykırı değer tespitini etkinleştir",
            "outlier_enable_help": "İstatistiksel olarak aşırı Cq değerlerini tespit eder.",
            "outlier_method_label": "Tespit yöntemi",
            "outlier_method_help": "Grubbs: normal dağılım için, tek aykırı değer. IQR: parametrik olmayan, çarpık dağılımlar için.",
            "outlier_alpha_label": "Anlamlılık düzeyi (α)",
            "outlier_alpha_help": "α = 0.05 standart değerdir. Düşük α = daha muhafazakâr.",
            "outlier_iqr_label": "IQR çarpanı (k)",
            "outlier_iqr_help": "k=1.5 = standart Tukey sınırları. k=3.0 = yalnızca aşırı aykırı değerler.",
            "outlier_expander": "ℹ️ qPCR'de aykırı değer tespiti hakkında",
            "grubbs_info": "ℹ️ **Grubbs testi gereksinimleri:** Her grup için minimum **n ≥ 3** replikat. Anlamlılık eşiği: **α = {alpha:.2f}**. Test normallik varsayar; n < 8 için normallik güvenilir biçimde değerlendirilemez — sonuçlar dikkatli yorumlanmalıdır. Gürültülü replikatların ΔCq hesabına yansımasını önlemek için **ham Cq değerlerine** (normalizasyon öncesi) uygulanması önerilir.",
            "outlier_excluded_no": "Hayır",
            "outlier_excluded_yes": "Evet",
            # Outlier stage selector 
            "outlier_stage_label": "🔬 Aykırı Değer Uygulama Aşaması",
            "outlier_stage_raw": "Ham Cq — normalizasyon öncesi (önerilen)",
            "outlier_stage_dct": "ΔCq — normalizasyon sonrası (eski davranış)",
            "outlier_stage_help": (
                "**Ham Ct (önerilen):** Aykırı değerler, ΔCq hesaplanmadan önce ham Cq değerlerine "
                "uygulanır. Her target ve referans gen için ayrı ayrı kontrol edilir. "
                "Gürültülü replikatların normalizasyona sızması engellenir.\n\n"
                "**ΔCq:** Aykırı değerler normalizasyon sonrası uygulanır (orijinal davranış)."
            ),
            # Distribution plot mode selector 
            "dist_plot_mode_label": "📊 Dağılım Grafiği — Görüntüleme Modu",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — önerilen",
            "dist_plot_dct":  "ΔCq  — ham normalize değerler",
            "dist_plot_ddct": "ΔΔCq  — kontrol ortalamasına göre",
            "dist_plot_help": (
                "**RQ (önerilen):** ΔCq → 2^(-ΔCt) dönüşümü. Yüksek değer = yüksek ekspresyon. "
                "Yüksek ΔCq = düşük ekspresyon paradoksunu ortadan kaldırır.\n\n"
                "**ΔCq:** Ham logaritmik değerler. Veri dağılımı ve normallik kontrolü için.\n\n"
                "**ΔΔCq:** Her örneğin ΔCq'si eksi kontrol grubu ortalaması. Kontrole göre değişimi gösterir."
            ),
            "unequal_n_warning": (
                "⚠️ **Eşit olmayan replikat sayısı — {group}:**  \n"
                "{details}  \nAnaliz **en kısa ortak uzunluk (n={min_n})** kullanılarak devam edecek.  \n"
                "Veri girişinizi kontrol edin — farklı n değerleri veri giriş hatası olabilir."
            ),
            # Sidebar
            "sidebar_load_example": "📂 Örnek Veri Yükle",
            "sidebar_example_loaded": "✅ Örnek veri yüklendi! Veri Girişi sekmesine geçin.",
            "sidebar_desktop_title": "### 💻 Masaüstü Uygulaması",
            "sidebar_desktop_btn": "⬇️ Masaüstü Uygulamasını İndir",
            "sidebar_opensource_title": "### 🔓 Açık Kaynak",
            "sidebar_opensource_body": "GeneQuantify açık kaynaklıdır (GPL-3.0).  \nKaynak kod GitHub'da mevcuttur:",
            "sidebar_github_btn": "⭐ GitHub'da Kaynak Kodu Görüntüle",
            "sidebar_scenarios_title": "📋 Doğrulama Senaryosu Yükle",
            "sidebar_scenario_select": "Senaryo seçin",
            "sidebar_load_scenario_btn": "▶ Senaryoyu Yükle",
            "sidebar_scenario_loaded": "✅ {s} yüklendi! Veri Girişi sekmesine geçin.",
            # Statistical decision
            "stat_decision_title": "🔬 İstatistiksel karar",
            "stat_decision_steps": "**Adım adım test seçimi:**",
            "stat_shapiro_title": "**1. Shapiro-Wilk normallik testi**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Normal değil",
            "stat_levene_title": "**2. Levene varyans homojenliği testi**",
            "stat_levene_skipped": "**2. Levene testi** — *atlandı* (normallik sağlanmadı; parametrik olmayan test kullanılacak)",
            "stat_equal_var": "Eşit varyans",
            "stat_unequal_var": "Eşitsiz varyans",
            "stat_selected_test": "**3. Seçilen test:**",
            "stat_reason": "**Gerekçe:**",
            "stat_result": "**Sonuç:**",
            "stat_reason_nonnormal": "Bir veya her iki grupta normal dağılım sağlanmadı",
            "stat_reason_normal_equal": "Her iki grup normal + eşit varyans",
            "stat_reason_normal_unequal": "Her iki grup normal + eşitsiz varyans (Levene p < 0.05)",
            "stat_multigroup_note": "⚠️ Not: ≥ 3 grup varsa, ANOVA / Kruskal-Wallis testi için aşağıdaki **Çoklu Grup Karşılaştırması** bölümüne bakın.",
            # Multi-group
            "multigroup_title": "## 📊 Çoklu Grup Karşılaştırma Analizi",
            "multigroup_expander": "ℹ️ Çoklu grup istatistiksel analizi hakkında",
            "multigroup_omnibus_test": "Omnibus Testi",
            "multigroup_pvalue": "p-değeri",
            "multigroup_result": "Sonuç",
            "multigroup_significant": "Anlamlı",
            "multigroup_not_significant": "Anlamlı değil",
            "multigroup_omnibus_ns": "ℹ️ Omnibus testi **anlamlı değil** (p ≥ 0.05). Post-hoc karşılaştırmalar bilgi amaçlı gösterilmektedir — genel grup etkisi tespit edilmedi.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Post-hoc sonuçlarını indir —",
            "multigroup_2group_note": "ℹ️ **Çoklu grup analizi uygulanamaz:** Yalnızca 2 grup tespit edildi (Kontrol + 1 hasta grubu). İkili istatistikler yukarıda raporlanmıştır.",
            "multigroup_decision_normal_equal": "✅ Normal dağılım + eşit varyans → **Tek yönlü ANOVA + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Normal dağılım + **eşitsiz varyans** → **Welch ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **Normal dağılım sağlanmadı** → **Kruskal-Wallis + Dunn post-hoc**",
            # Multi-gene correction
            "multigene_title": "### 🧬 Çoklu Gen Çoklu Karşılaştırma Düzeltmesi",
            "multigene_expander": "ℹ️ Bu neden gereklidir?",
            "multigene_sig_raw": "Anlamlı (ham)",
            "multigene_sig_bonf": "Anlamlı (Bonferroni)",
            "multigene_sig_fdr": "Anlamlı (FDR B-H)",
            "multigene_warning": "⚠️ Düzeltme sonrası, ham p < 0.05 eşiğinde anlamlı görünen {lost} sonuç FDR düzeltmesi sonrası anlamlılığını yitirdi. Çoklu gen analizlerinde düzeltilmiş p-değerlerini birincil sonuç olarak raporlayın.",
            "multigene_success": "✅ {n} anlamlı sonucun tamamı FDR düzeltmesi sonrasında da anlamlı kalmaktadır — bulgular çoklu karşılaştırmaya karşı güçlüdür.",
            "multigene_no_sig": "Ham p < 0.05 eşiğinde anlamlı ikili sonuç tespit edilmedi.",
            "multigene_dl_button": "📥 Düzeltilmiş p-değerlerini indir (CSV)",
            "multigene_chart_title": "Çoklu Gen p-değeri Düzeltmesi: Ham / Bonferroni / FDR",
            "multigene_fc_chart_title": "Çoklu Gen İfade Karşılaştırması",
            "multigene_1gene_note": "ℹ️ **Çoklu gen düzeltmesi:** Yalnızca 1 hedef gen analiz edildi — genler arası çoklu karşılaştırma düzeltmesi uygulanamaz.",
            "multigene_no_data": "Henüz p-değeri yok — hesaplama için yukarıya veri girin.",
            # Reference gene settings
            "ref_gene_section_title": "### 📚 Referans Gen Ayarları",
            "ref_gene_num_label": "Hedef gen başına referans gen sayısı",
            "ref_gene_num_help": "MIQE kılavuzları sağlam normalizasyon için ≥2 doğrulanmış referans gen önerir.",
            "ref_gene_1_warning": "⚠️ **Metodolojik not:** Tek referans gen kullanımı normalizasyon sağlamlığını kısıtlar. MIQE kılavuzları (Bustin et al. 2009) **≥2 referans gen** ve stabilite değerlendirmesi (geNorm/NormFinder) önermektedir.",
            "ref_gene_multi_success": "✅ {n} referans gen seçildi. Geometrik ortalama normalizasyonu ve geNorm M-değeri stabilitesi otomatik hesaplanacaktır.",
            "ref_gene_expander": "ℹ️ Çoklu referans normalizasyonu hakkında",
            # Standard curve calculator
            "sc_expander": "📐 Standart Eğri Hesaplayıcı — Dilüsyon serisinden E hesapla",
            "sc_gene_label": "Gen / Primer etiketi",
            "sc_num_points": "Dilüsyon noktası sayısı",
            "sc_dilution_factor_label": "**Dilüsyon faktörü** (örn. 10 katlı dilüsyon için 10)",
            "sc_dilution_factor_input": "Dilüsyon faktörü",
            "sc_start_conc_label": "**Başlangıç konsantrasyonu** (keyfi birim, örn. 1)",
            "sc_start_conc_input": "Başlangıç konsantrasyonu",
            "sc_enter_ct": "**Her dilüsyon için ortalama Cq girin:**",
            "sc_calc_button": "📊 Etkinliği Hesapla",
            "sc_slope": "Eğim",
            "sc_e_value": "E değeri",
            "sc_efficiency_pct": "Etkinlik %",
            "sc_excellent": "✅ Mükemmel! E={e:.4f} ({pct:.1f}%), R²={r2:.4f} — Bu E değerini aşağıdaki etkinlik bölümüne girin.",
            "sc_warning_r2": "⚠️ E kabul edilebilir ({pct:.1f}%) ancak R²={r2:.4f} < 0.99 — dilüsyon serinizi kontrol edin.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) kabul edilebilir aralığın dışında (90–110%). Primer tasarımınızı veya dilüsyon serinizi gözden geçirin.",
            "sc_chart_title": "Standart Eğri — {label} | E={e:.4f} ({pct:.1f}%), R²={r2:.4f}",
            "sc_xaxis": "log₁₀(Konsantrasyon)",
            "sc_data_points": "Veri noktaları",
            "sc_copy_hint": "💡 Aşağıdaki etkinlik girdilerine eğim **{slope:.4f}** veya E değeri **{e:.4f}** kopyalayın.",
            "sc_description": """\
    **Standart Eğri Hesaplayıcı nasıl kullanılır:**

    Seri dilüsyon Cq değerlerinizi aşağıya girin. Hesaplayıcı doğrusal regresyon uygulayarak eğim, R² ve amplifikasyon etkinliğini otomatik hesaplar.

    **Kullanım:**  
    1. Her primer için seri dilüsyonlarda qPCR çalıştırın (örn. seyreltilmemiş, 1:10, 1:100, 1:1000, 1:10000)  
    2. Her dilüsyon için ortalama Cq değerini girin  
    3. Eğim, E ve R² değerlerini okuyun  
    """,
            "ref_multi_description": """\
    **Geometrik ortalama normalizasyonu** (Vandesompele et al. 2002)  
    Normalizasyon faktörü (NF), her örnek için tüm referans genlerinin Cq değerlerinin aritmetik ortalamasıdır;  
    bu da ifade düzeylerinin geometrik ortalamasına karşılık gelir.  
    `NF_örnek = ortalama(Ct_ref1, Ct_ref2, ..., Ct_refN)` her örnek için  
    `ΔCq = Ct_hedef − NF`

    **geNorm M-değeri** (stabilite skoru)  
    Her referans gen için M, diğer tüm referans genlerine karşı log-oranlarının ortalama standart sapmasıdır.  
    **Düşük M = daha kararlı.** MIQE tavsiye edilen eşik: M < 0,5 (katı) veya M < 1,0 (kabul edilebilir).

    **CV (Varyasyon Katsayısı)**  
    `CV = (SS / ortalama) × 100%` tüm örneklerdeki ham Cq değerlerinin.  
    Düşük CV, daha az varyasyon ve referans olarak daha iyi kararlılık anlamına gelir.

    **Referans:** Vandesompele J et al. *Genome Biology* 2002; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **qPCR'de aykırı değer tespiti neden önemlidir?**

    Teknik değişkenlik qPCR'ye özgüdür: pipetleme hataları, hava kabarcığı oluşumu, inhibitör taşınması veya RNA kalite farklılıkları, replikat grubunun geri kalanıyla istatistiksel olarak uyumsuz Cq değerleri üretebilir.  
    Bu tür değerlerin dahil edilmesi varyansı şişirir, ortalamaları saptırır ve yanlış sonuçlara yol açabilir — özellikle küçük örneklem büyüklüklüeri olan klinik veri setlerinde.

    **Bu kısıtlamanın kritik hale geldiği durumlar:**
    - Küçük gruplar (n < 5): tek bir hatalı Ct ortalamayı önemli ölçüde kaydırır
    - Yüksek biyolojik değişkenlik (örn. tümör heterojenliği, klinik kohortlar)
    - Bir replikatın diğerlerinden > 0,5 Ct sapma gösterdiği teknik triplıkatlar
    - Ct > 35 olan düşük bolluklu hedefler, gürültünün baskın olduğu durumlar

    **Grubbs testi** *(Grubbs 1969)*  
    Normallik varsayar. En uç değerin istatistiksel olarak anlamlı bir aykırı değer olup olmadığını test eder (p < α). Başka aykırı değer bulunmayana kadar tekrarlanır.  
    En iyi: tek bir deneysel gruptan replikat Cq değerleri için.

    **IQR yöntemi** *(Tukey 1977)*  
    Parametrik olmayan. Q1 − k×IQR veya Q3 + k×IQR dışındaki değerleri işaretler.  
    En iyi: daha büyük gruplar veya normal olmayan dağılımlar için.

    **Önemli:** Aykırı değer dışlama **biyolojik veya teknik gerekçe** gerektirir.  
    Bu araç adayları işaretler — nihai karar her zaman araştırmacıya aittir.  
    Tüm dışlamalar kaydedilir ve PDF çıktısında raporlanır.

    **Referanslar:** Grubbs FE. *Technometrics* 1969; Tukey JW. *Exploratory Data Analysis* 1977;  
    Bustin SA et al. *Clin Chem* 2009 (MIQE kılavuzları).
    """,

            # ── PDF rapor stringleri ──────────────────────────────────────────────
            "pdf_cover_subtitle": "qPCR Gen Ekspresyonu Analiz Raporu",
            "pdf_generated": "Oluşturulma tarihi: {now}",
            "pdf_s1_title": "1. Yöntemler ve Analiz Ayarları",
            "pdf_s1_calc": "1.1 Hesaplama Yöntemleri",
            "pdf_s1_calc_body": "Kat değişimi hesabı için iki tamamlayıcı yöntem kullanıldı:",
            "pdf_s1_classic": "Klasik ΔΔCq (Livak & Schmittgen, 2001): ΔCq = Cq(hedef) - Ct(referans);  ΔΔCq = ΔCq(örnek) - ΔCt(kontrol);  Kat Değişimi = 2^(-ΔΔCt). Her iki gen için eşit amplifikasyon verimliliği varsayar (E ≈ 2.0).",
            "pdf_s1_pfaffl": "Pfaffl Yöntemi (Pfaffl, 2001): Oran = (E_hedef ^ ΔCq_hedef) / (E_ref ^ ΔCt_ref). Primer'e özgü verimlilikleri düzeltir; verimlilik farkı > %10 olduğunda önerilir.",
            "pdf_s1_norm": "1.2 Normalizasyon",
            "pdf_s1_norm_multi": "Çoklu referans gen (n={n}) kullanıldı. Normalizasyon faktörü (NF), her örnek için referans gen Cq değerlerinin aritmetik ortalaması olarak hesaplandı (geNorm yaklaşımı, Vandesompele et al. 2002). geNorm M-değerleri ve varyasyon katsayısı (CV%) hesaplandı.",
            "pdf_s1_norm_single": "Normalizasyon için tek referans gen kullanıldı. MIQE kılavuzları sağlam normalizasyon için ≥2 referans gen önermektedir.",
            "pdf_s1_eff": "1.3 Amplifikasyon Verimliliği",
            "pdf_s1_eff_range": "Kabul edilebilir verimlilik aralığı: E = 1.8-2.2 (%90-110%). Uygulanan verimlilik farkı eşiği: {thr}%.",
            "pdf_s1_outlier": "1.4 Aykırı Değer Tespiti",
            "pdf_s1_grubbs": "Grubbs testi (Grubbs 1969) uygulandı, alfa = {alpha}. Test, bir veri setindeki en uç değerin istatistiksel olarak anlamlı bir aykırı değer olup olmadığını t-dağılımı kritik değeri ile test eder. {n} örnek işaretlendi ve kullanıcı tarafından dışlama onaylandı.",
            "pdf_s1_iqr": "IQR yöntemi (Tukey 1977) uygulandı, çarpan k = {k}. [Q1 - k*IQR, Q3 + k*IQR] dışındaki değerler potansiyel aykırı değer olarak işaretlendi. {n} örnek dışlama için kullanıcı tarafından onaylandı.",
            "pdf_s1_outlier_warn": "UYARI: Aykırı değer dışlama biyolojik veya teknik gerekçe gerektirir. Dışlanan örnekler aşağıdaki veri tablosunda işaretlendi.",
            "pdf_s1_outlier_off": "Bu analiz için aykırı değer tespiti devre dışı bırakıldı.",
            "pdf_s2_title": "2. Giriş Verileri",
            "pdf_s2_body": "Aykırı değer işleme sonrası kullanıcı tarafından girilen ham Cq değerleri. 'Aykırı Değer Dışlandı' sütununda 'Evet' olan satırlar hesaplamalardan çıkarıldı.",
            "pdf_s3_title": "3. Gen Ekspresyonu Sonuçları",
            "pdf_s3_body": "Klasik ΔΔCq ve Pfaffl yöntemleriyle hesaplanan kat değişimi değerleri. Kat değişimi > 1 hasta grubunda kontrole göre yüksek ekspresyonu gösterir.",
            "pdf_s4_title": "4. İstatistiksel Analiz",
            "pdf_s4_body": "Kontrol ve hasta grupları arasındaki gen ekspresyon farklılıklarının istatistiksel anlamlılığı. Tüm testler ham ΔCq değil RQ (2^-ΔCt) değerleri üzerinden uygulandı; çünkü ΔCt logaritmik ölçekte olduğundan, ΔCt üzerinden t-testi biyolojik değişkenliği hafife alabilir. Test seçimi normallik (Shapiro-Wilk) ve varyans homojenliği (Levene) testlerine göre otomatik yapıldı. Anlamlılık eşiği: p < 0.05.",
            "pdf_s4_interp": "İstatistiksel Testlerin Yorumu",
            "pdf_s4_interp_body": "Student t-testi: Her iki grup normal dağılım ve eşit varyanstaysa kullanılır. Welch t-testi: Her iki grup normal fakat varyanslar eşit değilse kullanılır. Mann-Whitney U: Normallik varsayımı karşılanmadığında kullanılan parametrik olmayan test. p < 0.05 istatistiksel olarak anlamlı diferansiyel ekspresyonu gösterir.",
            "pdf_s5_title": "5. Göreli Miktar (RQ) Dağılım Grafikleri",
            "pdf_s5_body": "Her hedef gen için RQ (2^-ΔCq) değerlerinin dağılımı. Her nokta bir biyolojik replikatı temsil eder. Yatay çubuklar grup ortalamalarını gösterir. İstatistiksel testler de RQ değerleri üzerinden gerçekleştirilmiştir.",
            "pdf_s6_title": "6. Sonuçların Yorumlanması",
            "pdf_s6_fc": "6.1 Kat Değişimi Yorumu",
            "pdf_s6_choose": "6.2 ΔΔCq ve Pfaffl Arasında Seçim",
            "pdf_s6_choose_body": "Klasik ΔΔCq'yi şu durumlarda kullanın: Her iki genin verimliliği %90-110 aralığında ve aralarındaki fark %10'dan az. Pfaffl'ı şu durumlarda kullanın: Verimlilik farkı %10'u aşıyor ya da gen verimlilikleri ölçülmüş ve farklı. Her durumda her iki değeri de raporlayın.",
            "pdf_s6_stat": "6.3 İstatistiksel Test Seçimi Gerekçesi",
            "pdf_s6_stat_body": "Normallik değerlendirmesi için Shapiro-Wilk testi (küçük örneklemler, n < 50 için önerilir) kullanıldı. Varyans homojenliği için Levene testi uygulandı. Eşit varyanslı parametrik veriler için Student t-testi maksimum istatistiksel güç sağlar. Welch t-testi varyanslar farklı olduğunda daha sağlamdır. Mann-Whitney U normallik varsayılamadığında parametrik olmayan alternatiftir.",
            "pdf_s7_title": "7. Kaynaklar",
            "pdf_fc_interp_header": ["Kat Değişimi", "ΔΔCq", "Yorum", "Biyolojik Önem"],
            "pdf_fc_interp_rows": [
                [">2.0", "<-1.0", "Güçlü yukarı regülasyon", "Biyolojik olarak anlamlı kabul edilebilir"],
                ["1.5-2.0", "-1.0 ila -0.58", "Orta yukarı regülasyon", "İlgili olabilir; doğrulayın"],
                ["1.0-1.5", "-0.58 ila 0", "Zayıf yukarı regülasyon", "Tek başına genellikle önemsiz"],
                ["1.0", "0", "Değişim yok", "Diferansiyel ekspresyon yok"],
                ["0.67-1.0", "0 ila 0.58", "Zayıf aşağı regülasyon", "Tek başına genellikle önemsiz"],
                ["0.5-0.67", "0.58 ila 1.0", "Orta aşağı regülasyon", "İlgili olabilir; doğrulayın"],
                ["<0.5", ">1.0", "Güçlü aşağı regülasyon", "Biyolojik olarak anlamlı kabul edilebilir"],
            ],
            "pdf_stat_note": "Not: İstatistiksel anlamlılık (p < 0.05) ve biyolojik anlamlılık (kat değişimi büyüklüğü) birlikte değerlendirilmelidir.",
            "pdf_summary_param": "Parametre",
            "pdf_summary_val": "Değer",
            "pdf_summary_genes": "Analiz edilen hedef gen sayısı",
            "pdf_summary_groups": "Hasta grupları",
            "pdf_summary_samples": "Toplam örnek (satır)",
            "pdf_summary_excluded": "Aykırı değer dışlanan örnek",
            "pdf_summary_tests": "karşılaştırma",
            "pdf_summary_norm": "Normalizasyon yöntemi",
            "pdf_summary_norm_multi": "geNorm NF",
            "pdf_summary_norm_single": "Tek referans gen",
            "pdf_summary_methods": "Hesaplama yöntemleri",
            "pdf_summary_methods_val": "Klasik ΔΔCq + Pfaffl",
            "pdf_disclaimer": "Bu rapor GeneQuantify tarafından otomatik oluşturulmuştur. Tüm hesaplamalar MIQE kılavuzlarını (Bustin et al., Clin Chem 2009) izler.",
            "pdf_footer": "GeneQuantify — Yalnızca araştırma ve eğitim amaçlı. Klinik tanı için doğrulanmamıştır.",
            "pdf_fig1": "Şekil 1. Klasik ΔΔCq ve Pfaffl yöntemleri arasında kat değişimi karşılaştırması. Kesikli çizgi y=1'de kontrole göre değişim olmadığını gösterir.",
            "pdf_fig2": "Şekil 2. Tüm ikili karşılaştırmalar için p-değerleri. Kırmızı çubuklar istatistiksel olarak anlamlı sonuçları (p < 0.05) gösterir. Kesikli çizgi anlamlılık eşiğini işaretler.",
            "pdf_fig3": "Şekil. {gene} için RQ (2^-ΔCq) dağılımı. Noktalar = bireysel replikatlar; yatay çubuklar = grup ortalamaları.",
            "pdf_nochange": "Değişim Yok",
            "pdf_stat_cols": ["Hedef Gen", "Karşılaştırma", "Test Türü", "Kullanılan Test", "p-değeri", "Anlamlılık"],
            "pdf_res_cols": ["Hedef Gen", "Grup", "ΔCq Kontrol", "ΔCq Örnek", "ΔΔCq", "2^(-ΔΔCq)", "Pfaffl Oranı", "Regülasyon", "E hedef", "E ref"],
            "pdf_eff_cols": ["Gen", "E (hedef)", "Eff% (hedef)", "E (ref)", "Eff% (ref)", "Fark%", "Durum"],
            "pdf_eff_ok": "Kabul edilebilir",
            "pdf_eff_warn": "UYARI: Pfaffl kullanın",
            "pdf_outlier_col": "Aykırı Değer Dışlandı",
            "pdf_contact": "İletişim: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} kayıt hazır — PDF oluşturabilirsiniz.",
            # RDML / RDES import
            "rdml_expander":        "📂 RDML / RDES Dosyası İçe Aktar",
            "rdml_description":     "Cq değerlerini otomatik doldurmak için **RDML** (`.rdml`) veya **RDES** (`.tsv`/`.csv`/`.txt`) dosyası yükleyin.",
            "rdml_uploader":        "Dosya seçin",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX, Roche LightCycler vb.  RDES: sekmeyle ayrılmış tablo formatı.",
            "rdml_success":         "✅ {fmt} dosyası yüklendi — {n} reaksiyon bulundu.",
            "rdml_error":           "❌ {fmt} ayrıştırma hatası: {err}",
            "rdml_preview":         "Ayrıştırılan verileri önizle",
            "rdml_step1":           "**Adım 1 — Kontrol grubunu etiketleyin**",
            "rdml_ctrl_label":      "Kontrol örnek adı (virgülle ayrılmış alt dizeler)",
            "rdml_ctrl_help":       "Adı bu metni içeren tüm örnekler Kontrol grubu olarak işlenecektir.",
            "rdml_step2":           "**Adım 2 — Hasta gruplarını etiketleyin**",
            "rdml_n_pat":           "Hasta grubu sayısı",
            "rdml_pat_label":       "Hasta grubu {i} örnek adı (adları)",
            "rdml_pat_help":        "Virgülle ayrılmış alt dizeler. Eşleşen tüm örnekler bu gruba dahil edilir.",
            "rdml_apply":           "✅ {fmt} verilerini Veri Girişine Uygula",
            "rdml_apply_success":   "✅ {n} Cq değeri Veri Girişi sekmesine yüklendi! Kontrol edip ayarlayabilirsiniz.",
            "rdml_apply_warning":   "⚠️ Hiçbir değer eşleştirilemedi. Kontrol/hasta etiketlerinin yukarıdaki önizlemedeki örnek adlarıyla uyuştuğundan emin olun.",
        },

        "en": {
            "title": "🧬 GeneQuantify: Expression & CNV Analysis",
            "tab_data": "Data Entry",
            "tab_results": "Results",
            "tab_report": "Report",
            "subtitle": "Developed by B. Yalçınkaya",
            "patient_data_header": "📊 Enter Patient and Control Group Data",
            "num_target_genes": "🔹 Enter the Number of Target Genes",
            "num_patient_groups": "🔹 Enter the Number of Patient Groups",
            "sample_number": "Sample Number",
            "Grup": "Group",
            "x_axis_title": "Group Name",
            "ct_value": "Cq Value",
            "reference_ct": "Reference Cq",
            "delta_ct_control": "ΔCq (Control)",
            "delta_ct_patient": "ΔCq (Patient)",
            "warning_empty_input": "⚠️ Warning: Write data one below the other or copy-paste without empty cells from Excel.",
            "download_csv": "📥 Download CSV",
            "generate_pdf": "📥 Prepare PDF Report",
            "pdf_report": "Gene Expression Analysis Report",
            "nil_mine": "📊 Results",
            "gr_tbl": "📋 Input Data Table",
            "control_group": "🧬 Control Group",
            "ctrl_trgt_ct": "🟦 Control Group Target Gene {i} Cq Values",
            "ctrl_ref_ct": "🟦 Control Group Reference Gene {i} Cq Values",
            "hst_trgt_ct": "🩸 Patient Group Target Gene {j} Cq Values",
            "hst_ref_ct": "🩸 Patient Group Reference Gene {j} Cq Values",
            "warning_control_ct": "⚠️ Warning: Control Group {i} data should be entered line by line or copied from Excel without empty cells.",
            "warning_patient_cq": "⚠️ Warning: Enter patient group Cq values line by line or copy-paste from Excel without empty cells.",
            "target_gene": "Target Gene",
            "reference_gene": "Reference Gene",
            "target_ct": "Target Gene Cq", 
            "distribution_graph": "Distribution Graph",
            "error_missing_control_data": "⚠️ Error: Missing data for Target Gene {i} in the Control Group!",
            "control_group_avg": "Control Group Average",
            "avg": "Average",
            "control": "Control",
            "sample": "Sample",
            "patient": "Patient",
            "delta_ct_distribution": "ΔCq Distribution",
            "delta_ct_value": "ΔCq Value",
            "parametric": "Parametric",
            "non_parametric": "Nonparametric",
            "t_test": "t-test",
            "mann_whitney_u_test": "Mann-Whitney U test",
            "welch_t_test": "Welch t-test",
            "significant": "Significant",
            "insignificant": "Insignificant",
            "test_type": "Test Type",
            "test_method": "Test Method",
            "test_pvalue": "Test P-value",
            "significance": "Significance",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "Gene Expression Change (2^(-ΔΔCq))",
            "regulation_status": "Regulation Status",
            "no_change": "No Change",
            "upregulated": "Upregulated",
            "downregulated": "Downregulated",
            "report_title": "Gene Expression Analysis Report",
            "input_data_table": "Input Data Table",
            "results": "Results",
            "statistical_results": "📈 Statistical Results",
            "statistics": "Statistical Results",
            "statistical_evaluation": "Statistical Evaluation",
            "target_gene": "Target Gene",
            "patient_group": "🩸 Patient Group",
            "expression_change": "Gene Expression Change",
            "generate_pdf": "Generate PDF",
            "pdf_report": "Gene Expression Report",
            "error_no_data": "No data found, PDF could not be generated.",
            # Efficiency translations
            "efficiency_header": "🔬 Amplification Efficiency Validation",
            "efficiency_method": "Efficiency Input Method",
            "efficiency_manual": "Enter E value manually",
            "efficiency_slope": "Calculate from slope",
            "efficiency_target_label": "Target Gene {i} Efficiency (E)",
            "efficiency_ref_label": "Reference Gene {i} Efficiency (E)",
            "efficiency_target_slope_label": "Target Gene {i} Slope",
            "efficiency_ref_slope_label": "Reference Gene {i} Slope",
            "efficiency_threshold": "Acceptable efficiency difference threshold (%)",
            "efficiency_ok": "✅ Efficiency difference is acceptable ({diff:.1f}%)",
            "efficiency_warning": "⚠️ Efficiency difference exceeds threshold ({diff:.1f}%) — ΔΔCq method may not be reliable!",
            "efficiency_target_pct": "Target Gene Efficiency",
            "efficiency_ref_pct": "Reference Gene Efficiency",
            "efficiency_diff": "Difference",
            "pfaffl_result": "Pfaffl Ratio",
            "pfaffl_header": "Pfaffl Method Results",
            "classic_ddct": "Classic ΔΔCq Result (2^(-ΔΔCq))",
            "pfaffl_ratio": "Pfaffl Ratio",
            "method_comparison": "📊 Method Comparison",
            "efficiency_note": "Note: E=2.0 represents perfect efficiency (100%). Accepted range: 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "During the statistical evaluation process, data distribution was analyzed using the Shapiro-Wilk test. "
                "If normality was met, variance homogeneity between groups was checked with Levene's test. "
                "If variance was equal, an independent sample t-test was applied; otherwise, a Welch t-test was used. "
                "If normal distribution was not achieved, the non-parametric Mann-Whitney U test was applied. "
                "Significance was determined using the p < 0.05 criterion. "
                "For suggestions and support, Burhanettin Yalçinkaya - email: mailtoburhanettin@gmail.com"
            ),
            "outlier_section_title": "### 🔍 Outlier Detection Settings",
            "outlier_enable": "Enable outlier detection",
            "outlier_enable_help": "Detects statistically extreme Cq values that may reflect technical errors.",
            "outlier_method_label": "Detection method",
            "outlier_method_help": "Grubbs: best for normally distributed data, detects one outlier at a time. IQR: non-parametric, robust for skewed distributions.",
            "outlier_alpha_label": "Significance level (α)",
            "outlier_alpha_help": "α = 0.05 is standard. Lower α = more conservative (fewer outliers flagged).",
            "outlier_iqr_label": "IQR multiplier (k)",
            "outlier_iqr_help": "k=1.5 = standard Tukey fences. k=3.0 = extreme outliers only.",
            "outlier_expander": "ℹ️ About outlier detection in qPCR",
            "grubbs_info": "ℹ️ **Grubbs' test requirements:** Minimum **n ≥ 3** replicates per group. Significance threshold: **α = {alpha:.2f}**. The test assumes normality; for n < 8, normality cannot be reliably assessed — results should be interpreted with caution. Applying the test on **raw Cq values** (before normalization) is recommended to prevent noisy replicates from propagating into the ΔCq calculation.",
            "outlier_excluded_no": "No",
            "outlier_excluded_yes": "Yes",
            # Outlier stage selector
            "outlier_stage_label": "🔬 Outlier Detection Stage",
            "outlier_stage_raw": "Raw Cq — before normalization (recommended)",
            "outlier_stage_dct": "ΔCq — after normalization (previous behaviour)",
            "outlier_stage_help": (
                "**Raw Cq (recommended):** Outliers are flagged on raw Ct values before ΔCq is computed. "
                "Applied separately to target and each reference gene. Prevents noisy replicates from "
                "propagating into the normalization step.\n\n"
                "**ΔCq:** Outliers are flagged after normalization (original behaviour)."
            ),
            # Distribution plot mode selector
            "dist_plot_mode_label": "📊 Distribution Plot — Display Mode",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — recommended",
            "dist_plot_dct":  "ΔCq  — raw normalized values",
            "dist_plot_ddct": "ΔΔCq  — relative to control mean",
            "dist_plot_help": (
                "**RQ (recommended):** Converts ΔCq to 2^(-ΔCt). Higher value = higher expression. "
                "Avoids the counter-intuitive ΔCq paradox (high ΔCt = low expression).\n\n"
                "**ΔCq:** Raw normalized values on a log scale. Useful for checking data spread and normality.\n\n"
                "**ΔΔCq:** Each sample's ΔCq minus the control group mean ΔCt. "
                "Shows expression change relative to control on a log scale."
            ),
            "unequal_n_warning": (
                "⚠️ **Unequal replicate counts detected — {group}:**  \n"
                "{details}  \nAnalysis will proceed using the **shortest common length (n={min_n})**.  \n"
                "Please verify your input data — mismatched replicates may indicate a data entry error."
            ),
            # Sidebar
            "sidebar_load_example": "📂 Load Example Data",
            "sidebar_example_loaded": "✅ Example data loaded! Switch to Data Entry tab.",
            "sidebar_desktop_title": "### 💻 Desktop Application",
            "sidebar_desktop_btn": "⬇️ Download Desktop App",
            "sidebar_opensource_title": "### 🔓 Open Source",
            "sidebar_opensource_body": "GeneQuantify is open source (GPL-3.0).  \nSource code available on GitHub:",
            "sidebar_github_btn": "⭐ View Source on GitHub",
            "sidebar_scenarios_title": "📋 Load Validation Scenario",
            "sidebar_scenario_select": "Select scenario",
            "sidebar_load_scenario_btn": "▶ Load Scenario",
            "sidebar_scenario_loaded": "✅ {s} loaded! Switch to Data Entry tab.",
            "stat_decision_title": "🔬 Statistical decision",
            "stat_decision_steps": "**Step-by-step test selection:**",
            "stat_shapiro_title": "**1. Shapiro-Wilk normality test**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Non-normal",
            "stat_levene_title": "**2. Levene variance homogeneity test**",
            "stat_levene_skipped": "**2. Levene test** — *skipped* (normality not met; non-parametric test will be used)",
            "stat_equal_var": "Equal variances",
            "stat_unequal_var": "Unequal variances",
            "stat_selected_test": "**3. Selected test:**",
            "stat_reason": "**Reason:**",
            "stat_result": "**Result:**",
            "stat_reason_nonnormal": "Non-normal distribution in one or both groups",
            "stat_reason_normal_equal": "Both groups normal + equal variances",
            "stat_reason_normal_unequal": "Both groups normal + unequal variances (Levene p < 0.05)",
            "stat_multigroup_note": "⚠️ Note: When ≥ 3 groups are present, see the **Multi-Group Comparison** section below for ANOVA / Kruskal-Wallis testing with post-hoc correction.",
            "multigroup_title": "## 📊 Multi-Group Comparison Analysis",
            "multigroup_expander": "ℹ️ About multi-group statistical analysis",
            "multigroup_omnibus_test": "Omnibus Test",
            "multigroup_pvalue": "p-value",
            "multigroup_result": "Result",
            "multigroup_significant": "Significant",
            "multigroup_not_significant": "Not significant",
            "multigroup_omnibus_ns": "ℹ️ Omnibus test is **not significant** (p ≥ 0.05). Post-hoc comparisons are shown for completeness but should be interpreted with caution — no overall group effect was detected.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Download post-hoc results —",
            "multigroup_2group_note": "ℹ️ **Multi-group analysis not applicable:** Only 2 groups detected (Control + 1 patient group). Pairwise statistics are reported above.",
            "multigroup_decision_normal_equal": "✅ Normal distribution + equal variances → **One-way ANOVA + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Normal distribution + **unequal variances** → **Welch ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **Non-normal distribution** → **Kruskal-Wallis + Dunn post-hoc**",
            "multigene_title": "### 🧬 Multi-Gene Multiple Comparison Correction",
            "multigene_expander": "ℹ️ Why is this needed?",
            "multigene_sig_raw": "Significant (raw)",
            "multigene_sig_bonf": "Significant (Bonferroni)",
            "multigene_sig_fdr": "Significant (FDR B-H)",
            "multigene_warning": "⚠️ After correction, {lost} result(s) that appeared significant at raw p < 0.05 are no longer significant after FDR adjustment. Report corrected p-values as primary results in multi-gene analyses.",
            "multigene_success": "✅ All {n} significant result(s) remain significant after FDR correction — findings are robust to multiple testing.",
            "multigene_no_sig": "No significant pairwise results detected (raw p < 0.05).",
            "multigene_dl_button": "📥 Download corrected p-values (CSV)",
            "multigene_chart_title": "Multi-Gene p-value Correction: Raw vs Bonferroni vs FDR",
            "multigene_fc_chart_title": "Multi-Gene Expression Comparison",
            "multigene_1gene_note": "ℹ️ **Multi-gene correction:** Only 1 target gene analysed — multiple comparison correction across genes is not applicable.",
            "multigene_no_data": "No p-values available yet — enter data above to calculate corrections.",
            "ref_gene_section_title": "### 📚 Reference Gene Settings",
            "ref_gene_num_label": "Number of reference genes per target gene",
            "ref_gene_num_help": "MIQE guidelines recommend ≥2 validated reference genes for robust normalization.",
            "ref_gene_1_warning": "⚠️ **Methodological note:** Using a single reference gene is a meaningful constraint on normalization robustness. MIQE guidelines (Bustin et al. 2009) recommend using **≥ 2 validated reference genes** and assessing their stability with tools such as geNorm or NormFinder.",
            "ref_gene_multi_success": "✅ {n} reference genes selected. Geometric mean normalization and geNorm M-value stability will be calculated automatically.",
            "ref_gene_expander": "ℹ️ About multi-reference normalization",
            "sc_expander": "📐 Standard Curve Calculator — Calculate E from dilution series",
            "sc_gene_label": "Gene / Primer label",
            "sc_num_points": "Number of dilution points",
            "sc_dilution_factor_label": "**Dilution factor** (e.g. 10 for 10-fold dilutions)",
            "sc_dilution_factor_input": "Dilution factor",
            "sc_start_conc_label": "**Starting concentration** (arbitrary units, e.g. 1)",
            "sc_start_conc_input": "Starting concentration",
            "sc_enter_ct": "**Enter mean Cq for each dilution:**",
            "sc_calc_button": "📊 Calculate Efficiency",
            "sc_slope": "Slope",
            "sc_e_value": "E value",
            "sc_efficiency_pct": "Efficiency %",
            "sc_excellent": "✅ Excellent! E={e:.4f} ({pct:.1f}%), R²={r2:.4f} — Use this E value in the efficiency section below.",
            "sc_warning_r2": "⚠️ E is acceptable ({pct:.1f}%) but R²={r2:.4f} is below 0.99 — check your dilution series.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) is outside acceptable range (90–110%). Review your primer design or dilution series.",
            "sc_chart_title": "Standard Curve — {label} | E={e:.4f} ({pct:.1f}%), R²={r2:.4f}",
            "sc_xaxis": "log₁₀(Concentration)",
            "sc_data_points": "Data points",
            "sc_copy_hint": "💡 Copy slope **{slope:.4f}** or E value **{e:.4f}** into the efficiency inputs below.",
            "sc_description": """\
    Enter your serial dilution Ct values below. The calculator will fit a linear regression,
    compute the slope, R², and amplification efficiency automatically.

    **How to use:**  
    1. Run qPCR on serial dilutions (e.g. undiluted, 1:10, 1:100, 1:1000, 1:10000)  
    2. Enter the mean Ct for each dilution below  
    3. Read off slope, E, and R²  
    """,
            "ref_multi_description": """\
    **Geometric mean normalization** (Vandesompele et al. 2002)  
    The normalization factor (NF) is the arithmetic mean of Ct values across all reference genes per sample,
    which corresponds to the geometric mean of their expression levels.  
    `NF_sample = mean(Ct_ref1, Ct_ref2, ..., Ct_refN)` for each sample  
    `ΔCq = Ct_target − NF`

    **geNorm M-value** (stability score)  
    For each reference gene, M = average standard deviation of log-ratios against all other reference genes.  
    **Lower M = more stable.** MIQE-recommended threshold: M < 0.5 (strict) or M < 1.0 (acceptable).

    **CV (Coefficient of Variation)**  
    `CV = (SD / mean) × 100%` of raw Ct values across all samples.  
    Lower CV indicates less variation and better stability as a reference.

    **Reference:** Vandesompele J et al. *Genome Biology* 2002; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **Why outlier detection matters in qPCR**

    Technical variability is inherent to qPCR: pipetting errors, bubble formation, 
    inhibitor carry-over, or RNA quality variation can produce Ct values that are 
    statistically inconsistent with the rest of a replicate group. 
    Including such values inflates variance, biases means, and can lead to false 
    conclusions — particularly in clinical datasets with small sample sizes.

    **When this limitation becomes critical:**
    - Small groups (n < 5): a single erroneous Ct shifts the mean substantially
    - High biological variability (e.g. tumour heterogeneity, clinical cohorts)
    - Technical triplicates where one replicate diverges > 0.5 Ct from the others
    - Low-abundance targets with Ct > 35, where noise dominates

    **Grubbs test** *(Grubbs 1969)*  
    Assumes normality. Tests whether the most extreme value is a statistically 
    significant outlier (p < α). Iterates until no further outliers are found.  
    Best for: replicate Ct values from a single experimental group.

    **IQR method** *(Tukey 1977)*  
    Non-parametric. Flags values outside Q1 − k×IQR or Q3 + k×IQR.  
    Best for: larger groups or non-normal distributions.

    **Important:** Outlier exclusion requires **biological or technical justification**. 
    This tool flags candidates — the final decision always rests with the researcher.  
    All exclusions are logged and reported in the PDF output.

    **References:** Grubbs FE. *Technometrics* 1969; Tukey JW. *Exploratory Data Analysis* 1977;  
    Bustin SA et al. *Clin Chem* 2009 (MIQE guidelines).
    """,

            # ── PDF report strings ────────────────────────────────────────────────
            "pdf_cover_subtitle": "qPCR Gene Expression Analysis Report",
            "pdf_generated": "Generated: {now}",
            "pdf_s1_title": "1. Methods and Analysis Settings",
            "pdf_s1_calc": "1.1 Calculation Methods",
            "pdf_s1_calc_body": "Two complementary methods were applied for fold-change calculation:",
            "pdf_s1_classic": "Classic ΔΔCq (Livak & Schmittgen, 2001): ΔCq = Cq(target) - Cq(reference);  ΔΔCq = ΔCq(sample) - ΔCt(control);  Fold Change = 2^(-ΔΔCt). Assumes equal amplification efficiencies (E ≈ 2.0) for both genes.",
            "pdf_s1_pfaffl": "Pfaffl Method (Pfaffl, 2001): Ratio = (E_target ^ ΔCq_target) / (E_ref ^ ΔCt_ref). Corrects for primer-specific efficiencies; recommended when efficiency difference > 10%.",
            "pdf_s1_norm": "1.2 Normalization",
            "pdf_s1_norm_multi": "Multiple reference genes (n={n}) were used. Normalization factor (NF) was calculated as the arithmetic mean of reference gene Cq values per sample (geNorm approach, Vandesompele et al. 2002). geNorm M-values and CV% were computed.",
            "pdf_s1_norm_single": "A single reference gene was used. MIQE guidelines recommend ≥2 reference genes for robust normalization.",
            "pdf_s1_eff": "1.3 Amplification Efficiency",
            "pdf_s1_eff_range": "Acceptable efficiency range: E = 1.8-2.2 (90-110%). Efficiency difference threshold applied: {thr}%.",
            "pdf_s1_outlier": "1.4 Outlier Detection",
            "pdf_s1_grubbs": "Grubbs test (Grubbs 1969) applied at alpha = {alpha}. {n} sample(s) flagged and confirmed for exclusion by user.",
            "pdf_s1_iqr": "IQR method (Tukey 1977) applied with multiplier k = {k}. {n} sample(s) flagged and confirmed for exclusion by user.",
            "pdf_s1_outlier_warn": "WARNING: Outlier exclusion requires biological or technical justification. Excluded samples are flagged in the data table.",
            "pdf_s1_outlier_off": "Outlier detection was disabled for this analysis.",
            "pdf_s2_title": "2. Input Data",
            "pdf_s2_body": "Raw Cq values entered by the user, after outlier processing. Rows marked Yes in the Outlier Excluded column were removed from calculations.",
            "pdf_s3_title": "3. Gene Expression Results",
            "pdf_s3_body": "Fold change values calculated by both Classic ΔΔCq and Pfaffl methods. Fold change > 1 indicates higher expression in patient group relative to control.",
            "pdf_s4_title": "4. Statistical Analysis",
            "pdf_s4_body": "Statistical significance of gene expression differences between control and patient groups. All tests are performed on RQ values (2^-ΔCq) rather than raw ΔCt, because ΔCt is on a logarithmic scale and direct parametric testing on ΔCt underestimates biological variability. Test selection is automatic based on normality (Shapiro-Wilk) and variance homogeneity (Levene). Significance threshold: p < 0.05.",
            "pdf_s4_interp": "Interpretation of Statistical Tests",
            "pdf_s4_interp_body": "Student's t-test: Used when both groups are normal with equal variances. Welch's t-test: Used when groups are normal but variances differ. Mann-Whitney U: Non-parametric test when normality is violated. p < 0.05 = statistically significant differential expression.",
            "pdf_s5_title": "5. Relative Quantity (RQ) Distribution Plots",
            "pdf_s5_body": "RQ values (2^-ΔCq) per target gene across groups. Each dot = one biological replicate; horizontal bars = group means. Statistical tests are performed on RQ values, not raw ΔCt, to avoid underestimation of biological variability on the logarithmic scale.",
            "pdf_s6_title": "6. How to Interpret Your Results",
            "pdf_s6_fc": "6.1 Fold Change Interpretation",
            "pdf_s6_choose": "6.2 Choosing Between ΔΔCq and Pfaffl",
            "pdf_s6_choose_body": "Use Classic ΔΔCq when: both efficiencies are 90-110% and difference < 10%. Use Pfaffl when: efficiency difference > 10%. Always report both values.",
            "pdf_s6_stat": "6.3 Statistical Test Selection Rationale",
            "pdf_s6_stat_body": "Normality assessed using Shapiro-Wilk (recommended for n < 50). Variance homogeneity assessed using Levene's test. Parametric data with equal variances: Student's t-test. Unequal variances: Welch's t-test. Non-normal: Mann-Whitney U.",
            "pdf_s7_title": "7. References",
            "pdf_fc_interp_header": ["Fold Change", "ΔΔCq", "Interpretation", "Biological Significance"],
            "pdf_fc_interp_rows": [
                [">2.0", "<-1.0", "Strong upregulation", "Consider biologically relevant"],
                ["1.5-2.0", "-1.0 to -0.58", "Moderate upregulation", "May be relevant; verify"],
                ["1.0-1.5", "-0.58 to 0", "Weak upregulation", "Likely not significant alone"],
                ["1.0", "0", "No change", "No differential expression"],
                ["0.67-1.0", "0 to 0.58", "Weak downregulation", "Likely not significant alone"],
                ["0.5-0.67", "0.58 to 1.0", "Moderate downregulation", "May be relevant; verify"],
                ["<0.5", ">1.0", "Strong downregulation", "Consider biologically relevant"],
            ],
            "pdf_stat_note": "Note: Statistical significance (p < 0.05) and biological significance (fold change) should be considered together.",
            "pdf_summary_param": "Parameter",
            "pdf_summary_val": "Value",
            "pdf_summary_genes": "Target genes analyzed",
            "pdf_summary_groups": "Patient groups",
            "pdf_summary_samples": "Total samples (rows)",
            "pdf_summary_excluded": "Outlier-excluded samples",
            "pdf_summary_tests": "comparisons",
            "pdf_summary_norm": "Normalization method",
            "pdf_summary_norm_multi": "geNorm NF",
            "pdf_summary_norm_single": "Single reference gene",
            "pdf_summary_methods": "Calculation methods",
            "pdf_summary_methods_val": "Classic ΔΔCq + Pfaffl",
            "pdf_disclaimer": "This report was generated automatically by GeneQuantify. All calculations follow MIQE guidelines (Bustin et al., Clin Chem 2009).",
            "pdf_footer": "GeneQuantify — For research and educational use only. Not validated for clinical diagnostic purposes.",
            "pdf_fig1": "Figure 1. Fold change comparison: Classic ΔΔCq vs Pfaffl. Dashed line at y=1 = no change relative to control.",
            "pdf_fig2": "Figure 2. p-values for all comparisons. Red bars = significant (p < 0.05). Dashed line = significance threshold.",
            "pdf_fig3": "Figure. RQ (2^-ΔCq) distribution for {gene}. Points = individual replicates; horizontal bars = group means. Statistical tests performed on RQ values.",
            "pdf_nochange": "No Change",
            "pdf_stat_cols": ["Target Gene", "Comparison", "Test Type", "Test Method", "p-value", "Significance"],
            "pdf_res_cols": ["Target Gene", "Group", "ΔCq Control", "ΔCq Sample", "ΔΔCq", "2^(-ΔΔCq)", "Pfaffl Ratio", "Regulation", "E target", "E ref"],
            "pdf_eff_cols": ["Gene", "E (target)", "Eff% (target)", "E (ref)", "Eff% (ref)", "Diff%", "Status"],
            "pdf_eff_ok": "OK",
            "pdf_eff_warn": "WARNING: use Pfaffl",
            "pdf_outlier_col": "Outlier Excluded",
            "pdf_contact": "Contact: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} records ready — you can generate the PDF.",
            # RDML / RDES import
            "rdml_expander":        "📂 Import RDML / RDES File",
            "rdml_description":     "Upload an **RDML** (`.rdml`) or **RDES** (`.tsv`/`.csv`/`.txt`) file to auto-fill Cq values.",
            "rdml_uploader":        "Choose file",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX, Roche LightCycler, etc.  RDES: tab-separated spreadsheet format.",
            "rdml_success":         "✅ {fmt} file loaded — {n} reactions found.",
            "rdml_error":           "❌ {fmt} parse error: {err}",
            "rdml_preview":         "Preview parsed data",
            "rdml_step1":           "**Step 1 — Label your Control group**",
            "rdml_ctrl_label":      "Control sample name(s) (comma-separated substrings)",
            "rdml_ctrl_help":       "Any sample whose name contains this text will be treated as Control.",
            "rdml_step2":           "**Step 2 — Label your Patient groups**",
            "rdml_n_pat":           "Number of patient groups",
            "rdml_pat_label":       "Patient group {i} sample name(s)",
            "rdml_pat_help":        "Comma-separated substrings. All matching samples will be pooled into this group.",
            "rdml_apply":           "✅ Apply {fmt} import to Data Entry",
            "rdml_apply_success":   "✅ {n} Cq values loaded into Data Entry tab! Switch to review and adjust.",
            "rdml_apply_warning":   "⚠️ No values were mapped. Check that your labels match the sample names in the preview above.",
        },

        "de": {
            "title": "🧬 GeneQuantify: Expressions- und CNV-Analyse",
            "tab_data": "Dateneingabe",
            "tab_results": "Ergebnisse",
            "tab_report": "Bericht",
            "subtitle": "Entwickelt von B. Yalçınkaya",
            "patient_data_header": "📊 Geben Sie Patientendaten und Kontrollgruppen ein",
            "num_target_genes": "🔹 Geben Sie die Anzahl der Zielgene ein",
            "num_patient_groups": "🔹 Geben Sie die Anzahl der Patientengruppen ein",
            "sample_number": "Beispielnummer",
            "Grup": "Gruppe",
            "x_axis_title": "Gruppenname",
            "ct_value": "Cq-Wert",
            "reference_ct": "Referenz Cq",
            "delta_ct_control": "ΔCq (Kontrolle)",
            "delta_ct_patient": "ΔCq (Patientendaten)",
            "warning_empty_input": "⚠️ Warnung: Geben Sie die Daten untereinander ein oder kopieren Sie sie ohne leere Zellen aus Excel.",
            "download_csv": "📥 CSV herunterladen",
            "generate_pdf": "📥 PDF-Bericht erstellen",
            "pdf_report": "Genexpression-Analysebericht",
            "nil_mine": "📊 Ergebnisse",
            "gr_tbl": "📋 Eingabedaten Tabelle",
            "control_group": "🧬 Kontrollgruppe",
            "ctrl_trgt_ct": "🟦 Kontrollgruppe Zielgen {i} Cq-Werte",
            "ctrl_ref_ct": "🟦 Kontrollgruppe Referenz {i} Ct-Werte",
            "hst_trgt_ct": "🩸 Patientengruppe Zielgen {j} Cq-Werte",
            "hst_ref_ct": "🩸 Patientengruppe Referenz {j} Ct-Werte",
            "warning_control_ct": "⚠️ Achtung: Kontrollgruppe {i} Daten sollten untereinander eingegeben oder aus Excel ohne leere Zellen eingefügt werden.",
            "warning_patient_cq": "⚠️ Achtung: Geben Sie die Cq-Werte der Patientengruppe untereinander ein oder kopieren Sie sie aus Excel ohne leere Zellen.",
            "target_gene": "Zielgen",
            "reference_gene": "Referenzgen",
            "target_ct": "Zielgen Cq",
            "distribution_graph": "Verteilungsdiagramm",
            "error_missing_control_data": "⚠️ Fehler: Fehlende Daten für Zielgen {i} in der Kontrollgruppe!",
            "control_group_avg": "Durchschnitt der Kontrollgruppe",
            "avg": "Durchschnitt",
            "control": "Kontrolle",
            "sample": "Probe",
            "patient": "Patient",
            "delta_ct_distribution": "ΔCq-Verteilung",
            "delta_ct_value": "ΔCq-Wert",
            "parametric": "Parametrisch",
            "non_parametric": "Nicht parametrisch",
            "t_test": "t-Test",
            "mann_whitney_u_test": "Mann-Whitney U-Test",
            "welch_t_test": "Welch t-Test",
            "significant": "Signifikant",
            "insignificant": "Nicht signifikant",
            "test_type": "Testtyp",
            "test_method": "Verwendeter Test",
            "test_pvalue": "P-Wert",
            "significance": "Signifikanz",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "Genexpression Veränderung (2^(-ΔΔCq))",
            "regulation_status": "Regulierungsstatus",
            "no_change": "Keine Veränderung",
            "upregulated": "Hochreguliert",
            "downregulated": "Herunterreguliert",
            "report_title": "Genexpressionsanalysebericht",
            "input_data_table": "Eingabedatentabelle",
            "results": "Ergebnisse",
            "statistical_results": "📈 Statistische Ergebnisse",
            "statistics": "Statistische Ergebnisse",
            "statistical_evaluation": "Statistische Auswertung",
            "target_gene": "Zielgen",
            "patient_group": "🩸 Patientengruppe",
            "expression_change": "Genexpressionsänderung",
            "generate_pdf": "PDF Erstellen",
            "pdf_report": "Genexpressionsbericht",
            "error_no_data": "Keine Daten gefunden, PDF konnte nicht erstellt werden.",
            # Efficiency translations
            "efficiency_header": "🔬 Amplifikationseffizienz-Validierung",
            "efficiency_method": "Effizienzeingabemethode",
            "efficiency_manual": "E-Wert manuell eingeben",
            "efficiency_slope": "Aus Steigung berechnen",
            "efficiency_target_label": "Zielgen {i} Effizienz (E)",
            "efficiency_ref_label": "Referenzgen {i} Effizienz (E)",
            "efficiency_target_slope_label": "Zielgen {i} Steigung",
            "efficiency_ref_slope_label": "Referenzgen {i} Steigung",
            "efficiency_threshold": "Akzeptable Effizienzdifferenz-Schwelle (%)",
            "efficiency_ok": "✅ Effizienzdifferenz ist akzeptabel ({diff:.1f}%)",
            "efficiency_warning": "⚠️ Effizienzdifferenz überschreitet Schwelle ({diff:.1f}%) — ΔΔCq-Methode möglicherweise nicht zuverlässig!",
            "efficiency_target_pct": "Zielgen-Effizienz",
            "efficiency_ref_pct": "Referenzgen-Effizienz",
            "efficiency_diff": "Differenz",
            "pfaffl_result": "Pfaffl-Verhältnis",
            "pfaffl_header": "Pfaffl-Methode Ergebnisse",
            "classic_ddct": "Klassisches ΔΔCq-Ergebnis (2^(-ΔΔCq))",
            "pfaffl_ratio": "Pfaffl-Verhältnis",
            "method_comparison": "📊 Methodenvergleich",
            "efficiency_note": "Hinweis: E=2.0 steht für perfekte Effizienz (100%). Akzeptierter Bereich: 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "Während des statistischen Bewertungsprozesses wurde die Datenverteilung mit dem Shapiro-Wilk-Test analysiert. "
                "Wenn die Normalität erfüllt war, wurde die Varianzhomogenität zwischen den Gruppen mit dem Levene-Test überprüft. "
                "War die Varianz gleich, wurde ein unabhängiger Stichprobent-Test angewendet; andernfalls wurde ein Welch-T-Test verwendet. "
                "Wenn keine normale Verteilung vorlag, wurde der nicht-parametrische Mann-Whitney-U-Test angewendet. "
                "Die Signifikanz wurde anhand des Kriteriums p < 0,05 bestimmt. "
                "Für Vorschläge und Unterstützung, Burhanettin Yalçinkaya - E-Mail: mailtoburhanettin@gmail.com"
            ),
            "outlier_section_title": "### 🔍 Ausreißer-Erkennungseinstellungen",
            "outlier_enable": "Ausreißererkennung aktivieren",
            "outlier_enable_help": "Erkennt statistisch extreme Cq-Werte, die auf technische Fehler hinweisen können.",
            "outlier_method_label": "Erkennungsmethode",
            "outlier_method_help": "Grubbs: für normalverteilte Daten. IQR: nicht-parametrisch, robust bei schiefen Verteilungen.",
            "outlier_alpha_label": "Signifikanzniveau (α)",
            "outlier_alpha_help": "α = 0,05 ist Standard. Niedrigeres α = konservativer.",
            "outlier_iqr_label": "IQR-Multiplikator (k)",
            "outlier_iqr_help": "k=1,5 = Standard Tukey-Grenzen. k=3,0 = nur extreme Ausreißer.",
            "outlier_expander": "ℹ️ Über Ausreißererkennung in qPCR",
            "grubbs_info": "ℹ️ **Grubbs-Test-Anforderungen:** Mindestens **n ≥ 3** Replikate pro Gruppe. Signifikanzschwelle: **α = {alpha:.2f}**. Der Test setzt Normalverteilung voraus; bei n < 8 kann Normalität nicht zuverlässig geprüft werden. Die Anwendung auf **rohe Cq-Werte** (vor Normalisierung) wird empfohlen.",
            "outlier_excluded_no": "Nein",
            "outlier_excluded_yes": "Ja",
            # Outlier stage selector
            "outlier_stage_label": "🔬 Ausreißererkennung — Analysestufe",
            "outlier_stage_raw": "Roh-Cq — vor Normalisierung (empfohlen)",
            "outlier_stage_dct": "ΔCq — nach Normalisierung (bisheriges Verhalten)",
            "outlier_stage_help": (
                "**Roh-Ct (empfohlen):** Ausreißer werden vor der ΔCq-Berechnung erkannt. "
                "Für Zielgen und jedes Referenzgen separat angewendet.\n\n"
                "**ΔCq:** Ausreißer werden nach der Normalisierung erkannt (bisheriges Verhalten)."
            ),
            # Distribution plot mode selector
            "dist_plot_mode_label": "📊 Verteilungsdiagramm — Anzeigemodus",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — empfohlen",
            "dist_plot_dct":  "ΔCq  — rohe normalisierte Werte",
            "dist_plot_ddct": "ΔΔCq  — relativ zum Kontrollmittelwert",
            "dist_plot_help": (
                "**RQ (empfohlen):** Konvertiert ΔCq zu 2^(-ΔCt). Höherer Wert = höhere Expression.\n\n"
                "**ΔCq:** Rohe logarithmische Werte. Nützlich zur Überprüfung der Datenverteilung.\n\n"
                "**ΔΔCq:** ΔCq jeder Probe minus dem Kontrollgruppenmittelwert."
            ),
            "unequal_n_warning": (
                "⚠️ **Ungleiche Replikatanzahl erkannt — {group}:**  \n"
                "{details}  \nAnalyse wird mit der **kürzesten gemeinsamen Länge (n={min_n})** fortgesetzt.  \n"
                "Bitte überprüfen Sie Ihre Eingabedaten."
            ),
            # Sidebar
            "sidebar_load_example": "📂 Beispieldaten laden",
            "sidebar_example_loaded": "✅ Beispieldaten geladen! Wechseln Sie zur Dateneingabe-Registerkarte.",
            "sidebar_desktop_title": "### 💻 Desktop-Anwendung",
            "sidebar_desktop_btn": "⬇️ Desktop-App herunterladen",
            "sidebar_opensource_title": "### 🔓 Open Source",
            "sidebar_opensource_body": "GeneQuantify ist Open Source (GPL-3.0).  \nQuellcode auf GitHub verfügbar:",
            "sidebar_github_btn": "⭐ Quellcode auf GitHub ansehen",
            "sidebar_scenarios_title": "📋 Validierungsszenario laden",
            "sidebar_scenario_select": "Szenario auswählen",
            "sidebar_load_scenario_btn": "▶ Szenario laden",
            "sidebar_scenario_loaded": "✅ {s} geladen! Zur Dateneingabe wechseln.",
            "outlier_excluded_no": "Nein",
            "outlier_excluded_yes": "Ja",
            "stat_decision_title": "🔬 Statistische Entscheidung",
            "stat_decision_steps": "**Schrittweise Testauswahl:**",
            "stat_shapiro_title": "**1. Shapiro-Wilk-Normalitätstest**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Nicht normal",
            "stat_levene_title": "**2. Levene-Varianzhomogeintätstest**",
            "stat_levene_skipped": "**2. Levene-Test** — *übersprungen* (Normalität nicht erfüllt; nicht-parametrischer Test wird verwendet)",
            "stat_equal_var": "Gleiche Varianzen",
            "stat_unequal_var": "Ungleiche Varianzen",
            "stat_selected_test": "**3. Ausgewählter Test:**",
            "stat_reason": "**Grund:**",
            "stat_result": "**Ergebnis:**",
            "stat_reason_nonnormal": "Nicht-normale Verteilung in einer oder beiden Gruppen",
            "stat_reason_normal_equal": "Beide Gruppen normal + gleiche Varianzen",
            "stat_reason_normal_unequal": "Beide Gruppen normal + ungleiche Varianzen (Levene p < 0,05)",
            "stat_multigroup_note": "⚠️ Hinweis: Bei ≥ 3 Gruppen siehe Abschnitt **Mehrgruppen-Vergleich** unten für ANOVA / Kruskal-Wallis.",
            "multigroup_title": "## 📊 Mehrgruppen-Vergleichsanalyse",
            "multigroup_expander": "ℹ️ Über die Mehrgruppen-Statistikanalyse",
            "multigroup_omnibus_test": "Omnibus-Test",
            "multigroup_pvalue": "p-Wert",
            "multigroup_result": "Ergebnis",
            "multigroup_significant": "Signifikant",
            "multigroup_not_significant": "Nicht signifikant",
            "multigroup_omnibus_ns": "ℹ️ Omnibus-Test ist **nicht signifikant** (p ≥ 0,05). Post-hoc-Vergleiche werden zur Information angezeigt.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Post-hoc-Ergebnisse herunterladen —",
            "multigroup_2group_note": "ℹ️ **Mehrgruppen-Analyse nicht anwendbar:** Nur 2 Gruppen erkannt (Kontrolle + 1 Patientengruppe).",
            "multigroup_decision_normal_equal": "✅ Normalverteilung + gleiche Varianzen → **Einfaktorielle ANOVA + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Normalverteilung + **ungleiche Varianzen** → **Welch-ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **Keine Normalverteilung** → **Kruskal-Wallis + Dunn post-hoc**",
            "multigene_title": "### 🧬 Mehrgen-Mehrfachvergleichskorrektur",
            "multigene_expander": "ℹ️ Warum ist das notwendig?",
            "multigene_sig_raw": "Signifikant (roh)",
            "multigene_sig_bonf": "Signifikant (Bonferroni)",
            "multigene_sig_fdr": "Signifikant (FDR B-H)",
            "multigene_warning": "⚠️ Nach Korrektur sind {lost} Ergebnis(se) nach FDR-Anpassung nicht mehr signifikant. Korrigierte p-Werte als Hauptergebnisse berichten.",
            "multigene_success": "✅ Alle {n} signifikanten Ergebnis(se) bleiben nach FDR-Korrektur signifikant.",
            "multigene_no_sig": "Keine signifikanten paarweisen Ergebnisse erkannt (roh p < 0,05).",
            "multigene_dl_button": "📥 Korrigierte p-Werte herunterladen (CSV)",
            "multigene_chart_title": "Mehrgen p-Wert-Korrektur: Roh / Bonferroni / FDR",
            "multigene_fc_chart_title": "Mehrgen-Expressionsvergleich",
            "multigene_1gene_note": "ℹ️ **Mehrgen-Korrektur:** Nur 1 Zielgen analysiert — Mehrfachvergleichskorrektur nicht anwendbar.",
            "multigene_no_data": "Noch keine p-Werte — oben Daten eingeben.",
            "ref_gene_section_title": "### 📚 Referenzgen-Einstellungen",
            "ref_gene_num_label": "Anzahl der Referenzgene pro Zielgen",
            "ref_gene_num_help": "MIQE-Richtlinien empfehlen ≥2 validierte Referenzgene für eine robuste Normalisierung.",
            "ref_gene_1_warning": "⚠️ **Methodischer Hinweis:** Die Verwendung eines einzigen Referenzgens schränkt die Normalisierungsrobustheit ein. MIQE-Richtlinien (Bustin et al. 2009) empfehlen **≥2 validierte Referenzgene** und Stabilitätsbewertung (geNorm/NormFinder).",
            "ref_gene_multi_success": "✅ {n} Referenzgene ausgewählt. Geometrische Mittelnormalisierung und geNorm M-Wert-Stabilität werden automatisch berechnet.",
            "ref_gene_expander": "ℹ️ Über die Mehrfach-Referenznormalisierung",
            "sc_expander": "📐 Standardkurven-Rechner — E aus Verdünnungsreihe berechnen",
            "sc_gene_label": "Gen / Primer-Bezeichnung",
            "sc_num_points": "Anzahl der Verdünnungspunkte",
            "sc_dilution_factor_label": "**Verdünnungsfaktor** (z.B. 10 für 10-fache Verdünnung)",
            "sc_dilution_factor_input": "Verdünnungsfaktor",
            "sc_start_conc_label": "**Ausgangskonzentration** (beliebige Einheiten, z.B. 1)",
            "sc_start_conc_input": "Ausgangskonzentration",
            "sc_enter_ct": "**Mittleren Cq-Wert für jede Verdünnung eingeben:**",
            "sc_calc_button": "📊 Effizienz berechnen",
            "sc_slope": "Steigung",
            "sc_e_value": "E-Wert",
            "sc_efficiency_pct": "Effizienz %",
            "sc_excellent": "✅ Ausgezeichnet! E={e:.4f} ({pct:.1f}%), R²={r2:.4f} — Diesen E-Wert im Effizienzabschnitt unten verwenden.",
            "sc_warning_r2": "⚠️ E ist akzeptabel ({pct:.1f}%), aber R²={r2:.4f} < 0,99 — Verdünnungsreihe überprüfen.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) liegt außerhalb des akzeptablen Bereichs (90–110%). Primerdesign oder Verdünnungsreihe überprüfen.",
            "sc_chart_title": "Standardkurve — {label} | E={e:.4f} ({pct:.1f}%), R²={r2:.4f}",
            "sc_xaxis": "log₁₀(Konzentration)",
            "sc_data_points": "Datenpunkte",
            "sc_copy_hint": "💡 Steigung **{slope:.4f}** oder E-Wert **{e:.4f}** in die Effizienz-Eingaben unten kopieren.",
            "sc_description": """\
    Geben Sie die Ct-Werte Ihrer seriellen Verdünnung unten ein. Der Rechner passt eine lineare Regression an und berechnet Steigung, R² und Amplifikationseffizienz automatisch.

    **Verwendung:**  
    1. Führen Sie qPCR auf seriellen Verdünnungen durch (z.B. unverdünnt, 1:10, 1:100, 1:1000, 1:10000)  
    2. Geben Sie den mittleren Ct-Wert für jede Verdünnung ein  
    3. Lesen Sie Steigung, E und R² ab  
    """,
            "ref_multi_description": """\
    **Geometrische Mittelnormalisierung** (Vandesompele et al. 2002)  
    Der Normalisierungsfaktor (NF) ist das arithmetische Mittel der Ct-Werte über alle Referenzgene pro Probe,  
    was dem geometrischen Mittel ihrer Expressionsniveaus entspricht.  
    `NF_Probe = Mittel(Ct_ref1, Ct_ref2, ..., Ct_refN)` für jede Probe  
    `ΔCq = Ct_Ziel − NF`

    **geNorm M-Wert** (Stabilitätsscore)  
    Für jedes Referenzgen ist M die durchschnittliche Standardabweichung der Log-Verhältnisse gegenüber allen anderen Referenzgenen.  
    **Niedrigerer M = stabiler.** MIQE-empfohlener Schwellenwert: M < 0,5 (streng) oder M < 1,0 (akzeptabel).

    **CV (Variationskoeffizient)**  
    `CV = (SD / Mittel) × 100%` der rohen Ct-Werte über alle Proben.  
    Niedrigerer CV weist auf weniger Variation und bessere Stabilität als Referenz hin.

    **Referenz:** Vandesompele J et al. *Genome Biology* 2002; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **Warum Ausreißererkennung in qPCR wichtig ist**

    Technische Variabilität ist qPCR inhärent: Pipettierfehler, Blasenbildung, Inhibitor-Verschleppung oder RNA-Qualitätsschwankungen können Ct-Werte erzeugen, die statistisch inkonsistent mit dem Rest einer Replikatgruppe sind.  
    Das Einschließen solcher Werte erhöht die Varianz, verzerrt Mittelwerte und kann zu falschen Schlussfolgerungen führen — besonders in klinischen Datensätzen mit kleinen Stichprobengrößen.

    **Wann diese Einschränkung kritisch wird:**
    - Kleine Gruppen (n < 5): ein einziger fehlerhafter Ct verschiebt den Mittelwert erheblich
    - Hohe biologische Variabilität (z.B. Tumorheterogenität, klinische Kohorten)
    - Technische Triplikate, bei denen ein Replikat > 0,5 Ct von den anderen abweicht
    - Targets mit geringer Abundanz mit Ct > 35, wo Rauschen dominiert

    **Grubbs-Test** *(Grubbs 1969)*  
    Setzt Normalverteilung voraus. Testet, ob der extremste Wert ein statistisch signifikanter Ausreißer ist (p < α). Iteriert, bis keine weiteren Ausreißer gefunden werden.  
    Am besten für: Replikat-Ct-Werte aus einer einzelnen experimentellen Gruppe.

    **IQR-Methode** *(Tukey 1977)*  
    Nicht-parametrisch. Markiert Werte außerhalb Q1 − k×IQR oder Q3 + k×IQR.  
    Am besten für: größere Gruppen oder nicht-normale Verteilungen.

    **Wichtig:** Ausreißerausschluss erfordert **biologische oder technische Begründung**.  
    Dieses Tool markiert Kandidaten — die endgültige Entscheidung liegt immer beim Forscher.  
    Alle Ausschlüsse werden protokolliert und im PDF-Bericht gemeldet.

    **Referenzen:** Grubbs FE. *Technometrics* 1969; Tukey JW. *Exploratory Data Analysis* 1977;  
    Bustin SA et al. *Clin Chem* 2009 (MIQE-Richtlinien).
    """,

            # ── PDF-Berichtsstrings ───────────────────────────────────────────────
            "pdf_cover_subtitle": "qPCR-Genexpressionsanalysebericht",
            "pdf_generated": "Erstellt: {now}",
            "pdf_s1_title": "1. Methoden und Analyseeinstellungen",
            "pdf_s1_calc": "1.1 Berechnungsmethoden",
            "pdf_s1_calc_body": "Zwei komplementäre Methoden zur Berechnung des Fold-Change:",
            "pdf_s1_classic": "Klassische ΔΔCq-Methode (Livak & Schmittgen, 2001): Fold-Change = 2^(-ΔΔCq). Gleiche Effizienz vorausgesetzt.",
            "pdf_s1_pfaffl": "Pfaffl-Methode (Pfaffl, 2001): Verhältnis = (E_Ziel ^ ΔCq_Ziel) / (E_Ref ^ ΔCt_Ref). Empfohlen bei Effizienzunterschied > 10%.",
            "pdf_s1_norm": "1.2 Normalisierung",
            "pdf_s1_norm_multi": "Mehrere Referenzgene (n={n}) verwendet (geNorm, Vandesompele et al. 2002).",
            "pdf_s1_norm_single": "Ein Referenzgen verwendet. MIQE empfiehlt ≥2 Referenzgene.",
            "pdf_s1_eff": "1.3 Amplifikationseffizienz",
            "pdf_s1_eff_range": "Akzeptabler Effizienzbereich: E = 1,8-2,2 (90-110%). Schwellenwert: {thr}%.",
            "pdf_s1_outlier": "1.4 Ausreißererkennung",
            "pdf_s1_grubbs": "Grubbs-Test (1969), Alpha = {alpha}. {n} Probe(n) ausgeschlossen.",
            "pdf_s1_iqr": "IQR-Methode (Tukey 1977), k = {k}. {n} Probe(n) ausgeschlossen.",
            "pdf_s1_outlier_warn": "WARNUNG: Ausreißerausschluss erfordert biologische oder technische Begründung.",
            "pdf_s1_outlier_off": "Ausreißererkennung deaktiviert.",
            "pdf_s2_title": "2. Eingabedaten",
            "pdf_s2_body": "Roh-Cq-Werte nach der Ausreißerverarbeitung.",
            "pdf_s3_title": "3. Genexpressionsergebnisse",
            "pdf_s3_body": "Fold-Change berechnet mit Klassischer ΔΔCq- und Pfaffl-Methode.",
            "pdf_s4_title": "4. Statistische Analyse",
            "pdf_s4_body": "Statistische Signifikanz. Testauswahl automatisch (Shapiro-Wilk, Levene). p < 0,05.",
            "pdf_s4_interp": "Interpretation der Tests",
            "pdf_s4_interp_body": "Student-t: Normalverteilung, gleiche Varianzen. Welch-t: ungleiche Varianzen. Mann-Whitney U: nicht-normal.",
            "pdf_s5_title": "5. Delta-Cq-Verteilungsdiagramme",
            "pdf_s5_body": "ΔCq-Verteilung je Zielgen. Punkte = Replikate; Balken = Mittelwerte.",
            "pdf_s6_title": "6. Interpretation",
            "pdf_s6_fc": "6.1 Fold-Change-Interpretation",
            "pdf_s6_choose": "6.2 ΔΔCq vs. Pfaffl",
            "pdf_s6_choose_body": "ΔΔCq wenn Effizienzen 90-110% und Unterschied < 10%. Pfaffl wenn > 10%.",
            "pdf_s6_stat": "6.3 Testauswahl-Begründung",
            "pdf_s6_stat_body": "Normalität: Shapiro-Wilk. Varianzhomogenität: Levene. Student/Welch/Mann-Whitney je nach Ergebnis.",
            "pdf_s7_title": "7. Referenzen",
            "pdf_fc_interp_header": ["Fold-Change", "ΔΔCq", "Interpretation", "Biologische Bedeutung"],
            "pdf_fc_interp_rows": [
                [">2,0", "<-1,0", "Starke Hochregulation", "Biologisch relevant"],
                ["1,5-2,0", "-1,0 bis -0,58", "Mäßige Hochregulation", "Möglicherweise relevant"],
                ["1,0-1,5", "-0,58 bis 0", "Schwache Hochregulation", "Wahrscheinlich nicht signifikant"],
                ["1,0", "0", "Keine Änderung", "Keine differentielle Expression"],
                ["0,67-1,0", "0 bis 0,58", "Schwache Herunterregulation", "Wahrscheinlich nicht signifikant"],
                ["0,5-0,67", "0,58 bis 1,0", "Mäßige Herunterregulation", "Möglicherweise relevant"],
                ["<0,5", ">1,0", "Starke Herunterregulation", "Biologisch relevant"],
            ],
            "pdf_stat_note": "Hinweis: Statistische und biologische Signifikanz gemeinsam bewerten.",
            "pdf_summary_param": "Parameter", "pdf_summary_val": "Wert",
            "pdf_summary_genes": "Analysierte Zielgene", "pdf_summary_groups": "Patientengruppen",
            "pdf_summary_samples": "Proben gesamt", "pdf_summary_excluded": "Ausgeschlossene Ausreißer",
            "pdf_summary_tests": "Vergleiche", "pdf_summary_norm": "Normalisierungsmethode",
            "pdf_summary_norm_multi": "geNorm NF", "pdf_summary_norm_single": "Einzelnes Referenzgen",
            "pdf_summary_methods": "Berechnungsmethoden", "pdf_summary_methods_val": "Klassische ΔΔCq + Pfaffl",
            "pdf_disclaimer": "Dieser Bericht wurde automatisch von GeneQuantify erstellt (MIQE-Richtlinien).",
            "pdf_footer": "GeneQuantify — Nur für Forschung und Bildung. Nicht für klinische Diagnostik.",
            "pdf_fig1": "Abbildung 1. Fold-Change: Klassisches ΔΔCq vs. Pfaffl. Linie y=1 = keine Änderung.",
            "pdf_fig2": "Abbildung 2. p-Werte. Rote Balken = signifikant (p < 0,05).",
            "pdf_fig3": "Abbildung. ΔCq-Verteilung für {gene}.",
            "pdf_nochange": "Keine Änderung",
            "pdf_stat_cols": ["Zielgen", "Vergleich", "Testtyp", "Testmethode", "p-Wert", "Signifikanz"],
            "pdf_res_cols": ["Zielgen", "Gruppe", "ΔCq Kontrolle", "ΔCq Probe", "ΔΔCq", "2^(-ΔΔCq)", "Pfaffl-Verhältnis", "Regulation", "E Ziel", "E Ref"],
            "pdf_eff_cols": ["Gen", "E (Ziel)", "Eff% (Ziel)", "E (Ref)", "Eff% (Ref)", "Diff%", "Status"],
            "pdf_eff_ok": "OK", "pdf_eff_warn": "WARNUNG: Pfaffl verwenden",
            "pdf_outlier_col": "Ausreißer ausgeschlossen", "pdf_contact": "Kontakt: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} Einträge bereit — Sie können das PDF erstellen.",
            # RDML / RDES import
            "rdml_expander":        "📂 RDML / RDES-Datei importieren",
            "rdml_description":     "Laden Sie eine **RDML** (`.rdml`) oder **RDES** (`.tsv`/`.csv`/`.txt`) Datei hoch, um Cq-Werte automatisch einzufügen.",
            "rdml_uploader":        "Datei auswählen",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX, Roche LightCycler usw.  RDES: tabulatorgetrennte Tabellenkalkulation.",
            "rdml_success":         "✅ {fmt}-Datei geladen — {n} Reaktionen gefunden.",
            "rdml_error":           "❌ {fmt}-Fehler: {err}",
            "rdml_preview":         "Geparste Daten anzeigen",
            "rdml_step1":           "**Schritt 1 — Kontrollgruppe bezeichnen**",
            "rdml_ctrl_label":      "Kontrollprobenname(n) (kommagetrennte Teilstrings)",
            "rdml_ctrl_help":       "Alle Proben, deren Name diesen Text enthält, werden als Kontrolle behandelt.",
            "rdml_step2":           "**Schritt 2 — Patientengruppen bezeichnen**",
            "rdml_n_pat":           "Anzahl der Patientengruppen",
            "rdml_pat_label":       "Patientengruppe {i} Probenname(n)",
            "rdml_pat_help":        "Kommagetrennte Teilstrings. Alle passenden Proben werden in diese Gruppe zusammengeführt.",
            "rdml_apply":           "✅ {fmt}-Import auf Dateneingabe anwenden",
            "rdml_apply_success":   "✅ {n} Cq-Werte in den Dateneingabe-Tab geladen! Wechseln Sie dorthin zum Überprüfen.",
            "rdml_apply_warning":   "⚠️ Keine Werte zugeordnet. Prüfen Sie, ob Ihre Bezeichnungen mit den Probennamen in der Vorschau übereinstimmen.",
        },

        "fr": {
            "title": "🧬 GeneQuantify : Analyse de l'expression génique et des variations du nombre de copies (CNV)",
            "tab_data": "Saisie des données",
            "tab_results": "Résultats",
            "tab_report": "Rapport",
            "subtitle": "Développé par B. Yalçınkaya",
            "patient_data_header": "📊 Entrez les données des groupes patients et témoins",
            "num_target_genes": "🔹 Entrez le nombre de gènes cibles",
            "num_patient_groups": "🔹 Entrez le nombre de groupes de patients",
            "sample_number": "Numéro de l'échantillon",
            "Grup": "Groupe",
            "x_axis_title": "Nom du Groupe",
            "ct_value": "Valeur Cq",
            "reference_ct": "Cq de Référence",
            "delta_ct_control": "ΔCq (Contrôle)",
            "delta_ct_patient": "ΔCq (Patient)",
            "warning_empty_input": "⚠️ Avertissement : Entrez les données sous forme de liste ou copiez-collez sans cellules vides depuis Excel.",
            "download_csv": "📥 Télécharger CSV",
            "generate_pdf": "📥 Préparer le Rapport PDF",
            "pdf_report": "Rapport d'Analyse de l'Expression Génétique",
            "nil_mine": "📊 Résultats",
            "gr_tbl": "📋 Tableau des Données d'Entrée",
            "control_group": "🧬 Groupe Contrôle",
            "ctrl_trgt_ct": "🟦 Valeurs Cq du Gène Cible {i} pour le Groupe Contrôle",
            "ctrl_ref_ct": "🟦 Valeurs Ct du Gène Référence {i} pour le Groupe Contrôle",
            "hst_trgt_ct": "🩸 Valeurs Cq du Gène Cible {j} pour le Groupe Patient",
            "hst_ref_ct": "🩸 Valeurs Ct du Gène Référence {j} pour le Groupe Patient",
            "warning_control_ct": "⚠️ Avertissement : Les données du groupe témoin {i} doivent être saisies ligne par ligne ou copiées depuis Excel sans cellules vides.",
            "warning_patient_cq": "⚠️ Avertissement : Entrez les valeurs Cq du groupe patient ligne par ligne ou copiez-les depuis Excel sans cellules vides.",
            "target_gene": "Gène Cible",
            "reference_gene": "Gène Référence",
            "target_ct": "Cq du Gène Cible", 
            "distribution_graph": "Graphique de Distribution",
            "error_missing_control_data": "⚠️ Erreur : Données manquantes pour le Gène Cible {i} dans le Groupe Contrôle!",
            "control_group_avg": "Moyenne du Groupe Contrôle",
            "avg": "Moyenne",
            "control": "Contrôle",
            "sample": "Échantillon",
            "patient": "Patient",
            "delta_ct_distribution": "Distribution ΔCq",
            "delta_ct_value": "Valeur ΔCq",
            "parametric": "Paramétrique",
            "non_parametric": "Non paramétrique",
            "t_test": "Test t",
            "mann_whitney_u_test": "Test Mann-Whitney U",
            "welch_t_test": "Test t de Welch",
            "significant": "Significatif",
            "insignificant": "Non Significatif",
            "test_type": "Type de Test",
            "test_method": "Méthode de Test",
            "test_pvalue": "P-valeur du Test",
            "significance": "Signification",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "Changement de l'Expression Génétique (2^(-ΔΔCq))",
            "regulation_status": "Statut de Régulation",
            "no_change": "Aucun Changement",
            "upregulated": "Upregulé",
            "downregulated": "Downregulé",
            "report_title": "Rapport d'Analyse de l'Expression Génétique",
            "input_data_table": "Tableau des Données d'Entrée",
            "results": "Résultats",
            "statistical_results": "📈 Résultats Statistiques",
            "statistics": "Résultats statistiques",
            "statistical_evaluation": "Évaluation Statistique",
            "target_gene": "Gène Cible",
            "patient_group": "🩸 Groupe Patient",
            "expression_change": "Changement de l'Expression Génétique",
            "generate_pdf": "Générer le PDF",
            "pdf_report": "Rapport sur l'Expression Génétique",
            "error_no_data": "Aucune donnée trouvée, le PDF n'a pas pu être généré.",
            # Efficiency translations
            "efficiency_header": "🔬 Validation de l'Efficacité d'Amplification",
            "efficiency_method": "Méthode de saisie d'efficacité",
            "efficiency_manual": "Entrer la valeur E manuellement",
            "efficiency_slope": "Calculer à partir de la pente",
            "efficiency_target_label": "Efficacité du Gène Cible {i} (E)",
            "efficiency_ref_label": "Efficacité du Gène Référence {i} (E)",
            "efficiency_target_slope_label": "Pente du Gène Cible {i}",
            "efficiency_ref_slope_label": "Pente du Gène Référence {i}",
            "efficiency_threshold": "Seuil de différence d'efficacité acceptable (%)",
            "efficiency_ok": "✅ La différence d'efficacité est acceptable ({diff:.1f}%)",
            "efficiency_warning": "⚠️ La différence d'efficacité dépasse le seuil ({diff:.1f}%) — La méthode ΔΔCq peut ne pas être fiable!",
            "efficiency_target_pct": "Efficacité du Gène Cible",
            "efficiency_ref_pct": "Efficacité du Gène Référence",
            "efficiency_diff": "Différence",
            "pfaffl_result": "Rapport Pfaffl",
            "pfaffl_header": "Résultats de la Méthode Pfaffl",
            "classic_ddct": "Résultat ΔΔCq Classique (2^(-ΔΔCq))",
            "pfaffl_ratio": "Rapport Pfaffl",
            "method_comparison": "📊 Comparaison des Méthodes",
            "efficiency_note": "Note : E=2.0 représente une efficacité parfaite (100%). Plage acceptée : 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "Au cours du processus d'évaluation statistique, la répartition des données a été analysée à l'aide du test de Shapiro-Wilk. "
                "Si la normalité était remplie, l'homogénéité de la variance entre les groupes a été vérifiée à l'aide du test de Levene. "
                "Si la variance était égale, un test t pour échantillons indépendants a été appliqué, sinon, un test t de Welch a été utilisé. "
                "Si aucune distribution normale n'était atteinte, le test non paramétrique de Mann-Whitney U a été appliqué. "
                "La signification a été déterminée en utilisant le critère p < 0,05. "
                "Pour des suggestions et un soutien, Burhanettin Yalçınkaya - e-mail : mailtoburhanettin@gmail.com"
            ),
            "outlier_section_title": "### 🔍 Paramètres de détection des valeurs aberrantes",
            "outlier_enable": "Activer la détection des valeurs aberrantes",
            "outlier_enable_help": "Détecte les valeurs Cq statistiquement extrêmes pouvant indiquer des erreurs techniques.",
            "outlier_method_label": "Méthode de détection",
            "outlier_method_help": "Grubbs : pour les données normalement distribuées. IQR : non paramétrique, robuste pour les distributions asymétriques.",
            "outlier_alpha_label": "Niveau de signification (α)",
            "outlier_alpha_help": "α = 0,05 est standard. α plus bas = plus conservateur.",
            "outlier_iqr_label": "Multiplicateur IQR (k)",
            "outlier_iqr_help": "k=1,5 = clôtures de Tukey standard. k=3,0 = uniquement les valeurs extrêmes.",
            "outlier_expander": "ℹ️ À propos de la détection des valeurs aberrantes en qPCR",
            "grubbs_info": "ℹ️ **Conditions du test de Grubbs :** Minimum **n ≥ 3** réplicats par groupe. Seuil de signification : **α = {alpha:.2f}**. Le test suppose la normalité ; pour n < 8, la normalité ne peut pas être évaluée de manière fiable. L'application sur les **valeurs Cq brutes** (avant normalisation) est recommandée.",
            "outlier_excluded_no": "Non",
            "outlier_excluded_yes": "Oui",
            # Outlier stage selector
            "outlier_stage_label": "🔬 Étape de détection des valeurs aberrantes",
            "outlier_stage_raw": "Cq brut — avant normalisation (recommandé)",
            "outlier_stage_dct": "ΔCq — après normalisation (comportement précédent)",
            "outlier_stage_help": (
                "**Cq brut (recommandé):** Les valeurs aberrantes sont détectées sur les valeurs Cq brutes "
                "avant le calcul du ΔCq. Appliqué séparément au gène cible et à chaque gène de référence.\n\n"
                "**ΔCq:** Détection après normalisation (comportement original)."
            ),
            # Distribution plot mode selector
            "dist_plot_mode_label": "📊 Graphique de distribution — Mode d'affichage",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — recommandé",
            "dist_plot_dct":  "ΔCq  — valeurs normalisées brutes",
            "dist_plot_ddct": "ΔΔCq  — relatif à la moyenne du contrôle",
            "dist_plot_help": (
                "**RQ (recommandé):** Convertit ΔCq en 2^(-ΔCt). Valeur plus élevée = expression plus élevée.\n\n"
                "**ΔCq:** Valeurs logarithmiques brutes. Utile pour vérifier la distribution des données.\n\n"
                "**ΔΔCq:** ΔCq de chaque échantillon moins la moyenne du groupe contrôle."
            ),
            "unequal_n_warning": (
                "⚠️ **Nombre de réplicats inégal détecté — {group}:**  \n"
                "{details}  \nL'analyse utilisera la **longueur commune la plus courte (n={min_n})**.  \n"
                "Veuillez vérifier vos données d'entrée."
            ),
            # Sidebar
            "sidebar_load_example": "📂 Charger les données d'exemple",
            "sidebar_example_loaded": "✅ Données d'exemple chargées ! Allez à l'onglet Saisie des données.",
            "sidebar_desktop_title": "### 💻 Application de bureau",
            "sidebar_desktop_btn": "⬇️ Télécharger l'application de bureau",
            "sidebar_opensource_title": "### 🔓 Open Source",
            "sidebar_opensource_body": "GeneQuantify est open source (GPL-3.0).  \nCode source disponible sur GitHub :",
            "sidebar_github_btn": "⭐ Voir le code source sur GitHub",
            "sidebar_scenarios_title": "📋 Charger un scénario de validation",
            "sidebar_scenario_select": "Sélectionner un scénario",
            "sidebar_load_scenario_btn": "▶ Charger le scénario",
            "sidebar_scenario_loaded": "✅ {s} chargé ! Allez à la saisie des données.",
            "outlier_excluded_no": "Non",
            "outlier_excluded_yes": "Oui",
            "stat_decision_title": "🔬 Décision statistique",
            "stat_decision_steps": "**Sélection du test étape par étape :**",
            "stat_shapiro_title": "**1. Test de normalité de Shapiro-Wilk**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Non normal",
            "stat_levene_title": "**2. Test d'homogénéité des variances de Levene**",
            "stat_levene_skipped": "**2. Test de Levene** — *ignoré* (normalité non satisfaite ; test non paramétrique utilisé)",
            "stat_equal_var": "Variances égales",
            "stat_unequal_var": "Variances inégales",
            "stat_selected_test": "**3. Test sélectionné :**",
            "stat_reason": "**Raison :**",
            "stat_result": "**Résultat :**",
            "stat_reason_nonnormal": "Distribution non normale dans un ou les deux groupes",
            "stat_reason_normal_equal": "Les deux groupes normaux + variances égales",
            "stat_reason_normal_unequal": "Les deux groupes normaux + variances inégales (Levene p < 0,05)",
            "stat_multigroup_note": "⚠️ Remarque : Avec ≥ 3 groupes, voir la section **Comparaison multi-groupes** ci-dessous pour ANOVA / Kruskal-Wallis.",
            "multigroup_title": "## 📊 Analyse de comparaison multi-groupes",
            "multigroup_expander": "ℹ️ À propos de l'analyse statistique multi-groupes",
            "multigroup_omnibus_test": "Test omnibus",
            "multigroup_pvalue": "p-valeur",
            "multigroup_result": "Résultat",
            "multigroup_significant": "Significatif",
            "multigroup_not_significant": "Non significatif",
            "multigroup_omnibus_ns": "ℹ️ Le test omnibus est **non significatif** (p ≥ 0,05). Les comparaisons post-hoc sont affichées à titre indicatif.",
            "multigroup_posthoc_label": "**Post-hoc :**",
            "multigroup_dl_button": "📥 Télécharger les résultats post-hoc —",
            "multigroup_2group_note": "ℹ️ **Analyse multi-groupes non applicable :** Seulement 2 groupes détectés (Contrôle + 1 groupe patient).",
            "multigroup_decision_normal_equal": "✅ Distribution normale + variances égales → **ANOVA à un facteur + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Distribution normale + **variances inégales** → **ANOVA de Welch + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **Distribution non normale** → **Kruskal-Wallis + post-hoc Dunn**",
            "multigene_title": "### 🧬 Correction des comparaisons multiples multi-gènes",
            "multigene_expander": "ℹ️ Pourquoi est-ce nécessaire ?",
            "multigene_sig_raw": "Significatif (brut)",
            "multigene_sig_bonf": "Significatif (Bonferroni)",
            "multigene_sig_fdr": "Significatif (FDR B-H)",
            "multigene_warning": "⚠️ Après correction, {lost} résultat(s) ne sont plus significatifs après ajustement FDR. Rapportez les p-valeurs corrigées comme résultats principaux.",
            "multigene_success": "✅ Tous les {n} résultats significatifs restent significatifs après correction FDR.",
            "multigene_no_sig": "Aucun résultat pairwise significatif détecté (p brut < 0,05).",
            "multigene_dl_button": "📥 Télécharger les p-valeurs corrigées (CSV)",
            "multigene_chart_title": "Correction p-valeur multi-gènes : Brut / Bonferroni / FDR",
            "multigene_fc_chart_title": "Comparaison d'expression multi-gènes",
            "multigene_1gene_note": "ℹ️ **Correction multi-gènes :** Seulement 1 gène cible analysé — correction non applicable.",
            "multigene_no_data": "Pas encore de p-valeurs — entrez des données ci-dessus.",
            "ref_gene_section_title": "### 📚 Paramètres des gènes de référence",
            "ref_gene_num_label": "Nombre de gènes de référence par gène cible",
            "ref_gene_num_help": "Les directives MIQE recommandent ≥2 gènes de référence validés pour une normalisation robuste.",
            "ref_gene_1_warning": "⚠️ **Note méthodologique :** L'utilisation d'un seul gène de référence limite la robustesse de la normalisation. Les directives MIQE (Bustin et al. 2009) recommandent **≥2 gènes de référence validés** avec évaluation de la stabilité (geNorm/NormFinder).",
            "ref_gene_multi_success": "✅ {n} gènes de référence sélectionnés. La normalisation par moyenne géométrique et la stabilité geNorm M seront calculées automatiquement.",
            "ref_gene_expander": "ℹ️ À propos de la normalisation multi-référence",
            "sc_expander": "📐 Calculateur de courbe standard — Calculer E à partir d'une série de dilutions",
            "sc_gene_label": "Gène / étiquette d'amorce",
            "sc_num_points": "Nombre de points de dilution",
            "sc_dilution_factor_label": "**Facteur de dilution** (ex. 10 pour des dilutions 10 fois)",
            "sc_dilution_factor_input": "Facteur de dilution",
            "sc_start_conc_label": "**Concentration initiale** (unités arbitraires, ex. 1)",
            "sc_start_conc_input": "Concentration initiale",
            "sc_enter_ct": "**Entrez la valeur Cq moyenne pour chaque dilution :**",
            "sc_calc_button": "📊 Calculer l'efficacité",
            "sc_slope": "Pente",
            "sc_e_value": "Valeur E",
            "sc_efficiency_pct": "Efficacité %",
            "sc_excellent": "✅ Excellent ! E={e:.4f} ({pct:.1f}%), R²={r2:.4f} — Utilisez cette valeur E dans la section efficacité ci-dessous.",
            "sc_warning_r2": "⚠️ E est acceptable ({pct:.1f}%) mais R²={r2:.4f} < 0,99 — vérifiez votre série de dilutions.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) est hors de la plage acceptable (90–110%). Vérifiez la conception des amorces ou la série de dilutions.",
            "sc_chart_title": "Courbe standard — {label} | E={e:.4f} ({pct:.1f}%), R²={r2:.4f}",
            "sc_xaxis": "log₁₀(Concentration)",
            "sc_data_points": "Points de données",
            "sc_copy_hint": "💡 Copiez la pente **{slope:.4f}** ou la valeur E **{e:.4f}** dans les champs d'efficacité ci-dessous.",
            "sc_description": """\
    Entrez vos valeurs Ct de dilution en série ci-dessous. Le calculateur ajustera une régression linéaire et calculera automatiquement la pente, R² et l'efficacité d'amplification.

    **Comment utiliser :**  
    1. Effectuez la qPCR sur des dilutions en série (ex. non dilué, 1:10, 1:100, 1:1000, 1:10000)  
    2. Entrez la valeur Ct moyenne pour chaque dilution  
    3. Lisez la pente, E et R²  
    """,
            "ref_multi_description": """\
    **Normalisation par moyenne géométrique** (Vandesompele et al. 2002)  
    Le facteur de normalisation (NF) est la moyenne arithmétique des valeurs Ct de tous les gènes de référence par échantillon,  
    ce qui correspond à la moyenne géométrique de leurs niveaux d'expression.  
    `NF_échantillon = moyenne(Ct_ref1, Ct_ref2, ..., Ct_refN)` pour chaque échantillon  
    `ΔCq = Ct_cible − NF`

    **Valeur M de geNorm** (score de stabilité)  
    Pour chaque gène de référence, M = écart-type moyen des log-ratios par rapport à tous les autres gènes de référence.  
    **M plus bas = plus stable.** Seuil recommandé MIQE : M < 0,5 (strict) ou M < 1,0 (acceptable).

    **CV (Coefficient de Variation)**  
    `CV = (ET / moyenne) × 100%` des valeurs Cq brutes sur tous les échantillons.  
    Un CV plus faible indique moins de variation et une meilleure stabilité comme référence.

    **Référence :** Vandesompele J et al. *Genome Biology* 2002 ; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **Pourquoi la détection des valeurs aberrantes est importante en qPCR**

    La variabilité technique est inhérente à la qPCR : erreurs de pipetage, formation de bulles, contamination par des inhibiteurs ou variation de la qualité de l'ARN peuvent produire des valeurs Cq statistiquement incohérentes avec le reste d'un groupe de réplicats.  
    L'inclusion de telles valeurs gonfle la variance, biaise les moyennes et peut conduire à de fausses conclusions — particulièrement dans les jeux de données cliniques avec de petits effectifs.

    **Quand cette limitation devient critique :**
    - Petits groupes (n < 5) : un seul Ct erroné déplace substantiellement la moyenne
    - Variabilité biologique élevée (ex. hétérogénéité tumorale, cohortes cliniques)
    - Triplicats techniques où un réplicat diverge de > 0,5 Ct des autres
    - Cibles à faible abondance avec Ct > 35, où le bruit domine

    **Test de Grubbs** *(Grubbs 1969)*  
    Suppose la normalité. Teste si la valeur la plus extrême est un outlier statistiquement significatif (p < α). Itère jusqu'à ce qu'aucun autre outlier ne soit trouvé.  
    Meilleur pour : valeurs Ct répliquées d'un seul groupe expérimental.

    **Méthode IQR** *(Tukey 1977)*  
    Non paramétrique. Signale les valeurs en dehors de Q1 − k×IQR ou Q3 + k×IQR.  
    Meilleur pour : groupes plus importants ou distributions non normales.

    **Important :** L'exclusion des outliers nécessite une **justification biologique ou technique**.  
    Cet outil signale des candidats — la décision finale appartient toujours au chercheur.  
    Toutes les exclusions sont enregistrées et rapportées dans le rapport PDF.

    **Références :** Grubbs FE. *Technometrics* 1969 ; Tukey JW. *Exploratory Data Analysis* 1977 ;  
    Bustin SA et al. *Clin Chem* 2009 (directives MIQE).
    """,

            # ── Chaînes du rapport PDF ────────────────────────────────────────────
            "pdf_cover_subtitle": "Rapport d'analyse d'expression génique qPCR",
            "pdf_generated": "Généré le: {now}",
            "pdf_s1_title": "1. Méthodes et paramètres d'analyse",
            "pdf_s1_calc": "1.1 Méthodes de calcul",
            "pdf_s1_calc_body": "Deux méthodes complémentaires ont été appliquées pour le calcul du fold-change:",
            "pdf_s1_classic": "ΔΔCq classique (Livak & Schmittgen, 2001): ΔCq = Ct(cible) - Ct(référence);  ΔΔCt = ΔCt(échantillon) - ΔCt(contrôle);  Fold-Change = 2^(-ΔΔCt). Suppose des efficacités égales (E ≈ 2,0).",
            "pdf_s1_pfaffl": "Méthode Pfaffl (Pfaffl, 2001): Ratio = (E_cible ^ ΔCq_cible) / (E_réf ^ ΔCt_réf). Corrige les efficacités spécifiques; recommandé si différence > 10%.",
            "pdf_s1_norm": "1.2 Normalisation",
            "pdf_s1_norm_multi": "Gènes de référence multiples (n={n}) utilisés. NF = moyenne arithmétique des Cq des gènes de référence par échantillon (geNorm, Vandesompele et al. 2002).",
            "pdf_s1_norm_single": "Un seul gène de référence utilisé. Les directives MIQE recommandent ≥2 gènes de référence.",
            "pdf_s1_eff": "1.3 Efficacité d'amplification",
            "pdf_s1_eff_range": "Plage d'efficacité acceptable: E = 1,8-2,2 (90-110%). Seuil de différence appliqué: {thr}%.",
            "pdf_s1_outlier": "1.4 Détection des valeurs aberrantes",
            "pdf_s1_grubbs": "Test de Grubbs (Grubbs 1969) appliqué, alpha = {alpha}. {n} échantillon(s) signalé(s) et confirmé(s) par l'utilisateur.",
            "pdf_s1_iqr": "Méthode IQR (Tukey 1977), multiplicateur k = {k}. {n} échantillon(s) exclu(s).",
            "pdf_s1_outlier_warn": "AVERTISSEMENT: L'exclusion des valeurs aberrantes nécessite une justification biologique ou technique.",
            "pdf_s1_outlier_off": "Détection des valeurs aberrantes désactivée pour cette analyse.",
            "pdf_s2_title": "2. Données d'entrée",
            "pdf_s2_body": "Valeurs Cq brutes saisies par l'utilisateur après traitement des valeurs aberrantes.",
            "pdf_s3_title": "3. Résultats d'expression génique",
            "pdf_s3_body": "Valeurs de fold-change calculées par ΔΔCq classique et méthode Pfaffl. Fold-change > 1 = expression plus élevée dans le groupe patient.",
            "pdf_s4_title": "4. Analyse statistique",
            "pdf_s4_body": "Signification statistique des différences d'expression génique. Sélection automatique du test selon normalité (Shapiro-Wilk) et homogénéité des variances (Levene). Seuil: p < 0,05.",
            "pdf_s4_interp": "Interprétation des tests statistiques",
            "pdf_s4_interp_body": "t de Student: groupes normaux avec variances égales. t de Welch: normaux mais variances inégales. Mann-Whitney U: non-paramétrique. p < 0,05 = expression différentielle significative.",
            "pdf_s5_title": "5. Graphiques de distribution Delta Cq",
            "pdf_s5_body": "Distribution des valeurs ΔCq pour chaque gène cible. Chaque point = un réplicat. Barres horizontales = moyennes des groupes.",
            "pdf_s6_title": "6. Comment interpréter vos résultats",
            "pdf_s6_fc": "6.1 Interprétation du fold-change",
            "pdf_s6_choose": "6.2 Choisir entre ΔΔCq et Pfaffl",
            "pdf_s6_choose_body": "ΔΔCq classique si: efficacités 90-110% et différence < 10%. Pfaffl si: différence > 10%. Toujours rapporter les deux valeurs.",
            "pdf_s6_stat": "6.3 Justification du choix du test",
            "pdf_s6_stat_body": "Normalité: test de Shapiro-Wilk (n < 50). Homogénéité des variances: test de Levene. Paramétrique/variances égales: t de Student. Variances inégales: t de Welch. Non-normal: Mann-Whitney U.",
            "pdf_s7_title": "7. Références",
            "pdf_fc_interp_header": ["Fold-Change", "ΔΔCq", "Interprétation", "Signification biologique"],
            "pdf_fc_interp_rows": [
                [">2,0", "<-1,0", "Forte surexpression", "Biologiquement pertinent"],
                ["1,5-2,0", "-1,0 à -0,58", "Surexpression modérée", "Potentiellement pertinent"],
                ["1,0-1,5", "-0,58 à 0", "Faible surexpression", "Probablement non significatif seul"],
                ["1,0", "0", "Aucun changement", "Pas d'expression différentielle"],
                ["0,67-1,0", "0 à 0,58", "Faible sous-expression", "Probablement non significatif seul"],
                ["0,5-0,67", "0,58 à 1,0", "Sous-expression modérée", "Potentiellement pertinent"],
                ["<0,5", ">1,0", "Forte sous-expression", "Biologiquement pertinent"],
            ],
            "pdf_stat_note": "Note: La signification statistique et biologique doivent être évaluées ensemble.",
            "pdf_summary_param": "Paramètre",
            "pdf_summary_val": "Valeur",
            "pdf_summary_genes": "Gènes cibles analysés",
            "pdf_summary_groups": "Groupes de patients",
            "pdf_summary_samples": "Échantillons totaux",
            "pdf_summary_excluded": "Valeurs aberrantes exclues",
            "pdf_summary_tests": "comparaisons",
            "pdf_summary_norm": "Méthode de normalisation",
            "pdf_summary_norm_multi": "geNorm NF",
            "pdf_summary_norm_single": "Gène de référence unique",
            "pdf_summary_methods": "Méthodes de calcul",
            "pdf_summary_methods_val": "ΔΔCq classique + Pfaffl",
            "pdf_disclaimer": "Ce rapport a été généré automatiquement par GeneQuantify conformément aux directives MIQE.",
            "pdf_footer": "GeneQuantify — Usage recherche et éducation uniquement. Non validé pour usage clinique.",
            "pdf_fig1": "Figure 1. Comparaison du fold-change: ΔΔCq classique vs Pfaffl. Ligne pointillée y=1 = aucun changement.",
            "pdf_fig2": "Figure 2. Valeurs p de toutes les comparaisons. Barres rouges = significatif (p < 0,05).",
            "pdf_fig3": "Figure. Distribution ΔCq pour {gene}. Points = réplicats; barres = moyennes des groupes.",
            "pdf_nochange": "Aucun changement",
            "pdf_stat_cols": ["Gène cible", "Comparaison", "Type de test", "Méthode", "Valeur p", "Signification"],
            "pdf_res_cols": ["Gène cible", "Groupe", "ΔCq Contrôle", "ΔCq Échantillon", "ΔΔCq", "2^(-ΔΔCq)", "Ratio Pfaffl", "Régulation", "E cible", "E réf"],
            "pdf_eff_cols": ["Gène", "E (cible)", "Eff% (cible)", "E (réf)", "Eff% (réf)", "Diff%", "Statut"],
            "pdf_eff_ok": "OK",
            "pdf_eff_warn": "AVERTISSEMENT: utiliser Pfaffl",
            "pdf_outlier_col": "Valeur aberrante exclue",
            "pdf_contact": "Contact: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} enregistrements prêts — vous pouvez générer le PDF.",
            # RDML / RDES import
            "rdml_expander":        "📂 Importer un fichier RDML / RDES",
            "rdml_description":     "Importez un fichier **RDML** (`.rdml`) ou **RDES** (`.tsv`/`.csv`/`.txt`) pour remplir automatiquement les valeurs Cq.",
            "rdml_uploader":        "Choisir un fichier",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX, Roche LightCycler, etc.  RDES: tableau séparé par tabulations.",
            "rdml_success":         "✅ Fichier {fmt} chargé — {n} réactions trouvées.",
            "rdml_error":           "❌ Erreur d'analyse {fmt} : {err}",
            "rdml_preview":         "Aperçu des données analysées",
            "rdml_step1":           "**Étape 1 — Étiquetez votre groupe contrôle**",
            "rdml_ctrl_label":      "Nom(s) d'échantillon contrôle (sous-chaînes séparées par des virgules)",
            "rdml_ctrl_help":       "Tout échantillon dont le nom contient ce texte sera traité comme Contrôle.",
            "rdml_step2":           "**Étape 2 — Étiquetez vos groupes patients**",
            "rdml_n_pat":           "Nombre de groupes patients",
            "rdml_pat_label":       "Nom(s) d'échantillon du groupe patient {i}",
            "rdml_pat_help":        "Sous-chaînes séparées par des virgules. Tous les échantillons correspondants seront regroupés.",
            "rdml_apply":           "✅ Appliquer l'import {fmt} à la saisie des données",
            "rdml_apply_success":   "✅ {n} valeurs Cq chargées dans l'onglet Saisie ! Vérifiez et ajustez si nécessaire.",
            "rdml_apply_warning":   "⚠️ Aucune valeur correspondante. Vérifiez que vos étiquettes correspondent aux noms dans l'aperçu.",
        },

        "es": {
            "title": "🧬 GeneQuantify: Análisis de Expresión Génica y CNV",
            "tab_data": "Entrada de datos",
            "tab_results": "Resultados",
            "tab_report": "Informe",
            "subtitle": "Desarrollado por B. Yalçınkaya",
            "patient_data_header": "📊 Ingrese Datos de Grupos de Pacientes y de Control",
            "num_target_genes": "🔹 Ingrese el número de Genes Objetivo",
            "num_patient_groups": "🔹 Ingrese el número de Grupos de Pacientes",
            "sample_number": "Número de muestra",
            "Grup": "Grupo",
            "x_axis_title": "Nombre del Grupo",
            "ct_value": "Valor de Cq",
            "reference_ct": "Ct de Referencia",
            "delta_ct_control": "ΔCq (Control)",
            "delta_ct_patient": "ΔCq (Paciente)",
            "warning_empty_input": "⚠️ Advertencia: Ingrese los datos uno debajo del otro o cópielos sin celdas vacías desde Excel.",
            "download_csv": "📥 Descargar CSV",
            "generate_pdf": "📥 Preparar Informe en PDF",
            "pdf_report": "Informe de Análisis de Expresión Génica",
            "nil_mine": "📊 Resultados",
            "gr_tbl": "📋 Tabla de Datos de Entrada",
            "control_group": "🧬 Grupo Control",
            "ctrl_trgt_ct": "🟦 Valores Ct del Gen Objetivo {i} para el Grupo Control",
            "ctrl_ref_ct": "🟦 Valores Ct del Gen de Referencia {i} para el Grupo Control",
            "hst_trgt_ct": "🩸 Valores Ct del Gen Objetivo {j} para el Grupo Paciente",
            "hst_ref_ct": "🩸 Valores Ct del Gen de Referencia {j} para el Grupo Paciente",
            "warning_control_ct": "⚠️ Advertencia: Los datos del grupo control {i} deben ingresarse fila por fila o copiarse desde Excel sin celdas vacías.",
            "warning_patient_cq": "⚠️ Advertencia: Ingrese los valores de Ct del grupo paciente fila por fila o cópielos desde Excel sin celdas vacías.",
            "target_gene": "Gen Objetivo",
            "reference_gene": "Gen de Referencia",
            "target_ct": "Ct del Gen Objetivo", 
            "distribution_graph": "Gráfico de Distribución",
            "error_missing_control_data": "⚠️ Error: ¡Datos faltantes para el Gen Objetivo {i} en el Grupo Control!",
            "control_group_avg": "Promedio del Grupo Control",
            "avg": "Promedio",
            "control": "Control",
            "sample": "Muestra",
            "patient": "Paciente",
            "delta_ct_distribution": "Distribución ΔCq",
            "delta_ct_value": "Valor ΔCq",
            "parametric": "Paramétrico",
            "non_parametric": "No paramétrico",
            "t_test": "Test t",
            "mann_whitney_u_test": "Test Mann-Whitney U",
            "welch_t_test": "Test t de Welch",
            "significant": "Significativo",
            "insignificant": "No Significativo",
            "test_type": "Tipo de Test",
            "test_method": "Método de Test",
            "test_pvalue": "P-valor del Test",
            "significance": "Significación",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "Cambio de Expresión Génica (2^(-ΔΔCq))",
            "regulation_status": "Estado de Regulación",
            "no_change": "Sin Cambio",
            "upregulated": "Upregulado",
            "downregulated": "Downregulado",
            "report_title": "Informe de Análisis de Expresión Génica",
            "input_data_table": "Tabla de Datos de Entrada",
            "results": "Resultados",
            "statistical_results": "📈 Resultados Estadísticos",
            "statistics": "Resultados estadísticos",
            "statistical_evaluation": "Evaluación Estadística",
            "target_gene": "Gen Objetivo",
            "patient_group": "🩸 Grupo Paciente",
            "expression_change": "Cambio de Expresión Génica",
            "generate_pdf": "Generar PDF",
            "pdf_report": "Informe de Expresión Génica",
            "error_no_data": "No se encontraron datos, no se pudo generar el PDF.",
            # Efficiency translations
            "efficiency_header": "🔬 Validación de Eficiencia de Amplificación",
            "efficiency_method": "Método de entrada de eficiencia",
            "efficiency_manual": "Ingresar valor E manualmente",
            "efficiency_slope": "Calcular desde pendiente",
            "efficiency_target_label": "Eficiencia del Gen Objetivo {i} (E)",
            "efficiency_ref_label": "Eficiencia del Gen de Referencia {i} (E)",
            "efficiency_target_slope_label": "Pendiente del Gen Objetivo {i}",
            "efficiency_ref_slope_label": "Pendiente del Gen de Referencia {i}",
            "efficiency_threshold": "Umbral de diferencia de eficiencia aceptable (%)",
            "efficiency_ok": "✅ La diferencia de eficiencia es aceptable ({diff:.1f}%)",
            "efficiency_warning": "⚠️ La diferencia de eficiencia supera el umbral ({diff:.1f}%) — ¡El método ΔΔCq puede no ser confiable!",
            "efficiency_target_pct": "Eficiencia del Gen Objetivo",
            "efficiency_ref_pct": "Eficiencia del Gen de Referencia",
            "efficiency_diff": "Diferencia",
            "pfaffl_result": "Relación Pfaffl",
            "pfaffl_header": "Resultados del Método Pfaffl",
            "classic_ddct": "Resultado ΔΔCq Clásico (2^(-ΔΔCq))",
            "pfaffl_ratio": "Relación Pfaffl",
            "method_comparison": "📊 Comparación de Métodos",
            "efficiency_note": "Nota: E=2.0 representa eficiencia perfecta (100%). Rango aceptado: 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "Durante el proceso de evaluación estadística, se analizó la distribución de los datos mediante la prueba de Shapiro-Wilk. "
                "Si se cumplió la normalidad, se verificó la homogeneidad de varianza entre los grupos mediante la prueba de Levene. "
                "Si la varianza era igual, se aplicó la prueba t de muestras independientes; de lo contrario, se utilizó la prueba t de Welch. "
                "Si no se alcanzó una distribución normal, se aplicó la prueba no paramétrica Mann-Whitney U. "
                "La significancia se determinó utilizando el criterio p < 0.05. "
                "Para sugerencias y soporte, Burhanettin Yalçınkaya - correo electrónico: mailtoburhanettin@gmail.com"
            ),
            "outlier_section_title": "### 🔍 Configuración de detección de valores atípicos",
            "outlier_enable": "Activar detección de valores atípicos",
            "outlier_enable_help": "Detecta valores Ct estadísticamente extremos que pueden reflejar errores técnicos.",
            "outlier_method_label": "Método de detección",
            "outlier_method_help": "Grubbs: para datos normalmente distribuidos. IQR: no paramétrico, robusto para distribuciones asimétricas.",
            "outlier_alpha_label": "Nivel de significancia (α)",
            "outlier_alpha_help": "α = 0,05 es estándar. α más bajo = más conservador.",
            "outlier_iqr_label": "Multiplicador IQR (k)",
            "outlier_iqr_help": "k=1,5 = cercas de Tukey estándar. k=3,0 = solo valores extremos.",
            "outlier_expander": "ℹ️ Sobre la detección de valores atípicos en qPCR",
            "grubbs_info": "ℹ️ **Requisitos del test de Grubbs:** Mínimo **n ≥ 3** réplicas por grupo. Umbral de significancia: **α = {alpha:.2f}**. El test asume normalidad; para n < 8, la normalidad no puede evaluarse de forma confiable. Se recomienda aplicar el test a los **valores Cq brutos** (antes de la normalización).",
            "outlier_excluded_no": "No",
            "outlier_excluded_yes": "Sí",
            # Outlier stage selector
            "outlier_stage_label": "🔬 Etapa de detección de valores atípicos",
            "outlier_stage_raw": "Ct bruto — antes de la normalización (recomendado)",
            "outlier_stage_dct": "ΔCq — después de la normalización (comportamiento anterior)",
            "outlier_stage_help": (
                "**Ct bruto (recomendado):** Los valores atípicos se detectan en los valores Ct brutos "
                "antes del cálculo del ΔCq. Aplicado por separado al gen objetivo y a cada gen de referencia.\n\n"
                "**ΔCq:** Detección después de la normalización (comportamiento original)."
            ),
            # Distribution plot mode selector
            "dist_plot_mode_label": "📊 Gráfico de distribución — Modo de visualización",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — recomendado",
            "dist_plot_dct":  "ΔCq  — valores normalizados brutos",
            "dist_plot_ddct": "ΔΔCq  — relativo a la media del control",
            "dist_plot_help": (
                "**RQ (recomendado):** Convierte ΔCq a 2^(-ΔCt). Mayor valor = mayor expresión.\n\n"
                "**ΔCq:** Valores logarítmicos brutos. Útil para verificar la distribución.\n\n"
                "**ΔΔCq:** ΔCq de cada muestra menos la media del grupo control."
            ),
            "unequal_n_warning": (
                "⚠️ **Recuentos de réplicas desiguales detectados — {group}:**  \n"
                "{details}  \nEl análisis usará la **longitud común más corta (n={min_n})**.  \n"
                "Verifique sus datos de entrada."
            ),
            # Sidebar
            "sidebar_load_example": "📂 Cargar datos de ejemplo",
            "sidebar_example_loaded": "✅ ¡Datos de ejemplo cargados! Cambie a la pestaña de entrada de datos.",
            "sidebar_desktop_title": "### 💻 Aplicación de escritorio",
            "sidebar_desktop_btn": "⬇️ Descargar aplicación de escritorio",
            "sidebar_opensource_title": "### 🔓 Código abierto",
            "sidebar_opensource_body": "GeneQuantify es de código abierto (GPL-3.0).  \nCódigo fuente disponible en GitHub:",
            "sidebar_github_btn": "⭐ Ver código fuente en GitHub",
            "sidebar_scenarios_title": "📋 Cargar escenario de validación",
            "sidebar_scenario_select": "Seleccionar escenario",
            "sidebar_load_scenario_btn": "▶ Cargar escenario",
            "sidebar_scenario_loaded": "✅ {s} cargado. Vaya a Entrada de datos.",
            "outlier_excluded_no": "No",
            "outlier_excluded_yes": "Sí",
            "stat_decision_title": "🔬 Decisión estadística",
            "stat_decision_steps": "**Selección de prueba paso a paso:**",
            "stat_shapiro_title": "**1. Prueba de normalidad de Shapiro-Wilk**",
            "stat_normal": "Normal",
            "stat_nonnormal": "No normal",
            "stat_levene_title": "**2. Prueba de homogeneidad de varianzas de Levene**",
            "stat_levene_skipped": "**2. Prueba de Levene** — *omitida* (normalidad no cumplida; se usará prueba no paramétrica)",
            "stat_equal_var": "Varianzas iguales",
            "stat_unequal_var": "Varianzas desiguales",
            "stat_selected_test": "**3. Prueba seleccionada:**",
            "stat_reason": "**Razón:**",
            "stat_result": "**Resultado:**",
            "stat_reason_nonnormal": "Distribución no normal en uno o ambos grupos",
            "stat_reason_normal_equal": "Ambos grupos normales + varianzas iguales",
            "stat_reason_normal_unequal": "Ambos grupos normales + varianzas desiguales (Levene p < 0,05)",
            "stat_multigroup_note": "⚠️ Nota: Con ≥ 3 grupos, consulte la sección **Comparación multigrupo** para ANOVA / Kruskal-Wallis.",
            "multigroup_title": "## 📊 Análisis de comparación multigrupo",
            "multigroup_expander": "ℹ️ Sobre el análisis estadístico multigrupo",
            "multigroup_omnibus_test": "Prueba ómnibus",
            "multigroup_pvalue": "p-valor",
            "multigroup_result": "Resultado",
            "multigroup_significant": "Significativo",
            "multigroup_not_significant": "No significativo",
            "multigroup_omnibus_ns": "ℹ️ La prueba ómnibus es **no significativa** (p ≥ 0,05). Las comparaciones post-hoc se muestran a título informativo.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Descargar resultados post-hoc —",
            "multigroup_2group_note": "ℹ️ **Análisis multigrupo no aplicable:** Solo 2 grupos detectados (Control + 1 grupo paciente).",
            "multigroup_decision_normal_equal": "✅ Distribución normal + varianzas iguales → **ANOVA de un factor + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Distribución normal + **varianzas desiguales** → **ANOVA de Welch + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **Distribución no normal** → **Kruskal-Wallis + post-hoc Dunn**",
            "multigene_title": "### 🧬 Corrección de comparaciones múltiples multigénicas",
            "multigene_expander": "ℹ️ ¿Por qué es necesario?",
            "multigene_sig_raw": "Significativo (bruto)",
            "multigene_sig_bonf": "Significativo (Bonferroni)",
            "multigene_sig_fdr": "Significativo (FDR B-H)",
            "multigene_warning": "⚠️ Tras la corrección, {lost} resultado(s) ya no son significativos tras el ajuste FDR. Reporte los p-valores corregidos como resultados principales.",
            "multigene_success": "✅ Todos los {n} resultados significativos permanecen significativos tras la corrección FDR.",
            "multigene_no_sig": "No se detectaron resultados pairwise significativos (p bruto < 0,05).",
            "multigene_dl_button": "📥 Descargar p-valores corregidos (CSV)",
            "multigene_chart_title": "Corrección p-valor multigénica: Bruto / Bonferroni / FDR",
            "multigene_fc_chart_title": "Comparación de expresión multi-gen",
            "multigene_1gene_note": "ℹ️ **Corrección multigénica:** Solo 1 gen objetivo analizado — corrección no aplicable.",
            "multigene_no_data": "Aún no hay p-valores — ingrese datos arriba.",
            "ref_gene_section_title": "### 📚 Configuración de genes de referencia",
            "ref_gene_num_label": "Número de genes de referencia por gen objetivo",
            "ref_gene_num_help": "Las directrices MIQE recomiendan ≥2 genes de referencia validados para una normalización robusta.",
            "ref_gene_1_warning": "⚠️ **Nota metodológica:** El uso de un solo gen de referencia limita la robustez de la normalización. Las directrices MIQE (Bustin et al. 2009) recomiendan **≥2 genes de referencia validados** con evaluación de estabilidad (geNorm/NormFinder).",
            "ref_gene_multi_success": "✅ {n} genes de referencia seleccionados. La normalización por media geométrica y la estabilidad geNorm M se calcularán automáticamente.",
            "ref_gene_expander": "ℹ️ Sobre la normalización con múltiples referencias",
            "sc_expander": "📐 Calculadora de curva estándar — Calcular E a partir de serie de diluciones",
            "sc_gene_label": "Gen / etiqueta de cebador",
            "sc_num_points": "Número de puntos de dilución",
            "sc_dilution_factor_label": "**Factor de dilución** (ej. 10 para diluciones 10 veces)",
            "sc_dilution_factor_input": "Factor de dilución",
            "sc_start_conc_label": "**Concentración inicial** (unidades arbitrarias, ej. 1)",
            "sc_start_conc_input": "Concentración inicial",
            "sc_enter_ct": "**Ingrese el Cq medio para cada dilución:**",
            "sc_calc_button": "📊 Calcular eficiencia",
            "sc_slope": "Pendiente",
            "sc_e_value": "Valor E",
            "sc_efficiency_pct": "Eficiencia %",
            "sc_excellent": "✅ ¡Excelente! E={e:.4f} ({pct:.1f}%), R²={r2:.4f} — Use este valor E en la sección de eficiencia abajo.",
            "sc_warning_r2": "⚠️ E es aceptable ({pct:.1f}%) pero R²={r2:.4f} < 0,99 — verifique su serie de diluciones.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) está fuera del rango aceptable (90–110%). Revise el diseño de cebadores o la serie de diluciones.",
            "sc_chart_title": "Curva estándar — {label} | E={e:.4f} ({pct:.1f}%), R²={r2:.4f}",
            "sc_xaxis": "log₁₀(Concentración)",
            "sc_data_points": "Puntos de datos",
            "sc_copy_hint": "💡 Copie la pendiente **{slope:.4f}** o el valor E **{e:.4f}** en los campos de eficiencia abajo.",
            "sc_description": """\
    Ingrese sus valores Ct de dilución en serie a continuación. La calculadora ajustará una regresión lineal y calculará automáticamente la pendiente, R² y la eficiencia de amplificación.

    **Cómo usar:**  
    1. Realice qPCR en diluciones seriadas (ej. sin diluir, 1:10, 1:100, 1:1000, 1:10000)  
    2. Ingrese el Ct medio para cada dilución  
    3. Lea la pendiente, E y R²  
    """,
            "ref_multi_description": """\
    **Normalización por media geométrica** (Vandesompele et al. 2002)  
    El factor de normalización (NF) es la media aritmética de los valores Ct de todos los genes de referencia por muestra,  
    lo que corresponde a la media geométrica de sus niveles de expresión.  
    `NF_muestra = media(Ct_ref1, Ct_ref2, ..., Ct_refN)` para cada muestra  
    `ΔCq = Ct_objetivo − NF`

    **Valor M de geNorm** (puntuación de estabilidad)  
    Para cada gen de referencia, M = desviación estándar media de los log-ratios contra todos los demás genes de referencia.  
    **M más bajo = más estable.** Umbral recomendado MIQE: M < 0,5 (estricto) o M < 1,0 (aceptable).

    **CV (Coeficiente de Variación)**  
    `CV = (DE / media) × 100%` de los valores Ct brutos en todas las muestras.  
    Un CV más bajo indica menos variación y mejor estabilidad como referencia.

    **Referencia:** Vandesompele J et al. *Genome Biology* 2002; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **Por qué la detección de valores atípicos es importante en qPCR**

    La variabilidad técnica es inherente a la qPCR: errores de pipeteo, formación de burbujas, arrastre de inhibidores o variación en la calidad del ARN pueden producir valores Ct estadísticamente inconsistentes con el resto de un grupo de réplicas.  
    Incluir tales valores infla la varianza, sesga las medias y puede llevar a conclusiones falsas — particularmente en conjuntos de datos clínicos con tamaños de muestra pequeños.

    **Cuándo esta limitación se vuelve crítica:**
    - Grupos pequeños (n < 5): un único Ct erróneo desplaza sustancialmente la media
    - Alta variabilidad biológica (ej. heterogeneidad tumoral, cohortes clínicas)
    - Triplicados técnicos donde una réplica diverge > 0,5 Ct de las demás
    - Objetivos de baja abundancia con Ct > 35, donde el ruido domina

    **Prueba de Grubbs** *(Grubbs 1969)*  
    Asume normalidad. Prueba si el valor más extremo es un outlier estadísticamente significativo (p < α). Itera hasta que no se encuentren más outliers.  
    Mejor para: valores Ct replicados de un único grupo experimental.

    **Método IQR** *(Tukey 1977)*  
    No paramétrico. Señala valores fuera de Q1 − k×IQR o Q3 + k×IQR.  
    Mejor para: grupos más grandes o distribuciones no normales.

    **Importante:** La exclusión de outliers requiere **justificación biológica o técnica**.  
    Esta herramienta señala candidatos — la decisión final siempre recae en el investigador.  
    Todas las exclusiones se registran y reportan en el PDF.

    **Referencias:** Grubbs FE. *Technometrics* 1969; Tukey JW. *Exploratory Data Analysis* 1977;  
    Bustin SA et al. *Clin Chem* 2009 (directrices MIQE).
    """,

            # ── Cadenas del informe PDF ───────────────────────────────────────────
            "pdf_cover_subtitle": "Informe de análisis de expresión génica por qPCR",
            "pdf_generated": "Generado: {now}",
            "pdf_s1_title": "1. Métodos y configuración del análisis",
            "pdf_s1_calc": "1.1 Métodos de cálculo",
            "pdf_s1_calc_body": "Se aplicaron dos métodos complementarios para el cálculo del fold-change:",
            "pdf_s1_classic": "ΔΔCq clásico (Livak & Schmittgen, 2001): ΔCq = Ct(objetivo) - Ct(referencia);  ΔΔCt = ΔCt(muestra) - ΔCt(control);  Fold-Change = 2^(-ΔΔCt). Asume eficiencias iguales (E ≈ 2,0).",
            "pdf_s1_pfaffl": "Método Pfaffl (Pfaffl, 2001): Ratio = (E_objetivo ^ ΔCq_objetivo) / (E_ref ^ ΔCt_ref). Corrige eficiencias específicas; recomendado si diferencia > 10%.",
            "pdf_s1_norm": "1.2 Normalización",
            "pdf_s1_norm_multi": "Genes de referencia múltiples (n={n}) utilizados. NF calculado como media aritmética de Ct de referencia (geNorm, Vandesompele et al. 2002).",
            "pdf_s1_norm_single": "Un solo gen de referencia utilizado. Las directrices MIQE recomiendan ≥2 genes de referencia.",
            "pdf_s1_eff": "1.3 Eficiencia de amplificación",
            "pdf_s1_eff_range": "Rango aceptable: E = 1,8-2,2 (90-110%). Umbral de diferencia aplicado: {thr}%.",
            "pdf_s1_outlier": "1.4 Detección de valores atípicos",
            "pdf_s1_grubbs": "Prueba de Grubbs (Grubbs 1969), alpha = {alpha}. {n} muestra(s) marcada(s) y confirmada(s) por el usuario.",
            "pdf_s1_iqr": "Método IQR (Tukey 1977), multiplicador k = {k}. {n} muestra(s) excluida(s).",
            "pdf_s1_outlier_warn": "ADVERTENCIA: La exclusión de valores atípicos requiere justificación biológica o técnica.",
            "pdf_s1_outlier_off": "Detección de valores atípicos desactivada para este análisis.",
            "pdf_s2_title": "2. Datos de entrada",
            "pdf_s2_body": "Valores Ct brutos introducidos por el usuario tras el procesamiento de valores atípicos.",
            "pdf_s3_title": "3. Resultados de expresión génica",
            "pdf_s3_body": "Valores de fold-change calculados por ΔΔCq clásico y método Pfaffl. Fold-change > 1 = expresión mayor en el grupo paciente.",
            "pdf_s4_title": "4. Análisis estadístico",
            "pdf_s4_body": "Significación estadística de las diferencias de expresión. Selección automática según normalidad (Shapiro-Wilk) y homogeneidad de varianzas (Levene). Umbral: p < 0,05.",
            "pdf_s4_interp": "Interpretación de los tests estadísticos",
            "pdf_s4_interp_body": "t de Student: grupos normales con varianzas iguales. t de Welch: normales con varianzas desiguales. Mann-Whitney U: no paramétrico. p < 0,05 = expresión diferencial significativa.",
            "pdf_s5_title": "5. Gráficos de distribución Delta Ct",
            "pdf_s5_body": "Distribución de valores ΔCq por gen objetivo. Cada punto = un réplica. Barras horizontales = medias de grupo.",
            "pdf_s6_title": "6. Cómo interpretar los resultados",
            "pdf_s6_fc": "6.1 Interpretación del fold-change",
            "pdf_s6_choose": "6.2 Elección entre ΔΔCq y Pfaffl",
            "pdf_s6_choose_body": "ΔΔCq clásico si: eficiencias 90-110% y diferencia < 10%. Pfaffl si: diferencia > 10%. Reportar siempre ambos valores.",
            "pdf_s6_stat": "6.3 Justificación de la selección del test",
            "pdf_s6_stat_body": "Normalidad: Shapiro-Wilk (n < 50). Homogeneidad: Levene. Paramétrico/varianzas iguales: t de Student. Desiguales: Welch. No normal: Mann-Whitney U.",
            "pdf_s7_title": "7. Referencias",
            "pdf_fc_interp_header": ["Fold-Change", "ΔΔCq", "Interpretación", "Significado biológico"],
            "pdf_fc_interp_rows": [
                [">2,0", "<-1,0", "Fuerte sobreexpresión", "Considerar biológicamente relevante"],
                ["1,5-2,0", "-1,0 a -0,58", "Sobreexpresión moderada", "Puede ser relevante"],
                ["1,0-1,5", "-0,58 a 0", "Sobreexpresión débil", "Probablemente no significativo solo"],
                ["1,0", "0", "Sin cambio", "Sin expresión diferencial"],
                ["0,67-1,0", "0 a 0,58", "Subexpresión débil", "Probablemente no significativo solo"],
                ["0,5-0,67", "0,58 a 1,0", "Subexpresión moderada", "Puede ser relevante"],
                ["<0,5", ">1,0", "Fuerte subexpresión", "Considerar biológicamente relevante"],
            ],
            "pdf_stat_note": "Nota: Evaluar conjuntamente la significación estadística y biológica.",
            "pdf_summary_param": "Parámetro",
            "pdf_summary_val": "Valor",
            "pdf_summary_genes": "Genes objetivo analizados",
            "pdf_summary_groups": "Grupos de pacientes",
            "pdf_summary_samples": "Muestras totales",
            "pdf_summary_excluded": "Muestras excluidas",
            "pdf_summary_tests": "comparaciones",
            "pdf_summary_norm": "Método de normalización",
            "pdf_summary_norm_multi": "geNorm NF",
            "pdf_summary_norm_single": "Gen de referencia único",
            "pdf_summary_methods": "Métodos de cálculo",
            "pdf_summary_methods_val": "ΔΔCq clásico + Pfaffl",
            "pdf_disclaimer": "Este informe fue generado automáticamente por GeneQuantify siguiendo las directrices MIQE.",
            "pdf_footer": "GeneQuantify — Solo para investigación y educación. No validado para diagnóstico clínico.",
            "pdf_fig1": "Figura 1. Comparación fold-change: ΔΔCq clásico vs Pfaffl. Línea discontinua y=1 = sin cambio.",
            "pdf_fig2": "Figura 2. Valores p de todas las comparaciones. Barras rojas = significativo (p < 0,05).",
            "pdf_fig3": "Figura. Distribución ΔCq para {gene}. Puntos = réplicas; barras = medias de grupo.",
            "pdf_nochange": "Sin cambio",
            "pdf_stat_cols": ["Gen objetivo", "Comparación", "Tipo de test", "Método", "Valor p", "Significación"],
            "pdf_res_cols": ["Gen objetivo", "Grupo", "ΔCq Control", "ΔCq Muestra", "ΔΔCq", "2^(-ΔΔCq)", "Ratio Pfaffl", "Regulación", "E objetivo", "E ref"],
            "pdf_eff_cols": ["Gen", "E (objetivo)", "Eff% (objetivo)", "E (ref)", "Eff% (ref)", "Dif%", "Estado"],
            "pdf_eff_ok": "OK",
            "pdf_eff_warn": "ADVERTENCIA: usar Pfaffl",
            "pdf_outlier_col": "Valor atípico excluido",
            "pdf_contact": "Contacto: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} registros listos — puede generar el PDF.",
            # RDML / RDES import
            "rdml_expander":        "📂 Importar archivo RDML / RDES",
            "rdml_description":     "Cargue un archivo **RDML** (`.rdml`) o **RDES** (`.tsv`/`.csv`/`.txt`) para rellenar automáticamente los valores Cq.",
            "rdml_uploader":        "Seleccionar archivo",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX, Roche LightCycler, etc.  RDES: tabla separada por tabulaciones.",
            "rdml_success":         "✅ Archivo {fmt} cargado — {n} reacciones encontradas.",
            "rdml_error":           "❌ Error al analizar {fmt}: {err}",
            "rdml_preview":         "Vista previa de los datos analizados",
            "rdml_step1":           "**Paso 1 — Etiquete su grupo de control**",
            "rdml_ctrl_label":      "Nombre(s) de muestra de control (subcadenas separadas por comas)",
            "rdml_ctrl_help":       "Cualquier muestra cuyo nombre contenga este texto se tratará como Control.",
            "rdml_step2":           "**Paso 2 — Etiquete sus grupos de pacientes**",
            "rdml_n_pat":           "Número de grupos de pacientes",
            "rdml_pat_label":       "Nombre(s) de muestra del grupo de pacientes {i}",
            "rdml_pat_help":        "Subcadenas separadas por comas. Todas las muestras coincidentes se agruparán.",
            "rdml_apply":           "✅ Aplicar importación {fmt} a la entrada de datos",
            "rdml_apply_success":   "✅ {n} valores Cq cargados en la pestaña de entrada. ¡Revise y ajuste si es necesario!",
            "rdml_apply_warning":   "⚠️ No se mapearon valores. Compruebe que sus etiquetas coinciden con los nombres de muestra de la vista previa.",
        },

        "ar": {
            "title": "🧬 جين كوانتيفاي: تحليل التعبير الجيني وتغير عدد النسخ (CNV)",
            "tab_data": "إدخال البيانات",
            "tab_results": "النتائج",
            "tab_report": "التقرير",
            "subtitle": "تم تطويره بواسطة ب. يالجنكايا",
            "patient_data_header": "📊 إدخال بيانات مجموعة المرضى ومجموعة التحكم",
            "num_target_genes": "🔹 إدخال عدد الجينات المستهدفة",
            "num_patient_groups": "🔹 إدخال عدد مجموعات المرضى",
            "sample_number": "رقم العينة",
            "Grup": "مجموعة",
            "x_axis_title": "اسم المجموعة",
            "ct_value": "قيمة Cq",
            "reference_ct": "قيمة Ct المرجعية",
            "delta_ct_control": "ΔCq (التحكم)",
            "delta_ct_patient": "ΔCq (المريض)",
            "warning_empty_input": "⚠️ تحذير: أدخل البيانات واحدًا تلو الآخر أو انسخها دون خلايا فارغة من Excel.",
            "download_csv": "📥 تحميل CSV",
            "generate_pdf": "📥 إعداد تقرير PDF",
            "pdf_report": "تقرير تحليل التعبير الجيني",
            "nil_mine": "📊 النتائج",
            "gr_tbl": "📋 جدول بيانات الإدخال",
            "control_group": "🧬 مجموعة التحكم",
            "ctrl_trgt_ct": "🟦 قيم Ct الجين المستهدف {i} لمجموعة التحكم",
            "ctrl_ref_ct": "🟦 قيم Ct الجين المرجعي {i} لمجموعة التحكم",
            "hst_trgt_ct": "🩸 قيم Ct الجين المستهدف {j} لمجموعة المرضى",
            "hst_ref_ct": "🩸 قيم Ct الجين المرجعي {j} لمجموعة المرضى",
            "warning_control_ct": "⚠️ تحذير: يجب إدخال بيانات مجموعة التحكم {i} سطرًا بسطر أو نسخها من Excel دون خلايا فارغة.",
            "warning_patient_cq": "⚠️ تحذير: أدخل قيم Ct لمجموعة المرضى سطرًا بسطر أو انسخها من Excel دون خلايا فارغة.",
            "target_gene": "الجين المستهدف",
            "reference_gene": "الجين المرجعي",
            "target_ct": "قيمة Ct الجين المستهدف", 
            "distribution_graph": "رسم بياني للتوزيع",
            "error_missing_control_data": "⚠️ خطأ: بيانات مفقودة للجين المستهدف {i} في مجموعة التحكم!",
            "control_group_avg": "متوسط مجموعة التحكم",
            "avg": "متوسط",
            "control": "التحكم",
            "sample": "عينة",
            "patient": "مريض",
            "delta_ct_distribution": "توزيع ΔCq",
            "delta_ct_value": "قيمة ΔCq",
            "parametric": "معلمي",
            "non_parametric": "غير معلمي",
            "t_test": "اختبار t",
            "mann_whitney_u_test": "اختبار مان-ويتني U",
            "welch_t_test": "اختبار ويلش t",
            "significant": "مهم",
            "insignificant": "غير مهم",
            "test_type": "نوع الاختبار",
            "test_method": "طريقة الاختبار",
            "test_pvalue": "قيمة P للاختبار",
            "significance": "الدلالة",
            "delta_delta_ct": "ΔΔCq",
            "gene_expression_change": "تغيير التعبير الجيني (2^(-ΔΔCq))",
            "regulation_status": "حالة التنظيم",
            "no_change": "لا تغيير",
            "upregulated": "مرتفع التنظيم",
            "downregulated": "منخفض التنظيم",
            "report_title": "تقرير تحليل التعبير الجيني",
            "input_data_table": "جدول بيانات الإدخال",
            "results": "النتائج",
            "statistical_results": "📈 النتائج الإحصائية",
            "statistics": "النتائج الإحصائية",
            "statistical_evaluation": "التقييم الإحصائي",
            "target_gene": "الجين المستهدف",
            "patient_group": "🩸 مجموعة المرضى",
            "expression_change": "تغيير التعبير الجيني",
            "generate_pdf": "توليد تقرير PDF",
            "pdf_report": "تقرير التعبير الجيني",
            "error_no_data": "لم يتم العثور على بيانات، لم يتم إنشاء التقرير PDF.",
            # Efficiency translations
            "efficiency_header": "🔬 التحقق من كفاءة التضخيم",
            "efficiency_method": "طريقة إدخال الكفاءة",
            "efficiency_manual": "إدخال قيمة E يدويًا",
            "efficiency_slope": "الحساب من الانحدار",
            "efficiency_target_label": "كفاءة الجين المستهدف {i} (E)",
            "efficiency_ref_label": "كفاءة الجين المرجعي {i} (E)",
            "efficiency_target_slope_label": "انحدار الجين المستهدف {i}",
            "efficiency_ref_slope_label": "انحدار الجين المرجعي {i}",
            "efficiency_threshold": "عتبة فرق الكفاءة المقبول (%)",
            "efficiency_ok": "✅ فرق الكفاءة مقبول ({diff:.1f}%)",
            "efficiency_warning": "⚠️ فرق الكفاءة يتجاوز العتبة ({diff:.1f}%) — قد لا تكون طريقة ΔΔCq موثوقة!",
            "efficiency_target_pct": "كفاءة الجين المستهدف",
            "efficiency_ref_pct": "كفاءة الجين المرجعي",
            "efficiency_diff": "الفرق",
            "pfaffl_result": "نسبة Pfaffl",
            "pfaffl_header": "نتائج طريقة Pfaffl",
            "classic_ddct": "نتيجة ΔΔCq الكلاسيكية (2^(-ΔΔCq))",
            "pfaffl_ratio": "نسبة Pfaffl",
            "method_comparison": "📊 مقارنة الطرق",
            "efficiency_note": "ملاحظة: E=2.0 تمثل الكفاءة المثالية (100%). النطاق المقبول: 1.8–2.2 (90–110%)",
            "statistical_explanation": (
                "أثناء عملية التقييم الإحصائي، تم تحليل توزيع البيانات باستخدام اختبار شابيرو-ويلك. "
                "إذا تم تحقيق التوزيع الطبيعي، تم التحقق من تجانس التباين بين المجموعات باستخدام اختبار ليفين. "
                "إذا كانت التباين متساويًا، تم تطبيق اختبار t للعينة المستقلة، وإذا لم يكن كذلك، تم استخدام اختبار t ويلش. "
                "إذا لم يتم تحقيق التوزيع الطبيعي، تم تطبيق اختبار مان-ويتني U غير المعلمي. "
                "تم تحديد الدلالة باستخدام المعيار p < 0.05. "
                "للاقتراحات والدعم، بورهانيتين يالجنكايا - البريد الإلكتروني: mailtoburhanettin@gmail.com"
            ),
            "outlier_section_title": "### 🔍 إعدادات اكتشاف القيم الشاذة",
            "outlier_enable": "تفعيل اكتشاف القيم الشاذة",
            "outlier_enable_help": "يكتشف قيم Ct المتطرفة إحصائياً التي قد تعكس أخطاء تقنية.",
            "outlier_method_label": "طريقة الاكتشاف",
            "outlier_method_help": "Grubbs: للبيانات الموزعة طبيعياً. IQR: غير معلمي، قوي للتوزيعات غير المتماثلة.",
            "outlier_alpha_label": "مستوى الدلالة (α)",
            "outlier_alpha_help": "α = 0.05 هو المعيار. α أقل = أكثر تحفظاً.",
            "outlier_iqr_label": "مُضاعف IQR (k)",
            "outlier_iqr_help": "k=1.5 = حدود Tukey القياسية. k=3.0 = القيم الشاذة الشديدة فقط.",
            "outlier_expander": "ℹ️ حول اكتشاف القيم الشاذة في qPCR",
            "grubbs_info": "ℹ️ **متطلبات اختبار Grubbs:** الحد الأدنى **n ≥ 3** مكررات لكل مجموعة. عتبة الأهمية: **α = {alpha:.2f}**. يفترض الاختبار التوزيع الطبيعي؛ لـ n < 8، لا يمكن تقييم الطبيعية بشكل موثوق. يُنصح بتطبيق الاختبار على **قيم Cq الخام** (قبل التطبيع).",
            "outlier_excluded_no": "لا",
            "outlier_excluded_yes": "نعم",
            # Outlier stage selector
            "outlier_stage_label": "🔬 مرحلة اكتشاف القيم الشاذة",
            "outlier_stage_raw": "Ct الخام — قبل التطبيع (موصى به)",
            "outlier_stage_dct": "ΔCq — بعد التطبيع (السلوك السابق)",
            "outlier_stage_help": (
                "**Ct الخام (موصى به):** يتم اكتشاف القيم الشاذة على قيم Ct الخام قبل حساب ΔCq. "
                "يُطبَّق بشكل منفصل على الجين المستهدف وكل جين مرجعي.\n\n"
                "**ΔCq:** الاكتشاف بعد التطبيع (السلوك الأصلي)."
            ),
            # Distribution plot mode selector
            "dist_plot_mode_label": "📊 مخطط التوزيع — وضع العرض",
            "dist_plot_rq":   "RQ (2^-ΔCq)  — موصى به",
            "dist_plot_dct":  "ΔCq  — القيم المطبَّعة الخام",
            "dist_plot_ddct": "ΔΔCq  — بالنسبة لمتوسط المجموعة الضابطة",
            "dist_plot_help": (
                "**RQ (موصى به):** يحوّل ΔCq إلى 2^(-ΔCt). قيمة أعلى = تعبير أعلى.\n\n"
                "**ΔCq:** قيم لوغاريتمية خام. مفيد للتحقق من توزيع البيانات.\n\n"
                "**ΔΔCq:** ΔCq لكل عينة ناقص متوسط مجموعة التحكم."
            ),
            "unequal_n_warning": (
                "⚠️ **تم اكتشاف أعداد متكررة غير متساوية — {group}:**  \n"
                "{details}  \nسيستمر التحليل باستخدام **أقصر طول مشترك (n={min_n})**.  \n"
                "يرجى التحقق من بيانات الإدخال."
            ),
            # Sidebar
            "sidebar_load_example": "📂 تحميل البيانات النموذجية",
            "sidebar_example_loaded": "✅ تم تحميل البيانات النموذجية! انتقل إلى تبويب إدخال البيانات.",
            "sidebar_desktop_title": "### 💻 تطبيق سطح المكتب",
            "sidebar_desktop_btn": "⬇️ تنزيل تطبيق سطح المكتب",
            "sidebar_opensource_title": "### 🔓 مفتوح المصدر",
            "sidebar_opensource_body": "GeneQuantify مفتوح المصدر (GPL-3.0).  \nالكود المصدري متاح على GitHub:",
            "sidebar_github_btn": "⭐ عرض الكود المصدري على GitHub",
            "sidebar_scenarios_title": "📋 تحميل سيناريو التحقق",
            "sidebar_scenario_select": "اختر سيناريو",
            "sidebar_load_scenario_btn": "▶ تحميل السيناريو",
            "sidebar_scenario_loaded": "✅ تم تحميل {s}! انتقل إلى تبويب إدخال البيانات.",
            "outlier_excluded_no": "لا",
            "outlier_excluded_yes": "نعم",
            "stat_decision_title": "🔬 القرار الإحصائي",
            "stat_decision_steps": "**اختيار الاختبار خطوة بخطوة:**",
            "stat_shapiro_title": "**1. اختبار شابيرو-ويلك للتوزيع الطبيعي**",
            "stat_normal": "طبيعي",
            "stat_nonnormal": "غير طبيعي",
            "stat_levene_title": "**2. اختبار ليفين لتجانس التباين**",
            "stat_levene_skipped": "**2. اختبار ليفين** — *تم تخطيه* (لم يتحقق التوزيع الطبيعي؛ سيُستخدم اختبار غير معلمي)",
            "stat_equal_var": "تبايانات متساوية",
            "stat_unequal_var": "تبايانات غير متساوية",
            "stat_selected_test": "**3. الاختبار المختار:**",
            "stat_reason": "**السبب:**",
            "stat_result": "**النتيجة:**",
            "stat_reason_nonnormal": "توزيع غير طبيعي في مجموعة واحدة أو كلتيهما",
            "stat_reason_normal_equal": "كلا المجموعتين طبيعيتان + تبايانات متساوية",
            "stat_reason_normal_unequal": "كلا المجموعتين طبيعيتان + تبايانات غير متساوية (Levene p < 0.05)",
            "stat_multigroup_note": "⚠️ ملاحظة: مع ≥ 3 مجموعات، راجع قسم **المقارنة متعددة المجموعات** أدناه لاختبار ANOVA / كروسكال-واليس.",
            "multigroup_title": "## 📊 تحليل مقارنة متعددة المجموعات",
            "multigroup_expander": "ℹ️ حول التحليل الإحصائي متعدد المجموعات",
            "multigroup_omnibus_test": "اختبار شامل",
            "multigroup_pvalue": "قيمة p",
            "multigroup_result": "النتيجة",
            "multigroup_significant": "دال",
            "multigroup_not_significant": "غير دال",
            "multigroup_omnibus_ns": "ℹ️ الاختبار الشامل **غير دال** (p ≥ 0.05). المقارنات البعدية معروضة للاطلاع فقط.",
            "multigroup_posthoc_label": "**ما بعد الاختبار:**",
            "multigroup_dl_button": "📥 تحميل نتائج ما بعد الاختبار —",
            "multigroup_2group_note": "ℹ️ **تحليل متعدد المجموعات غير قابل للتطبيق:** تم الكشف عن مجموعتين فقط (مجموعة التحكم + مجموعة مرضى واحدة).",
            "multigroup_decision_normal_equal": "✅ توزيع طبيعي + تبايانات متساوية → **ANOVA أحادي الاتجاه + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ توزيع طبيعي + **تبايانات غير متساوية** → **Welch ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ **توزيع غير طبيعي** → **كروسكال-واليس + Dunn**",
            "multigene_title": "### 🧬 تصحيح المقارنات المتعددة متعدد الجينات",
            "multigene_expander": "ℹ️ لماذا هذا ضروري؟",
            "multigene_sig_raw": "دال (خام)",
            "multigene_sig_bonf": "دال (بونفيروني)",
            "multigene_sig_fdr": "دال (FDR B-H)",
            "multigene_warning": "⚠️ بعد التصحيح، {lost} نتيجة لم تعد دالة بعد تعديل FDR. أبلغ عن قيم p المصححة كنتائج رئيسية.",
            "multigene_success": "✅ جميع {n} النتائج الدالة لا تزال دالة بعد تصحيح FDR.",
            "multigene_no_sig": "لم يتم اكتشاف نتائج زوجية دالة (p خام < 0.05).",
            "multigene_dl_button": "📥 تحميل قيم p المصححة (CSV)",
            "multigene_chart_title": "تصحيح قيمة p متعدد الجينات: خام / بونفيروني / FDR",
            "multigene_fc_chart_title": "مقارنة التعبير الجيني المتعدد",
            "multigene_1gene_note": "ℹ️ **تصحيح متعدد الجينات:** تم تحليل جين مستهدف واحد فقط — التصحيح غير قابل للتطبيق.",
            "multigene_no_data": "لا توجد قيم p بعد — أدخل البيانات أعلاه.",
            "ref_gene_section_title": "### 📚 إعدادات الجين المرجعي",
            "ref_gene_num_label": "عدد الجينات المرجعية لكل جين مستهدف",
            "ref_gene_num_help": "توصي إرشادات MIQE بـ ≥2 جين مرجعي معتمد لتطبيع قوي.",
            "ref_gene_1_warning": "⚠️ **ملاحظة منهجية:** استخدام جين مرجعي واحد يحد من متانة التطبيع. توصي إرشادات MIQE (Bustin et al. 2009) بـ **≥2 جين مرجعي معتمد** مع تقييم الاستقرار (geNorm/NormFinder).",
            "ref_gene_multi_success": "✅ تم اختيار {n} جينات مرجعية. سيتم حساب التطبيع بالوسط الهندسي وقيمة M لـ geNorm تلقائياً.",
            "ref_gene_expander": "ℹ️ حول التطبيع متعدد المراجع",
            "sc_expander": "📐 حاسبة المنحنى المعياري — احسب E من سلسلة التخفيف",
            "sc_gene_label": "الجين / تسمية البادئ",
            "sc_num_points": "عدد نقاط التخفيف",
            "sc_dilution_factor_label": "**عامل التخفيف** (مثال: 10 للتخفيف العشري)",
            "sc_dilution_factor_input": "عامل التخفيف",
            "sc_start_conc_label": "**التركيز الابتدائي** (وحدات اعتباطية، مثال: 1)",
            "sc_start_conc_input": "التركيز الابتدائي",
            "sc_enter_ct": "**أدخل متوسط Cq لكل تخفيف:**",
            "sc_calc_button": "📊 احسب الكفاءة",
            "sc_slope": "الميل",
            "sc_e_value": "قيمة E",
            "sc_efficiency_pct": "الكفاءة %",
            "sc_excellent": "✅ ممتاز! E={e:.4f} ({pct:.1f}%)، R²={r2:.4f} — استخدم هذه القيمة في قسم الكفاءة أدناه.",
            "sc_warning_r2": "⚠️ E مقبولة ({pct:.1f}%) لكن R²={r2:.4f} < 0.99 — تحقق من سلسلة التخفيف.",
            "sc_error_range": "❌ E={e:.4f} ({pct:.1f}%) خارج النطاق المقبول (90–110%). راجع تصميم البادئ أو سلسلة التخفيف.",
            "sc_chart_title": "المنحنى المعياري — {label} | E={e:.4f} ({pct:.1f}%)، R²={r2:.4f}",
            "sc_xaxis": "log₁₀(التركيز)",
            "sc_data_points": "نقاط البيانات",
            "sc_copy_hint": "💡 انسخ الميل **{slope:.4f}** أو قيمة E **{e:.4f}** في حقول الكفاءة أدناه.",
            "sc_description": """\
    أدخل قيم Ct لتخفيفاتك التسلسلية أدناه. سيطبق الحاسب انحداراً خطياً ويحسب الميل وR² وكفاءة التضخيم تلقائياً.

    **كيفية الاستخدام:**  
    1. قم بتشغيل qPCR على تخفيفات تسلسلية (مثل غير مخفف، 1:10، 1:100، 1:1000، 1:10000)  
    2. أدخل متوسط Ct لكل تخفيف  
    3. اقرأ الميل وE وR²  
    """,
            "ref_multi_description": """\
    **التطبيع بالوسط الهندسي** (Vandesompele et al. 2002)  
    عامل التطبيع (NF) هو المتوسط الحسابي لقيم Ct عبر جميع الجينات المرجعية لكل عينة،  
    وهو ما يتوافق مع الوسط الهندسي لمستويات تعبيرها.  
    `NF_عينة = متوسط(Ct_ref1, Ct_ref2, ..., Ct_refN)` لكل عينة  
    `ΔCq = Ct_المستهدف − NF`

    **قيمة M لـ geNorm** (درجة الاستقرار)  
    لكل جين مرجعي، M = متوسط الانحراف المعياري للنسب اللوغاريتمية مقابل جميع الجينات المرجعية الأخرى.  
    **M أقل = أكثر استقراراً.** العتبة الموصى بها من MIQE: M < 0.5 (صارم) أو M < 1.0 (مقبول).

    **CV (معامل الاختلاف)**  
    `CV = (الانحراف المعياري / المتوسط) × 100%` لقيم Ct الخام عبر جميع العينات.  
    CV أقل يشير إلى تباين أقل واستقرار أفضل كمرجع.

    **مرجع:** Vandesompele J et al. *Genome Biology* 2002; Bustin SA et al. *Clin Chem* 2009 (MIQE).
    """,
            "outlier_description": """\
    **لماذا يهم اكتشاف القيم الشاذة في qPCR**

    التباين التقني متأصل في qPCR: أخطاء السحب بالماصة، وتكوين الفقاعات، وانتقال المثبطات، أو تباين جودة RNA يمكن أن تنتج قيم Ct غير متسقة إحصائياً مع بقية مجموعة الطبعات.  
    تضمين هذه القيم يضخم التباين، ويحيز المتوسطات، ويمكن أن يؤدي إلى استنتاجات خاطئة — خاصة في مجموعات البيانات السريرية ذات الأحجام الصغيرة.

    **عندما تصبح هذه القيود حرجة:**
    - مجموعات صغيرة (n < 5): Ct واحد خاطئ يزيح المتوسط بشكل كبير
    - تباين بيولوجي عالٍ (مثل عدم تجانس الورم، الدراسات السريرية)
    - طبعات ثلاثية تقنية حيث تنحرف طبعة واحدة > 0.5 Ct عن الأخريات
    - أهداف منخفضة الوفرة مع Ct > 35، حيث يسود الضوضاء

    **اختبار Grubbs** *(Grubbs 1969)*  
    يفترض التوزيع الطبيعي. يختبر ما إذا كانت القيمة الأكثر تطرفاً تمثل قيمة شاذة ذات دلالة إحصائية (p < α). يتكرر حتى لا يجد المزيد من القيم الشاذة.  
    الأفضل لـ: قيم Ct المكررة من مجموعة تجريبية واحدة.

    **طريقة IQR** *(Tukey 1977)*  
    غير معلمية. تعلم القيم خارج Q1 − k×IQR أو Q3 + k×IQR.  
    الأفضل لـ: مجموعات أكبر أو توزيعات غير طبيعية.

    **مهم:** يتطلب استبعاد القيم الشاذة **مبرراً بيولوجياً أو تقنياً**.  
    تُعلم هذه الأداة المرشحين — القرار النهائي يعود دائماً للباحث.  
    يتم تسجيل جميع الاستبعادات والإبلاغ عنها في تقرير PDF.

    **المراجع:** Grubbs FE. *Technometrics* 1969; Tukey JW. *Exploratory Data Analysis* 1977;  
    Bustin SA et al. *Clin Chem* 2009 (إرشادات MIQE).
    """,

            # ── سلاسل تقرير PDF ───────────────────────────────────────────────────
            "pdf_cover_subtitle": "تقرير تحليل التعبير الجيني بـ qPCR",
            "pdf_generated": "تم الإنشاء: {now}",
            "pdf_s1_title": "1. الطرق وإعدادات التحليل",
            "pdf_s1_calc": "1.1 طرق الحساب",
            "pdf_s1_calc_body": "طُبِّقت طريقتان متكاملتان لحساب نسبة التضخيم:",
            "pdf_s1_classic": "طريقة ΔΔCq الكلاسيكية: نسبة التضخيم = 2^(-ΔΔCq). تفترض كفاءة متساوية.",
            "pdf_s1_pfaffl": "طريقة Pfaffl: النسبة = (E_الهدف ^ ΔCq_الهدف) / (E_مرجع ^ ΔCt_مرجع). موصى بها عند اختلاف الكفاءة > 10%.",
            "pdf_s1_norm": "1.2 التطبيع",
            "pdf_s1_norm_multi": "استُخدمت جينات مرجعية متعددة (n={n}) (geNorm, Vandesompele et al. 2002).",
            "pdf_s1_norm_single": "استُخدم جين مرجعي واحد. توصي MIQE باستخدام ≥2 جين.",
            "pdf_s1_eff": "1.3 كفاءة التضخيم",
            "pdf_s1_eff_range": "النطاق المقبول: E = 1.8-2.2 (90-110%). عتبة الفارق: {thr}%.",
            "pdf_s1_outlier": "1.4 اكتشاف القيم الشاذة",
            "pdf_s1_grubbs": "اختبار Grubbs (1969) عند alpha = {alpha}. {n} عينة مستبعدة.",
            "pdf_s1_iqr": "طريقة IQR (Tukey 1977) بمعامل k = {k}. {n} عينة مستبعدة.",
            "pdf_s1_outlier_warn": "تحذير: يستلزم الاستبعاد مبرراً بيولوجياً أو تقنياً.",
            "pdf_s1_outlier_off": "تم تعطيل اكتشاف القيم الشاذة.",
            "pdf_s2_title": "2. بيانات الإدخال",
            "pdf_s2_body": "قيم Ct الخام بعد معالجة القيم الشاذة.",
            "pdf_s3_title": "3. نتائج التعبير الجيني",
            "pdf_s3_body": "نسب التضخيم المحسوبة بطريقتي ΔΔCq الكلاسيكية و Pfaffl.",
            "pdf_s4_title": "4. التحليل الإحصائي",
            "pdf_s4_body": "الدلالة الإحصائية. اختيار الاختبار تلقائياً (Shapiro-Wilk، Levene). p < 0.05.",
            "pdf_s4_interp": "تفسير الاختبارات",
            "pdf_s4_interp_body": "t للطلاب: متساويا التباين. Welch: غير متساويا التباين. Mann-Whitney U: لامعلمي.",
            "pdf_s5_title": "5. مخططات توزيع Delta Ct",
            "pdf_s5_body": "توزيع قيم ΔCq. كل نقطة = مكرر. الأشرطة = المتوسطات.",
            "pdf_s6_title": "6. تفسير النتائج",
            "pdf_s6_fc": "6.1 تفسير نسبة التضخيم",
            "pdf_s6_choose": "6.2 الاختيار بين ΔΔCq و Pfaffl",
            "pdf_s6_choose_body": "ΔΔCq إذا كانت الكفاءتان 90-110% والفارق < 10%. Pfaffl إذا كان > 10%.",
            "pdf_s6_stat": "6.3 مبررات اختيار الاختبار",
            "pdf_s6_stat_body": "التوزيع الطبيعي: Shapiro-Wilk. تجانس التباين: Levene. Student/Welch/Mann-Whitney حسب النتيجة.",
            "pdf_s7_title": "7. المراجع",
            "pdf_fc_interp_header": ["نسبة التضخيم", "ΔΔCq", "التفسير", "الأهمية البيولوجية"],
            "pdf_fc_interp_rows": [
                [">2.0", "<-1.0", "زيادة تعبير قوية", "مهم بيولوجياً"],
                ["1.5-2.0", "-1.0 إلى -0.58", "زيادة معتدلة", "قد يكون مهماً"],
                ["1.0-1.5", "-0.58 إلى 0", "زيادة طفيفة", "غير مهم منفرداً"],
                ["1.0", "0", "لا تغيير", "لا تعبير تفاضلي"],
                ["0.67-1.0", "0 إلى 0.58", "انخفاض طفيف", "غير مهم منفرداً"],
                ["0.5-0.67", "0.58 إلى 1.0", "انخفاض معتدل", "قد يكون مهماً"],
                ["<0.5", ">1.0", "انخفاض قوي", "مهم بيولوجياً"],
            ],
            "pdf_stat_note": "ملاحظة: يجب تقييم الدلالة الإحصائية والبيولوجية معاً.",
            "pdf_summary_param": "المعلمة", "pdf_summary_val": "القيمة",
            "pdf_summary_genes": "الجينات الهدف", "pdf_summary_groups": "مجموعات المرضى",
            "pdf_summary_samples": "إجمالي العينات", "pdf_summary_excluded": "العينات المستبعدة",
            "pdf_summary_tests": "مقارنات", "pdf_summary_norm": "طريقة التطبيع",
            "pdf_summary_norm_multi": "geNorm NF", "pdf_summary_norm_single": "جين مرجعي واحد",
            "pdf_summary_methods": "طرق الحساب", "pdf_summary_methods_val": "ΔΔCq الكلاسيكي + Pfaffl",
            "pdf_disclaimer": "تم إنشاء هذا التقرير تلقائياً بواسطة GeneQuantify وفق إرشادات MIQE.",
            "pdf_footer": "GeneQuantify — للبحث والتعليم فقط. غير مُصادَق لأغراض التشخيص السريري.",
            "pdf_fig1": "شكل 1. مقارنة نسبة التضخيم. الخط المتقطع y=1 = لا تغيير.",
            "pdf_fig2": "شكل 2. قيم p. الأشرطة الحمراء = دالة (p < 0.05).",
            "pdf_fig3": "شكل. توزيع ΔCq لـ {gene}.",
            "pdf_nochange": "لا تغيير",
            "pdf_stat_cols": ["الجين الهدف", "المقارنة", "نوع الاختبار", "الاختبار", "قيمة p", "الدلالة"],
            "pdf_res_cols": ["الجين الهدف", "المجموعة", "ΔCq الضبط", "ΔCq العينة", "ΔΔCq", "2^(-ΔΔCq)", "نسبة Pfaffl", "التنظيم", "E الهدف", "E المرجع"],
            "pdf_eff_cols": ["الجين", "E (الهدف)", "Eff% (الهدف)", "E (المرجع)", "Eff% (المرجع)", "الفارق%", "الحالة"],
            "pdf_eff_ok": "مقبول", "pdf_eff_warn": "تحذير: استخدم Pfaffl",
            "pdf_outlier_col": "قيمة شاذة مستبعدة", "pdf_contact": "التواصل: mailtoburhanettin@gmail.com",
            "pdf_ready": "{n} سجلات جاهزة — يمكنك إنشاء تقرير PDF.",
            # RDML / RDES import
            "rdml_expander":        "📂 استيراد ملف RDML / RDES",
            "rdml_description":     "ارفع ملف **RDML** (`.rdml`) أو **RDES** (`.tsv`/`.csv`/`.txt`) لملء قيم Cq تلقائيًا.",
            "rdml_uploader":        "اختر ملفًا",
            "rdml_uploader_help":   "RDML: Bio-Rad CFX، Roche LightCycler، إلخ.  RDES: جدول مفصول بعلامات تبويب.",
            "rdml_success":         "✅ تم تحميل ملف {fmt} — تم العثور على {n} تفاعل.",
            "rdml_error":           "❌ خطأ في تحليل {fmt}: {err}",
            "rdml_preview":         "معاينة البيانات المحللة",
            "rdml_step1":           "**الخطوة 1 — حدد مجموعة التحكم**",
            "rdml_ctrl_label":      "اسم (أسماء) عينة التحكم (سلاسل فرعية مفصولة بفواصل)",
            "rdml_ctrl_help":       "أي عينة يحتوي اسمها على هذا النص ستُعامَل كمجموعة تحكم.",
            "rdml_step2":           "**الخطوة 2 — حدد مجموعات المرضى**",
            "rdml_n_pat":           "عدد مجموعات المرضى",
            "rdml_pat_label":       "اسم (أسماء) عينات مجموعة المرضى {i}",
            "rdml_pat_help":        "سلاسل فرعية مفصولة بفواصل. سيتم تجميع جميع العينات المطابقة في هذه المجموعة.",
            "rdml_apply":           "✅ تطبيق استيراد {fmt} على إدخال البيانات",
            "rdml_apply_success":   "✅ تم تحميل {n} قيمة Cq في تبويب إدخال البيانات! انتقل إليه للمراجعة والتعديل.",
            "rdml_apply_warning":   "⚠️ لم يتم تعيين أي قيم. تحقق من أن التسميات تتطابق مع أسماء العينات في المعاينة.",
        }
    }

                  
    # ═══════════════════════════════════════════════════════════════════════════════
    # RDML / RDES SIDEBAR
    # ═══════════════════════════════════════════════════════════════════════════════
    _t = translations[language_code]

    # ──Scenario-based example data loader ────────
    st.sidebar.markdown("---")
    st.sidebar.markdown(f"### {_t.get('sidebar_scenarios_title', '📋 Load Validation Scenario')}")

    # Scenario definitions — all 7 validation datasets from Supplementary Data S1
    SCENARIOS = {
        "S1 — Basic (n=3, t-test)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description": "1 gene, 1 group, n=3. Strong upregulation (~4.1x). n<8 → t-test assumed.",
            "control_target_ct_0": "23.15\n22.98\n23.42",
            "control_reference_ct_0_0": "18.22\n18.05\n18.38",
            "sample_target_ct_0_0": "21.05\n20.88\n21.23",
            "sample_reference_ct_0_0_0": "18.15\n17.98\n18.28",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
        "S2 — Multi-gene + dual ref (n=4)": {
            "gene_count": 3, "patient_count": 2, "num_ref_genes": 2,
            "outlier_method": "IQR", "outlier_enabled": True,
            "description": "3 genes, 2 groups, n=4. Dual reference (geNorm). Gene 2: Pfaffl recommended (E diff >10%).",
            # Gene 1 control
            "control_target_ct_0": "23.88\n24.12\n23.95\n24.32",
            "control_reference_ct_0_0": "18.32\n18.55\n18.44\n18.68",
            "control_reference_ct_0_1": "19.95\n20.18\n20.05\n20.28",
            # Gene 2 control
            "control_target_ct_1": "19.05\n19.28\n19.15\n19.38",
            "control_reference_ct_1_0": "18.32\n18.55\n18.44\n18.68",
            "control_reference_ct_1_1": "19.95\n20.18\n20.05\n20.28",
            # Gene 3 control
            "control_target_ct_2": "24.92\n25.15\n25.02\n25.28",
            "control_reference_ct_2_0": "18.32\n18.55\n18.44\n18.68",
            "control_reference_ct_2_1": "19.95\n20.18\n20.05\n20.28",
            # Gene 1 Group 1
            "sample_target_ct_0_0": "21.50\n21.82\n21.65\n21.95",
            "sample_reference_ct_0_0_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_0_0_1": "19.98\n20.22\n20.08\n20.30",
            # Gene 1 Group 2
            "sample_target_ct_0_1": "22.45\n22.68\n22.55\n22.78",
            "sample_reference_ct_0_1_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_0_1_1": "19.98\n20.22\n20.08\n20.30",
            # Gene 2 Group 1
            "sample_target_ct_1_0": "22.10\n22.38\n22.22\n22.48",
            "sample_reference_ct_1_0_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_1_0_1": "19.98\n20.22\n20.08\n20.30",
            # Gene 2 Group 2
            "sample_target_ct_1_1": "23.75\n24.02\n23.88\n24.15",
            "sample_reference_ct_1_1_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_1_1_1": "19.98\n20.22\n20.08\n20.30",
            # Gene 3 Group 1
            "sample_target_ct_2_0": "25.05\n25.28\n25.15\n25.38",
            "sample_reference_ct_2_0_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_2_0_1": "19.98\n20.22\n20.08\n20.30",
            # Gene 3 Group 2
            "sample_target_ct_2_1": "25.02\n25.25\n25.12\n25.35",
            "sample_reference_ct_2_1_0": "18.35\n18.58\n18.47\n18.70",
            "sample_reference_ct_2_1_1": "19.98\n20.22\n20.08\n20.30",
            "target_E_0": 2.0, "ref_E_0": 2.0,
            "target_E_1": 2.103, "ref_E_1": 1.952,
            "target_E_2": 1.99, "ref_E_2": 2.0,
        },
        "S3 — Outlier detection (Grubbs, n=6)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True, "grubbs_alpha": 0.05,
            "description": "1 gene, 1 group, n=6. Sample 5 is an outlier (Cq=27.82). Grubbs on raw Cq detects it.",
            "control_target_ct_0": "23.12\n22.95\n23.38\n23.05\n27.82\n23.22",
            "control_reference_ct_0_0": "18.15\n17.98\n18.32\n18.08\n18.22\n18.12",
            "sample_target_ct_0_0": "21.05\n20.88\n21.23\n20.95\n21.15\n20.72",
            "sample_reference_ct_0_0_0": "18.15\n17.98\n18.28\n18.05\n18.20\n17.88",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
        "S4 — ANOVA 3 groups (n=5)": {
            "gene_count": 1, "patient_count": 3, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description": "1 gene, 3 groups, n=5. One-way ANOVA + Tukey HSD. Group 3 = no change.",
            "control_target_ct_0": "24.12\n23.95\n24.38\n24.05\n24.22",
            "control_reference_ct_0_0": "18.50\n18.38\n18.72\n18.42\n18.55",
            # Group 1: mild upregulation
            "sample_target_ct_0_0": "23.28\n23.05\n23.52\n23.18\n23.38",
            "sample_reference_ct_0_0_0": "18.52\n18.40\n18.75\n18.45\n18.58",
            # Group 2: strong upregulation
            "sample_target_ct_0_1": "21.05\n20.88\n21.23\n20.95\n21.15",
            "sample_reference_ct_0_1_0": "18.52\n18.40\n18.75\n18.45\n18.58",
            # Group 3: no change
            "sample_target_ct_0_2": "24.10\n23.92\n24.35\n24.02\n24.18",
            "sample_reference_ct_0_2_0": "18.52\n18.40\n18.75\n18.45\n18.58",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
        "S1B — Student t-test (n=8, equal var)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description": "n=8, normal distribution, equal variance → Student t-test. FC ≈ 4.15.",
            "control_target_ct_0": "23.15\n22.98\n23.42\n23.05\n23.28\n22.88\n23.52\n23.18",
            "control_reference_ct_0_0": "18.22\n18.05\n18.38\n18.12\n18.30\n17.95\n18.45\n18.18",
            "sample_target_ct_0_0": "21.05\n20.88\n21.23\n20.95\n21.15\n20.72\n21.38\n21.02",
            "sample_reference_ct_0_0_0": "18.15\n17.98\n18.28\n18.05\n18.20\n17.88\n18.35\n18.10",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
        "S1C — Welch t-test (n=8, unequal var)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description": "n=8, normal distribution, unequal variance → Welch t-test. FC ≈ 3.86.",
            "control_target_ct_0": "23.15\n22.98\n23.42\n23.05\n23.28\n22.88\n23.52\n23.18",
            "control_reference_ct_0_0": "18.22\n18.05\n18.38\n18.12\n18.30\n17.95\n18.45\n18.18",
            "sample_target_ct_0_0": "21.20\n20.50\n21.80\n20.85\n21.55\n20.65\n21.90\n20.75",
            "sample_reference_ct_0_0_0": "18.15\n17.98\n18.28\n18.05\n18.20\n17.88\n18.35\n18.10",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
        "S1D — Mann-Whitney U (n=8, bimodal)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description": "n=8, bimodal (responder/non-responder) → Mann-Whitney U. FC ≈ 12.43.",
            "control_target_ct_0": "23.15\n22.98\n23.42\n23.05\n23.28\n22.88\n23.52\n23.18",
            "control_reference_ct_0_0": "18.22\n18.05\n18.38\n18.12\n18.30\n17.95\n18.45\n18.18",
            "sample_target_ct_0_0": "20.05\n19.88\n20.23\n17.50\n20.15\n19.72\n20.38\n17.80",
            "sample_reference_ct_0_0_0": "18.15\n17.98\n18.28\n18.05\n18.20\n17.88\n18.35\n18.10",
            "target_E_0": 2.0, "ref_E_0": 2.0,
        },
    }

    st.sidebar.markdown(
        "<div style='font-size:12px;font-weight:600;color:#1a237e;margin-bottom:2px;'>"
        f"📋 {_t.get('sidebar_scenarios_title','Validation Scenarios')}</div>",
        unsafe_allow_html=True
    )
    selected_scenario = st.sidebar.selectbox(
        _t.get('sidebar_scenario_select', 'Select scenario'),
        options=["—"] + list(SCENARIOS.keys()),
        key="scenario_selector",
        label_visibility="collapsed"
    )

    if selected_scenario != "—":
        sc = SCENARIOS[selected_scenario]
        st.sidebar.caption(sc.get("description", ""))
        if st.sidebar.button(_t.get('sidebar_load_scenario_btn', '▶ Load Scenario'), key="load_scenario_btn", use_container_width=True):
            for key, val in sc.items():
                if key == "description":
                    continue
                st.session_state[key] = val
            st.sidebar.success(_t.get('sidebar_scenario_loaded', f'✅ {selected_scenario} loaded!').format(s=selected_scenario))

    st.sidebar.divider()
    with st.sidebar.expander(_t.get("rdml_expander", "📂 Import RDML / RDES"), expanded=False):
        st.markdown(_t.get("rdml_description", "Upload an RDML or RDES file to auto-fill Cq values."))
        imported_file = st.file_uploader(
            _t.get("rdml_uploader", "Choose file"),
            type=["rdml", "tsv", "csv", "txt"],
            key="rdml_rdes_uploader",
            help=_t.get("rdml_uploader_help", "RDML: Bio-Rad CFX, Roche LightCycler, etc.  RDES: tab-separated format."),
        )
        if imported_file is not None:
            file_bytes = imported_file.read()
            fname = imported_file.name.lower()
            if fname.endswith(".rdml"):
                import_df, import_err = parse_rdml(file_bytes)
                fmt_label = "RDML"
            else:
                import_df, import_err = parse_rdes(file_bytes)
                fmt_label = "RDES"
            if import_err:
                st.error(_t.get("rdml_error", "❌ {fmt} parse error: {err}").format(fmt=fmt_label, err=import_err))
                import_df = None
            if import_df is not None:
                st.success(_t.get("rdml_success", "✅ {fmt} file loaded — {n} reactions found.").format(fmt=fmt_label, n=len(import_df)))
                with st.expander(_t.get("rdml_preview", "Preview parsed data"), expanded=False):
                    st.dataframe(import_df, use_container_width=True)
                all_samples = sorted(import_df["Sample"].unique())
                st.markdown(_t.get("rdml_step1", "**Step 1 — Label your Control group**"))
                ctrl_label = st.text_input(
                    _t.get("rdml_ctrl_label", "Control sample name(s) (comma-separated substrings)"),
                    value=", ".join([s for s in all_samples[:1]]),
                    key="rdml_ctrl_label_input",
                    help=_t.get("rdml_ctrl_help", "Any sample whose name contains this text will be treated as Control.")
                )
                st.markdown(_t.get("rdml_step2", "**Step 2 — Label your Patient groups**"))
                n_pat_grps = st.number_input(
                    _t.get("rdml_n_pat", "Number of patient groups"),
                    min_value=1, max_value=10, value=1, step=1, key="rdml_n_pat"
                )
                patient_labels = []
                for pg in range(int(n_pat_grps)):
                    default_pat = all_samples[pg + 1] if pg + 1 < len(all_samples) else ""
                    pat_lbl = st.text_input(
                        _t.get("rdml_pat_label", "Patient group {i} sample name(s)").format(i=pg+1),
                        value=default_pat,
                        key=f"rdml_pat_{pg}",
                        help=_t.get("rdml_pat_help", "Comma-separated substrings.")
                    )
                    patient_labels.append(pat_lbl)
                if st.button(_t.get("rdml_apply", "✅ Apply {fmt} import to Data Entry").format(fmt=fmt_label), key="rdml_apply_btn"):
                    n_filled = apply_import_to_session(import_df, ctrl_label, patient_labels)
                    if n_filled > 0:
                        st.success(_t.get("rdml_apply_success", "✅ {n} Cq values loaded!").format(n=n_filled))
                    else:
                        st.warning(_t.get("rdml_apply_warning", "⚠️ No values were mapped. Check your labels."))

    # ═══════════════════════════════════════════════════════════════════════════════
    # ANA BAŞLIK
    # ═══════════════════════════════════════════════════════════════════════════════
    _title_txt = _t.get('title', '🧬 GeneQuantify')
    _sub_txt   = _t.get('subtitle', 'Developed by B. Yalçınkaya')
    _title_parts = _title_txt.split(' ', 1)
    _header_emoji = _title_parts[0]
    _header_title_text = _title_parts[1] if len(_title_parts) > 1 else _title_txt
    st.markdown(
        f"""
        <div style="background:linear-gradient(90deg,#1a237e,#3949ab);
                    color:white;padding:16px 18px;border-radius:8px;margin-bottom:8px;
                    display:flex;align-items:center;gap:12px;">
            <span style="font-size:28px;line-height:1;flex-shrink:0;">{_header_emoji}</span>
            <div style="display:flex;flex-direction:column;justify-content:center;">
                <span style="font-size:20px;font-weight:800;line-height:1.3;">{_header_title_text}</span>
                <span style="font-size:11px;opacity:0.75;margin-top:3px;line-height:1.3;">{_sub_txt}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    tab_data, tab_results, tab_report = st.tabs([
        f"📥 {_t.get('tab_data', 'Veri Girişi')}",
        f"📊 {_t.get('tab_results', 'Sonuçlar')}",
        f"📄 {_t.get('tab_report', 'Rapor')}",
    ])

    # ─── HELPER FUNCTIONS ─────────────────────────────────────────────────────────
    def parse_input_data(input_data):
        values = [x.replace(",", ".").strip() for x in input_data.split() if x.strip()]
        return np.array([float(x) for x in values if x])


    def apply_import_to_session(df, ctrl_sample_label, patient_labels):
        """
        Given a parsed DataFrame (columns: Sample, SampleType, Target, TargetType, Cq),
        fill st.session_state keys that GeneQuantify's text_area widgets read from.

        Mapping logic:
          - Rows where SampleType in ('ntc','nac','ntp','std','opt') → skipped
          - ctrl_sample_label: sample name(s) that belong to the control group
            (comma-separated string; matches by substring if needed)
          - patient_labels: list of sample names for each patient group
          - TargetType == 'ref' → reference gene; 'toi' → target gene
        """
        if df is None:
            return 0

        # Clean: drop rows without a Cq value or with Cq == -1
        df = df.dropna(subset=["Cq"]).copy()
        df = df[df["Cq"] != -1.0]

        ctrl_keywords = [s.strip() for s in ctrl_sample_label.split(",") if s.strip()]

        def is_ctrl(sample_name):
            return any(kw.lower() in sample_name.lower() for kw in ctrl_keywords)

        # Separate target genes and reference genes
        targets = sorted(df[df["TargetType"] == "toi"]["Target"].unique())
        refs    = sorted(df[df["TargetType"] == "ref"]["Target"].unique())

        count = 0
        for gene_i, target_name in enumerate(targets):
            tgt_df = df[(df["Target"] == target_name) & (df["TargetType"] == "toi")]

            # Control group — target gene
            ctrl_cqs = tgt_df[tgt_df["Sample"].apply(is_ctrl)]["Cq"].dropna().tolist()
            if ctrl_cqs:
                st.session_state[f"control_target_ct_{gene_i}"] = "\n".join(f"{v:.4f}" for v in ctrl_cqs)
                count += len(ctrl_cqs)

            # Control group — reference genes
            for ref_i, ref_name in enumerate(refs):
                ref_df = df[(df["Target"] == ref_name) & (df["TargetType"] == "ref")]
                ctrl_ref_cqs = ref_df[ref_df["Sample"].apply(is_ctrl)]["Cq"].dropna().tolist()
                if ctrl_ref_cqs:
                    st.session_state[f"control_reference_ct_{gene_i}_{ref_i}"] = "\n".join(f"{v:.4f}" for v in ctrl_ref_cqs)
                    count += len(ctrl_ref_cqs)

            # Patient groups
            for grp_i, pat_label in enumerate(patient_labels):
                pat_keywords = [s.strip() for s in pat_label.split(",") if s.strip()]

                def is_pat(sample_name, kws=pat_keywords):
                    return any(kw.lower() in sample_name.lower() for kw in kws)

                pat_cqs = tgt_df[tgt_df["Sample"].apply(is_pat)]["Cq"].dropna().tolist()
                if pat_cqs:
                    st.session_state[f"sample_target_ct_{gene_i}_{grp_i}"] = "\n".join(f"{v:.4f}" for v in pat_cqs)
                    count += len(pat_cqs)

                for ref_i, ref_name in enumerate(refs):
                    ref_df = df[(df["Target"] == ref_name) & (df["TargetType"] == "ref")]
                    pat_ref_cqs = ref_df[ref_df["Sample"].apply(is_pat)]["Cq"].dropna().tolist()
                    if pat_ref_cqs:
                        st.session_state[f"sample_reference_ct_{gene_i}_{grp_i}_{ref_i}"] = "\n".join(f"{v:.4f}" for v in pat_ref_cqs)
                        count += len(pat_ref_cqs)
        return count

    def compute_genorm_m(ref_ct_matrix):
        n_refs, n_samples = ref_ct_matrix.shape
        if n_refs < 2:
            return np.array([0.0])
        m_values = []
        for i in range(n_refs):
            pairwise_vars = []
            for j in range(n_refs):
                if i == j:
                    continue
                ratio = ref_ct_matrix[i] - ref_ct_matrix[j]
                pairwise_vars.append(np.std(ratio, ddof=1) if len(ratio) > 1 else 0.0)
            m_values.append(np.mean(pairwise_vars))
        return np.array(m_values)

    def compute_cv(ct_values):
        if len(ct_values) < 2 or np.mean(ct_values) == 0:
            return 0.0
        return (np.std(ct_values, ddof=1) / np.mean(ct_values)) * 100

    def geometric_mean_ct(ct_arrays):
        stacked = np.vstack(ct_arrays)
        return np.mean(stacked, axis=0)

    # ─── OUTLIER DETECTION FUNCTIONS ─────────────────────────────────────────────
    def detect_outliers_grubbs(data, alpha=0.05):
        data = np.array(data, dtype=float)
        n = len(data)
        if n < 3:
            return []
        outlier_indices = []
        working = data.copy()
        original_indices = list(range(n))
        while len(working) >= 3:
            mean_w = np.mean(working)
            std_w  = np.std(working, ddof=1)
            if std_w == 0:
                break
            g_vals = np.abs(working - mean_w) / std_w
            max_idx = np.argmax(g_vals)
            G = g_vals[max_idx]
            t_crit = stats.t.ppf(1 - alpha / (2 * len(working)), df=len(working) - 2)
            G_crit = ((len(working) - 1) / np.sqrt(len(working))) * \
                     np.sqrt(t_crit**2 / (len(working) - 2 + t_crit**2))
            if G > G_crit:
                outlier_indices.append(original_indices[max_idx])
                original_indices.pop(max_idx)
                working = np.delete(working, max_idx)
            else:
                break
        return outlier_indices

    def detect_outliers_iqr(data, multiplier=1.5):
        data = np.array(data, dtype=float)
        q1, q3 = np.percentile(data, [25, 75])
        iqr = q3 - q1
        lower = q1 - multiplier * iqr
        upper = q3 + multiplier * iqr
        return [i for i, v in enumerate(data) if v < lower or v > upper]

    def render_outlier_ui(data, label, key_prefix, method):
        data = np.array(data, dtype=float)
        if method == "Grubbs":
            detected = detect_outliers_grubbs(data)
        else:
            detected = detect_outliers_iqr(data)
        if not detected:
            return data, []
        st.warning(
            f"⚠️ **Potential outlier(s) detected in {label}** "
            f"({method} test): Sample(s) **{[i+1 for i in detected]}** "
            f"— values: **{[round(data[i], 3) for i in detected]}**\n\n"
            f"Select which samples to exclude from analysis:"
        )
        excluded = []
        for idx in detected:
            confirm = st.checkbox(
                f"Exclude Sample {idx+1}  (Ct = {data[idx]:.3f}) from {label}",
                value=False,
                key=f"{key_prefix}_excl_{idx}"
            )
            if confirm:
                excluded.append(idx)
        if excluded:
            cleaned = np.delete(data, excluded)
            st.info(
                f"ℹ️ {len(excluded)} sample(s) excluded from {label}. "
                f"Remaining n = {len(cleaned)}. "
                f"Excluded values will be flagged in the results table and PDF report."
            )
            return cleaned, excluded
        return data, []

    # ─────────────────────────────────────────────────────────────────────────────

    input_values_table = []
    data = []
    stats_data = []

    last_control_delta_ct = None
    last_gene_index = None

    control_group = "Control"
    target_gene = _t.get('target_gene', '')
    reference_gene = _t.get('reference_gene', '')
    ct_value = _t.get('ct_value', '')
    patient_group = _t.get('patient_group', '')

    # ═══════════════════════════════════════════════════════════════════════════════
    # SEKME 1: VERİ GİRİŞİ  (tüm girişler bu tab içinde)
    # ═══════════════════════════════════════════════════════════════════════════════

    # ── Genel CSS: boşlukları azalt, kartları güzelleştir ─────────────────────────
    st.markdown("""
    <style>
    /* Bölümler arası boşluğu azalt */
    /* padding-top ayarı yukarıda (satır ~209) merkezi olarak tanımlı — burada tekrar ezilmiyor */
    div[data-testid="stVerticalBlock"] > div { gap: 0.4rem; }
    /* st.info kutusunu küçült */
    div[data-testid="stAlert"] { padding: 6px 12px !important; font-size: 12px !important; }
    /* number_input ve radio margin azalt */
    div[data-testid="stNumberInput"] { margin-bottom: 0 !important; }
    div[data-testid="stRadio"] { margin-bottom: 0 !important; }
    </style>
    """, unsafe_allow_html=True)

    with tab_data:

        # ── KART 1: Çalışma Tasarımı ──────────────────────────────────────────────
        with st.container(border=True):
            st.markdown(f"**⚙️ {_t.get('tab_data', 'Data Entry')} — Study Design**")
            sd_c1, sd_c2, sd_c3 = st.columns(3)
            with sd_c1:
                num_target_genes = st.number_input(
                    _t.get('num_target_genes', '🔹 Target Genes'),
                    min_value=1, step=1, key="gene_count",
                )
            with sd_c2:
                num_patient_groups = st.number_input(
                    _t.get('num_patient_groups', '🔹 Patient Groups'),
                    min_value=1, step=1, key="patient_count",
                )

            with sd_c3:
                num_ref_genes = st.number_input(
                    _t.get('ref_gene_num_label', '🔹 Reference Genes'),
                    min_value=1, max_value=10, step=1,
                    key="num_ref_genes",
                    help=_t.get('ref_gene_num_help', 'MIQE: ≥2 reference genes recommended'),
                    **({} if "num_ref_genes" in st.session_state else {"value": 1})
                )
            # Referans gen durumu — tek satır
            if num_ref_genes == 1:
                st.caption(_t.get('ref_gene_1_warning', '⚠️ Single reference gene — MIQE recommends ≥2.').replace("**Methodological note:** ", "").replace("**", ""))
            else:
                st.caption(_t.get('ref_gene_multi_success', '✅ {n} reference genes selected.').format(n=num_ref_genes))
            if num_ref_genes > 1:
                with st.expander(_t.get('ref_gene_expander', 'ℹ️ About multi-reference normalization'), expanded=False):
                    st.markdown(_t.get('ref_multi_description', ''))

        # ── KART 2: Aykırı Değer + Stage ─────────────────────────────────────────
        with st.container(border=True):
            st.markdown(f"**🔍 {_t.get('outlier_section_title', 'Outlier Detection').replace('### ', '').replace('🔍 ', '')}**")
            out_c1, out_c2, out_c3 = st.columns([1, 2, 2])
            with out_c1:
                outlier_enabled = st.checkbox(
                    _t.get('outlier_enable', 'Enable'),
                    key="outlier_enabled",
                    help=_t.get('outlier_enable_help', ''),
                    **({} if "outlier_enabled" in st.session_state else {"value": True})
                )
                outlier_method = st.radio(
                    _t.get('outlier_method_label', 'Method'),
                    options=["Grubbs", "IQR"], key="outlier_method",
                    help=_t.get('outlier_method_help', '')
                )
            with out_c2:
                if outlier_method == "Grubbs":
                    grubbs_alpha = st.number_input(
                        _t.get('outlier_alpha_label', 'Significance level (α)'),
                        min_value=0.01, max_value=0.10, step=0.01, format="%.2f",
                        key="grubbs_alpha", help=_t.get('outlier_alpha_help', ''),
                        **({} if "grubbs_alpha" in st.session_state else {"value": 0.05})
                    )
                    iqr_multiplier = 1.5
                    st.caption(_t.get('grubbs_info', 'ℹ️ min n ≥ 3, normality assumed').format(alpha=grubbs_alpha))
                else:
                    iqr_multiplier = st.number_input(
                        _t.get('outlier_iqr_label', 'IQR multiplier (k)'),
                        min_value=1.0, max_value=3.0, step=0.25, format="%.2f",
                        key="iqr_mult", help=_t.get('outlier_iqr_help', ''),
                        **({} if "iqr_mult" in st.session_state else {"value": 1.5})
                    )
                    grubbs_alpha = 0.05
                    st.caption("ℹ️ Q1−k×IQR / Q3+k×IQR")
            with out_c3:
                outlier_stage = st.radio(
                    _t.get('outlier_stage_label', '🔬 Detection Stage'),
                    options=[
                        _t.get('outlier_stage_raw', 'Raw Cq — before normalization (recommended)'),
                        _t.get('outlier_stage_dct', 'ΔCq — after normalization'),
                    ],
                    index=0, key="outlier_stage",
                    help=_t.get('outlier_stage_help', '')
                )
            with st.expander(_t.get('outlier_expander', 'ℹ️ About outlier detection in qPCR'), expanded=False):
                st.markdown(_t.get('outlier_description', ''))

        outlier_on_raw = st.session_state.get("outlier_stage", _t.get('outlier_stage_raw', 'Raw Cq — before normalization (recommended)')) == _t.get('outlier_stage_raw', 'Raw Cq — before normalization (recommended)')

        # ── KART 3: Amplifikasyon Verimliliği ─────────────────────────────────────
        with st.container(border=True):
            st.markdown(f"**🔬 {_t.get('efficiency_header', 'Amplification Efficiency').replace('🔬 ', '')}**")
            eff_c1, eff_c2, eff_c3 = st.columns([2, 2, 3])
            with eff_c1:
                efficiency_method = st.radio(
                    _t.get('efficiency_method', 'Input Method'),
                    options=[_t.get('efficiency_manual', 'Manual E value'), _t.get('efficiency_slope', 'Calculate from slope')],
                    key="eff_method", horizontal=True
                )
            with eff_c2:
                efficiency_threshold = st.number_input(
                    _t.get('efficiency_threshold', 'Max diff threshold (%)'),
                    min_value=1.0, max_value=50.0, value=10.0, step=0.5, key="eff_threshold",
                    help=_t.get('efficiency_note', 'Recommended: 10% (MIQE guidelines).')
                )
            with eff_c3:
                st.caption(_t.get('efficiency_note', 'E=2.0 = perfect (100%). Accepted: 1.8–2.2 (90–110%).'))
            with st.expander(_t.get('ref_gene_expander', 'ℹ️ How to obtain Efficiency (E)').replace('📚 ', ''), expanded=False):
                st.markdown(_t.get('ref_multi_description', '') or
                    "**Method 1 — Standard Curve:** `E = 10^(-1 / slope)`  \n"
                    "**Method 2 — Software:** LinRegPCR, qBase+, CFX Maestro  \n"
                    "**Accepted range:** E = 1.8–2.2 (90–110%)"
                )

        st.divider()

        # ── Standart eğri hesaplayıcı ─────────────────────────────────────────────
        with st.expander(_t.get('sc_expander', ''), expanded=False):
            st.markdown(_t.get('sc_description', ''))
            sc_c1, sc_c2 = st.columns(2)
            with sc_c1:
                sc_gene_label = st.text_input(_t.get('sc_gene_label', ''), value="Target Gene 1", key="sc_label")
                sc_num_points = st.number_input(_t.get('sc_num_points', ''), min_value=3, max_value=10, value=5, step=1, key="sc_npts")
            with sc_c2:
                st.markdown(_t.get('sc_dilution_factor_label', ''))
                sc_dilution_factor = st.number_input(_t.get('sc_dilution_factor_input', ''), min_value=2, max_value=100, value=10, step=1, key="sc_dilfactor")
                st.markdown(_t.get('sc_start_conc_label', ''))
                sc_start_conc = st.number_input(_t.get('sc_start_conc_input', ''), min_value=0.0001, value=1.0, format="%.4f", key="sc_startconc")
            st.markdown(_t.get('sc_enter_ct', ''))
            sc_ct_cols = st.columns(min(sc_num_points, 5))
            sc_ct_values = []
            sc_log_concs = []
            for pt in range(sc_num_points):
                conc = sc_start_conc / (sc_dilution_factor ** pt)
                log_c = np.log10(conc)
                with sc_ct_cols[pt % 5]:
                    ct_val = st.number_input(f"Dil. {pt+1} (log={log_c:.2f})", value=18.0 + pt * 3.32, step=0.01, format="%.2f", key=f"sc_ct_{pt}")
                sc_ct_values.append(ct_val)
                sc_log_concs.append(log_c)
            if st.button(_t.get('sc_calc_button', ''), key="sc_calc"):
                sc_log_concs_arr = np.array(sc_log_concs)
                sc_ct_arr = np.array(sc_ct_values)
                slope_val, intercept_val, r_val, p_val, se_val = stats.linregress(sc_log_concs_arr, sc_ct_arr)
                r2 = r_val ** 2
                E_calc = 10 ** (-1.0 / slope_val) if slope_val != 0 else float('nan')
                E_pct = (E_calc - 1) * 100
                rc1, rc2, rc3, rc4 = st.columns(4)
                rc1.metric(_t.get('sc_slope', ''), f"{slope_val:.4f}")
                rc2.metric(_t.get('sc_e_value', ''), f"{E_calc:.4f}")
                rc3.metric(_t.get('sc_efficiency_pct', ''), f"{E_pct:.1f}%")
                rc4.metric("R²", f"{r2:.4f}")
                if 1.8 <= E_calc <= 2.2 and r2 >= 0.99:
                    st.success(_t.get('sc_excellent', '').format(e=E_calc, pct=E_pct, r2=r2))
                elif 1.8 <= E_calc <= 2.2:
                    st.warning(_t.get('sc_warning_r2', '').format(pct=E_pct, r2=r2))
                else:
                    st.error(_t.get('sc_error_range', '').format(e=E_calc, pct=E_pct))
                st.info(_t.get('sc_copy_hint', '').format(slope=slope_val, e=E_calc))

        # ── Per-gen efficiency girişi ─────────────────────────────────────────────
        gene_efficiencies = {}
        use_slope = (efficiency_method == _t.get('efficiency_slope', ''))
        for i in range(num_target_genes):
            with st.expander(f"🔬 {_t.get('target_gene', "")} {i+1} — Efficiency", expanded=(i == 0)):
                ec1, ec2 = st.columns(2)
                with ec1:
                    if use_slope:
                        target_slope = st.number_input(_t.get('efficiency_target_slope_label', '').format(i=i+1), value=-3.32, step=0.01, format="%.4f", key=f"target_slope_{i}")
                        target_E = 10 ** (-1.0 / target_slope) if target_slope != 0 else 2.0
                        st.markdown(f"**E (target) = {target_E:.4f}** ({(target_E-1)*100:.1f}%)")
                    else:
                        target_E = st.number_input(_t.get('efficiency_target_label', '').format(i=i+1), min_value=1.0, max_value=3.0, value=2.0, step=0.01, format="%.4f", key=f"target_E_{i}")
                        st.markdown(f"**{(target_E-1)*100:.1f}%**")
                with ec2:
                    if use_slope:
                        ref_slope = st.number_input(_t.get('efficiency_ref_slope_label', '').format(i=i+1), value=-3.32, step=0.01, format="%.4f", key=f"ref_slope_{i}")
                        ref_E = 10 ** (-1.0 / ref_slope) if ref_slope != 0 else 2.0
                        st.markdown(f"**E (ref) = {ref_E:.4f}** ({(ref_E-1)*100:.1f}%)")
                    else:
                        ref_E = st.number_input(_t.get('efficiency_ref_label', '').format(i=i+1), min_value=1.0, max_value=3.0, value=2.0, step=0.01, format="%.4f", key=f"ref_E_{i}")
                        st.markdown(f"**{(ref_E-1)*100:.1f}%**")
                diff = abs((target_E-1)*100 - (ref_E-1)*100)
                if diff <= efficiency_threshold:
                    st.success(_t.get('efficiency_ok', '').format(diff=diff))
                else:
                    st.warning(_t.get('efficiency_warning', '').format(diff=diff))
                gene_efficiencies[i] = {"target_E": target_E, "ref_E": ref_E}

        st.divider()

        # ── Gen verisi giriş başlığı ───────────────────────────────────────────────
        st.markdown(
            f"<div style='font-size:15px;font-weight:700;color:#1a237e;margin-bottom:6px;'>"
            f"📥 {_t.get('patient_data_header', 'Enter Patient and Control Group Data')}"
            f"</div>",
            unsafe_allow_html=True
        )

        # Kontrol + Hasta Grubu Veri Giriş Döngüsü
        for i in range(num_target_genes):
            st.markdown(
                f"<h4 style='margin-top:12px;margin-bottom:4px;color:#283593;'>"
                f"🧬 {_t.get('target_gene', 'Target Gene')} {i+1}</h4>",
                unsafe_allow_html=True
            )

            control_target_ct = st.text_area(
                f"Control {i+1} - {_t.get('target_gene', '')} {i+1} - {_t.get('ct_value', '')}",
                value=st.session_state.get(f"control_target_ct_{i}", ""),
                key=f"control_target_ct_{i}"
            )

            # ── Multi-reference gene input (Control) ─────────────────────────────────
            ctrl_ref_arrays = []
            ctrl_ref_names  = []
            all_ctrl_refs_valid = True

            for r in range(num_ref_genes):
                ref_label = f"Ref Gene {r+1}" if num_ref_genes > 1 else _t.get('reference_gene', '')
                ctrl_ref_ct_raw = st.text_area(
                    f"Control {i+1} — {ref_label} {i+1} — {_t.get('ct_value', "")}",
                    value=st.session_state.get(f"control_reference_ct_{i}_{r}", ""),
                    key=f"control_reference_ct_{i}_{r}"
                )
                parsed = parse_input_data(ctrl_ref_ct_raw)
                if len(parsed) == 0:
                    all_ctrl_refs_valid = False
                else:
                    ctrl_ref_arrays.append(parsed)
                    ctrl_ref_names.append(f"Ref Gene {r+1}")

            control_target_ct_values = np.array(parse_input_data(control_target_ct))

            if len(control_target_ct_values) == 0 or not all_ctrl_refs_valid or len(ctrl_ref_arrays) == 0:
                st.error(_t.get('warning_control_ct', '').format(i=i+1))
                continue

            # Trim all arrays to common length
            # warn user if n differs between target and reference genes
            min_control_len = min(len(control_target_ct_values), *[len(a) for a in ctrl_ref_arrays])
            all_ctrl_lengths = [len(control_target_ct_values)] + [len(a) for a in ctrl_ref_arrays]
            if len(set(all_ctrl_lengths)) > 1:
                details = f"Target Gene: n={len(control_target_ct_values)}" + \
                          "".join([f", Ref Gene {r+1}: n={len(ctrl_ref_arrays[r])}" for r in range(len(ctrl_ref_arrays))])
                st.warning(_t.get('unequal_n_warning', '').format(
                    group=f"Control Group {i+1}",
                    details=details,
                    min_n=min_control_len
                ))
            control_target_ct_values = control_target_ct_values[:min_control_len]
            ctrl_ref_arrays = [a[:min_control_len] for a in ctrl_ref_arrays]

            # ── Outlier detection — Raw Cq stage (BEFORE normalization) ──────────────
            # When outlier_on_raw is True, Grubbs/IQR is applied to raw Ct values
            # separately for target gene and each reference gene before ΔCq is computed.
            # This prevents noisy replicates from propagating into normalization.
            ctrl_excluded_target = []

            if outlier_enabled and outlier_on_raw:
                # --- Target Ct outlier check ---
                if len(control_target_ct_values) >= 3:
                    detected_raw_tgt = detect_outliers_grubbs(control_target_ct_values, alpha=grubbs_alpha) \
                                       if outlier_method == "Grubbs" \
                                       else detect_outliers_iqr(control_target_ct_values, multiplier=iqr_multiplier)
                    if detected_raw_tgt:
                        control_target_ct_values, ctrl_excluded_target = render_outlier_ui(
                            control_target_ct_values,
                            f"Control Group {i+1} — Target Gene {i+1} (Raw Cq)",
                            f"ctrl_raw_tgt_{i}",
                            outlier_method
                        )
                        if ctrl_excluded_target:
                            keep_indices = [k for k in range(min_control_len) if k not in ctrl_excluded_target]
                            ctrl_ref_arrays = [a[keep_indices] for a in ctrl_ref_arrays]
                            min_control_len = len(keep_indices)

                # --- Reference gene Ct outlier check (each ref gene separately) ---
                for r in range(len(ctrl_ref_arrays)):
                    if len(ctrl_ref_arrays[r]) >= 3:
                        detected_raw_ref = detect_outliers_grubbs(ctrl_ref_arrays[r], alpha=grubbs_alpha) \
                                           if outlier_method == "Grubbs" \
                                           else detect_outliers_iqr(ctrl_ref_arrays[r], multiplier=iqr_multiplier)
                        if detected_raw_ref:
                            cleaned_ref, excl_ref = render_outlier_ui(
                                ctrl_ref_arrays[r],
                                f"Control Group {i+1} — Reference Gene {r+1} (Raw Cq)",
                                f"ctrl_raw_ref_{i}_{r}",
                                outlier_method
                            )
                            if excl_ref:
                                # Remove same indices from target and all other refs
                                keep_ref = [k for k in range(len(ctrl_ref_arrays[r])) if k not in excl_ref]
                                control_target_ct_values = control_target_ct_values[keep_ref]
                                ctrl_ref_arrays = [a[keep_ref] for a in ctrl_ref_arrays]
                                min_control_len = len(keep_ref)

            # ── Outlier detection — Control Target Ct (ΔCq stage fallback) ──────────
            elif outlier_enabled and not outlier_on_raw and len(control_target_ct_values) >= 3:
                detected_ctrl_tgt = detect_outliers_grubbs(control_target_ct_values, alpha=grubbs_alpha) \
                                    if outlier_method == "Grubbs" \
                                    else detect_outliers_iqr(control_target_ct_values, multiplier=iqr_multiplier)
                if detected_ctrl_tgt:
                    control_target_ct_values, ctrl_excluded_target = render_outlier_ui(
                        control_target_ct_values,
                        f"Control Group {i+1} — Target Gene {i+1}",
                        f"ctrl_tgt_{i}",
                        outlier_method
                    )
                    if ctrl_excluded_target:
                        keep_indices = [k for k in range(min_control_len) if k not in ctrl_excluded_target]
                        ctrl_ref_arrays = [a[keep_indices] for a in ctrl_ref_arrays]
                        min_control_len = len(keep_indices)

            # ── geNorm + CV stability (shown when ≥2 ref genes) ──────────────────────
            if num_ref_genes >= 2:
                ref_matrix = np.vstack(ctrl_ref_arrays)   # (n_refs, n_samples)
                m_values   = compute_genorm_m(ref_matrix)
                cv_values  = [compute_cv(a) for a in ctrl_ref_arrays]

                unstable_ctrl = [r for r, m in enumerate(m_values) if m >= 1.0]
                borderline_ctrl = [r for r, m in enumerate(m_values) if 0.5 <= m < 1.0]

                st.markdown(f"##### 📊 " + _t.get('genorm_title', 'Reference Gene Stability') + f" — " + _t.get('control_group', 'Control Group') + f" {i+1}")
                stab_cols = st.columns(num_ref_genes)
                for r, col in enumerate(stab_cols):
                    m_ok = m_values[r] < 1.0
                    cv_ok = cv_values[r] < 5.0
                    with col:
                        st.metric(
                            label=f"Ref Gene {r+1}",
                            value=f"M = {m_values[r]:.3f}",
                            delta=f"CV = {cv_values[r]:.2f}%"
                        )
                        if m_ok and cv_ok:
                            st.caption("✅ " + _t.get("stable", "Stable"))
                        elif m_ok or cv_ok:
                            st.caption("⚠️ " + _t.get("borderline", "Borderline"))
                        else:
                            st.caption("❌ " + _t.get("unstable", "Unstable") + " — M ≥ 1.0")

                # Stability bar chart
                fig_stab = go.Figure()
                fig_stab.add_trace(go.Bar(
                    name="geNorm M-value",
                    x=[f"Ref {r+1}" for r in range(num_ref_genes)],
                    y=m_values,
                    marker_color=["#2ecc71" if m < 0.5 else "#f39c12" if m < 1.0 else "#e74c3c" for m in m_values],
                    text=[f"{m:.3f}" for m in m_values],
                    textposition="outside"
                ))
                fig_stab.add_hline(y=0.5, line_dash="dot", line_color="green",
                                   annotation_text="M=0.5 (strict)", annotation_position="right")
                fig_stab.add_hline(y=1.0, line_dash="dash", line_color="orange",
                                   annotation_text="M=1.0 (acceptable)", annotation_position="right")
                fig_stab.update_layout(
                    title=f"geNorm M-value — {_t.get('control_group', 'Control Group').replace('🧬 ','')} {i+1}",
                    yaxis_title=_t.get('m_value', 'M-value') + " (lower = more stable)",
                    height=280
                )
                st.plotly_chart(fig_stab, use_container_width=True, key=f"stab_ctrl_{i}")

                # ── Stability warnings ────────────────────────────────────────────────
                _ctrl_grp_lbl = f"{_t.get('control_group','Control Group').replace('🧬 ','')} {i+1}"
                if unstable_ctrl:
                    unstable_names = ", ".join([f"{_t.get('ref_gene','Ref Gene')} {r+1}" for r in unstable_ctrl])
                    st.warning(
                        f"⚠️ **{_t.get('unstable','Unstable')} — {_ctrl_grp_lbl}: {unstable_names}**\n\n"
                        f"geNorm M ≥ 1.0 — expression varies considerably. Normalization may be distorted.\n\n"
                        f"- Verify Ct values for {unstable_names}\n"
                        f"- Consider replacing with a more stable reference gene\n"
                        f"- Ref: Vandesompele et al. *Genome Biology* 2002"
                    )
                elif borderline_ctrl:
                    borderline_names = ", ".join([f"{_t.get('ref_gene','Ref Gene')} {r+1}" for r in borderline_ctrl])
                    st.info(
                        f"ℹ️ **{_t.get('borderline','Borderline')} — {_ctrl_grp_lbl}: {borderline_names}** (M = 0.5–1.0)\n\n"
                        f"{_t.get('stability','Stability')} acceptable per MIQE. Consider adding a third reference gene."
                    )
                else:
                    st.success(
                        f"✅ {_t.get('stable','All stable')} — {_ctrl_grp_lbl} (M < 0.5). "
                        f"Normalization quality is excellent."
                    )

            # ── Compute normalization factor (geometric mean of refs) ─────────────────
            # Re-sync min_control_len to actual array lengths after any outlier removal
            min_control_len = min(len(control_target_ct_values), *[len(a) for a in ctrl_ref_arrays])
            control_target_ct_values = control_target_ct_values[:min_control_len]
            ctrl_ref_arrays = [a[:min_control_len] for a in ctrl_ref_arrays]

            ctrl_norm_factor = geometric_mean_ct(ctrl_ref_arrays)   # per-sample NF
            control_delta_ct = control_target_ct_values - ctrl_norm_factor

            # For table: show first ref gene Ct as representative; NF shown separately
            control_reference_ct_values = ctrl_ref_arrays[0]   # kept for legacy table column

            average_control_delta_ct = np.mean(control_delta_ct) if len(control_delta_ct) > 0 else None
            sample_counter = 1

            for idx in range(min_control_len):
                row = {
                    "__sample_num__": sample_counter,
                    "__target_gene__": f"Gene {i+1}",
                    "Grup": "Control",
                    "__target_ct__": control_target_ct_values[idx],
                    "__ref_ct__": round(ctrl_norm_factor[idx], 4),
                    "__dct_ctrl__": round(control_delta_ct[idx], 4),
                    "Outlier Excluded": "No"
                }
                if num_ref_genes > 1:
                    for r, arr in enumerate(ctrl_ref_arrays):
                        row[f"Ref Gene {r+1} Ct"] = arr[idx]
                input_values_table.append(row)
                sample_counter += 1

            # Log excluded outliers as separate flagged rows
            for ex_idx in ctrl_excluded_target:
                input_values_table.append({
                    "__sample_num__": f"{ex_idx + 1} ⚠️",
                    "__target_gene__": f"Gene {i+1}",
                    "Grup": "Control",
                    "__target_ct__": "EXCLUDED",
                    "__ref_ct__": "EXCLUDED",
                    "__dct_ctrl__": "EXCLUDED",
                    "Outlier Excluded": f"Yes ({outlier_method})"
                })

            for j in range(num_patient_groups):
                st.markdown(
                    f"<h4>{_t.get('patient_group', "")} {j+1} - {_t.get('target_gene', "")} {i+1}</h4>",
                    unsafe_allow_html=True
                )

                sample_target_ct = st.text_area(
                    f"{_t.get('patient_group', "")} {j+1} - {_t.get('target_gene', "")} {i+1} - {_t.get('ct_value', "")}",
                    value=st.session_state.get(f"sample_target_ct_{i}_{j}", ""),
                    key=f"sample_target_ct_{i}_{j}"
                )

                # ── Multi-reference gene input (Patient) ──────────────────────────────
                smp_ref_arrays = []
                all_smp_refs_valid = True

                for r in range(num_ref_genes):
                    ref_label = f"Ref Gene {r+1}" if num_ref_genes > 1 else _t.get('reference_gene', '')
                    smp_ref_ct_raw = st.text_area(
                        f"{_t.get('patient_group', "")} {j+1} — {ref_label} {i+1} — {_t.get('ct_value', "")}",
                        value=st.session_state.get(f"sample_reference_ct_{i}_{j}_{r}", ""),
                        key=f"sample_reference_ct_{i}_{j}_{r}"
                    )
                    parsed = parse_input_data(smp_ref_ct_raw)
                    if len(parsed) == 0:
                        all_smp_refs_valid = False
                    else:
                        smp_ref_arrays.append(parsed)

                sample_target_ct_values = np.array(parse_input_data(sample_target_ct))

                if len(sample_target_ct_values) == 0 or not all_smp_refs_valid or len(smp_ref_arrays) == 0:
                    st.error(_t.get('warning_patient_cq', '').format(j=j+1))
                    continue

                # warn if n differs between target and reference genes
                min_sample_len = min(len(sample_target_ct_values), *[len(a) for a in smp_ref_arrays])
                all_smp_lengths = [len(sample_target_ct_values)] + [len(a) for a in smp_ref_arrays]
                if len(set(all_smp_lengths)) > 1:
                    details = f"Target Gene: n={len(sample_target_ct_values)}" + \
                              "".join([f", Ref Gene {r+1}: n={len(smp_ref_arrays[r])}" for r in range(len(smp_ref_arrays))])
                    st.warning(_t.get('unequal_n_warning', '').format(
                        group=f"{_t.get('patient_group', "")} {j+1}, Gene {i+1}",
                        details=details,
                        min_n=min_sample_len
                    ))
                sample_target_ct_values = sample_target_ct_values[:min_sample_len]
                smp_ref_arrays = [a[:min_sample_len] for a in smp_ref_arrays]

                # ── Outlier detection — Raw Cq stage (BEFORE normalization) ──────────
                # same logic as control group above
                smp_excluded_target = []

                if outlier_enabled and outlier_on_raw:
                    # --- Target Ct ---
                    if len(sample_target_ct_values) >= 3:
                        detected_raw_smp_tgt = detect_outliers_grubbs(sample_target_ct_values, alpha=grubbs_alpha) \
                                               if outlier_method == "Grubbs" \
                                               else detect_outliers_iqr(sample_target_ct_values, multiplier=iqr_multiplier)
                        if detected_raw_smp_tgt:
                            sample_target_ct_values, smp_excluded_target = render_outlier_ui(
                                sample_target_ct_values,
                                f"{_t.get('patient_group', "")} {j+1} — Target Gene {i+1} (Raw Cq)",
                                f"smp_raw_tgt_{i}_{j}",
                                outlier_method
                            )
                            if smp_excluded_target:
                                keep_indices_smp = [k for k in range(min_sample_len) if k not in smp_excluded_target]
                                smp_ref_arrays = [a[keep_indices_smp] for a in smp_ref_arrays]
                                min_sample_len = len(keep_indices_smp)

                    # --- Reference gene Ct (each separately) ---
                    for r in range(len(smp_ref_arrays)):
                        if len(smp_ref_arrays[r]) >= 3:
                            detected_raw_smp_ref = detect_outliers_grubbs(smp_ref_arrays[r], alpha=grubbs_alpha) \
                                                   if outlier_method == "Grubbs" \
                                                   else detect_outliers_iqr(smp_ref_arrays[r], multiplier=iqr_multiplier)
                            if detected_raw_smp_ref:
                                cleaned_smp_ref, excl_smp_ref = render_outlier_ui(
                                    smp_ref_arrays[r],
                                    f"{_t.get('patient_group', "")} {j+1} — Reference Gene {r+1} (Raw Cq)",
                                    f"smp_raw_ref_{i}_{j}_{r}",
                                    outlier_method
                                )
                                if excl_smp_ref:
                                    keep_ref_smp = [k for k in range(len(smp_ref_arrays[r])) if k not in excl_smp_ref]
                                    sample_target_ct_values = sample_target_ct_values[keep_ref_smp]
                                    smp_ref_arrays = [a[keep_ref_smp] for a in smp_ref_arrays]
                                    min_sample_len = len(keep_ref_smp)

                # ── Outlier detection — Patient Target Ct (ΔCq stage fallback) ──────
                elif outlier_enabled and not outlier_on_raw and len(sample_target_ct_values) >= 3:
                    detected_smp_tgt = detect_outliers_grubbs(sample_target_ct_values, alpha=grubbs_alpha) \
                                       if outlier_method == "Grubbs" \
                                       else detect_outliers_iqr(sample_target_ct_values, multiplier=iqr_multiplier)
                    if detected_smp_tgt:
                        sample_target_ct_values, smp_excluded_target = render_outlier_ui(
                            sample_target_ct_values,
                            f"{_t.get('patient_group', "")} {j+1} — Target Gene {i+1}",
                            f"smp_tgt_{i}_{j}",
                            outlier_method
                        )
                        if smp_excluded_target:
                            keep_indices_smp = [k for k in range(min_sample_len) if k not in smp_excluded_target]
                            smp_ref_arrays = [a[keep_indices_smp] for a in smp_ref_arrays]
                            min_sample_len = len(keep_indices_smp)

                # ── geNorm + CV stability (Patient, shown when ≥2 ref genes) ─────────
                if num_ref_genes >= 2:
                    smp_ref_matrix = np.vstack(smp_ref_arrays)
                    smp_m_values   = compute_genorm_m(smp_ref_matrix)
                    smp_cv_values  = [compute_cv(a) for a in smp_ref_arrays]

                    unstable_smp   = [r for r, m in enumerate(smp_m_values) if m >= 1.0]
                    borderline_smp = [r for r, m in enumerate(smp_m_values) if 0.5 <= m < 1.0]

                    st.markdown(f"##### 📊 Reference Gene Stability — {_t.get('patient_group', "")} {j+1}")
                    smp_stab_cols = st.columns(num_ref_genes)
                    for r, col in enumerate(smp_stab_cols):
                        m_ok = smp_m_values[r] < 1.0
                        cv_ok = smp_cv_values[r] < 5.0
                        with col:
                            st.metric(
                                label=f"Ref Gene {r+1}",
                                value=f"M = {smp_m_values[r]:.3f}",
                                delta=f"CV = {smp_cv_values[r]:.2f}%"
                            )
                            if m_ok and cv_ok:
                                st.caption("✅ " + _t.get("stable", "Stable"))
                            elif m_ok or cv_ok:
                                st.caption("⚠️ " + _t.get("borderline", "Borderline"))
                            else:
                                st.caption("❌ " + _t.get("unstable", "Unstable") + " — M ≥ 1.0")

                    # Stability bar chart (patient)
                    fig_stab_smp = go.Figure()
                    fig_stab_smp.add_trace(go.Bar(
                        name="geNorm M-value",
                        x=[f"Ref {r+1}" for r in range(num_ref_genes)],
                        y=smp_m_values,
                        marker_color=["#2ecc71" if m < 0.5 else "#f39c12" if m < 1.0 else "#e74c3c" for m in smp_m_values],
                        text=[f"{m:.3f}" for m in smp_m_values],
                        textposition="outside"
                    ))
                    fig_stab_smp.add_hline(y=0.5, line_dash="dot", line_color="green",
                                       annotation_text="M=0.5 (strict)", annotation_position="right")
                    fig_stab_smp.add_hline(y=1.0, line_dash="dash", line_color="orange",
                                       annotation_text="M=1.0 (acceptable)", annotation_position="right")
                    fig_stab_smp.update_layout(
                        title=f"geNorm M-value — {_t.get('patient_group', '').replace('🩸 ','')} {j+1}",
                        yaxis_title=_t.get('m_value', 'M-value') + " (lower = more stable)",
                        height=280
                    )
                    st.plotly_chart(fig_stab_smp, use_container_width=True, key=f"stab_smp_{i}_{j}")

                    # ── Stability warnings (patient) ──────────────────────────────────
                    _smp_grp_lbl = f"{_t.get('patient_group','').replace('🩸 ','')} {j+1}"
                    if unstable_smp:
                        unstable_names = ", ".join([f"{_t.get('ref_gene','Ref Gene')} {r+1}" for r in unstable_smp])
                        st.warning(
                            f"⚠️ **{_t.get('unstable','Unstable')} — {_smp_grp_lbl}: {unstable_names}**\n\n"
                            f"geNorm M ≥ 1.0 — normalization reliability compromised. Interpret with caution.\n\n"
                            f"- Check for outliers or data entry errors\n"
                            f"- Validate {unstable_names} in this sample group\n"
                            f"- Consider replacing with a validated reference gene"
                        )
                    elif borderline_smp:
                        borderline_names = ", ".join([f"{_t.get('ref_gene','Ref Gene')} {r+1}" for r in borderline_smp])
                        st.info(
                            f"ℹ️ **{_t.get('borderline','Borderline')} — {_smp_grp_lbl}: {borderline_names}** (M = 0.5–1.0)\n\n"
                            f"{_t.get('stability','Stability')} within MIQE range. Consider a third reference gene."
                        )
                    else:
                        st.success(
                            f"✅ {_t.get('stable','Stable')} — {_smp_grp_lbl} (M < 0.5)."
                        )

                # ── Normalization factor & ΔCq ────────────────────────────────────────
                # Re-sync lengths after any outlier removal
                min_sample_len = min(len(sample_target_ct_values), *[len(a) for a in smp_ref_arrays])
                sample_target_ct_values = sample_target_ct_values[:min_sample_len]
                smp_ref_arrays = [a[:min_sample_len] for a in smp_ref_arrays]

                smp_norm_factor = geometric_mean_ct(smp_ref_arrays)
                sample_delta_ct = sample_target_ct_values - smp_norm_factor
                sample_reference_ct_values = smp_ref_arrays[0]

                average_sample_delta_ct = np.mean(sample_delta_ct) if len(sample_delta_ct) > 0 else None

                sample_counter = 1
                for idx in range(min_sample_len):
                    row = {
                        "__sample_num__": sample_counter,
                        "__target_gene__": f"Gene {i+1}",
                        "Grup": f"Group {j+1}",
                        "__target_ct__": sample_target_ct_values[idx],
                        "__ref_ct__": round(smp_norm_factor[idx], 4),
                        "__dct_patient__": round(sample_delta_ct[idx], 4),
                        "Outlier Excluded": "No"
                    }
                    if num_ref_genes > 1:
                        for r, arr in enumerate(smp_ref_arrays):
                            row[f"Ref Gene {r+1} Ct"] = arr[idx]
                    input_values_table.append(row)
                    sample_counter += 1

                # Log excluded outliers as flagged rows
                for ex_idx in smp_excluded_target:
                    input_values_table.append({
                        "__sample_num__": f"{ex_idx + 1} ⚠️",
                        "__target_gene__": f"Gene {i+1}",
                        "Grup": f"Group {j+1}",
                        "__target_ct__": "EXCLUDED",
                        "__ref_ct__": "EXCLUDED",
                        "__dct_patient__": "EXCLUDED",
                        "Outlier Excluded": f"Yes ({outlier_method})"
                    })

                # ΔΔCq ve Gen Ekspresyon Değişimi Hesaplama
                if average_control_delta_ct is not None and average_sample_delta_ct is not None:
                    delta_delta_ct = average_sample_delta_ct - average_control_delta_ct
                    expression_change = 2 ** (-delta_delta_ct)

                    # ── Pfaffl Calculation ──────────────────────────────────────────
                    eff = gene_efficiencies.get(i, {"target_E": 2.0, "ref_E": 2.0})
                    E_target = eff["target_E"]
                    E_ref = eff["ref_E"]

                    avg_ctrl_target = np.mean(control_target_ct_values)
                    avg_ctrl_ref    = np.mean(ctrl_norm_factor)
                    avg_smp_target  = np.mean(sample_target_ct_values)
                    avg_smp_ref     = np.mean(smp_norm_factor)

                    delta_ct_target_pfaffl = avg_ctrl_target - avg_smp_target
                    delta_ct_ref_pfaffl    = avg_ctrl_ref    - avg_smp_ref

                    pfaffl_ratio = (E_target ** delta_ct_target_pfaffl) / (E_ref ** delta_ct_ref_pfaffl)
                    # ────────────────────────────────────────────────────────────────
            
                    # Regulation: literatür standardı FC > 1.5 = up, < 0.67 = down
                    # (bazı çalışmalar 1.2/0.83 kullanır — burada 1.5/0.67)
                    if expression_change >= 1.5:
                        regulation_status = _t.get('upregulated', '')
                    elif expression_change <= 0.67:
                        regulation_status = _t.get('downregulated', '')
                    else:
                        regulation_status = _t.get('no_change', '')

                    # Pfaffl regulation
                    if pfaffl_ratio >= 1.5:
                        pfaffl_regulation = _t.get('upregulated', '')
                    elif pfaffl_ratio <= 0.67:
                        pfaffl_regulation = _t.get('downregulated', '')
                    else:
                        pfaffl_regulation = _t.get('no_change', '')

                    # ── Method comparison display ─────────────────────────────────
                    st.markdown(f"#### {_t.get('method_comparison', "")} — {_t.get('target_gene', "")} {i+1} / {_t.get('patient_group', "")} {j+1}")
                    comp_col1, comp_col2 = st.columns(2)
                    with comp_col1:
                        st.metric(
                            label=_t.get('classic_ddct', ''),
                            value=f"{expression_change:.4f}",
                            delta=regulation_status
                        )
                    with comp_col2:
                        st.metric(
                            label=_t.get('pfaffl_ratio', ''),
                            value=f"{pfaffl_ratio:.4f}",
                            delta=pfaffl_regulation
                        )
                    # ─────────────────────────────────────────────────────────────

                    # ── Per-group pairwise stats (control vs this patient group) ────
                    # Statistical tests are now performed on RQ values (2^-ΔCq) instead of
                    # raw ΔCt values. ΔCt values are on a logarithmic scale; performing
                    # t-tests directly on ΔCt underestimates biological variability and can
                    # produce false significant differences compared to linear RQ-based tests.
                    control_rq = 2 ** (-np.array(control_delta_ct))
                    sample_rq  = 2 ** (-np.array(sample_delta_ct))

                    n_ctrl = len(control_rq)
                    n_smp  = len(sample_rq)

                    # n < 2 ise istatistik hesaplanamaz
                    if n_ctrl < 2 or n_smp < 2:
                        test_pvalue = float('nan')
                        test_method = "N/A (n < 2)"
                        test_type   = "—"
                        significance = "—"
                        equal_variance = True
                        control_normal = True
                        sample_normal  = True
                        shapiro_control = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                        shapiro_sample  = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                        levene_test     = type('LV', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                    else:
                        _MIN_N_SHAPIRO = 8

                        if n_ctrl >= _MIN_N_SHAPIRO and n_smp >= _MIN_N_SHAPIRO:
                            shapiro_control = stats.shapiro(control_rq)
                            shapiro_sample  = stats.shapiro(sample_rq)
                            control_normal  = shapiro_control.pvalue > 0.05
                            sample_normal   = shapiro_sample.pvalue  > 0.05
                        else:
                            shapiro_control = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                            shapiro_sample  = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                            control_normal  = True
                            sample_normal   = True

                        try:
                            levene_test    = stats.levene(control_rq, sample_rq)
                            equal_variance = (levene_test.pvalue > 0.05) if not np.isnan(levene_test.pvalue) else True
                        except Exception:
                            levene_test    = type('LV', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                            equal_variance = True

                        try:
                            if control_normal and sample_normal:
                                if equal_variance:
                                    test_pvalue = stats.ttest_ind(control_rq, sample_rq).pvalue
                                    test_method = _t.get('t_test', '')
                                else:
                                    test_pvalue = stats.ttest_ind(control_rq, sample_rq, equal_var=False).pvalue
                                    test_method = _t.get('welch_t_test', '')
                                test_type = _t.get('parametric', '')
                            else:
                                test_pvalue = stats.mannwhitneyu(control_rq, sample_rq,
                                                                  alternative='two-sided').pvalue
                                test_method = _t.get('mann_whitney_u_test', '')
                                test_type   = _t.get('non_parametric', '')
                        except Exception:
                            test_pvalue = float('nan')
                            test_method = "Error"
                            test_type   = "—"

                        significance = _t.get('significant', '') if (not np.isnan(test_pvalue) and test_pvalue < 0.05) \
                                       else (_t.get('insignificant', '') if not np.isnan(test_pvalue) else "—")

                    # ── Decision pathway display ──────────────────────────────────
                    with st.expander(
                        f"{_t.get('stat_decision_title', "")} — "
                        f"{_t.get('target_gene', "")} {i+1} / "
                        f"Group {j+1}",
                        expanded=False
                    ):
                        st.markdown(_t.get('stat_decision_steps', ''))

                        sw_ctrl_sym = "✅" if control_normal else "❌"
                        sw_smp_sym  = "✅" if sample_normal  else "❌"

                        if n_ctrl >= _MIN_N_SHAPIRO and n_smp >= _MIN_N_SHAPIRO:
                            st.markdown(
                                f"{_t.get('stat_shapiro_title', "")}  \n"
                                f"- Control: W={shapiro_control.statistic:.4f}, "
                                f"p={shapiro_control.pvalue:.4f} {sw_ctrl_sym} "
                                f"{_t.get('stat_normal', "") if control_normal else _t.get('stat_nonnormal', "")}  \n"
                                f"- {_t.get('patient_group', "")} {j+1}: "
                                f"W={shapiro_sample.statistic:.4f}, "
                                f"p={shapiro_sample.pvalue:.4f} {sw_smp_sym} "
                                f"{_t.get('stat_normal', "") if sample_normal else _t.get('stat_nonnormal', "")}"
                            )
                        else:
                            st.info(
                                f"ℹ️ **Shapiro-Wilk atlandı** — n={min(n_ctrl, n_smp)} "
                                f"(gerekli minimum: {_MIN_N_SHAPIRO}). "
                                f"Küçük örneklemde Shapiro-Wilk güvenilir sonuç vermez; "
                                f"normallik varsayılarak parametrik test uygulandı."
                            )

                        if control_normal and sample_normal:
                            lev_sym = "✅" if equal_variance else "⚠️"
                            st.markdown(
                                f"{_t.get('stat_levene_title', "")}  \n"
                                f"- F={levene_test.statistic:.4f}, p={levene_test.pvalue:.4f} "
                                f"{lev_sym} {_t.get('stat_equal_var', "") if equal_variance else _t.get('stat_unequal_var', "")}"
                            )
                        else:
                            st.markdown(_t.get('stat_levene_skipped', ''))

                        if not control_normal or not sample_normal:
                            reason = _t.get('stat_reason_nonnormal', '')
                        elif equal_variance:
                            reason = _t.get('stat_reason_normal_equal', '')
                        else:
                            reason = _t.get('stat_reason_normal_unequal', '')

                        st.success(
                            f"{_t.get('stat_selected_test', "")} {test_method}  \n"
                            f"{_t.get('stat_reason', "")} {reason}  \n"
                            f"{_t.get('stat_result', "")} p = {test_pvalue:.4f} → **{significance}**"
                        )

                        if num_patient_groups >= 2:
                            st.caption(_t.get('stat_multigroup_note', ''))
                    # ─────────────────────────────────────────────────────────────

                    stats_data.append({
                        "__target_gene__":   f"Gene {i+1}",
                        "__patient_group__": f"Group {j+1}",
                        "__test_type__":     test_type,
                        "__test_method__":   test_method,
                        "__pvalue__":   test_pvalue,
                        "__significance__":  significance,
                        "Comparison": f"Control vs Group {j+1}"
                    })

                    data.append({
                        "__target_gene__":         f"Gene {i+1}",
                        "__patient_group__":       f"Group {j+1}",
                        "__ddct__":      delta_delta_ct,
                        "__fc__": expression_change,
                        "__pfaffl__":        pfaffl_ratio,
                        "E target":                                          round(E_target, 4),
                        "E ref":                                             round(E_ref, 4),
                        "__regulation__":   regulation_status,
                        "__dct_ctrl__":    average_control_delta_ct,
                        "__dct_patient__":    average_sample_delta_ct
                    })

    # ─── MULTI-GROUP ANALYSIS (3+ patient groups per target gene) ────────────────
    # Collect all ΔCq arrays per target gene for omnibus testing
    multigroup_results = []   # records for display / PDF

    for i in range(num_target_genes):
        # Pull per-group ΔCt values from stats_data provenance via data dict
        # Re-derive from input_values_table (source of truth after outlier removal)
        gene_label = f"Gene {i+1}"

        ctrl_dct = [
            float(d["__dct_ctrl__"])
            for d in input_values_table
            if d.get("Grup") == "Control"
            and d.get("__target_gene__") == gene_label
            and d.get("__dct_ctrl__") not in ("EXCLUDED", None)
            and d.get("Outlier Excluded", "No") == "No"
        ]

        patient_dcts = {}
        for j in range(num_patient_groups):
            pg_label = f"Group {j+1}"
            vals = [
                float(d["__dct_patient__"])
                for d in input_values_table
                if d.get("Grup") == pg_label
                and d.get("__target_gene__") == gene_label
                and d.get("__dct_patient__") not in ("EXCLUDED", None)
                and d.get("Outlier Excluded", "No") == "No"
            ]
            if vals:
                patient_dcts[pg_label] = vals

        if not ctrl_dct or not patient_dcts:
            continue

        # Convert ΔCq lists to RQ (2^-ΔCt) for all statistical tests
        all_groups_dct  = [ctrl_dct] + list(patient_dcts.values())
        all_groups      = [list(2 ** (-np.array(g))) for g in all_groups_dct]
        all_group_names = ["Control"] + list(patient_dcts.keys())
        n_groups        = len(all_groups)

        if n_groups < 3:
            # 2-group: already handled above — just note it
            multigroup_results.append({
                "gene": gene_label,
                "n_groups": n_groups,
                "note": "2-group comparison — pairwise test already reported above.",
                "omnibus_test": "—", "omnibus_p": None,
                "posthoc": [], "correction": "—"
            })
            continue

        # ── Omnibus test selection ────────────────────────────────────────────────
        normality_ok  = all(
            (len(g) < 8 or stats.shapiro(g).pvalue > 0.05)
            for g in all_groups if len(g) >= 3
        )
        levene_p      = stats.levene(*all_groups).pvalue if n_groups >= 2 else 1.0
        variance_ok   = levene_p > 0.05

        if normality_ok and variance_ok:
            omnibus_stat, omnibus_p = stats.f_oneway(*all_groups)
            omnibus_test  = "One-way ANOVA"
            omnibus_type  = "parametric"
            posthoc_method = "Tukey HSD"
        elif normality_ok and not variance_ok:
            # Welch ANOVA (scipy ≥ 1.11) — fallback to regular ANOVA if unavailable
            try:
                from scipy.stats import alexandergovern
                result = alexandergovern(*all_groups)
                omnibus_p    = result.pvalue
                omnibus_stat = result.statistic
            except Exception:
                omnibus_stat, omnibus_p = stats.f_oneway(*all_groups)
            omnibus_test   = "Welch ANOVA (unequal variances)"
            omnibus_type   = "parametric"
            posthoc_method = "Games-Howell (approx. via pairwise Welch t-test + FDR)"
        else:
            omnibus_stat, omnibus_p = stats.kruskal(*all_groups)
            omnibus_test   = "Kruskal-Wallis"
            omnibus_type   = "non-parametric"
            posthoc_method = "Dunn (pairwise Mann-Whitney U)"

        omnibus_sig = _t.get('multigroup_significant', '') if omnibus_p < 0.05 else _t.get('multigroup_not_significant', '')

        # ── Post-hoc pairwise comparisons ────────────────────────────────────────
        pairs      = []
        raw_pvals  = []

        for a in range(n_groups):
            for b in range(a + 1, n_groups):
                g_a, g_b = all_groups[a], all_groups[b]
                if omnibus_type == "parametric" and variance_ok:
                    p = stats.ttest_ind(g_a, g_b).pvalue
                elif omnibus_type == "parametric" and not variance_ok:
                    p = stats.ttest_ind(g_a, g_b, equal_var=False).pvalue
                else:
                    p = stats.mannwhitneyu(g_a, g_b, alternative="two-sided").pvalue
                pairs.append((all_group_names[a], all_group_names[b]))
                raw_pvals.append(p)

        # ── Multiple comparison correction ───────────────────────────────────────
        n_tests = len(raw_pvals)
        bonf_pvals = [min(p * n_tests, 1.0) for p in raw_pvals]

        # FDR Benjamini-Hochberg
        ranked     = sorted(range(n_tests), key=lambda k: raw_pvals[k])
        fdr_pvals  = [1.0] * n_tests
        for rank, idx in enumerate(ranked):
            fdr_pvals[idx] = min(raw_pvals[idx] * n_tests / (rank + 1), 1.0)
        # Enforce monotonicity
        for k in range(n_tests - 2, -1, -1):
            fdr_pvals[ranked[k]] = min(fdr_pvals[ranked[k]], fdr_pvals[ranked[k + 1]])

        posthoc_rows = []
        for idx, (pa, pb) in enumerate(pairs):
            posthoc_rows.append({
                "Comparison":        f"{pa} vs {pb}",
                "Raw p":             round(raw_pvals[idx], 4),
                "Bonferroni p":      round(bonf_pvals[idx], 4),
                "FDR p (B-H)":       round(fdr_pvals[idx], 4),
                "Sig (raw)":         "✅" if raw_pvals[idx]  < 0.05 else "—",
                "Sig (Bonferroni)":  "✅" if bonf_pvals[idx] < 0.05 else "—",
                "Sig (FDR)":         "✅" if fdr_pvals[idx]  < 0.05 else "—",
            })

        multigroup_results.append({
            "gene":          gene_label,
            "n_groups":      n_groups,
            "omnibus_test":  omnibus_test,
            "omnibus_type":  omnibus_type,
            "omnibus_p":     omnibus_p,
            "omnibus_sig":   omnibus_sig,
            "posthoc_method": posthoc_method,
            "posthoc_rows":  posthoc_rows,
            "normality_ok":  normality_ok,
            "variance_ok":   variance_ok,
            "note":          None
        })

    # ─────────────────────────────────────────────────────────────────────────────

    # ═══════════════════════════════════════════════════════════════════════════════
    # SEKME 2: SONUÇLAR
    # ═══════════════════════════════════════════════════════════════════════════════
    with tab_results:

        # ── Multi-group display ───────────────────────────────────────────────────
        if any(r["n_groups"] >= 3 for r in multigroup_results):
            st.markdown("---")
            st.markdown(_t.get('multigroup_title', ''))

            with st.expander(_t.get('multigroup_expander', ''), expanded=False):
                st.markdown("""
    **When is multi-group analysis applied?**  
    Automatically activated when **≥ 3 groups** (control + 2 or more patient groups) are present for a target gene.  
    This addresses the limitation of pairwise-only testing, which inflates Type I error when multiple comparisons are made without correction.

    **Test selection logic (automatic):**

    | Condition | Test |
    |---|---|
    | All groups normal + equal variances | One-way ANOVA → Tukey HSD |
    | All groups normal + unequal variances | Welch ANOVA → Games-Howell |
    | Any group non-normal | Kruskal-Wallis → Dunn (Mann-Whitney U) |

    **Multiple comparison correction:**
    - **Bonferroni**: conservative, controls family-wise error rate (FWER). Best when few comparisons.
    - **FDR (Benjamini-Hochberg)**: controls false discovery rate. Better power for many comparisons.

    **Recommendation:** Report both, discuss which is more appropriate for your study design.  
    **Reference:** Dunn OJ. *J Am Stat Assoc* 1961; Benjamini & Hochberg. *J R Stat Soc B* 1995.
    """)

            for res in multigroup_results:
                if res["n_groups"] < 3:
                    continue

                st.markdown(f"### 🧬 {res['gene']} — {res['n_groups']} {_t.get('patient_group', "").replace('🩸 ', '')}")

                if res["normality_ok"] and res["variance_ok"]:
                    st.success(_t.get('multigroup_decision_normal_equal', ''))
                elif res["normality_ok"] and not res["variance_ok"]:
                    st.warning(_t.get('multigroup_decision_normal_unequal', ''))
                else:
                    st.warning(_t.get('multigroup_decision_nonnormal', ''))

                omni_col1, omni_col2, omni_col3 = st.columns(3)
                omni_col1.metric(_t.get('multigroup_omnibus_test', ''), res["omnibus_test"])
                omni_col2.metric(_t.get('multigroup_pvalue', ''), f"{res['omnibus_p']:.4f}")
                omni_col3.metric(_t.get('multigroup_result', ''), res["omnibus_sig"])

                if res["omnibus_p"] >= 0.05:
                    st.info(_t.get('multigroup_omnibus_ns', ''))

                st.markdown(f"{_t.get('multigroup_posthoc_label', "")} {res['posthoc_method']} — Bonferroni & FDR")
                ph_df = pd.DataFrame(res["posthoc_rows"])
                st.dataframe(ph_df, use_container_width=True)

                fig_ph = go.Figure()
                comparisons = [r["Comparison"] for r in res["posthoc_rows"]]
                fig_ph.add_trace(go.Bar(name="Raw p", x=comparisons, y=[r["Raw p"] for r in res["posthoc_rows"]], marker_color="#4C72B0"))
                fig_ph.add_trace(go.Bar(name="Bonferroni p", x=comparisons, y=[r["Bonferroni p"] for r in res["posthoc_rows"]], marker_color="#DD8452"))
                fig_ph.add_trace(go.Bar(name="FDR p (B-H)", x=comparisons, y=[r["FDR p (B-H)"] for r in res["posthoc_rows"]], marker_color="#55A868"))
                fig_ph.add_hline(y=0.05, line_dash="dash", line_color="red", annotation_text="a = 0.05", annotation_position="right")
                fig_ph.update_layout(barmode="group", title=f"{res['gene']} — Post-hoc p-values", yaxis_title="p-value", height=350)
                st.plotly_chart(fig_ph, use_container_width=True, key=f"posthoc_{res['gene']}")

                ph_csv = ph_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label=f"{_t.get('multigroup_dl_button', "")} {res['gene']}",
                    data=ph_csv,
                    file_name=f"posthoc_{res['gene'].replace(' ', '_')}.csv",
                    mime="text/csv",
                    key=f"ph_dl_{res['gene']}"
                )

        elif num_patient_groups >= 2 and multigroup_results:
            st.markdown("---")
            st.info(_t.get('multigroup_2group_note', ''))


        if input_values_table:
            st.subheader(f" {_t.get('gr_tbl', "")}")
            # Rename fixed keys to translated column headers for display
            _ivt_rename = {
                "__sample_num__":   _t.get("sample_number", "Sample #"),
                "__target_gene__":  _t.get("target_gene",   "Gene"),
                "Grup":             _t.get("Grup",          "Group"),
                "__target_ct__":    _t.get("target_ct",     "Target Cq"),
                "__ref_ct__":       _t.get("reference_ct",  "Ref Ct"),
                "__dct_ctrl__":     _t.get("delta_ct_control", "ΔCq Control"),
                "__dct_patient__":  _t.get("delta_ct_patient", "ΔCq Patient"),
                "Outlier Excluded": _t.get("pdf_outlier_col", "Outlier Excluded"),
            }
            input_df = pd.DataFrame(input_values_table).rename(columns=_ivt_rename)
            st.write(input_df)
            csv = input_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label=_t.get('download_csv', ''),
                data=csv, file_name="giris_verileri.csv", mime="text/csv",
                key="dl_input_csv")

        # Sonuçlar Tablosunu Göster
        if data:
            st.subheader(f" {_t.get('nil_mine', "")}")
            _data_rename = {
                "__target_gene__":  _t.get("target_gene",   "Gene"),
                "__patient_group__":_t.get("patient_group", "Group"),
                "__ddct__":         _t.get("delta_delta_ct","ΔΔCq"),
                "__fc__":           _t.get("gene_expression_change", "2^(-ΔΔCq)"),
                "__pfaffl__":       _t.get("pfaffl_ratio",  "Pfaffl"),
                "__regulation__":   _t.get("regulation_status", "Regulation"),
                "__dct_ctrl__":     _t.get("delta_ct_control", "ΔCq Control"),
                "__dct_patient__":  _t.get("delta_ct_patient", "ΔCq Patient"),
            }
            df = pd.DataFrame(data).rename(columns=_data_rename)
            st.write(df)

        # İstatistik Sonuçları
        if stats_data:
            st.subheader(f" {_t.get('statistical_results', "")}")
            _stats_rename = {
                "__target_gene__":  _t.get("target_gene",  "Gene"),
                "__patient_group__":_t.get("patient_group","Group"),
                "__test_type__":    _t.get("test_type",    "Test Type"),
                "__test_method__":  _t.get("test_method",  "Test Method"),
                "__pvalue__":       _t.get("test_pvalue",  "p-value"),
                "__significance__": _t.get("significance", "Significance"),
            }
            stats_df = pd.DataFrame(stats_data).rename(columns=_stats_rename)
            st.write(stats_df)
            csv_stats = stats_df.to_csv(index=False).encode("utf-8")
            st.download_button(
                label=_t.get('download_csv', ''),
                data=csv_stats,
                file_name="istatistik_sonuclari.csv",
                mime="text/csv",
                key="dl_stats_csv")

        # ─── MULTI-GENE P-VALUE CORRECTION ───────────────────────────────────────────
        if stats_data and num_target_genes >= 2:
            st.markdown("---")
            st.markdown(_t.get('multigene_title', ''))

            with st.expander(_t.get('multigene_expander', ''), expanded=False):
                st.markdown("""
    When testing **multiple target genes** simultaneously, the probability of obtaining 
    at least one false positive increases with the number of tests performed 
    (family-wise error inflation). For example, testing 5 genes at α = 0.05 gives a 
    ~23% chance of at least one spurious significant result by chance alone.

    **Standard practice in multi-gene expression studies requires correction:**

    | Method | Controls | Best for |
    |---|---|---|
    | **Bonferroni** | Family-wise error rate (FWER) | Few genes, conservative |
    | **FDR (Benjamini-Hochberg)** | False discovery rate | Many genes, more power |

    **References:** Benjamini & Hochberg. *J R Stat Soc B* 1995;  
    Ge Y et al. *Bioinformatics* 2003; Storey JD. *J R Stat Soc B* 2002.
    """)

            pval_key  = "__pvalue__"
            gene_key  = "__target_gene__"
            group_key = "__patient_group__"

            correction_rows = [
                {
                    "Gene":   r[gene_key],
                    "Group":  r[group_key],
                    "Raw p":  r[pval_key],
                    "Test":   r.get("__test_method__", "—"),
                }
                for r in stats_data
                if r.get(pval_key) is not None
            ]

            if not correction_rows:
                st.info(_t.get('multigene_no_data', ''))
            else:
                n_tests   = len(correction_rows)
                raw_pvals = [r["Raw p"] for r in correction_rows]

                bonf = [min(p * n_tests, 1.0) for p in raw_pvals]

                ranked = sorted(range(n_tests), key=lambda k: raw_pvals[k])
                fdr    = [1.0] * n_tests
                for rank, idx in enumerate(ranked):
                    fdr[idx] = min(raw_pvals[idx] * n_tests / (rank + 1), 1.0)
                for k in range(n_tests - 2, -1, -1):
                    fdr[ranked[k]] = min(fdr[ranked[k]], fdr[ranked[k + 1]])

                for idx, row in enumerate(correction_rows):
                    row["Bonferroni p"]     = round(bonf[idx], 4)
                    row["FDR p (B-H)"]      = round(fdr[idx],  4)
                    row["Sig (raw)"]        = "✅" if raw_pvals[idx] < 0.05 else "—"
                    row["Sig (Bonferroni)"] = "✅" if bonf[idx]      < 0.05 else "—"
                    row["Sig (FDR)"]        = "✅" if fdr[idx]        < 0.05 else "—"

                corr_df = pd.DataFrame(correction_rows)
                st.dataframe(corr_df, use_container_width=True)

                n_raw_sig  = sum(1 for p in raw_pvals if p < 0.05)
                n_bonf_sig = sum(1 for p in bonf       if p < 0.05)
                n_fdr_sig  = sum(1 for p in fdr         if p < 0.05)

                sum_col1, sum_col2, sum_col3 = st.columns(3)
                sum_col1.metric(_t.get('multigene_sig_raw', ''),  f"{n_raw_sig} / {n_tests}")
                sum_col2.metric(_t.get('multigene_sig_bonf', ''), f"{n_bonf_sig} / {n_tests}")
                sum_col3.metric(_t.get('multigene_sig_fdr', ''),  f"{n_fdr_sig} / {n_tests}")

                if n_raw_sig > n_fdr_sig:
                    st.warning(_t.get('multigene_warning', '').format(lost=n_raw_sig - n_fdr_sig))
                elif n_raw_sig == n_fdr_sig and n_raw_sig > 0:
                    st.success(_t.get('multigene_success', '').format(n=n_raw_sig))
                elif n_raw_sig == 0:
                    st.info(_t.get('multigene_no_sig', ''))

                fig_corr = go.Figure()
                labels = [f"{r['Gene']} / {r['Group']}" for r in correction_rows]
                fig_corr.add_trace(go.Bar(name="Raw p",        x=labels, y=raw_pvals, marker_color="#4C72B0"))
                fig_corr.add_trace(go.Bar(name="Bonferroni p", x=labels, y=bonf,      marker_color="#DD8452"))
                fig_corr.add_trace(go.Bar(name="FDR p (B-H)",  x=labels, y=fdr,       marker_color="#55A868"))
                fig_corr.add_hline(y=0.05, line_dash="dash", line_color="red", annotation_text="a = 0.05", annotation_position="right")
                fig_corr.update_layout(
                    barmode="group",
                    title=_t.get('multigene_chart_title', ''),
                    yaxis_title="p-value",
                    xaxis_title=f"{_t.get('target_gene', "")} / {_t.get('patient_group', "")}",
                    height=380
                )
                st.plotly_chart(fig_corr, use_container_width=True, key="multigene_corr_chart")

                corr_csv = corr_df.to_csv(index=False).encode("utf-8")
                st.download_button(
                    label=_t.get('multigene_dl_button', ''),
                    data=corr_csv,
                    file_name="multi_gene_correction.csv",
                    mime="text/csv",
                    key="multigene_corr_dl"
                )

        elif stats_data and num_target_genes == 1:
            st.markdown("---")
            st.info(_t.get('multigene_1gene_note', ''))

        # ── Çoklu Gen Karşılaştırma Grafiği ──────────────────────────────────────
        st.markdown("---")
        if data and num_target_genes >= 2:
            st.subheader(f"📊 {_t.get('multigene_fc_chart_title', 'Multi-Gene Expression Comparison')}")

            # Collect fold changes per gene per group
            fc_key  = "__fc__"
            pf_key  = "__pfaffl__"
            tg_key2 = "__target_gene__"
            pg_key2 = "__patient_group__"
            reg_key = "__regulation__"

            method_choice = st.radio(
                "Method",
                ["2^(-ΔΔCq)", "Pfaffl"],
                horizontal=True,
                key="multigene_chart_method"
            )

            # Build matrix: genes × groups
            genes  = sorted(set(r[tg_key2] for r in data))
            groups = sorted(set(r[pg_key2] for r in data))
            palette = ['#3f51b5','#e91e63','#009688','#ff9800','#9c27b0','#795548']

            fig_multi = go.Figure()
            for gi, gene in enumerate(genes):
                y_vals = []
                for grp in groups:
                    match = [r for r in data if r[tg_key2] == gene and r[pg_key2] == grp]
                    if match:
                        val = match[0][fc_key] if method_choice == "2^(-ΔΔCq)" else match[0][pf_key]
                        y_vals.append(round(val, 4) if isinstance(val, float) else 0)
                    else:
                        y_vals.append(0)
                fig_multi.add_trace(go.Bar(
                    name=gene,
                    x=groups,
                    y=y_vals,
                    marker_color=palette[gi % len(palette)],
                    text=[f"{v:.3f}" for v in y_vals],
                    textposition='outside',
                ))

            fig_multi.add_hline(y=1, line_dash="dash", line_color="black",
                                line_width=1, annotation_text="No change (1.0)",
                                annotation_position="right")
            fig_multi.update_layout(
                barmode='group',
                title=f"Gene Expression Fold Change — {method_choice}",
                xaxis_title=_t.get('patient_group', ''),
                yaxis_title=f"Fold Change ({method_choice})",
                legend_title=_t.get('target_gene', ''),
                height=420,
                plot_bgcolor='white',
                yaxis=dict(gridcolor='#eeeeee'),
            )
            st.plotly_chart(fig_multi, use_container_width=True, key="multigene_fc_chart")

            # Second chart: log2 fold change heatmap-style grouped bar
            if st.checkbox("Show log2 scale", key="multigene_log2"):
                import math
                fig_log = go.Figure()
                for gi, gene in enumerate(genes):
                    y_log = []
                    for grp in groups:
                        match = [r for r in data if r[tg_key2] == gene and r[pg_key2] == grp]
                        if match:
                            val = match[0][fc_key] if method_choice == "2^(-ΔΔCq)" else match[0][pf_key]
                            y_log.append(round(math.log2(val), 4) if isinstance(val, float) and val > 0 else 0)
                        else:
                            y_log.append(0)
                    fig_log.add_trace(go.Bar(
                        name=gene, x=groups, y=y_log,
                        marker_color=palette[gi % len(palette)],
                        text=[f"{v:.3f}" for v in y_log],
                        textposition='outside',
                    ))
                fig_log.add_hline(y=0, line_dash="dash", line_color="black", line_width=1)
                fig_log.update_layout(
                    barmode='group',
                    title=f"Gene Expression log2(Fold Change) — {method_choice}",
                    xaxis_title=_t.get('patient_group', ''),
                    yaxis_title="log2(Fold Change)",
                    legend_title=_t.get('target_gene', ''),
                    height=420, plot_bgcolor='white',
                    yaxis=dict(gridcolor='#eeeeee', zeroline=True, zerolinecolor='black'),
                )
                st.plotly_chart(fig_log, use_container_width=True, key="multigene_fc_log2")
                st.caption("log2 > 0 = upregulated, log2 < 0 = downregulated, log2 = 0 = no change")

        # ── Dağılım Grafikleri ────────────────────────────────────────────────────
        st.markdown("---")

        # Allow user to choose which values to display in the distribution plot:
        # RQ (2^-ΔCq), raw ΔCt, or ΔΔCq (relative to control mean).
        plot_mode = st.radio(
            _t.get('dist_plot_mode_label', ''),
            options=[
                _t.get('dist_plot_rq', ''),
                _t.get('dist_plot_dct', ''),
                _t.get('dist_plot_ddct', ''),
            ],
            index=0,
            horizontal=True,
            key="dist_plot_mode",
            help=_t.get('dist_plot_help', '')
        )

        # Map selected option back to mode identifier
        if plot_mode == _t.get('dist_plot_rq', ''):
            _plot_mode_id = "RQ"
        elif plot_mode == _t.get('dist_plot_ddct', ''):
            _plot_mode_id = "DDCT"
        else:
            _plot_mode_id = "DCT"

        for i in range(num_target_genes):
            st.subheader(f"{_t.get('target_gene', "")} {i+1} - {_t.get('distribution_graph', "")}")

            control_target_ct_values = [
                d["__target_ct__"] 
                for d in input_values_table
                if d["Grup"] == "Control" and
                   d["__target_gene__"] == f"Gene {i+1}" and
                   d.get("__target_ct__") not in ("EXCLUDED", None) and
                   d.get("Outlier Excluded", "No") == "No"
            ]
            control_reference_ct_values = [
                d["__ref_ct__"] 
                for d in input_values_table
                if d["Grup"] == "Control" and
                   d["__target_gene__"] == f"Gene {i+1}" and
                   d.get("__ref_ct__") not in ("EXCLUDED", None) and
                   d.get("Outlier Excluded", "No") == "No"
            ]

            if len(control_target_ct_values) == 0 or len(control_reference_ct_values) == 0:
                st.error(f" {_t.get('error_missing_control_data', "").format(i=i+1)}")
                continue

            control_delta_ct = np.array(control_target_ct_values, dtype=float) - np.array(control_reference_ct_values, dtype=float)
            average_control_delta_ct = np.mean(control_delta_ct)

            # ── Convert values based on selected plot mode ────────────────────────
            def _transform(dct_array, mode, ctrl_mean):
                if mode == "RQ":
                    return 2 ** (-np.array(dct_array))
                elif mode == "DDCT":
                    return np.array(dct_array) - ctrl_mean
                else:
                    return np.array(dct_array)

            def _yaxis_label(mode):
                if mode == "RQ":    return "RQ (2^-ΔCq)"
                elif mode == "DDCT": return "ΔΔCq (vs control mean)"
                else:                return "ΔCq"

            ctrl_plot_vals = _transform(control_delta_ct, _plot_mode_id, average_control_delta_ct)
            avg_ctrl_plot  = float(np.mean(ctrl_plot_vals))

            fig = go.Figure()
            fig.add_trace(go.Scatter(
                x=[0.8, 1.2],
                y=[avg_ctrl_plot, avg_ctrl_plot],
                mode='lines',
                line=dict(color='black', width=4),
                name=_t.get('control_group_avg', '')
            ))

            for j in range(num_patient_groups):
                sample_dct_raw = [
                    float(d["__dct_patient__"])
                    for d in input_values_table
                    if d["Grup"] == f"Group {j+1}" and
                       d["__target_gene__"] == f"Gene {i+1}" and
                       d.get("__dct_patient__") not in ("EXCLUDED", None) and
                       d.get("Outlier Excluded", "No") == "No"
                ]
                if not sample_dct_raw:
                    continue
                smp_plot_vals   = _transform(sample_dct_raw, _plot_mode_id, average_control_delta_ct)
                avg_smp_plot    = float(np.mean(smp_plot_vals))
                fig.add_trace(go.Scatter(
                    x=[(j + 1.8), (j + 2.2)],
                    y=[avg_smp_plot, avg_smp_plot],
                    mode='lines',
                    line=dict(color='black', width=4),
                    name=f"{_t.get('patient_group', "")} {j+1} {_t.get('avg', "")}"
                ))

            fig.add_trace(go.Scatter(
                x=np.ones(len(ctrl_plot_vals)) + np.random.uniform(-0.05, 0.05, len(ctrl_plot_vals)),
                y=ctrl_plot_vals,
                mode='markers',
                name="Control",
                marker=dict(color='blue'),
                text=[f"Control — {_yaxis_label(plot_mode)}={v:.4f}, replicate {idx+1}"
                      for idx, v in enumerate(ctrl_plot_vals)],
                hoverinfo='text'
            ))

            for j in range(num_patient_groups):
                sample_dct_raw = [
                    float(d["__dct_patient__"])
                    for d in input_values_table
                    if d["Grup"] == f"Group {j+1}" and
                       d["__target_gene__"] == f"Gene {i+1}" and
                       d.get("__dct_patient__") not in ("EXCLUDED", None) and
                       d.get("Outlier Excluded", "No") == "No"
                ]
                if not sample_dct_raw:
                    continue
                smp_plot_vals = _transform(sample_dct_raw, _plot_mode_id, average_control_delta_ct)
                fig.add_trace(go.Scatter(
                    x=np.ones(len(smp_plot_vals)) * (j + 2) + np.random.uniform(-0.05, 0.05, len(smp_plot_vals)),
                    y=smp_plot_vals,
                    mode='markers',
                    name=f"Group {j+1}",
                    marker=dict(color='red'),
                    text=[f"Group {j+1} — {_yaxis_label(plot_mode)}={v:.4f}, replicate {idx+1}"
                          for idx, v in enumerate(smp_plot_vals)],
                    hoverinfo='text'
                ))

            fig.update_layout(
                title=f"{_t.get('target_gene', "")} {i+1} — {_yaxis_label(plot_mode)} Distribution",
                xaxis=dict(
                    tickvals=[1] + [j + 2 for j in range(num_patient_groups)],
                    ticktext=["Control"] + [f"Group {j+1}" for j in range(num_patient_groups)],
                    title=_t.get('x_axis_title', "")
                ),
                yaxis=dict(title=_yaxis_label(plot_mode)),
                showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True, key=f"dist_chart_{i}")

    # ═══════════════════════════════════════════════════════════════════════════════
    # SEKME 3: RAPOR
    # ═══════════════════════════════════════════════════════════════════════════════

    # ── PDF Font Sistemi ─────────────────────────────────────────────────────────
    # Streamlit Cloud için: packages.txt'e fonts-noto ve fonts-noto-extra ekleyin
    # requirements.txt'e: arabic-reshaper>=3.0.0  python-bidi>=0.4.2 ekleyin

    def _find_font(candidates):
        """İlk bulunan geçerli font yolunu döndür."""
        import glob as _glob
        for p in candidates:
            if os.path.exists(p):
                return p
        # Sistem genelinde TTF ara
        all_ttf = _glob.glob('/usr/share/fonts/**/*.ttf', recursive=True)
        return all_ttf[0] if all_ttf else None

    # Noto Sans: Türkçe, Fransızca, Almanca, İspanyolca
    _NOTO_REGULAR = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
        '/usr/share/fonts/opentype/noto/NotoSans-Regular.otf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',          # Streamlit fallback
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',          # son çare
    ])
    _NOTO_BOLD = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf',
        '/usr/share/fonts/opentype/noto/NotoSans-Bold.otf',
        '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    ])
    # Noto Sans Arabic: Arapça
    _NOTO_ARABIC = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
        '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Regular.ttf',
        '/usr/share/fonts/opentype/noto/NotoSansArabic-Regular.otf',
        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',      # fallback
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
    ])
    _NOTO_ARABIC_BOLD = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSansArabic-Bold.ttf',
        '/usr/share/fonts/truetype/noto/NotoNaskhArabic-Bold.ttf',
        '/usr/share/fonts/truetype/noto/NotoSansArabic-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
    ])

    from reportlab.pdfbase.pdfmetrics import registerFontFamily as _regFamily

    try:
        pdfmetrics.registerFont(TTFont('NotoSans',      _NOTO_REGULAR))
        pdfmetrics.registerFont(TTFont('NotoSans-Bold', _NOTO_BOLD))
        _regFamily('NotoSans', normal='NotoSans', bold='NotoSans-Bold',
                   italic='NotoSans', boldItalic='NotoSans-Bold')
        PDF_FONT      = 'NotoSans'
        PDF_FONT_BOLD = 'NotoSans-Bold'
    except Exception:
        PDF_FONT      = 'Helvetica'
        PDF_FONT_BOLD = 'Helvetica-Bold'

    # Arapça için ayrı font kaydı
    _arabic_font_ok = False
    if _NOTO_ARABIC and _NOTO_ARABIC != _NOTO_REGULAR:
        try:
            pdfmetrics.registerFont(TTFont('NotoArabic',      _NOTO_ARABIC))
            pdfmetrics.registerFont(TTFont('NotoArabic-Bold', _NOTO_ARABIC_BOLD or _NOTO_ARABIC))
            _regFamily('NotoArabic', normal='NotoArabic', bold='NotoArabic-Bold',
                       italic='NotoArabic', boldItalic='NotoArabic-Bold')
            ARABIC_FONT      = 'NotoArabic'
            ARABIC_FONT_BOLD = 'NotoArabic-Bold'
            _arabic_font_ok  = True
        except Exception:
            pass

    if not _arabic_font_ok:
        ARABIC_FONT      = PDF_FONT        # fallback: NotoSans veya Helvetica
        ARABIC_FONT_BOLD = PDF_FONT_BOLD

    # Eski kod uyumluluğu için alias
    REGISTERED_FONT      = PDF_FONT
    REGISTERED_FONT_BOLD = PDF_FONT_BOLD

    # matplotlib da Noto kullan
    try:
        import matplotlib as _mpl
        _noto_name = 'Noto Sans' if 'NotoSans' in PDF_FONT else 'DejaVu Sans'
        _mpl.rcParams['font.family'] = _noto_name
        _mpl.rcParams['axes.unicode_minus'] = False
    except Exception:
        pass

    def safe_str(text, lang='en'):
        """PDF için metni hazırla: XML kaçış + Arapça reshape/bidi."""
        if not isinstance(text, str):
            text = str(text)
        if lang == 'ar':
            try:
                import arabic_reshaper
                from bidi.algorithm import get_display
                text = get_display(arabic_reshaper.reshape(text))
            except ImportError:
                pass  # paket yoksa olduğu gibi bırak
        return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    def get_pdf_fonts(lang):
        """Dile göre (normal_font, bold_font) döndür."""
        if lang == 'ar':
            return ARABIC_FONT, ARABIC_FONT_BOLD
        return PDF_FONT, PDF_FONT_BOLD

    def create_pdf(results, stat_rows, input_df, language_code, multigroup_results=None):
        T   = translations[language_code]  # shorthand
        RTL = language_code == 'ar'        # sağdan sola dil mi?
        fn, fnb = get_pdf_fonts(language_code)

        buffer = BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            leftMargin=50, rightMargin=50, topMargin=60, bottomMargin=50
        )
        elements = []
        styles = getSampleStyleSheet()

        # Metin hizalaması: Arapça için sağ, diğerleri için sol/orta
        _body_align  = 2 if RTL else 0   # 0=left, 1=center, 2=right
        _title_align = 1                  # başlıklar her zaman ortalı

        title_style   = ParagraphStyle('RT',  parent=styles['Title'],   fontName=fnb, fontSize=20, textColor=colors.HexColor('#1a237e'), spaceAfter=6,  alignment=_title_align)
        sub_style     = ParagraphStyle('RS',  parent=styles['Normal'],  fontName=fn,  fontSize=10, textColor=colors.HexColor('#555555'), spaceAfter=4,  alignment=_title_align)
        h1_style      = ParagraphStyle('H1',  parent=styles['Heading1'],fontName=fnb, fontSize=13, textColor=colors.HexColor('#1a237e'), spaceBefore=16,spaceAfter=5,  alignment=_body_align)
        h2_style      = ParagraphStyle('H2',  parent=styles['Heading2'],fontName=fnb, fontSize=11, textColor=colors.HexColor('#283593'), spaceBefore=10,spaceAfter=4,  alignment=_body_align)
        body_style    = ParagraphStyle('BD',  parent=styles['Normal'],  fontName=fn,  fontSize=9,  leading=13, spaceAfter=4, alignment=_body_align)
        small_style   = ParagraphStyle('SM',  parent=styles['Normal'],  fontName=fn,  fontSize=8,  leading=11, textColor=colors.HexColor('#444444'), alignment=_body_align)
        caption_style = ParagraphStyle('CA',  parent=styles['Normal'],  fontName=fn,  fontSize=8,  textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=6)
        info_style    = ParagraphStyle('IN',  parent=styles['Normal'],  fontName=fn,  fontSize=9,  leading=13, backColor=colors.HexColor('#e8f4fd'), borderPad=6, leftIndent=8, rightIndent=8, spaceAfter=6, alignment=_body_align)
        warn_style    = ParagraphStyle('WN',  parent=styles['Normal'],  fontName=fn,  fontSize=9,  leading=13, backColor=colors.HexColor('#fff8e1'), borderPad=6, leftIndent=8, rightIndent=8, spaceAfter=6, alignment=_body_align)

        def hr():
            from reportlab.platypus import HRFlowable
            return HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=8, spaceBefore=4)

        def s(key, **kw):
            """Get translated string, format with kwargs, make PDF-safe + Arabic reshape."""
            txt = T.get(key, key)
            if kw:
                try: txt = txt.format(**kw)
                except Exception: pass
            return safe_str(txt, lang=language_code)

        from reportlab.platypus import Flowable as _Flowable

        def make_table(rows, col_widths=None, header=True):
            if not rows: return Spacer(1,1)
            styled_rows = []
            _cell_align = 2 if RTL else 1
            for ri, row in enumerate(rows):
                styled_row = []
                for cell in row:
                    # Hücre zaten ReportLab nesnesi ise (Paragraph vb.) olduğu gibi kullan
                    if isinstance(cell, _Flowable):
                        styled_row.append(cell)
                        continue
                    cell_str = safe_str(
                        str(cell) if not isinstance(cell, str) else cell,
                        lang=language_code
                    )
                    if ri == 0 and header:
                        p = Paragraph(cell_str, ParagraphStyle('TH', fontName=fnb, fontSize=7,
                                      textColor=colors.white, alignment=1))
                    else:
                        p = Paragraph(cell_str, ParagraphStyle('TD', fontName=fn, fontSize=7,
                                      alignment=_cell_align))
                    styled_row.append(p)
                styled_rows.append(styled_row)
            tbl = Table(styled_rows, colWidths=col_widths, repeatRows=1 if header else 0)
            tbl_style = [
                ('FONTNAME',    (0,0),(-1,-1), fn),
                ('ALIGN',       (0,0),(-1,-1), 'CENTER'),
                ('VALIGN',      (0,0),(-1,-1), 'MIDDLE'),
                ('GRID',        (0,0),(-1,-1), 0.3, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white, colors.HexColor('#f5f7ff')]),
                ('TOPPADDING',  (0,0),(-1,-1), 4),
                ('BOTTOMPADDING',(0,0),(-1,-1), 4),
            ]
            if header:
                tbl_style += [('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1a237e'))]
            tbl.setStyle(TableStyle(tbl_style))
            return tbl

        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── COVER PAGE ────────────────────────────────────────────────────────────
        elements.append(Spacer(1, 40))
        elements.append(Paragraph(safe_str("GeneQuantify"), title_style))
        elements.append(Paragraph(s('pdf_cover_subtitle'), sub_style))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(s('pdf_generated', now=now), sub_style))
        elements.append(Paragraph(
            safe_str("Cite: Yalçınkaya B (2026). GeneQuantify. Mol Cell Biochem. https://doi.org/10.1007/s11010-026-05621-y"),
            sub_style
        ))
        elements.append(Spacer(1, 20))
        elements.append(hr())
        elements.append(Spacer(1, 10))

        n_genes    = len(set(r.get("__target_gene__", '') for r in results))
        n_groups   = len(set(r.get("__patient_group__", '') for r in results))
        n_samples  = len(input_df)
        n_excluded = sum(1 for _, row in input_df.iterrows()
                         if str(row.get('Outlier Excluded', 'No')).startswith('Yes'))
        norm_method = s('pdf_summary_norm_multi') if num_ref_genes > 1 else s('pdf_summary_norm_single')

        summary_rows = [
            [s('pdf_summary_param'), s('pdf_summary_val')],
            [s('pdf_summary_genes'),   str(n_genes)],
            [s('pdf_summary_groups'),  str(n_groups)],
            [s('pdf_summary_samples'), str(n_samples)],
            [s('pdf_summary_excluded'),str(n_excluded)],
            [s('pdf_summary_tests'),   f"{len(stat_rows)} {s('pdf_summary_tests')}"],
            [s('pdf_summary_norm'),    norm_method],
            [s('pdf_summary_methods'), s('pdf_summary_methods_val')],
        ]
        elements.append(make_table(summary_rows, col_widths=[260, 200]))
        elements.append(Spacer(1, 14))
        elements.append(Paragraph(s('pdf_disclaimer'), small_style))
        elements.append(PageBreak())

        # ── SECTION 1: METHODS ────────────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s1_title'), h1_style))
        elements.append(hr())

        elements.append(Paragraph(s('pdf_s1_calc'), h2_style))
        elements.append(Paragraph(s('pdf_s1_calc_body'), body_style))
        elements.append(Paragraph(s('pdf_s1_classic'), body_style))
        elements.append(Paragraph(s('pdf_s1_pfaffl'), body_style))

        elements.append(Paragraph(s('pdf_s1_norm'), h2_style))
        if num_ref_genes > 1:
            elements.append(Paragraph(s('pdf_s1_norm_multi', n=num_ref_genes), body_style))
        else:
            elements.append(Paragraph(s('pdf_s1_norm_single'), body_style))

        elements.append(Paragraph(s('pdf_s1_eff'), h2_style))
        eff_cols = T.get('pdf_eff_cols', ['Gene','E(t)','Eff%(t)','E(r)','Eff%(r)','Diff%','Status'])
        eff_rows = [eff_cols]
        for i, eff in gene_efficiencies.items():
            e_t = eff["target_E"]; e_r = eff["ref_E"]
            t_pct = (e_t-1)*100; r_pct = (e_r-1)*100; diff = abs(t_pct-r_pct)
            status = s('pdf_eff_ok') if diff <= efficiency_threshold else s('pdf_eff_warn')
            eff_rows.append([f"{T.get('target_gene','Gene')} {i+1}",
                             f"{e_t:.4f}", f"{t_pct:.1f}%",
                             f"{e_r:.4f}", f"{r_pct:.1f}%",
                             f"{diff:.1f}%", status])
        cw7 = (letter[0]-100)/7
        elements.append(make_table(eff_rows, col_widths=[cw7]*7))
        elements.append(Paragraph(s('pdf_s1_eff_range', thr=efficiency_threshold), small_style))

        elements.append(Paragraph(s('pdf_s1_outlier'), h2_style))
        if outlier_enabled:
            if outlier_method == "Grubbs":
                elements.append(Paragraph(s('pdf_s1_grubbs', alpha=grubbs_alpha, n=n_excluded), body_style))
            else:
                elements.append(Paragraph(s('pdf_s1_iqr', k=iqr_multiplier, n=n_excluded), body_style))
            if n_excluded > 0:
                elements.append(Paragraph(s('pdf_s1_outlier_warn'), warn_style))
        else:
            elements.append(Paragraph(s('pdf_s1_outlier_off'), body_style))
        elements.append(PageBreak())

        # ── SECTION 2: INPUT DATA ─────────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s2_title'), h1_style))
        elements.append(hr())
        elements.append(Paragraph(s('pdf_s2_body'), body_style))
        elements.append(Spacer(1, 6))

        if not input_df.empty:
            cols = input_df.columns.tolist()
            page_w = letter[0] - 100
            cw = page_w / max(len(cols), 1)

            # Header satırı
            tbl_rows = [cols]

            for _, row in input_df.iterrows():
                is_excl = str(row.get('Outlier Excluded', 'No')).startswith('Yes')
                row_cells = []
                for v in row.tolist():
                    cell_str = safe_str(str(v) if v is not None else '', lang=language_code)
                    style = ParagraphStyle(
                        'EX' if is_excl else 'TD',
                        fontName=fn, fontSize=7, alignment=1,
                        textColor=colors.HexColor('#cc0000') if is_excl else colors.black
                    )
                    row_cells.append(Paragraph(cell_str, style))
                tbl_rows.append(row_cells)

            elements.append(make_table(tbl_rows, col_widths=[cw]*len(cols)))
        elements.append(PageBreak())

        # ── SECTION 3: RESULTS ────────────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s3_title'), h1_style))
        elements.append(hr())
        elements.append(Paragraph(s('pdf_s3_body'), body_style))
        elements.append(Spacer(1, 6))

        res_cols = T.get('pdf_res_cols', ['Gene','Group','ΔCq Ctrl','ΔCq Sample','ΔΔCq','2^(-ΔΔCq)','Pfaffl','Reg','Et','Er'])
        res_rows = [res_cols]
        for r in results:
            ddc = r.get("__ddct__", '')
            fc  = r.get("__fc__", '')
            pf  = r.get("__pfaffl__", '')
            dcc = r.get("__dct_ctrl__", '')
            dcs = r.get("__dct_patient__", '')
            et  = r.get('E target', ''); er = r.get('E ref', '')
            res_rows.append([
                str(r.get("__target_gene__",'')),
                str(r.get("__patient_group__",'')),
                f"{dcc:.4f}" if isinstance(dcc, float) else str(dcc),
                f"{dcs:.4f}" if isinstance(dcs, float) else str(dcs),
                f"{ddc:.4f}" if isinstance(ddc, float) else str(ddc),
                f"{fc:.4f}"  if isinstance(fc,  float) else str(fc),
                f"{pf:.4f}"  if isinstance(pf,  float) else str(pf),
                str(r.get("__regulation__", s('pdf_nochange'))),
                str(et), str(er)
            ])
        cw10 = (letter[0]-100)/10
        elements.append(make_table(res_rows, col_widths=[cw10]*10))
        elements.append(Spacer(1, 8))

        # Fold change bar chart
        if results:
            try:
                fig_fc, ax_fc = plt.subplots(figsize=(7, 3.5))
                labels_fc = [f"{r.get("__target_gene__",'')} /\n{r.get("__patient_group__",'')}" for r in results]
                vals_2  = [r.get("__fc__", 0) for r in results]
                vals_pf = [r.get("__pfaffl__", 0) for r in results]
                xr = range(len(labels_fc)); w = 0.35
                b1 = ax_fc.bar([i-w/2 for i in xr], vals_2,  width=w, label='2^(-ΔΔCq)', color='#3f51b5', alpha=0.85)
                b2 = ax_fc.bar([i+w/2 for i in xr], vals_pf, width=w, label='Pfaffl',    color='#e91e63', alpha=0.85)
                ax_fc.axhline(y=1, color='black', linestyle='--', linewidth=0.8, alpha=0.6)
                ax_fc.set_xticks(list(xr)); ax_fc.set_xticklabels(labels_fc, fontsize=7)
                ax_fc.set_ylabel('Fold Change', fontsize=9)
                ax_fc.set_title('Gene Expression Fold Change', fontsize=10, fontweight='bold')
                ax_fc.legend(fontsize=8)
                ax_fc.spines['top'].set_visible(False); ax_fc.spines['right'].set_visible(False)
                for bar in [*b1, *b2]:
                    ax_fc.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.02,
                               f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=6)
                plt.tight_layout()
                ib = BytesIO(); plt.savefig(ib, format='png', dpi=150, bbox_inches='tight'); plt.close(); ib.seek(0)
                elements.append(RLImage(ib, width=460, height=230))
                elements.append(Paragraph(s('pdf_fig1'), caption_style))
            except Exception:
                pass
        elements.append(PageBreak())

        # ── SECTION 4: STATISTICS ─────────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s4_title'), h1_style))
        elements.append(hr())
        elements.append(Paragraph(s('pdf_s4_body'), body_style))
        elements.append(Spacer(1, 6))

        # Determine if any gene has 3+ groups (multigroup scenario)
        _has_multigroup = (
            multigroup_results is not None
            and any(r.get("n_groups", 0) >= 3 for r in multigroup_results)
        )

        if _has_multigroup:
            # ── 4a. Pairwise Control comparisons (t-test) for all genes ──────────
            # Still show pairwise tests for genes with only 2 groups
            _pairwise_rows = [r for r in stat_rows]
            _2group_genes = set()
            if multigroup_results:
                for _mg in multigroup_results:
                    if _mg.get("n_groups", 0) < 3:
                        _2group_genes.add(_mg.get("gene", ""))

            _pairwise_show = [r for r in _pairwise_rows if r.get("__target_gene__", "") in _2group_genes]

            if _pairwise_show:
                elements.append(Paragraph(
                    safe_str("4.1 Pairwise Comparisons (Control vs Group) — 2-Group Genes"),
                    h2_style
                ))
                stat_cols = T.get('pdf_stat_cols', ['Gene','Comparison','Type','Method','p','Sig'])
                _pw_table = [stat_cols]
                for _r in _pairwise_show:
                    _pw_table.append([
                        str(_r.get("__target_gene__", '')),
                        str(_r.get('Comparison', '')),
                        str(_r.get("__test_type__", '')),
                        str(_r.get("__test_method__", '')),
                        f"{_r.get('__pvalue__', 0):.4f}",
                        str(_r.get("__significance__", '')),
                    ])
                cw6 = (letter[0]-100)/6
                elements.append(make_table(_pw_table, col_widths=[cw6]*6))
                elements.append(Spacer(1, 10))

            # ── 4b. Multi-group ANOVA/Kruskal-Wallis results ─────────────────────
            elements.append(Paragraph(
                safe_str("4.2 Multi-Group Comparison (≥3 Groups) — Omnibus + Post-hoc"),
                h2_style
            ))
            elements.append(Paragraph(
                safe_str(
                    "All inferential tests are performed on RQ values (2^\u2212\u0394Cq). "
                    "With \u22653 groups, an omnibus test (One-way ANOVA or Kruskal-Wallis) is applied first, "
                    "followed by pairwise post-hoc comparisons with Bonferroni and FDR correction. "
                    "Test selection is automatic based on normality (Shapiro-Wilk, n\u22658) and variance homogeneity (Levene)."
                ),
                body_style
            ))
            elements.append(Spacer(1, 6))

            for _mg in (multigroup_results or []):
                if _mg.get("n_groups", 0) < 3:
                    continue
                _gene = _mg.get("gene", "")
                _omni_test = _mg.get("omnibus_test", "—")
                _omni_p    = _mg.get("omnibus_p", None)
                _omni_sig  = _mg.get("omnibus_sig", "—")
                _norm_ok   = _mg.get("normality_ok", True)
                _var_ok    = _mg.get("variance_ok", True)
                _posthoc   = _mg.get("posthoc_method", "—")

                # Decision pathway text
                if _norm_ok and _var_ok:
                    _decision_txt = "Normal distribution + equal variances \u2192 One-way ANOVA + Tukey HSD"
                elif _norm_ok and not _var_ok:
                    _decision_txt = "Normal distribution + unequal variances \u2192 Welch ANOVA + Games-Howell"
                else:
                    _decision_txt = "Non-normal distribution \u2192 Kruskal-Wallis + Dunn (Mann-Whitney U)"

                elements.append(Paragraph(safe_str(f"\u25b6 {_gene}"), h2_style))
                elements.append(Paragraph(safe_str(f"Test selection: {_decision_txt}"), body_style))

                _omni_p_str = f"{_omni_p:.4f}" if _omni_p is not None else "—"
                _omni_table = [
                    [safe_str("Omnibus Test"), safe_str("p-value"), safe_str("Result")],
                    [safe_str(_omni_test), safe_str(_omni_p_str), safe_str(_omni_sig)],
                ]
                elements.append(make_table(_omni_table, col_widths=[240, 120, 100]))
                elements.append(Spacer(1, 6))

                # Post-hoc table
                _ph_rows = _mg.get("posthoc_rows", [])
                if _ph_rows:
                    elements.append(Paragraph(
                        safe_str(f"Post-hoc comparisons ({_posthoc}) — RQ-based:"),
                        body_style
                    ))
                    _ph_header = [
                        safe_str("Comparison"),
                        safe_str("Raw p"),
                        safe_str("Bonferroni p"),
                        safe_str("FDR p (B-H)"),
                        safe_str("Sig (raw)"),
                        safe_str("Sig (FDR)"),
                    ]
                    _ph_table_rows = [_ph_header]
                    # collect all pvals for chart
                    _ph_labels_chart = []
                    _ph_pvals_chart  = []
                    for _ph in _ph_rows:
                        _raw_p = _ph.get("Raw p", 1)
                        _sig_raw = "Sig" if _raw_p < 0.05 else "n.s."
                        _fdr_p   = _ph.get("FDR p (B-H)", 1)
                        _sig_fdr = "Sig" if _fdr_p < 0.05 else "n.s."
                        _ph_table_rows.append([
                            safe_str(str(_ph.get("Comparison",""))),
                            safe_str(f"{_raw_p:.4f}"),
                            safe_str(f"{_ph.get('Bonferroni p', 1):.4f}"),
                            safe_str(f"{_fdr_p:.4f}"),
                            safe_str(_sig_raw),
                            safe_str(_sig_fdr),
                        ])
                        _ph_labels_chart.append(f"{_gene} / {_ph.get('Comparison','')}")
                        _ph_pvals_chart.append(_raw_p)

                    cw6b = (letter[0]-100)/6
                    elements.append(make_table(_ph_table_rows, col_widths=[cw6b]*6))
                    elements.append(Spacer(1, 6))

                    # p-value bar chart for this gene's post-hoc
                    try:
                        _fig_mg, _ax_mg = plt.subplots(figsize=(7, max(2.5, 0.4 * len(_ph_labels_chart) + 1)))
                        _bar_c = ['#e53935' if p < 0.05 else '#90a4ae' for p in _ph_pvals_chart]
                        _ax_mg.barh(_ph_labels_chart, _ph_pvals_chart, color=_bar_c, alpha=0.85)
                        _ax_mg.axvline(x=0.05, color='black', linestyle='--', linewidth=0.9)
                        _ax_mg.set_xlabel('p-value (raw)', fontsize=9)
                        _ax_mg.set_title(f'Post-hoc p-values — {_gene}', fontsize=10, fontweight='bold')
                        for _ii, _vv in enumerate(_ph_pvals_chart):
                            _ax_mg.text(min(_vv + 0.002, 0.045), _ii, f'{_vv:.4f}', va='center', fontsize=7)
                        _ax_mg.spines['top'].set_visible(False); _ax_mg.spines['right'].set_visible(False)
                        plt.tight_layout()
                        _ib_mg = BytesIO()
                        plt.savefig(_ib_mg, format='png', dpi=150, bbox_inches='tight')
                        plt.close()
                        _ib_mg.seek(0)
                        _img_h = max(130, 40 * len(_ph_labels_chart) + 60)
                        elements.append(RLImage(_ib_mg, width=460, height=min(_img_h, 260)))
                        elements.append(Paragraph(
                            safe_str(
                                f"Figure. Post-hoc p-values for {_gene}. "
                                "Red bars = significant (p < 0.05). Dashed line = significance threshold."
                            ),
                            caption_style
                        ))
                    except Exception:
                        pass

                elements.append(Spacer(1, 10))

        else:
            # ── Standard 2-group: pairwise stat table ─────────────────────────────
            stat_cols = T.get('pdf_stat_cols', ['Gene','Comparison','Type','Method','p','Sig'])
            stat_table_rows = [stat_cols]
            for st_row in stat_rows:
                stat_table_rows.append([
                    str(st_row.get("__target_gene__", '')),
                    str(st_row.get('Comparison', '')),
                    str(st_row.get("__test_type__", '')),
                    str(st_row.get("__test_method__", '')),
                    f"{st_row.get('__pvalue__', 0):.4f}",
                    str(st_row.get("__significance__", '')),
                ])
            cw6 = (letter[0]-100)/6
            elements.append(make_table(stat_table_rows, col_widths=[cw6]*6))
            elements.append(Spacer(1, 8))

            # p-value chart
            if stat_rows:
                try:
                    fig_p, ax_p = plt.subplots(figsize=(7, max(2.5, 0.4 * len(stat_rows) + 1)))
                    labels_p = [f"{sr.get('__target_gene__','')} / {sr.get('Comparison','')}" for sr in stat_rows]
                    pvals = [sr.get("__pvalue__", 1) for sr in stat_rows]
                    bar_colors = ['#e53935' if p < 0.05 else '#90a4ae' for p in pvals]
                    ax_p.barh(labels_p, pvals, color=bar_colors, alpha=0.85)
                    ax_p.axvline(x=0.05, color='black', linestyle='--', linewidth=0.9)
                    ax_p.set_xlabel('p-value', fontsize=9)
                    ax_p.set_title('Statistical Test p-values', fontsize=10, fontweight='bold')
                    for i, v in enumerate(pvals):
                        ax_p.text(v+0.005, i, f'{v:.4f}', va='center', fontsize=7)
                    ax_p.spines['top'].set_visible(False); ax_p.spines['right'].set_visible(False)
                    plt.tight_layout()
                    ib2 = BytesIO(); plt.savefig(ib2, format='png', dpi=150, bbox_inches='tight'); plt.close(); ib2.seek(0)
                    elements.append(RLImage(ib2, width=460, height=200))
                    elements.append(Paragraph(s('pdf_fig2'), caption_style))
                except Exception:
                    pass

        elements.append(Spacer(1, 10))
        elements.append(Paragraph(s('pdf_s4_interp'), h2_style))
        elements.append(Paragraph(s('pdf_s4_interp_body'), body_style))
        elements.append(PageBreak())

        # ── SECTION 5: DELTA CT PLOTS ─────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s5_title'), h1_style))
        elements.append(hr())
        elements.append(Paragraph(s('pdf_s5_body'), body_style))
        elements.append(Spacer(1, 8))

        tg_key_  = "__target_gene__"
        dcp_key_ = "__dct_patient__"
        palette  = ['#3f51b5','#e91e63','#009688','#ff9800','#9c27b0']

        for i in range(num_target_genes):
            gene_label = f"Gene {i+1}"          # sabit değer — sütun adı değil
            try:
                fig_d, ax_d = plt.subplots(figsize=(6, 3.2))
                all_vals = []; all_labels = []
                ctrl_dct_vals = [
                    float(d["__dct_ctrl__"]) for d in input_values_table
                    if d.get(tg_key_) == gene_label
                    and d.get("__dct_ctrl__") not in ("EXCLUDED", None)
                    and d.get("Outlier Excluded", "No") == "No"
                ]
                # Convert ΔCq to RQ = 2^(-ΔCt) for visualization.
                # Plotting raw ΔCt is misleading because higher ΔCt = lower expression,
                # which is counter-intuitive. RQ values reflect actual expression levels.
                ctrl_vals = [2 ** (-v) for v in ctrl_dct_vals]
                if ctrl_vals:
                    all_vals.append(ctrl_vals)
                    all_labels.append(T['control_group'])
                for j in range(num_patient_groups):
                    pg = f"Group {j+1}"
                    smp_dct_vals = [float(d[dcp_key_]) for d in input_values_table
                          if d.get(tg_key_) == gene_label
                          and d.get("Grup") == pg
                          and d.get(dcp_key_) not in ("EXCLUDED", None)
                          and d.get("Outlier Excluded","No") == "No"]
                    sv = [2 ** (-v) for v in smp_dct_vals]
                    if sv:
                        all_vals.append(sv); all_labels.append(pg)
                for k, (vals, lbl) in enumerate(zip(all_vals, all_labels)):
                    col = palette[k % len(palette)]
                    jitter = np.random.uniform(-0.08, 0.08, len(vals))
                    ax_d.scatter([k+1+j for j in jitter], vals, color=col, alpha=0.75, s=28, zorder=3)
                    ax_d.hlines(np.mean(vals), k+0.75, k+1.25, colors='black', linewidths=2, zorder=4)
                ax_d.set_xticks(range(1, len(all_labels)+1))
                ax_d.set_xticklabels(all_labels, fontsize=8)
                ax_d.set_ylabel('RQ (2^-ΔCq)', fontsize=9)
                ax_d.set_title(f'{gene_label} — Relative Quantity (RQ)', fontsize=10, fontweight='bold')
                ax_d.spines['top'].set_visible(False); ax_d.spines['right'].set_visible(False)
                plt.tight_layout()
                ib3 = BytesIO(); plt.savefig(ib3, format='png', dpi=150, bbox_inches='tight'); plt.close(); ib3.seek(0)
                elements.append(RLImage(ib3, width=420, height=210))
                elements.append(Paragraph(s('pdf_fig3', gene=gene_label), caption_style))
                elements.append(Spacer(1, 10))
            except Exception:
                pass
        elements.append(PageBreak())

        # ── SECTION 6: INTERPRETATION ─────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s6_title'), h1_style))
        elements.append(hr())

        elements.append(Paragraph(s('pdf_s6_fc'), h2_style))
        fc_hdr  = T.get('pdf_fc_interp_header', ['FC','ΔΔCq','Interpretation','Significance'])
        fc_rows_data = T.get('pdf_fc_interp_rows', [])
        elements.append(make_table([fc_hdr] + fc_rows_data, col_widths=[(letter[0]-100)/4]*4))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(s('pdf_stat_note'), info_style))

        elements.append(Paragraph(s('pdf_s6_choose'), h2_style))
        elements.append(Paragraph(s('pdf_s6_choose_body'), body_style))

        elements.append(Paragraph(s('pdf_s6_stat'), h2_style))
        elements.append(Paragraph(s('pdf_s6_stat_body'), body_style))
        elements.append(PageBreak())

        # ── SECTION 7: REFERENCES ─────────────────────────────────────────────────
        elements.append(Paragraph(s('pdf_s7_title'), h1_style))
        elements.append(hr())
        refs = [
            "Yalçınkaya B (2026). GeneQuantify: a web-based tool for qPCR gene expression and copy number variation analysis. Molecular and Cellular Biochemistry. https://doi.org/10.1007/s11010-026-05621-y",
            "Livak KJ & Schmittgen TD (2001). Methods, 25(4), 402-408. (ΔΔCq)",
            "Pfaffl MW (2001). Nucleic Acids Research, 29(9), e45. (Pfaffl)",
            "Vandesompele J et al. (2002). Genome Biology, 3(7). (geNorm)",
            "Bustin SA et al. (2009). Clinical Chemistry, 55(4), 611-622. (MIQE)",
            "Grubbs FE (1969). Technometrics, 11(1), 1-21.",
            "Tukey JW (1977). Exploratory Data Analysis. Addison-Wesley.",
            "Benjamini Y & Hochberg Y (1995). J Royal Stat Soc B, 57(1), 289-300. (FDR)",
        ]
        for ref in refs:
            elements.append(Paragraph(safe_str(f"• {ref}"), small_style))
            elements.append(Spacer(1, 3))

        elements.append(Spacer(1, 16))
        elements.append(hr())
        elements.append(Paragraph(
            safe_str(f"{s('pdf_footer')} | {s('pdf_generated', now=now)} | {s('pdf_contact')}"),
            small_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer



    with tab_report:
        st.markdown(f"### 📄 {_t.get('pdf_report', "")}")
        st.markdown("---")
        if not input_values_table:
            st.info(_t.get('error_no_data', ''))
        else:
            st.success('✅ ' + _t.get('pdf_ready', '{n} records ready').format(n=len(input_values_table)))
            if st.button(f"📥 {_t.get('generate_pdf', "")}", key="pdf_btn"):
                pdf_buffer = create_pdf(data, stats_data, pd.DataFrame(input_values_table), language_code, multigroup_results=multigroup_results)
                st.download_button(
                    label=f"⬇️ {_t.get('pdf_report', "")}",
                    data=pdf_buffer,
                    file_name="gen_ekspresyon_raporu.pdf",
                    mime="application/pdf",
                    key="pdf_dl"
                )

    st.sidebar.divider()
    c1, c2 = st.sidebar.columns(2)
    with c1:
        st.sidebar.link_button(
            "⬇️ Desktop",
            "https://drive.google.com/file/d/1zxmAKWm-cV_W2dCMCtb-momEau75UpXg/view?usp=sharing",
            use_container_width=True
        )
    with c2:
        st.sidebar.link_button(
            "⭐ GitHub",
            "https://github.com/burhanettiny/GeneQuantify",
            use_container_width=True
        )
    st.sidebar.caption("GeneQuantify — GPL-3.0 | mailtoburhanettin@gmail.com")

    st.sidebar.link_button(
        "📄 Cite this paper (Mol Cell Biochem)",
        "https://link.springer.com/article/10.1007/s11010-026-05621-y",
        use_container_width=True
    )
    st.sidebar.caption("Yalçınkaya B (2026). Mol Cell Biochem. doi:10.1007/s11010-026-05621-y")


else:  # analysis_mode == "dPCR"
    APP_VERSION = ABSOLUTEGENE_VERSION  # alias: preserves all internal references from the original AbsoluteGene source
    # ── Sidebar logo (absolutegene.png, same repo folder as this script) ──────────
    try:
        import base64 as _b64
        _logo_path = "absolutegene.png"
        if os.path.exists(_logo_path):
            with open(_logo_path, "rb") as _f:
                _logo_b64 = _b64.b64encode(_f.read()).decode()
            st.sidebar.markdown(
                f"<div style='text-align:center;padding:4px 0 6px 0;'>"
                f"<img src='data:image/png;base64,{_logo_b64}' width='180' style='border-radius:8px;'/>"
                f"</div>",
                unsafe_allow_html=True
            )
    except Exception:
        pass

    flags = {"Türkçe": "🇹🇷", "English": "🇬🇧"}
    default_index = list(flags.keys()).index(st.session_state.language) if st.session_state.language in flags else 1

    selected_language = st.sidebar.selectbox(
        "🌐 Language / Dil",
        options=[f"{flags[lang]} {lang}" for lang in flags],
        index=default_index,
        label_visibility="collapsed"
    )
    selected_language_name = selected_language.split(' ', 1)[1]
    language_map = {"Türkçe": "tr", "English": "en"}
    language_code = language_map.get(selected_language_name, "en")

    # ═══════════════════════════════════════════════════════════════════════════════
    # TRANSLATIONS (TR / EN)
    # ═══════════════════════════════════════════════════════════════════════════════
    translations = {
        "tr": {
            "title": "🧪 AbsoluteGene: Dijital PCR (dPCR/ddPCR) Gen Ekspresyonu ve CNV Analizi",
            "subtitle": "B. Yalçınkaya tarafından geliştirildi — GeneQuantify'nin dijital PCR versiyonu",
            "tab_data": "Veri Girişi",
            "tab_results": "Sonuçlar",
            "tab_report": "Rapor",
            "study_design": "⚙️ Çalışma Tasarımı",
            "num_target_genes": "🔹 Hedef Gen Sayısı",
            "num_patient_groups": "🔹 Hasta Grubu Sayısı",
            "num_ref_genes": "🔹 Referans Gen Sayısı",
            "ref_gene_help": "dMIQE kılavuzları sağlam normalizasyon için ≥1 doğrulanmış referans lokusu önerir; ≥2 önerilir.",
            "ploidy_label": "🔹 Referans Lokus Ploidi Sayısı",
            "ploidy_help": "Kopya sayısı varyasyonu (CNV) hesabı için referans lokusun bilinen kopya sayısı (diploid organizmada genelde 2). ⚠️ Önemli varsayım: bu değer, TÜM örneklerde (kontrol dahil) referans lokusun tam olarak bu sayıda kopyaya sahip olduğunu varsayar. Referans lokusun kendisi bir anöploidi veya CNV'den etkileniyorsa, tüm CN hesapları sessizce yanlış olur — referans lokusu seçerken bunu doğrulayın.",
            "partition_vol_label": "🔹 Partisyon Hacmi (nL)",
            "partition_vol_help": "Damlacık/kuyucuk başına hacim. Bio-Rad QX200: ~0.85 nL, QX ONE: ~0.7 nL, Qiagen QIAcuity: plaka tipine göre değişir. Sadece kopya/µL hesaplaması için kullanılır; oran/CNV hesabını etkilemez.",
            "qc_min_partitions": "🔹 Minimum Kabul Edilebilir Partisyon Sayısı (QC)",
            "qc_min_partitions_help": "Bu değerin altındaki replikatlar düşük kalite olarak işaretlenir (dMIQE önerisi: ddPCR için tipik olarak ≥10.000).",
            "outlier_section_title": "### 🔍 Aykırı Değer Tespiti Ayarları",
            "outlier_enable": "Aykırı değer tespitini etkinleştir",
            "outlier_enable_help": "λ (partisyon başına kopya sayısı) değerlerindeki istatistiksel olarak aşırı replikatları tespit eder.",
            "outlier_method_label": "Tespit yöntemi",
            "outlier_method_help": "Grubbs: normal dağılım için. IQR: parametrik olmayan, çarpık dağılımlar için.",
            "grubbs_power_warning": "⚠️ Grubbs testinin gücü küçük n'de (dPCR'de tipik olan 3-5 replikat) düşüktür — gerçek bir aykırı değeri kaçırma olasılığı yüksektir. 'Aykırı değer bulunamadı' sonucunu 'veri kesinlikle temiz' olarak yorumlamayın.",
            "outlier_alpha_label": "Anlamlılık düzeyi (α)",
            "outlier_iqr_label": "IQR çarpanı (k)",
            "patient_data_header": "📥 Hasta ve Kontrol Grubu Partisyon Verisi Girin",
            "target_gene": "Hedef Gen",
            "reference_gene": "Referans Gen",
            "control_group": "🧬 Kontrol Grubu",
            "patient_group": "🩸 Hasta Grubu",
            "positive_partitions": "Pozitif Partisyon Sayısı",
            "total_partitions": "Toplam Kabul Edilen Partisyon Sayısı",
            "input_format_info": "ℹ️ Her satıra bir replikat gelecek şekilde girin. 'Pozitif' ve 'Toplam' kutularındaki satır sayıları eşleşmelidir.",
            "warning_empty_input": "⚠️ Dikkat: Verileri alt alta yazın; Pozitif ve Toplam kutularındaki satır sayısı eşit olmalı.",
            "warning_field_empty": "⚠️ Bir veya birden fazla gerekli alan boş bırakılmış (Hedef Pozitif/Toplam veya Referans Gen Pozitif/Toplam). Lütfen tüm kutulara veri girin.",
            "warning_all_excluded_qc": "⚠️ Girilen tüm replikatlar **düşük partisyon sayısı (QC eşiği: {thr})** nedeniyle elendi — hiçbiri analiz edilemedi. Eşiği düşürmeyi (Çalışma Tasarımı ayarlarından) veya verileri kontrol etmeyi düşünün.",
            "warning_all_excluded_saturated": "⚠️ Girilen tüm replikatlar **doygun** (tüm partisyonlar pozitif, λ hesaplanamıyor) — hiçbiri analiz edilemedi. Örneğin daha fazla seyreltilmesi gerekebilir.",
            "warning_all_excluded_outlier": "⚠️ Girilen tüm replikatlar **aykırı değer olarak işaretlenip dışlandı** — hiçbiri analiz için kalmadı. Aykırı değer onay kutularınızı kontrol edin.",
            "warning_all_excluded_mixed": "⚠️ Girilen tüm replikatlar elendi (QC/doygunluk/aykırı değer kombinasyonu nedeniyle) — hiçbiri analiz edilemedi. Detaylar için Sonuçlar sekmesindeki Giriş Verileri Tablosu'na bakın.",
            "qc_fail_warning": "⚠️ **Düşük partisyon sayısı uyarısı:** {n} replikat, minimum eşik ({thr}) altında kabul edilen partisyona sahip. Bu replikatlar analiz için düşük güvenilirlikte olabilir.",
            "saturation_warning": "⚠️ **Doygunluk uyarısı:** {n} replikatta tüm partisyonlar pozitif (p≥1.0) — λ hesaplanamıyor. Örneğin daha fazla seyreltilmesi gerekiyor.",
            "gr_tbl": "📋 Giriş Verileri Tablosu (λ ve %95 GA dahil)",
            "nil_mine": "📊 Sonuçlar",
            "download_csv": "📥 CSV İndir",
            "generate_pdf": "📥 PDF Raporu Hazırla",
            "pdf_report": "Dijital PCR Analiz Raporu",
            "sample_number": "Replikat No",
            "lambda_col": "λ (kopya/partisyon)",
            "ci_low_col": "%95 GA Alt",
            "ci_high_col": "%95 GA Üst",
            "conc_col": "Konsantrasyon (kopya/µL)",
            "qc_col": "QC Durumu",
            "qc_pass": "✅ Geçti",
            "qc_fail": "❌ Düşük n",
            "qc_saturated": "❌ Doygun",
            "ratio_col": "Normalize Oran (Hedef/Referans)",
            "cn_col": "Kopya Sayısı (CN)",
            "fc_col": "Kat Değişimi (vs Kontrol)",
            "regulation_status": "Regülasyon Durumu",
            "no_change": "Değişim Yok",
            "upregulated": "Yukarı Regüle / Kazanım",
            "downregulated": "Aşağı Regüle / Kayıp",
            "outlier_excluded_no": "Hayır",
            "outlier_excluded_yes": "Evet",
            "genorm_title": "Referans Lokus Stabilitesi",
            "stable": "Kararlı",
            "borderline": "Sınırda",
            "unstable": "Kararsız",
            "m_value": "M-değeri",
            "method_comparison": "📊 Sonuç Özeti",
            "statistical_results": "📈 İstatistiksel Sonuçlar",
            "test_type": "Test Türü",
            "test_method": "Kullanılan Test",
            "test_pvalue": "Test P-değeri",
            "significance": "Anlamlılık",
            "significant": "Anlamlı",
            "insignificant": "Anlamsız",
            "parametric": "Parametrik",
            "non_parametric": "Nonparametrik",
            "t_test": "t-test",
            "welch_t_test": "Welch t-testi",
            "mann_whitney_u_test": "Mann-Whitney U testi",
            "stat_decision_title": "🔬 İstatistiksel karar",
            "stat_decision_steps": "**Adım adım test seçimi:**",
            "stat_shapiro_title": "**1. Shapiro-Wilk normallik testi**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Normal değil",
            "stat_levene_title": "**2. Levene varyans homojenliği testi**",
            "stat_levene_skipped": "**2. Levene testi** — *atlandı* (normallik sağlanmadı)",
            "stat_equal_var": "Eşit varyans",
            "stat_unequal_var": "Eşitsiz varyans",
            "stat_selected_test": "**3. Seçilen test:**",
            "stat_reason": "**Gerekçe:**",
            "stat_result": "**Sonuç:**",
            "stat_reason_nonnormal": "Bir veya her iki grupta normal dağılım sağlanmadı",
            "stat_reason_normal_equal": "Her iki grup normal + eşit varyans",
            "stat_reason_normal_unequal": "Her iki grup normal + eşitsiz varyans",
            "stat_multigroup_note": "⚠️ Not: ≥3 grup varsa, aşağıdaki Çoklu Grup Karşılaştırması bölümüne bakın.",
            "multigroup_title": "## 📊 Çoklu Grup Karşılaştırma Analizi",
            "multigroup_omnibus_test": "Omnibus Testi",
            "multigroup_pvalue": "p-değeri",
            "multigroup_result": "Sonuç",
            "multigroup_significant": "Anlamlı",
            "multigroup_not_significant": "Anlamlı değil",
            "multigroup_omnibus_ns": "ℹ️ Omnibus testi anlamlı değil (p ≥ 0.05). Post-hoc karşılaştırmalar bilgi amaçlıdır.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Post-hoc sonuçlarını indir —",
            "multigroup_2group_note": "ℹ️ Yalnızca 2 grup tespit edildi. İkili istatistikler yukarıda raporlanmıştır.",
            "multigroup_decision_normal_equal": "✅ Normal + eşit varyans → **Tek yönlü ANOVA + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Normal + eşitsiz varyans → **Welch ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ Normal değil → **Kruskal-Wallis + Dunn post-hoc**",
            "multigene_title": "### 🧬 Çoklu Gen Çoklu Karşılaştırma Düzeltmesi",
            "multigene_sig_raw": "Anlamlı (ham)",
            "multigene_sig_bonf": "Anlamlı (Bonferroni)",
            "multigene_sig_fdr": "Anlamlı (FDR B-H)",
            "multigene_warning": "⚠️ Düzeltme sonrası {lost} sonuç FDR düzeltmesi sonrası anlamlılığını yitirdi.",
            "multigene_success": "✅ {n} anlamlı sonucun tamamı FDR düzeltmesi sonrasında da anlamlı kalmaktadır.",
            "multigene_no_sig": "Ham p < 0.05 eşiğinde anlamlı ikili sonuç tespit edilmedi.",
            "multigene_dl_button": "📥 Düzeltilmiş p-değerlerini indir (CSV)",
            "multigene_1gene_note": "ℹ️ Yalnızca 1 hedef gen analiz edildi — genler arası düzeltme uygulanamaz.",
            "distribution_graph": "Dağılım Grafiği",
            "x_axis_title": "Grup Adı",
            "dist_plot_mode_label": "📊 Dağılım Grafiği — Görüntüleme Modu",
            "dist_plot_ratio": "Normalize Oran (Hedef/Referans) — önerilen",
            "dist_plot_lambda": "λ (kopya/partisyon) — ham",
            "dist_plot_fc": "Kat Değişimi — kontrole göre",
            "error_no_data": "Veri bulunamadı, PDF oluşturulamadı.",
            "pdf_ready": "{n} kayıt hazır — PDF oluşturabilirsiniz.",
            "sidebar_github_btn": "⭐ GitHub'da Kaynak Kodu Görüntüle",
            "sidebar_sister_tool": "🧬 qPCR için: GeneQuantify",
            "rdml_expander": "ℹ️ Veri girişi hakkında",
            "guide_btn": "📘 Kullanım Kılavuzu",
            "sidebar_example_title": "📋 Örnek Veri Yükle",
            "sidebar_example_select": "Senaryo seçin",
            "sidebar_example_load_btn": "▶ Senaryoyu Yükle",
            "sidebar_example_loaded": "✅ {s} yüklendi! Veri Girişi sekmesine geçin.",
            "ntc_expander": "🧫 NTC (Negatif Kontrol) / LOD-LOQ — opsiyonel",
            "ntc_description": "Tespit sınırını (LOD) ve ölçüm sınırını (LOQ) hesaplamak için No-Template Control (şablon içermeyen kontrol) replikatlarınızı girin. Boş bırakılırsa bu adım atlanır.",
            "ntc_positive_label": "NTC Pozitif Partisyon Sayısı",
            "ntc_total_label": "NTC Toplam Kabul Edilen Partisyon Sayısı",
            "ntc_calc_btn": "📊 LOD/LOQ Hesapla",
            "lod_result_title": "Tespit Sınırı (LOD) ve Ölçüm Sınırı (LOQ) — {gene}",
            "lod_label": "LOD (kopya/µL)",
            "loq_label": "LOQ (kopya/µL)",
            "ntc_zero_note": "ℹ️ NTC replikatlarında hiç pozitif partisyon tespit edilmedi (n={n} partisyon havuzlandı). LOD, 'üçler kuralı' (rule of three) ile üst güven sınırından hesaplandı: 3/n.",
            "ntc_contamination_warning": "⚠️ NTC replikatlarında arka plan sinyali tespit edildi (λ={lam:.5f}, {pos}/{tot} pozitif partisyon). Bu, reaktif kontaminasyonu veya çapraz kontaminasyon olabilir — LOD, bu arka plan seviyesi + güvenlik payı olarak hesaplandı.",
            "below_lod_flag": "❌ LOD altı",
            "between_lod_loq_flag": "⚠️ LOD-LOQ arası",
            "above_loq_flag": "✅ LOQ üstü",
            "lod_qc_col": "LOD/LOQ Durumu",
            "loq_heuristic_note": "Not: LOQ, yaygın kullanılan bir kural olarak LOD'un 3 katı şeklinde hesaplanmıştır. Daha kesin bir LOQ için tekrarlanabilirlik (CV%) temelli deneysel doğrulama önerilir.",
            "empirical_loq_label": "Ampirik LOQ Çapraz Kontrolü (CV%-tabanlı)",
            "empirical_loq_pass": "✅ Tekrarlanabilirlik kabul edilebilir",
            "empirical_loq_fail": "⚠️ Düşük tekrarlanabilirlik (CV>%25)",
            "csv_import_expander": "📂 Cihaz CSV Dosyası İçe Aktar (QuantaSoft / QX Manager / QIAcuity)",
            "csv_import_description": "Bio-Rad QuantaSoft/QX Manager veya QIAGEN QIAcuity yazılımından dışa aktardığınız CSV dosyasını yükleyin. Sütunlar otomatik tanınmaya çalışılır; gerekirse manuel eşleştirme yapabilirsiniz.",
            "csv_uploader": "CSV dosyası seçin",
            "csv_parse_error": "❌ Dosya ayrıştırma hatası: {err}",
            "csv_col_mapping_title": "**Sütun Eşleştirme**",
            "csv_col_sample": "Örnek Adı Sütunu",
            "csv_col_target": "Hedef/Assay Adı Sütunu",
            "csv_col_positives": "Pozitif Partisyon Sütunu",
            "csv_col_total": "Toplam Partisyon Sütunu",
            "csv_col_auto_detected": "✅ Otomatik tespit edildi",
            "csv_col_manual_needed": "⚠️ Otomatik tespit edilemedi — lütfen manuel seçin",
            "csv_preview_title": "Ayrıştırılan verileri önizle",
            "csv_assay_assignment_title": "**Assay (Hedef/Referans Gen) Ataması**",
            "csv_target_assays_label": "Hedef gen assay(ler)i",
            "csv_ref_assays_label": "Referans gen assay(ler)i",
            "csv_group_assignment_title": "**Grup Ataması**",
            "csv_ctrl_label": "Kontrol örnek adı (virgülle ayrılmış alt dizeler)",
            "csv_n_patient_groups": "Hasta grubu sayısı",
            "csv_patient_label": "Hasta grubu {i} örnek adı (adları)",
            "csv_apply_btn": "✅ İçe Aktarımı Veri Girişine Uygula",
            "csv_apply_success": "✅ {n} değer Veri Girişi sekmesine yüklendi! Kontrol edip ayarlayabilirsiniz.",
            "csv_apply_warning": "⚠️ Hiçbir değer eşleştirilemedi. Kontrol/hasta etiketlerinin örnek adlarıyla uyuştuğundan emin olun.",
            "download_example_csv": "📄 Örnek CSV Dosyasını İndir",
            "dilution_expander": "🧪 Optimal Seyreltme / Dinamik Aralık Hesaplayıcısı",
            "dilution_description": "Poisson tabanlı kantifikasyonun en hassas olduğu aralık partisyon başına ~1.0–2.0 kopyadır (λ). Bu hesaplayıcı, mevcut ölçümünüze göre önerilen seyreltme faktörünü bulur.",
            "dilution_mode_label": "Giriş yöntemi",
            "dilution_mode_counts": "Gözlenen partisyon sayılarından",
            "dilution_mode_lambda": "Bilinen λ değerinden",
            "dilution_positive_label": "Pozitif Partisyon Sayısı",
            "dilution_total_label": "Toplam Partisyon Sayısı",
            "dilution_lambda_label": "Mevcut λ (kopya/partisyon)",
            "dilution_target_label": "Hedef λ (kopya/partisyon)",
            "dilution_calc_btn": "📊 Seyreltme Öner",
            "dilution_result_optimal": "✅ **Optimal aralıkta** (λ={lam:.3f}). Ek seyreltme gerekmez.",
            "dilution_result_too_low": "⚠️ **Çok seyreltik** (λ={lam:.3f}). Bir sonraki hazırlığınızda seyreltme faktörünüzü **mevcut faktörün {factor:.3f} katı** yapın (yani daha az seyreltin).",
            "dilution_result_too_high": "⚠️ **Optimal aralığın üzerinde** (λ={lam:.3f}). Bir sonraki hazırlığınızda seyreltme faktörünüzü **mevcut faktörün {factor:.3f} katı** yapın (yani daha fazla seyreltin).",
            "dilution_result_saturated": "❌ **Doygunluğa yakın/riskli** (λ={lam:.3f}). Bir sonraki hazırlığınızda seyreltme faktörünüzü **mevcut faktörün {factor:.3f} katı** yapın.",
            "dilution_factor_note": "Bu hesaplayıcı, orijinal seyreltme faktörünüzü bilmediği için mutlak bir seyreltme oranı değil, **göreli bir çarpan** önerir: bir sonraki hazırlığınızda kullandığınız seyreltme faktörünü bu çarpanla çarpın. Örnek: şu an 1:50 seyreltiyorsanız ve önerilen çarpan 2.0 ise, bir sonraki denemede 1:100 seyreltin.",
            "multiplex_expander": "🔀 Multipleks Küme Dönüştürücü (2 Renk)",
            "multiplex_description": "İki hedefin (örn. hedef gen + referans gen) **aynı partisyonlarda** çift renkli (multipleks) ölçüldüğü ddPCR verileri için. 2×2 küme sayılarını girin — bu araç, aralarındaki korelasyonu doğru şekilde hesaba katarak (bağımsızlık varsayımı YAPMADAN) marjinal pozitif/toplam sayılarını ve oranı hesaplar.",
            "multiplex_n11_label": "Çift Pozitif (Hedef+ Ref+)",
            "multiplex_n10_label": "Yalnızca Hedef Pozitif (Hedef+ Ref-)",
            "multiplex_n01_label": "Yalnızca Referans Pozitif (Hedef- Ref+)",
            "multiplex_n00_label": "Çift Negatif (Hedef- Ref-)",
            "multiplex_calc_btn": "📊 Dönüştür ve Hesapla",
            "multiplex_result_title": "**Sonuç:**",
            "multiplex_pos_target_label": "Hedef Pozitif Partisyon (marjinal)",
            "multiplex_pos_ref_label": "Referans Pozitif Partisyon (marjinal)",
            "multiplex_total_label": "Toplam Partisyon",
            "multiplex_ratio_covaware": "Oran (kovaryans-farkında) %95 GA",
            "multiplex_ratio_indep": "Oran (bağımsızlık varsayımıyla) %95 GA — karşılaştırma amaçlı",
            "multiplex_paste_hint": "💡 Yukarıdaki marjinal pozitif/toplam değerlerini ana Veri Girişi alanlarına kopyalayabilirsiniz. Kovaryans-farkında güven aralığı, aynı partisyonlardan gelen verinin korelasyonunu doğru yansıtır ve genellikle bağımsızlık varsayımından daha dar/isabetlidir.",
            "qc_panel_title": "🩺 Veri Kalitesi Özeti",
            "qc_panel_total": "Toplam Replikat",
            "qc_panel_qc_fail": "Düşük Partisyon (QC)",
            "qc_panel_saturated": "Doygun",
            "qc_panel_outlier": "Aykırı Değer (Dışlanan)",
            "qc_panel_below_lod": "LOD Altı Sonuç",
            "qc_panel_high_cv": "Yüksek CV (>%25)",
            "qc_panel_verdict_good": "✅ **İyi** — veri seti sağlıklı görünüyor, önemli bir kalite sorunu tespit edilmedi.",
            "qc_panel_verdict_caution": "⚠️ **Dikkat** — bazı replikatlarda kalite bayrakları var. Sonuçları yorumlarken bu noktaları göz önünde bulundurun.",
            "qc_panel_verdict_poor": "❌ **Gözden geçirin** — veri setinde önemli oranda kalite sorunu bayrağı var (QC hatası, doygunluk veya yüksek değişkenlik). Rapor öncesi verileri kontrol etmenizi öneririz.",
            "qc_panel_no_data": "Henüz veri girilmedi.",
            "project_expander": "💾 Proje Kaydet / Yükle",
            "project_description": "Tüm veri girişinizi (çalışma tasarımı, kontrol/hasta grubu verileri, NTC) bir JSON dosyası olarak kaydedip daha sonra kaldığınız yerden devam edebilirsiniz.",
            "project_export_btn": "📥 Projeyi Dışa Aktar (JSON)",
            "project_import_uploader": "Proje dosyası yükleyin (.json)",
            "project_import_success": "✅ Proje yüklendi ({n} alan geri yüklendi)! Veri Girişi sekmesine geçin.",
            "project_import_error": "❌ Proje dosyası okunamadı: {err}",
            "history_expander": "📚 Oturum Geçmişi (Anlık Görüntüler)",
            "history_description": "Mevcut çalışma durumunuzun adlandırılmış anlık görüntülerini bu oturum içinde kaydedip geri yükleyebilirsiniz. ⚠️ Bunlar yalnızca bu tarayıcı oturumunda saklanır — sayfayı yenilerseniz kaybolur. Kalıcı saklama için yukarıdaki 'Projeyi Dışa Aktar' ile JSON indirin.",
            "history_name_label": "Anlık görüntü adı",
            "history_save_btn": "📌 Şu Anki Durumu Kaydet",
            "history_saved_success": "✅ '{name}' kaydedildi.",
            "history_select_label": "Kayıtlı anlık görüntüler",
            "history_restore_btn": "↩️ Geri Yükle",
            "history_download_btn": "📥 İndir",
            "history_delete_btn": "🗑️ Sil",
            "history_restore_success": "✅ Anlık görüntü geri yüklendi ({n} alan). Veri Girişi sekmesine geçin.",
            "history_no_snapshots": "Henüz kaydedilmiş anlık görüntü yok.",
            "advanced_mode_label": "🎓 Gelişmiş Mod",
            "advanced_mode_help": "Klinik Araçlar ve CRM Üretimi sekmeleri ile Multipleks Dönüştürücü ve NTC/LOD-LOQ gibi ileri düzey araçları gösterir. Kapalıyken sadece temel dPCR analiz akışı (Veri Girişi → Sonuçlar → Rapor) görünür.",
            "simple_mode_caption": "ℹ️ Basit Mod aktif — sadece temel analiz akışı görünüyor. İleri düzey araçlar (Klinik/CRM/Multipleks) için Gelişmiş Modu açın.",
            "advanced_gate_title": "🎓 Bu Gelişmiş Bir Araçtır",
            "advanced_gate_message": "Bu sekme, ileri düzey istatistiksel/metrolojik araçlar içerir (ölçüm belirsizliği, homojenlik testi, yöntem karşılaştırma vb.). Kullanmak için sol kenar çubuğundaki **Gelişmiş Mod**'u açın.",
            "batch_pdf_btn": "📥 Tarama Raporunu PDF Olarak Hazırla",
            "batch_pdf_report": "Toplu Numune Tarama Raporu",
            "batch_pdf_description": "Bu rapor, her numune için Poisson istatistiğine dayalı λ, normalize oran, %95 güven aralığı ve beklenen orana göre sınıflandırmayı içerir. Sınıflandırma, örneğin oran güven aralığının beklenen (değişim-yok) değeri içerip içermediğine göre belirlenir.",
            "vaf_pdf_btn": "📥 VAF Raporunu PDF Olarak Hazırla",
            "vaf_pdf_report": "VAF / Mutasyon Fraksiyonu Raporu",
            "vaf_pdf_description": "Bu rapor, her örnek için mutant ve wild-type assay λ değerlerini, Fraksiyonel Bolluk (%FA) ve delta-method %95 güven aralığını içerir.",
            "tab_clinical": "Klinik Araçlar",
            "tab_crm": "CRM Üretimi",
            "clinical_title": "🩺 Klinik Doğrulama Araçları",
            "clinical_description": "Ölçüm belirsizliği, referans değişim değeri, hassasiyet çalışması ve yöntem karşılaştırması gibi klinik/tanısal doğrulama için gerekli araçlar. Bu araçlar araştırma ve yöntem doğrulama amaçlıdır; klinik tanı kararı için doğrulanmamıştır.",
            "clinical_mode_label": "Araç seçin",
            "clinical_mode_mu": "📐 Ölçüm Belirsizliği (MU) Bütçesi",
            "clinical_mode_rcv": "📈 Referans Değişim Değeri (RCV)",
            "clinical_mode_precision": "🔁 Hassasiyet Çalışması",
            "clinical_mode_comparison": "⚖️ Yöntem Karşılaştırma",
            # MU Budget
            "mu_description": "GUM (Ölçüm Belirsizliği İfadesi Kılavuzu) yaklaşımıyla, bağımsız belirsizlik kaynakları karesel olarak birleştirilir ve bir kapsam faktörü (k) ile çarpılarak genişletilmiş belirsizlik elde edilir.",
            "mu_poisson_source": "Poisson Sayım İstatistiği Belirsizliği (%)",
            "mu_poisson_help": "Otomatik hesaplanabilir (replikat CV%'den) veya manuel girilebilir.",
            "mu_poisson_auto": "Sonuçlardan otomatik al",
            "mu_poisson_auto_replicate": "📊 Replikat CV%'sinden (gözlenen toplam varyasyon)",
            "mu_poisson_auto_theoretical": "🎯 Teorik Poisson SE'den (sadece sayım istatistiği)",
            "mu_double_count_warning": "⚠️ Bu değer replikatlar arası **gözlenen toplam varyasyonu** yansıtır (pipetleme + run-to-run + Poisson sayımı hepsi dahil). Aşağıya ayrıca pipetleme/hassasiyet belirsizliği eklerseniz **aynı varyasyon kaynağını iki kez saymış olabilirsiniz**. Bunun yerine 'Teorik Poisson SE' seçeneğini kullanmayı düşünün.",
            "mu_double_count_error": "❌ **Çifte sayım riski:** 'Replikat CV%'sinden' modunu seçtiniz VE ek olarak pipetleme/hassasiyet belirsizliği girdiniz. Replikat CV zaten bu kaynakları içerebilir. Ya pipetleme/hassasiyeti sıfırlayın (yalnızca replikat CV kullanın) ya da 'Teorik Poisson SE' moduna geçip diğer kaynakları ayrı ayrı ekleyin.",
            "mu_theoretical_note": "Bu değer, yalnızca partisyon sayım istatistiğinden (Poisson) türetilmiştir — pipetleme veya biyolojik varyasyon içermez. Bu nedenle aşağıdaki diğer bileşenlerle güvenle birleştirilebilir.",
            "mu_no_lambda_cache": "ℹ️ Bu sonuç için λ verisi önbellekte bulunamadı — lütfen önce Veri Girişi sekmesinde hesaplama yapın.",
            "unbalanced_design_note": "ℹ️ Dengesiz tasarım tespit edildi (grup başına replikat sayıları: {n_list}). Hiçbir veri atılmadı — varyans bileşenleri, dengesiz tek-yönlü rastgele-etki ANOVA için standart yöntem-momentleri düzeltmesiyle (Searle ve ark. 1992) hesaplandı.",
            "mu_poisson_manual": "Manuel gir",
            "mu_pipetting_label": "Pipetleme / Seyreltme Belirsizliği (%)",
            "mu_pipetting_help": "Tipik olarak iyi pipetleme tekniğinde %1-2 CV. Kendi doğrulama verinize göre girin.",
            "mu_precision_label": "Run-arası Hassasiyet Belirsizliği (%, opsiyonel)",
            "mu_precision_help": "Ayrı bir hassasiyet çalışmasından elde edilen günler-arası CV%. Bilinmiyorsa 0 bırakın.",
            "mu_gene_select": "Sonuç seçin (Gen / Grup)",
            "mu_k_label": "Kapsam Faktörü (k)",
            "mu_calc_btn": "📊 Belirsizlik Bütçesini Hesapla",
            "mu_result_title": "Sonuç: {value} ± {U:.2f}% (k={k}, ~%95 güven düzeyi)",
            "mu_budget_table_title": "**Belirsizlik Bütçesi Tablosu**",
            "mu_col_source": "Kaynak",
            "mu_col_contribution": "Katkı (%, göreli standart belirsizlik)",
            "mu_col_combined": "Birleşik Standart Belirsizlik (u_c)",
            "mu_col_expanded": "Genişletilmiş Belirsizlik (U)",
            # RCV
            "rcv_description": "İki seri sonuç arasındaki farkın, analitik ve biyolojik değişkenlikten beklenenin ötesinde istatistiksel olarak anlamlı olup olmadığını belirler (Fraser & Harris, Crit Rev Clin Lab Sci 1989).",
            "rcv_result1_label": "Sonuç 1 (önceki, örn. %VAF veya kopya/µL)",
            "rcv_result2_label": "Sonuç 2 (sonraki)",
            "rcv_cv_analytical_label": "Analitik CV (%)",
            "rcv_cv_analytical_help": "Bu ölçümün tekrarlanabilirlik CV%'si (replikat CV veya MU bütçesinden).",
            "rcv_cv_biological_label": "Biyolojik Değişkenlik CV (%)",
            "rcv_cv_biological_help": "Analitin bilinen birey-içi biyolojik değişkenliği (literatürden — analite özgüdür).",
            "rcv_z_label": "z-değeri (güven düzeyi)",
            "rcv_calc_btn": "📊 RCV Hesapla",
            "rcv_result_significant": "🔴 **Anlamlı değişim** — %{change:+.2f} değişim, RCV eşiği olan %{rcv:.2f}'yi aşıyor. Bu değişim ölçüm gürültüsüyle açıklanamaz.",
            "rcv_result_not_significant": "🟢 **Anlamlı değişim yok** — %{change:+.2f} değişim, RCV eşiği olan %{rcv:.2f} içinde. Analitik/biyolojik gürültü ile tutarlı.",
            # Precision study
            "precision_description": "Gün-içi (tekrarlanabilirlik) ve günler-arası (ara hassasiyet) varyans bileşenlerini iç içe (nested) ANOVA ile ayırır (CLSI EP05-A3 yaklaşımı).",
            "precision_n_days": "Gün/Çalışma Sayısı",
            "precision_day_label": "Gün {i} — Replikat Değerleri (her satıra bir değer)",
            "precision_calc_btn": "📊 Hassasiyeti Hesapla",
            "precision_repeatability_cv": "Tekrarlanabilirlik CV (%)",
            "precision_between_day_cv": "Günler-arası CV (%)",
            "precision_total_cv": "Toplam (Ara Hassasiyet) CV (%)",
            "precision_grand_mean": "Genel Ortalama",
            # Method comparison
            "comparison_description": "İki ölçüm yöntemi arasındaki uyumu Bland-Altman analizi ve Deming regresyonu ile değerlendirir.",
            "comparison_method1_label": "Yöntem 1 Değerleri (örn. dPCR — her satıra bir değer)",
            "comparison_method2_label": "Yöntem 2 Değerleri (örn. qPCR — eşleştirilmiş, aynı sırada)",
            "comparison_variance_ratio_label": "Varyans Oranı (λ, Deming için)",
            "comparison_calc_btn": "📊 Karşılaştırmayı Hesapla",
            "comparison_bias_label": "Ortalama Fark (Bias)",
            "comparison_loa_label": "Uyum Sınırları (%95)",
            "comparison_deming_label": "Deming Regresyonu",
            "comparison_ba_chart_title": "Bland-Altman Grafiği",
            "comparison_deming_chart_title": "Deming Regresyon Grafiği",
            "comparison_deming_ci_label": "Deming %95 GA (bootstrap, n=1000)",
            "comparison_pb_label": "Passing-Bablok",
            "comparison_pb_note": "ℹ️ Passing-Bablok, hata varyans oranı hakkında varsayım gerektirmeyen, aykırı değerlere karşı dayanıklı, parametrik olmayan bir yöntemdir (CLSI EP09 tarafından tercih edilir). Varyans oranını bilmiyorsanız bu sonuca Deming'den daha fazla güvenin.",
            "comparison_normality_ok": "✅ Farklar normal dağılıyor (Shapiro-Wilk p={p:.4f}) — %95 uyum sınırları geçerli.",
            "comparison_normality_warn": "⚠️ Farklar normal dağılmıyor olabilir (Shapiro-Wilk p={p:.4f}) — ±1.96 SD uyum sınırları güvenilir olmayabilir; dikkatli yorumlayın.",
            "comparison_prop_bias_ok": "✅ Anlamlı orantısal bias tespit edilmedi (eğim p={p:.4f}).",
            "comparison_prop_bias_warn": "⚠️ **Orantısal bias tespit edildi** (fark-ortalama eğimi={slope:.4f}, p={p:.4f}). Tek bir 'ortalama fark' değeri yanıltıcı olabilir — fark, ölçüm aralığına göre sistematik olarak değişiyor.",
            # CRM Production
            "crm_title": "🏭 Sertifikalı Referans Malzeme (CRM) Üretim Araçları",
            "crm_description": "ISO Guide 35 / ISO 17034 genel ilkelerine dayalı homojenlik testi, stabilite testi, atanmış değer & belirsizlik bütçesi ve sertifika (CoA) oluşturucu. Bu araçlar araştırma amaçlıdır; resmi bir CRM üretim/onay süreci yerine geçmez.",
            "crm_mode_label": "Araç seçin",
            "crm_mode_homogeneity": "🧪 Homojenlik Testi",
            "crm_mode_stability": "⏳ Stabilite Testi",
            "crm_mode_uncertainty": "📐 Atanmış Değer & Belirsizlik Bütçesi",
            "crm_mode_coa": "📜 Sertifika (CoA) Oluşturucu",
            # Homogeneity
            "homog_description": "Bir üretim partisindeki çok sayıda ünite/viyal arasında tek yönlü ANOVA ile homojenlik testi (Linsinger ve ark. 2001). F ≤ F_kritik ise partinin homojen olduğu kabul edilir.",
            "homog_n_units": "Ünite/Viyal Sayısı",
            "homog_unit_label": "Ünite {i} — Replikat Değerleri (her satıra bir değer)",
            "homog_calc_btn": "📊 Homojenliği Test Et",
            "homog_result_homogeneous": "✅ **Homojen** (F={F:.3f} ≤ F_kritik={fcrit:.3f}, p={p:.4f}). Partide istatistiksel olarak anlamlı bir üniteler-arası fark tespit edilmedi.",
            "homog_result_inhomogeneous": "❌ **Homojen Değil** (F={F:.3f} > F_kritik={fcrit:.3f}, p={p:.4f}). Üniteler arasında istatistiksel olarak anlamlı fark var — parti gözden geçirilmeli.",
            "homog_ubb_label": "Homojenlik Belirsizlik Katkısı (u_bb)",
            "homog_grand_mean_label": "Genel Ortalama",
            # Stability
            "stab_description": "Zaman içindeki ölçüm değerlerine doğrusal regresyon uygulayarak anlamlı bir bozunma trendi olup olmadığını test eder (ISO Guide 35 klasik stabilite yaklaşımı).",
            "stab_time_label": "Zaman Noktaları (örn. gün/ay — her satıra bir değer)",
            "stab_value_label": "Ölçülen Değerler (zaman noktalarıyla eşleştirilmiş, aynı sırada)",
            "stab_duration_label": "İncelenen Süre (raf ömrü, zaman noktalarıyla aynı birimde)",
            "stab_calc_btn": "📊 Stabiliteyi Test Et",
            "stab_result_stable": "✅ **Kararlı** (eğim p={p:.4f} ≥ 0.05). Zaman içinde istatistiksel olarak anlamlı bir trend tespit edilmedi.",
            "stab_result_unstable": "⚠️ **Anlamlı Trend Tespit Edildi** (eğim p={p:.4f} < 0.05). Materyal zaman içinde değişim gösteriyor olabilir — raf ömrü iddiasını gözden geçirin.",
            "stab_ustab_label": "Stabilite Belirsizlik Katkısı (u_stab)",
            "stab_slope_label": "Eğim (birim/zaman)",
            "stab_chart_title": "Stabilite — Zamana Karşı Değer",
            # Uncertainty budget
            "unc_description": "Karakterizasyon, homojenlik ve stabilite belirsizlik bileşenlerini GUM yaklaşımıyla (karesel toplam) birleştirerek nihai genişletilmiş belirsizliği hesaplar. Homojenlik/Stabilite sekmelerinde hesaplama yaptıysanız değerler otomatik doldurulur.",
            "unc_assigned_value_label": "Atanmış Değer",
            "unc_u_char_label": "Karakterizasyon Belirsizliği (u_char, mutlak birim)",
            "unc_u_char_help": "Atanmış değeri belirlemek için kullanılan ölçümlerin standart belirsizliği (örn. ortalamanın standart hatası).",
            "unc_u_bb_label": "Homojenlik Belirsizliği (u_bb)",
            "unc_u_stab_label": "Stabilite Belirsizliği (u_stab)",
            "unc_k_label": "Kapsam Faktörü (k)",
            "unc_use_cached": "Homojenlik/Stabilite sekmelerinden otomatik doldur",
            "unc_calc_btn": "📊 Belirsizlik Bütçesini Hesapla",
            "unc_result_title": "Atanmış Değer = {value:.5f} ± {U:.5f} (k={k}, %{urel:.2f} göreli)",
            "unc_no_cache": "ℹ️ Önceden hesaplanmış Homojenlik/Stabilite sonucu bulunamadı — manuel giriş kullanılacak.",
            # CoA
            "coa_description": "Atanmış değer, genişletilmiş belirsizlik ve izlenebilirlik bilgilerini içeren resmi tarzda bir Analiz Sertifikası (CoA) PDF'i oluşturur.",
            "coa_material_name": "Materyal Adı",
            "coa_lot_number": "Parti/Lot Numarası",
            "coa_producer": "Üretici/Laboratuvar",
            "coa_assigned_value_label": "Atanmış Değer",
            "coa_unit_label": "Birim",
            "coa_expanded_unc_label": "Genişletilmiş Belirsizlik (U)",
            "coa_k_label": "Kapsam Faktörü (k)",
            "coa_traceability_label": "İzlenebilirlik İfadesi",
            "coa_traceability_default": "Bu değerin izlenebilirliği, dijital PCR ile Poisson istatistiğine dayalı mutlak kantifikasyon yoluyla SI birimine (mol) dayanmaktadır.",
            "coa_validity_label": "Geçerlilik Tarihi / Raf Ömrü",
            "coa_generate_btn": "📥 Sertifikayı Oluştur (PDF)",
            "coa_download_btn": "⬇️ Analiz Sertifikası (CoA)",
            "dilution_factor_field_label": "Seyreltme Faktörü",
            "dilution_factor_field_help": "Bu örnek, dPCR reaksiyonuna eklenmeden önce seyreltildiyse buraya seyreltme faktörünü girin (örn. 1:100 seyreltme için 100). Varsayılan 1 = seyreltme yok. Not: Oran/CN/Kat Değişimi seyreltmeden etkilenmez (hedef ve referans eşit seyrelir); yalnızca mutlak konsantrasyon (kopya/µL) düzeltilir.",
            "dilution_settings_expander": "🧪 Seyreltme Ayarları (opsiyonel — stok konsantrasyonu için)",
            "dilution_mode_selector_label": "Nasıl belirtmek istersiniz?",
            "dilution_mode_volumes": "💧 Hacimlerden hesapla (önerilen)",
            "dilution_mode_manual": "🔢 Seyreltme faktörünü doğrudan gir",
            "dilution_rxn_vol_label": "Reaksiyon Hacmi (µL)",
            "dilution_rxn_vol_help": "Toplam dPCR reaksiyon karışımının hacmi (örn. Bio-Rad QX200 için tipik olarak 20-22 µL).",
            "dilution_tmpl_vol_label": "Eklenen Örnek/Şablon Hacmi (µL)",
            "dilution_tmpl_vol_help": "Reaksiyona eklediğiniz örnek (DNA/RNA) hacmi (örn. tipik olarak 2-9 µL).",
            "dilution_pre_dilution_label": "Ön-Seyreltme Faktörü (varsa)",
            "dilution_pre_dilution_help": "Reaksiyona eklemeden önce örneği ayrıca seyrelttiyseniz buraya o seyreltme faktörünü girin (örn. 1:10 ön-seyreltme için 10). Seyreltmediyseniz 1 bırakın.",
            "dilution_computed_factor": "📐 Hesaplanan toplam seyreltme faktörü: **{factor:.2f}×**  \n(Reaksiyon/Şablon = {rxn:.1f}/{tmpl:.1f} = {ratio:.2f}× {pre_txt})",
            "dilution_computed_factor_pre_txt": "× Ön-seyreltme {pre:.1f}×",
            "dilution_no_pre": "(ön-seyreltme yok)",
            "stock_conc_col": "Stok Konsantrasyonu (kopya/µL)",
            "dynamic_range_warning_label": "Dinamik Aralık Durumu",
            "dynamic_range_ok": "✅ Optimal aralıkta",
            "dynamic_range_low": "⚠️ Seyreltik (λ<0.05)",
            "dynamic_range_high": "⚠️ Yüksek (λ>3)",
            "dynamic_range_saturated": "❌ Doygunluk riski (λ>4)",
            "dynamic_range_qc_panel_label": "Dinamik Aralık Uyarısı",
            "lims_export_title": "🗂️ LIMS Uyumlu Dışa Aktarım",
            "lims_export_description": "Sonuçları yaygın LIMS içe aktarma şablonlarıyla uyumlu düz (flat) bir CSV formatında dışa aktarır. Her satır bir test sonucunu temsil eder.",
            "lims_operator_label": "Operatör",
            "lims_instrument_label": "Cihaz",
            "lims_run_date_label": "Çalışma Tarihi",
            "lims_export_btn": "📥 LIMS Formatında Dışa Aktar (CSV)",
            "excel_export_title": "📊 Biçimlendirilmiş Excel Raporu",
            "excel_export_description": "Özet, Giriş Verileri, Sonuçlar ve İstatistik sayfalarını içeren, renk kodlu (yukarı regüle=kırmızı, aşağı regüle=mavi, dışlanan replikat=sarı, anlamlı sonuç=yeşil) biçimlendirilmiş bir .xlsx dosyası indirin.",
            "excel_export_btn": "📥 Excel Raporunu İndir (.xlsx)",
            "lims_col_sample_id": "Sample_ID",
            "lims_col_test_code": "Test_Code",
            "lims_col_analyte": "Analyte",
            "lims_col_result_value": "Result_Value",
            "lims_col_result_unit": "Result_Unit",
            "lims_col_reference": "Reference_Value",
            "lims_col_flag": "Flag",
            "lims_col_result_date": "Result_Date",
            "lims_col_instrument": "Instrument",
            "lims_col_operator": "Operator",
            "lims_col_comments": "Comments",
            "crm_mode_equivalence": "⚖️ Parti-Parti Karşılaştırma (Lot Eşdeğerliği)",
            "equiv_description": "İki Tek Yönlü Test (TOST) prosedürüyle iki parti/lot arasındaki eşdeğerliği değerlendirir (Schuirmann 1987). Ortalama farkın %90 güven aralığı, tanımlanan eşdeğerlik marjı içinde kalırsa partiler eşdeğer kabul edilir.",
            "equiv_lot1_label": "Parti 1 (Referans) Değerleri (her satıra bir değer)",
            "equiv_lot2_label": "Parti 2 (Yeni/Test) Değerleri (her satıra bir değer)",
            "equiv_margin_label": "Eşdeğerlik Marjı (%, ± )",
            "equiv_calc_btn": "📊 Eşdeğerliği Test Et",
            "equiv_result_equivalent": "✅ **Eşdeğer** — fark %{diff:+.2f} (90% GA: %{lo:.2f} ile %{hi:.2f}), ±%{margin:.1f} marjı içinde.",
            "equiv_result_not_equivalent": "❌ **Eşdeğer Değil** — fark %{diff:+.2f} (90% GA: %{lo:.2f} ile %{hi:.2f}), ±%{margin:.1f} marjını aşıyor.",
            "equiv_mean1_label": "Parti 1 Ortalama",
            "equiv_mean2_label": "Parti 2 Ortalama",
            "equiv_chart_title": "Parti Karşılaştırması",
            "tab_batch": "Toplu Tarama",
            "batch_title": "🔬 Toplu Numune Tarama (CNV Taraması)",
            "batch_description": "Çok sayıda numuneyi (örn. bir kohort) tek bir referans/beklenen orana karşı hızlıca taramak için tasarlanmıştır. Her numune için ayrı replikat grupları oluşturmak yerine, CSV dosyanızı yükleyip her örnek için tek bir Poisson tabanlı güven aralığı hesaplanır.",
            "batch_uploader": "Toplu tarama CSV dosyası yükleyin",
            "batch_assay_title": "**Assay Ataması**",
            "batch_target_label": "Hedef gen assay'i",
            "batch_ref_label": "Referans gen assay(ler)i (birden fazla seçilirse partisyonlar havuzlanır)",
            "batch_settings_title": "**Tarama Ayarları**",
            "batch_expected_ratio": "Beklenen Oran (değişim yok referansı)",
            "batch_ploidy": "Referans Ploidi",
            "batch_run_btn": "▶ Taramayı Çalıştır",
            "batch_results_title": "📋 Tarama Sonuçları",
            "batch_col_sample": "Örnek",
            "batch_col_lambda_t": "λ Hedef",
            "batch_col_lambda_r": "λ Referans",
            "batch_col_ratio": "Oran",
            "batch_col_ci": "%95 GA (Oran)",
            "batch_col_cn": "Kopya Sayısı (CN)",
            "batch_col_conc": "Konsantrasyon (kopya/µL)",
            "batch_col_class": "Sınıflandırma",
            "batch_class_gain": "📈 Kazanım",
            "batch_class_loss": "📉 Kayıp",
            "batch_class_normal": "➖ Normal",
            "batch_chart_title": "Numuneler Arası Kopya Sayısı Taraması",
            "batch_download_btn": "📥 Tarama Sonuçlarını İndir (CSV)",
            "batch_no_data": "Sonuç yok — önce dosya yükleyip taramayı çalıştırın.",
            "batch_n_flagged": "{n} / {total} numune beklenen orandan istatistiksel olarak anlamlı şekilde farklı (%95 GA değişim-yok değerini içermiyor).",
            "batch_multi_gene_help": "Birden fazla hedef gen seçerseniz, sonuçlar sekmesinde tüm genler × örnekler için bir ısı haritası otomatik olarak oluşturulur.",
            "batch_heatmap_title": "🌡️ Çoklu Gen Isı Haritası",
            "batch_heatmap_description": "Seçilen tüm hedef genler için örnekler arası Kopya Sayısı veya Oran değerlerini tek bir ızgarada gösterir — kohort genelinde örüntüleri (örn. belirli örneklerde çoklu gen kazanımı) hızlıca tespit etmeye yardımcı olur.",
            "batch_heatmap_metric_label": "Gösterilecek metrik",
            "tab_vaf": "VAF Hesaplayıcı",
            "vaf_title": "🧬 VAF / Mutasyon Fraksiyonu Hesaplayıcı",
            "vaf_description": "Likit biyopsi / ctDNA izleme için tasarlanmıştır. Mutant ve Wild-type (yabanıl tip) assay'lerini içeren bir CSV yükleyin; her örnek için Fraksiyonel Bolluk (FA%, varyant allel fraksiyonu) ve %95 güven aralığı hesaplanır.",
            "vaf_uploader": "VAF CSV dosyası yükleyin",
            "vaf_assay_title": "**Assay Ataması**",
            "vaf_mutant_label": "Mutant assay",
            "vaf_wt_label": "Wild-type (yabanıl tip) assay",
            "vaf_run_btn": "▶ VAF Hesapla",
            "vaf_results_title": "📋 VAF Sonuçları",
            "vaf_col_sample": "Örnek",
            "vaf_col_lambda_mut": "λ Mutant",
            "vaf_col_lambda_wt": "λ Wild-type",
            "vaf_col_fa": "FA (%)",
            "vaf_col_ci": "%95 GA (FA%)",
            "vaf_col_conc_mut": "Mutant Konsantrasyonu (kopya/µL)",
            "vaf_col_detected": "Tespit Durumu",
            "vaf_detected_yes": "✅ Tespit Edildi",
            "vaf_detected_no": "❌ Tespit Edilemedi",
            "vaf_chart_title": "Örnekler Arası Fraksiyonel Bolluk (VAF%)",
            "vaf_download_btn": "📥 VAF Sonuçlarını İndir (CSV)",
            "vaf_no_data": "Sonuç yok — önce dosya yükleyip hesaplamayı çalıştırın.",
            "vaf_method_note": "ℹ️ FA = λ(mutant) / (λ(mutant) + λ(wild-type)). %95 GA delta-method ile hesaplanmıştır (Hindson et al. 2013, Anal Chem). Bu hesaplama, mutant ve wild-type assay'lerinin aynı partisyon setinde (multipleks) veya eşdeğer koşullarda ölçüldüğünü varsayar.",
            "vaf_n_detected": "{n} / {total} örnekte mutant alel tespit edildi (≥1 pozitif partisyon).",
        },
        "en": {
            "title": "🧪 AbsoluteGene: Digital PCR (dPCR/ddPCR) Gene Expression & CNV Analysis",
            "subtitle": "Developed by B. Yalçınkaya — the digital PCR companion to GeneQuantify",
            "tab_data": "Data Entry",
            "tab_results": "Results",
            "tab_report": "Report",
            "study_design": "⚙️ Study Design",
            "num_target_genes": "🔹 Number of Target Genes",
            "num_patient_groups": "🔹 Number of Patient Groups",
            "num_ref_genes": "🔹 Number of Reference Genes",
            "ref_gene_help": "dMIQE guidelines recommend ≥1 validated reference locus for normalization; ≥2 is preferred.",
            "ploidy_label": "🔹 Reference Locus Ploidy",
            "ploidy_help": "Known copy number of the reference locus, used for copy number variation (CNV) calculation (typically 2 for a diploid organism). ⚠️ Important assumption: this value assumes the reference locus has EXACTLY this copy number in ALL samples (including controls). If the reference locus itself is affected by aneuploidy or CNV, all downstream CN calculations will be silently wrong — verify this when choosing your reference locus.",
            "partition_vol_label": "🔹 Partition Volume (nL)",
            "partition_vol_help": "Volume per droplet/well. Bio-Rad QX200: ~0.85 nL, QX ONE: ~0.7 nL, Qiagen QIAcuity: varies by plate type. Only affects copies/µL, not the ratio/CNV calculation.",
            "qc_min_partitions": "🔹 Minimum Accepted Partitions (QC)",
            "qc_min_partitions_help": "Replicates below this value are flagged as low quality (dMIQE recommendation: typically ≥10,000 for ddPCR).",
            "outlier_section_title": "### 🔍 Outlier Detection Settings",
            "outlier_enable": "Enable outlier detection",
            "outlier_enable_help": "Detects statistically extreme replicates in λ (copies per partition).",
            "outlier_method_label": "Detection method",
            "outlier_method_help": "Grubbs: for normally distributed data. IQR: non-parametric, robust for skewed distributions.",
            "grubbs_power_warning": "⚠️ Grubbs' test has low power at small n (3-5 replicates, typical for dPCR) — it may easily miss a genuine outlier. Don't interpret 'no outlier found' as 'the data is definitely clean'.",
            "outlier_alpha_label": "Significance level (α)",
            "outlier_iqr_label": "IQR multiplier (k)",
            "patient_data_header": "📥 Enter Patient and Control Group Partition Data",
            "target_gene": "Target Gene",
            "reference_gene": "Reference Gene",
            "control_group": "🧬 Control Group",
            "patient_group": "🩸 Patient Group",
            "positive_partitions": "Positive Partition Count",
            "total_partitions": "Total Accepted Partition Count",
            "input_format_info": "ℹ️ Enter one replicate per line. The number of lines in 'Positive' and 'Total' boxes must match.",
            "warning_empty_input": "⚠️ Warning: Enter data one value per line; the number of lines in Positive and Total boxes must be equal.",
            "warning_field_empty": "⚠️ One or more required fields are empty (Target Positive/Total or a Reference Gene Positive/Total). Please enter data in all boxes.",
            "warning_all_excluded_qc": "⚠️ All entered replicates were excluded due to **low partition count (QC threshold: {thr})** — none could be analyzed. Consider lowering the threshold (in Study Design settings) or checking your data.",
            "warning_all_excluded_saturated": "⚠️ All entered replicates are **saturated** (all partitions positive, λ cannot be calculated) — none could be analyzed. The sample may need further dilution.",
            "warning_all_excluded_outlier": "⚠️ All entered replicates were **flagged and excluded as outliers** — none remain for analysis. Check your outlier exclusion checkboxes.",
            "warning_all_excluded_mixed": "⚠️ All entered replicates were excluded (a mix of QC/saturation/outlier reasons) — none could be analyzed. See the Input Data Table in the Results tab for details.",
            "qc_fail_warning": "⚠️ **Low partition count warning:** {n} replicate(s) have accepted partitions below the minimum threshold ({thr}). These replicates may have reduced reliability.",
            "saturation_warning": "⚠️ **Saturation warning:** {n} replicate(s) have all partitions positive (p≥1.0) — λ cannot be calculated. Further dilution is recommended.",
            "gr_tbl": "📋 Input Data Table (incl. λ and 95% CI)",
            "nil_mine": "📊 Results",
            "download_csv": "📥 Download CSV",
            "generate_pdf": "📥 Prepare PDF Report",
            "pdf_report": "Digital PCR Analysis Report",
            "sample_number": "Replicate #",
            "lambda_col": "λ (copies/partition)",
            "ci_low_col": "95% CI Lower",
            "ci_high_col": "95% CI Upper",
            "conc_col": "Concentration (copies/µL)",
            "qc_col": "QC Status",
            "qc_pass": "✅ Passed",
            "qc_fail": "❌ Low n",
            "qc_saturated": "❌ Saturated",
            "ratio_col": "Normalized Ratio (Target/Reference)",
            "cn_col": "Copy Number (CN)",
            "fc_col": "Fold Change (vs Control)",
            "regulation_status": "Regulation Status",
            "no_change": "No Change",
            "upregulated": "Upregulated / Gain",
            "downregulated": "Downregulated / Loss",
            "outlier_excluded_no": "No",
            "outlier_excluded_yes": "Yes",
            "genorm_title": "Reference Locus Stability",
            "stable": "Stable",
            "borderline": "Borderline",
            "unstable": "Unstable",
            "m_value": "M-value",
            "method_comparison": "📊 Result Summary",
            "statistical_results": "📈 Statistical Results",
            "test_type": "Test Type",
            "test_method": "Test Method",
            "test_pvalue": "Test P-value",
            "significance": "Significance",
            "significant": "Significant",
            "insignificant": "Insignificant",
            "parametric": "Parametric",
            "non_parametric": "Nonparametric",
            "t_test": "t-test",
            "welch_t_test": "Welch t-test",
            "mann_whitney_u_test": "Mann-Whitney U test",
            "stat_decision_title": "🔬 Statistical decision",
            "stat_decision_steps": "**Step-by-step test selection:**",
            "stat_shapiro_title": "**1. Shapiro-Wilk normality test**",
            "stat_normal": "Normal",
            "stat_nonnormal": "Non-normal",
            "stat_levene_title": "**2. Levene variance homogeneity test**",
            "stat_levene_skipped": "**2. Levene test** — *skipped* (normality not met)",
            "stat_equal_var": "Equal variances",
            "stat_unequal_var": "Unequal variances",
            "stat_selected_test": "**3. Selected test:**",
            "stat_reason": "**Reason:**",
            "stat_result": "**Result:**",
            "stat_reason_nonnormal": "Non-normal distribution in one or both groups",
            "stat_reason_normal_equal": "Both groups normal + equal variances",
            "stat_reason_normal_unequal": "Both groups normal + unequal variances",
            "stat_multigroup_note": "⚠️ Note: When ≥3 groups are present, see the Multi-Group Comparison section below.",
            "multigroup_title": "## 📊 Multi-Group Comparison Analysis",
            "multigroup_omnibus_test": "Omnibus Test",
            "multigroup_pvalue": "p-value",
            "multigroup_result": "Result",
            "multigroup_significant": "Significant",
            "multigroup_not_significant": "Not significant",
            "multigroup_omnibus_ns": "ℹ️ Omnibus test is not significant (p ≥ 0.05). Post-hoc comparisons are shown for completeness.",
            "multigroup_posthoc_label": "**Post-hoc:**",
            "multigroup_dl_button": "📥 Download post-hoc results —",
            "multigroup_2group_note": "ℹ️ Only 2 groups detected. Pairwise statistics are reported above.",
            "multigroup_decision_normal_equal": "✅ Normal + equal variances → **One-way ANOVA + Tukey HSD**",
            "multigroup_decision_normal_unequal": "⚠️ Normal + unequal variances → **Welch ANOVA + Games-Howell**",
            "multigroup_decision_nonnormal": "⚠️ Non-normal → **Kruskal-Wallis + Dunn post-hoc**",
            "multigene_title": "### 🧬 Multi-Gene Multiple Comparison Correction",
            "multigene_sig_raw": "Significant (raw)",
            "multigene_sig_bonf": "Significant (Bonferroni)",
            "multigene_sig_fdr": "Significant (FDR B-H)",
            "multigene_warning": "⚠️ After correction, {lost} result(s) no longer significant after FDR adjustment.",
            "multigene_success": "✅ All {n} significant result(s) remain significant after FDR correction.",
            "multigene_no_sig": "No significant pairwise results detected (raw p < 0.05).",
            "multigene_dl_button": "📥 Download corrected p-values (CSV)",
            "multigene_1gene_note": "ℹ️ Only 1 target gene analysed — multi-gene correction not applicable.",
            "distribution_graph": "Distribution Graph",
            "x_axis_title": "Group Name",
            "dist_plot_mode_label": "📊 Distribution Plot — Display Mode",
            "dist_plot_ratio": "Normalized Ratio (Target/Reference) — recommended",
            "dist_plot_lambda": "λ (copies/partition) — raw",
            "dist_plot_fc": "Fold Change — relative to control",
            "error_no_data": "No data found, PDF could not be generated.",
            "pdf_ready": "{n} records ready — you can generate the PDF.",
            "sidebar_github_btn": "⭐ View Source on GitHub",
            "sidebar_sister_tool": "🧬 For qPCR: GeneQuantify",
            "rdml_expander": "ℹ️ About data entry",
            "guide_btn": "📘 User Guide",
            "sidebar_example_title": "📋 Load Example Data",
            "sidebar_example_select": "Select scenario",
            "sidebar_example_load_btn": "▶ Load Scenario",
            "sidebar_example_loaded": "✅ {s} loaded! Switch to the Data Entry tab.",
            "ntc_expander": "🧫 NTC (No-Template Control) / LOD-LOQ — optional",
            "ntc_description": "Enter your No-Template Control replicates to calculate the limit of detection (LOD) and limit of quantification (LOQ). Leave blank to skip this step.",
            "ntc_positive_label": "NTC Positive Partition Count",
            "ntc_total_label": "NTC Total Accepted Partition Count",
            "ntc_calc_btn": "📊 Calculate LOD/LOQ",
            "lod_result_title": "Limit of Detection (LOD) and Limit of Quantification (LOQ) — {gene}",
            "lod_label": "LOD (copies/µL)",
            "loq_label": "LOQ (copies/µL)",
            "ntc_zero_note": "ℹ️ No positive partitions detected across NTC replicates (n={n} partitions pooled). LOD was calculated using the rule-of-three upper confidence bound: 3/n.",
            "ntc_contamination_warning": "⚠️ Background signal detected in NTC replicates (λ={lam:.5f}, {pos}/{tot} positive partitions). This may indicate reagent or cross-contamination — LOD was calculated as this background level plus a safety margin.",
            "below_lod_flag": "❌ Below LOD",
            "between_lod_loq_flag": "⚠️ Between LOD-LOQ",
            "above_loq_flag": "✅ Above LOQ",
            "lod_qc_col": "LOD/LOQ Status",
            "loq_heuristic_note": "Note: LOQ is calculated as 3× LOD, a commonly used heuristic. For a more rigorous LOQ, empirical validation based on reproducibility (CV%) is recommended.",
            "empirical_loq_label": "Empirical LOQ Cross-Check (CV%-based)",
            "empirical_loq_pass": "✅ Reproducibility acceptable",
            "empirical_loq_fail": "⚠️ Low reproducibility (CV>25%)",
            "csv_import_expander": "📂 Import Instrument CSV (QuantaSoft / QX Manager / QIAcuity)",
            "csv_import_description": "Upload a CSV export from Bio-Rad QuantaSoft/QX Manager or QIAGEN QIAcuity software. Columns are auto-detected where possible; you can manually map any that aren't recognized.",
            "csv_uploader": "Choose CSV file",
            "csv_parse_error": "❌ File parse error: {err}",
            "csv_col_mapping_title": "**Column Mapping**",
            "csv_col_sample": "Sample Name Column",
            "csv_col_target": "Target/Assay Name Column",
            "csv_col_positives": "Positive Partitions Column",
            "csv_col_total": "Total Partitions Column",
            "csv_col_auto_detected": "✅ Auto-detected",
            "csv_col_manual_needed": "⚠️ Not auto-detected — please select manually",
            "csv_preview_title": "Preview parsed data",
            "csv_assay_assignment_title": "**Assay (Target/Reference Gene) Assignment**",
            "csv_target_assays_label": "Target gene assay(s)",
            "csv_ref_assays_label": "Reference gene assay(s)",
            "csv_group_assignment_title": "**Group Assignment**",
            "csv_ctrl_label": "Control sample name(s) (comma-separated substrings)",
            "csv_n_patient_groups": "Number of patient groups",
            "csv_patient_label": "Patient group {i} sample name(s)",
            "csv_apply_btn": "✅ Apply Import to Data Entry",
            "csv_apply_success": "✅ {n} value(s) loaded into the Data Entry tab! Switch to review and adjust.",
            "csv_apply_warning": "⚠️ No values were mapped. Check that your control/patient labels match the sample names.",
            "download_example_csv": "📄 Download Example CSV",
            "dilution_expander": "🧪 Optimal Dilution / Dynamic Range Calculator",
            "dilution_description": "Poisson-based quantification is most precise in the range of ~1.0–2.0 copies per partition (λ). This calculator recommends a dilution factor based on your current measurement.",
            "dilution_mode_label": "Input method",
            "dilution_mode_counts": "From observed partition counts",
            "dilution_mode_lambda": "From known λ value",
            "dilution_positive_label": "Positive Partition Count",
            "dilution_total_label": "Total Partition Count",
            "dilution_lambda_label": "Current λ (copies/partition)",
            "dilution_target_label": "Target λ (copies/partition)",
            "dilution_calc_btn": "📊 Recommend Dilution",
            "dilution_result_optimal": "✅ **In the optimal range** (λ={lam:.3f}). No further dilution needed.",
            "dilution_result_too_low": "⚠️ **Too dilute** (λ={lam:.3f}). For your next prep, set your dilution factor to **{factor:.3f}× your current factor** (i.e., dilute less).",
            "dilution_result_too_high": "⚠️ **Above the optimal range** (λ={lam:.3f}). For your next prep, set your dilution factor to **{factor:.3f}× your current factor** (i.e., dilute more).",
            "dilution_result_saturated": "❌ **Near/at saturation risk** (λ={lam:.3f}). For your next prep, set your dilution factor to **{factor:.3f}× your current factor**.",
            "dilution_factor_note": "Since this calculator doesn't know your original dilution factor, it recommends a **relative multiplier**: multiply the dilution factor you used by this value for your next prep. Example: if you currently dilute 1:50 and the recommended factor is 2.0, try 1:100 next time.",
            "multiplex_expander": "🔀 2-Color Multiplex Cluster Converter",
            "multiplex_description": "For ddPCR data where two targets (e.g. target gene + reference gene) are measured in **multiplex on the same partitions**. Enter the 2×2 cluster counts — this tool computes the marginal positive/total counts and the ratio while correctly accounting for the correlation between channels (WITHOUT assuming independence).",
            "multiplex_n11_label": "Double Positive (Target+ Ref+)",
            "multiplex_n10_label": "Target-Positive Only (Target+ Ref-)",
            "multiplex_n01_label": "Reference-Positive Only (Target- Ref+)",
            "multiplex_n00_label": "Double Negative (Target- Ref-)",
            "multiplex_calc_btn": "📊 Convert and Calculate",
            "multiplex_result_title": "**Result:**",
            "multiplex_pos_target_label": "Target Positive Partitions (marginal)",
            "multiplex_pos_ref_label": "Reference Positive Partitions (marginal)",
            "multiplex_total_label": "Total Partitions",
            "multiplex_ratio_covaware": "Ratio (covariance-aware) 95% CI",
            "multiplex_ratio_indep": "Ratio (independence assumption) 95% CI — for comparison",
            "multiplex_paste_hint": "💡 You can copy the marginal positive/total values above into the main Data Entry fields. The covariance-aware confidence interval correctly reflects the correlation from shared partitions and is typically narrower/more accurate than the independence assumption.",
            "qc_panel_title": "🩺 Data Quality Summary",
            "qc_panel_total": "Total Replicates",
            "qc_panel_qc_fail": "Low Partition (QC)",
            "qc_panel_saturated": "Saturated",
            "qc_panel_outlier": "Outlier (Excluded)",
            "qc_panel_below_lod": "Below-LOD Results",
            "qc_panel_high_cv": "High CV (>25%)",
            "qc_panel_verdict_good": "✅ **Good** — the dataset looks healthy, no significant quality issues detected.",
            "qc_panel_verdict_caution": "⚠️ **Caution** — some replicates have quality flags. Consider these when interpreting results.",
            "qc_panel_verdict_poor": "❌ **Review recommended** — a significant proportion of quality flags (QC failures, saturation, or high variability) were detected. We recommend reviewing the data before reporting.",
            "qc_panel_no_data": "No data entered yet.",
            "project_expander": "💾 Save / Load Project",
            "project_description": "Save all your data entry (study design, control/patient group data, NTC) as a JSON file, then resume where you left off later.",
            "project_export_btn": "📥 Export Project (JSON)",
            "project_import_uploader": "Upload project file (.json)",
            "project_import_success": "✅ Project loaded ({n} fields restored)! Switch to the Data Entry tab.",
            "project_import_error": "❌ Could not read project file: {err}",
            "history_expander": "📚 Session History (Snapshots)",
            "history_description": "Save and restore named snapshots of your current work state within this session. ⚠️ These are only stored in this browser session — they're lost if you refresh the page. For permanent storage, use 'Export Project' above to download a JSON file.",
            "history_name_label": "Snapshot name",
            "history_save_btn": "📌 Save Current State",
            "history_saved_success": "✅ '{name}' saved.",
            "history_select_label": "Saved snapshots",
            "history_restore_btn": "↩️ Restore",
            "history_download_btn": "📥 Download",
            "history_delete_btn": "🗑️ Delete",
            "history_restore_success": "✅ Snapshot restored ({n} fields). Switch to the Data Entry tab.",
            "history_no_snapshots": "No snapshots saved yet.",
            "advanced_mode_label": "🎓 Advanced Mode",
            "advanced_mode_help": "Shows the Clinical Tools and CRM Production tabs, plus advanced tools like the Multiplex Converter and NTC/LOD-LOQ. When off, only the basic dPCR analysis flow (Data Entry → Results → Report) is shown.",
            "simple_mode_caption": "ℹ️ Simple Mode is active — only the basic analysis flow is shown. Turn on Advanced Mode for clinical/CRM/multiplex tools.",
            "advanced_gate_title": "🎓 This is an Advanced Tool",
            "advanced_gate_message": "This tab contains advanced statistical/metrology tools (measurement uncertainty, homogeneity testing, method comparison, etc.). Turn on **Advanced Mode** in the left sidebar to use it.",
            "batch_pdf_btn": "📥 Prepare Screening Report (PDF)",
            "batch_pdf_report": "Batch Sample Screening Report",
            "batch_pdf_description": "This report includes, for each sample, the Poisson-derived λ, normalized ratio, 95% confidence interval, and classification relative to the expected ratio. Classification is based on whether the sample's ratio confidence interval includes the expected (no-change) value.",
            "vaf_pdf_btn": "📥 Prepare VAF Report (PDF)",
            "vaf_pdf_report": "VAF / Mutation Fraction Report",
            "vaf_pdf_description": "This report includes, for each sample, the mutant and wild-type assay λ values, Fractional Abundance (FA%), and the delta-method 95% confidence interval.",
            "tab_clinical": "Clinical Tools",
            "tab_crm": "CRM Production",
            "clinical_title": "🩺 Clinical Validation Tools",
            "clinical_description": "Tools for clinical/diagnostic validation: measurement uncertainty, reference change value, precision study, and method comparison. These tools are for research and method-validation purposes; not validated for clinical diagnostic decision-making.",
            "clinical_mode_label": "Select tool",
            "clinical_mode_mu": "📐 Measurement Uncertainty (MU) Budget",
            "clinical_mode_rcv": "📈 Reference Change Value (RCV)",
            "clinical_mode_precision": "🔁 Precision Study",
            "clinical_mode_comparison": "⚖️ Method Comparison",
            "mu_description": "Using the GUM (Guide to the Expression of Uncertainty in Measurement) approach, independent uncertainty sources are combined in quadrature and multiplied by a coverage factor (k) to obtain the expanded uncertainty.",
            "mu_poisson_source": "Poisson Counting Statistics Uncertainty (%)",
            "mu_poisson_help": "Can be auto-calculated (from replicate CV%) or entered manually.",
            "mu_poisson_auto": "Auto-fetch from results",
            "mu_poisson_auto_replicate": "📊 From Replicate CV% (total observed variability)",
            "mu_poisson_auto_theoretical": "🎯 From Theoretical Poisson SE (counting statistics only)",
            "mu_double_count_warning": "⚠️ This value reflects the **total observed variability** between replicates (pipetting + run-to-run + Poisson counting, all combined). If you also add pipetting/precision uncertainty below, you may be **double-counting the same source of variation**. Consider using the 'Theoretical Poisson SE' option instead.",
            "mu_double_count_error": "❌ **Double-counting risk:** You selected 'From Replicate CV%' AND entered additional pipetting/precision uncertainty. Replicate CV may already include these sources. Either zero out pipetting/precision (use replicate CV alone), or switch to 'Theoretical Poisson SE' mode and add the other components separately.",
            "mu_theoretical_note": "This value is derived purely from partition counting statistics (Poisson) — it does not include pipetting or biological variation. It can therefore be safely combined with the other components below.",
            "mu_no_lambda_cache": "ℹ️ No cached λ data found for this result — please calculate it in the Data Entry tab first.",
            "unbalanced_design_note": "ℹ️ Unbalanced design detected (replicates per group: {n_list}). No data was discarded — variance components were calculated using the standard method-of-moments correction for unbalanced one-way random-effects ANOVA (Searle et al. 1992).",
            "mu_poisson_manual": "Enter manually",
            "mu_pipetting_label": "Pipetting / Dilution Uncertainty (%)",
            "mu_pipetting_help": "Typically ~1-2% CV for good pipetting technique. Enter based on your own validation data.",
            "mu_precision_label": "Inter-run Precision Uncertainty (%, optional)",
            "mu_precision_help": "Between-day CV% from a separate precision study. Leave at 0 if unknown.",
            "mu_gene_select": "Select result (Gene / Group)",
            "mu_k_label": "Coverage Factor (k)",
            "mu_calc_btn": "📊 Calculate Uncertainty Budget",
            "mu_result_title": "Result: {value} ± {U:.2f}% (k={k}, ~95% confidence level)",
            "mu_budget_table_title": "**Uncertainty Budget Table**",
            "mu_col_source": "Source",
            "mu_col_contribution": "Contribution (%, relative standard uncertainty)",
            "mu_col_combined": "Combined Standard Uncertainty (u_c)",
            "mu_col_expanded": "Expanded Uncertainty (U)",
            "rcv_description": "Determines whether the difference between two serial results is statistically significant beyond what would be expected from analytical and biological variability alone (Fraser & Harris, Crit Rev Clin Lab Sci 1989).",
            "rcv_result1_label": "Result 1 (earlier, e.g. %VAF or copies/µL)",
            "rcv_result2_label": "Result 2 (later)",
            "rcv_cv_analytical_label": "Analytical CV (%)",
            "rcv_cv_analytical_help": "The repeatability CV% of this measurement (replicate CV or from the MU budget).",
            "rcv_cv_biological_label": "Biological Variation CV (%)",
            "rcv_cv_biological_help": "The known within-subject biological variation of the analyte (from literature — analyte-specific).",
            "rcv_z_label": "z-value (confidence level)",
            "rcv_calc_btn": "📊 Calculate RCV",
            "rcv_result_significant": "🔴 **Significant change** — {change:+.2f}% change exceeds the RCV threshold of {rcv:.2f}%. This change cannot be explained by measurement noise alone.",
            "rcv_result_not_significant": "🟢 **No significant change** — {change:+.2f}% change is within the RCV threshold of {rcv:.2f}%. Consistent with analytical/biological noise.",
            "precision_description": "Separates within-day (repeatability) and between-day (intermediate precision) variance components using nested ANOVA (CLSI EP05-A3 approach).",
            "precision_n_days": "Number of Days/Runs",
            "precision_day_label": "Day {i} — Replicate Values (one per line)",
            "precision_calc_btn": "📊 Calculate Precision",
            "precision_repeatability_cv": "Repeatability CV (%)",
            "precision_between_day_cv": "Between-Day CV (%)",
            "precision_total_cv": "Total (Intermediate Precision) CV (%)",
            "precision_grand_mean": "Grand Mean",
            "comparison_description": "Assesses agreement between two measurement methods using Bland-Altman analysis and Deming regression.",
            "comparison_method1_label": "Method 1 Values (e.g. dPCR — one per line)",
            "comparison_method2_label": "Method 2 Values (e.g. qPCR — paired, same order)",
            "comparison_variance_ratio_label": "Variance Ratio (λ, for Deming)",
            "comparison_calc_btn": "📊 Calculate Comparison",
            "comparison_bias_label": "Mean Difference (Bias)",
            "comparison_loa_label": "Limits of Agreement (95%)",
            "comparison_deming_label": "Deming Regression",
            "comparison_ba_chart_title": "Bland-Altman Plot",
            "comparison_deming_chart_title": "Deming Regression Plot",
            "comparison_deming_ci_label": "Deming 95% CI (bootstrap, n=1000)",
            "comparison_pb_label": "Passing-Bablok",
            "comparison_pb_note": "ℹ️ Passing-Bablok is a non-parametric method that requires no assumption about the error-variance ratio and is robust to outliers (preferred by CLSI EP09). Trust this result more than Deming's if you don't know the true variance ratio.",
            "comparison_normality_ok": "✅ Differences are approximately normal (Shapiro-Wilk p={p:.4f}) — the 95% limits of agreement are valid.",
            "comparison_normality_warn": "⚠️ Differences may not be normally distributed (Shapiro-Wilk p={p:.4f}) — the ±1.96 SD limits of agreement may not be reliable; interpret with caution.",
            "comparison_prop_bias_ok": "✅ No significant proportional bias detected (slope p={p:.4f}).",
            "comparison_prop_bias_warn": "⚠️ **Proportional bias detected** (difference-vs-mean slope={slope:.4f}, p={p:.4f}). A single 'mean difference' figure may be misleading — the disagreement changes systematically across the measurement range.",
            # CRM Production
            "crm_title": "🏭 Certified Reference Material (CRM) Production Tools",
            "crm_description": "Homogeneity testing, stability testing, assigned value & uncertainty budget, and certificate (CoA) generation, based on the general principles of ISO Guide 35 / ISO 17034. These tools are for research purposes and do not replace a formal CRM production/certification process.",
            "crm_mode_label": "Select tool",
            "crm_mode_homogeneity": "🧪 Homogeneity Testing",
            "crm_mode_stability": "⏳ Stability Testing",
            "crm_mode_uncertainty": "📐 Assigned Value & Uncertainty Budget",
            "crm_mode_coa": "📜 Certificate (CoA) Generator",
            "homog_description": "Tests homogeneity across multiple units/vials from a production batch using one-way ANOVA (Linsinger et al. 2001). The batch is considered homogeneous if F ≤ F_critical.",
            "homog_n_units": "Number of Units/Vials",
            "homog_unit_label": "Unit {i} — Replicate Values (one per line)",
            "homog_calc_btn": "📊 Test Homogeneity",
            "homog_result_homogeneous": "✅ **Homogeneous** (F={F:.3f} ≤ F_critical={fcrit:.3f}, p={p:.4f}). No statistically significant between-unit difference was detected.",
            "homog_result_inhomogeneous": "❌ **Not Homogeneous** (F={F:.3f} > F_critical={fcrit:.3f}, p={p:.4f}). A statistically significant between-unit difference was found — the batch should be reviewed.",
            "homog_ubb_label": "Homogeneity Uncertainty Contribution (u_bb)",
            "homog_grand_mean_label": "Grand Mean",
            "stab_description": "Applies linear regression to measurements over time to test for a significant degradation trend (ISO Guide 35 classical stability approach).",
            "stab_time_label": "Time Points (e.g. day/month — one per line)",
            "stab_value_label": "Measured Values (paired with time points, same order)",
            "stab_duration_label": "Study Duration (shelf life, same units as time points)",
            "stab_calc_btn": "📊 Test Stability",
            "stab_result_stable": "✅ **Stable** (slope p={p:.4f} ≥ 0.05). No statistically significant trend over time was detected.",
            "stab_result_unstable": "⚠️ **Significant Trend Detected** (slope p={p:.4f} < 0.05). The material may be changing over time — review the shelf-life claim.",
            "stab_ustab_label": "Stability Uncertainty Contribution (u_stab)",
            "stab_slope_label": "Slope (unit/time)",
            "stab_chart_title": "Stability — Value vs Time",
            "unc_description": "Combines characterization, homogeneity, and stability uncertainty components using the GUM approach (sum of squares) to compute the final expanded uncertainty. If you've run the Homogeneity/Stability tools, values are auto-filled.",
            "unc_assigned_value_label": "Assigned Value",
            "unc_u_char_label": "Characterization Uncertainty (u_char, absolute units)",
            "unc_u_char_help": "Standard uncertainty of the measurements used to determine the assigned value (e.g. standard error of the mean).",
            "unc_u_bb_label": "Homogeneity Uncertainty (u_bb)",
            "unc_u_stab_label": "Stability Uncertainty (u_stab)",
            "unc_k_label": "Coverage Factor (k)",
            "unc_use_cached": "Auto-fill from Homogeneity/Stability tabs",
            "unc_calc_btn": "📊 Calculate Uncertainty Budget",
            "unc_result_title": "Assigned Value = {value:.5f} ± {U:.5f} (k={k}, {urel:.2f}% relative)",
            "unc_no_cache": "ℹ️ No previously calculated Homogeneity/Stability result found — manual input will be used.",
            "coa_description": "Generates a formal-style Certificate of Analysis (CoA) PDF containing the assigned value, expanded uncertainty, and traceability information.",
            "coa_material_name": "Material Name",
            "coa_lot_number": "Batch/Lot Number",
            "coa_producer": "Producer/Laboratory",
            "coa_assigned_value_label": "Assigned Value",
            "coa_unit_label": "Unit",
            "coa_expanded_unc_label": "Expanded Uncertainty (U)",
            "coa_k_label": "Coverage Factor (k)",
            "coa_traceability_label": "Traceability Statement",
            "coa_traceability_default": "Traceability of this value is based on absolute quantification via digital PCR using Poisson statistics, traceable to the SI unit (mole).",
            "coa_validity_label": "Expiry Date / Shelf Life",
            "coa_generate_btn": "📥 Generate Certificate (PDF)",
            "coa_download_btn": "⬇️ Certificate of Analysis (CoA)",
            "dilution_factor_field_label": "Dilution Factor",
            "dilution_factor_field_help": "If this sample was diluted before being added to the dPCR reaction, enter the dilution factor here (e.g. 100 for a 1:100 dilution). Default 1 = no dilution. Note: Ratio/CN/Fold Change are unaffected by dilution (target and reference are diluted equally); only the absolute concentration (copies/µL) is corrected.",
            "dilution_settings_expander": "🧪 Dilution Settings (optional — for stock concentration)",
            "dilution_mode_selector_label": "How would you like to specify this?",
            "dilution_mode_volumes": "💧 Calculate from volumes (recommended)",
            "dilution_mode_manual": "🔢 Enter dilution factor directly",
            "dilution_rxn_vol_label": "Reaction Volume (µL)",
            "dilution_rxn_vol_help": "Total volume of the dPCR reaction mix (e.g. typically 20-22 µL for Bio-Rad QX200).",
            "dilution_tmpl_vol_label": "Sample/Template Volume Added (µL)",
            "dilution_tmpl_vol_help": "Volume of sample (DNA/RNA) you added to the reaction (e.g. typically 2-9 µL).",
            "dilution_pre_dilution_label": "Pre-Dilution Factor (if any)",
            "dilution_pre_dilution_help": "If you diluted the sample separately before adding it to the reaction, enter that dilution factor here (e.g. 10 for a 1:10 pre-dilution). Leave at 1 if you didn't pre-dilute.",
            "dilution_computed_factor": "📐 Calculated total dilution factor: **{factor:.2f}×**  \n(Reaction/Template = {rxn:.1f}/{tmpl:.1f} = {ratio:.2f}× {pre_txt})",
            "dilution_computed_factor_pre_txt": "× Pre-dilution {pre:.1f}×",
            "dilution_no_pre": "(no pre-dilution)",
            "stock_conc_col": "Stock Concentration (copies/µL)",
            "dynamic_range_warning_label": "Dynamic Range Status",
            "dynamic_range_ok": "✅ Optimal range",
            "dynamic_range_low": "⚠️ Too dilute (λ<0.05)",
            "dynamic_range_high": "⚠️ High (λ>3)",
            "dynamic_range_saturated": "❌ Saturation risk (λ>4)",
            "dynamic_range_qc_panel_label": "Dynamic Range Warning",
            "lims_export_title": "🗂️ LIMS-Compatible Export",
            "lims_export_description": "Exports results as a flat CSV compatible with common LIMS import templates. Each row represents one test result.",
            "lims_operator_label": "Operator",
            "lims_instrument_label": "Instrument",
            "lims_run_date_label": "Run Date",
            "lims_export_btn": "📥 Export in LIMS Format (CSV)",
            "excel_export_title": "📊 Formatted Excel Report",
            "excel_export_description": "Download a formatted .xlsx file with Summary, Input Data, Results, and Statistics sheets, color-coded (upregulated=red, downregulated=blue, excluded replicate=yellow, significant result=green).",
            "excel_export_btn": "📥 Download Excel Report (.xlsx)",
            "lims_col_sample_id": "Sample_ID",
            "lims_col_test_code": "Test_Code",
            "lims_col_analyte": "Analyte",
            "lims_col_result_value": "Result_Value",
            "lims_col_result_unit": "Result_Unit",
            "lims_col_reference": "Reference_Value",
            "lims_col_flag": "Flag",
            "lims_col_result_date": "Result_Date",
            "lims_col_instrument": "Instrument",
            "lims_col_operator": "Operator",
            "lims_col_comments": "Comments",
            "crm_mode_equivalence": "⚖️ Batch-to-Batch Comparison (Lot Equivalence)",
            "equiv_description": "Assesses equivalence between two lots/batches using the Two One-Sided Tests (TOST) procedure (Schuirmann 1987). Lots are considered equivalent if the 90% confidence interval of the mean difference falls within the defined equivalence margin.",
            "equiv_lot1_label": "Lot 1 (Reference) Values (one per line)",
            "equiv_lot2_label": "Lot 2 (New/Test) Values (one per line)",
            "equiv_margin_label": "Equivalence Margin (%, ±)",
            "equiv_calc_btn": "📊 Test Equivalence",
            "equiv_result_equivalent": "✅ **Equivalent** — difference {diff:+.2f}% (90% CI: {lo:.2f}% to {hi:.2f}%), within the ±{margin:.1f}% margin.",
            "equiv_result_not_equivalent": "❌ **Not Equivalent** — difference {diff:+.2f}% (90% CI: {lo:.2f}% to {hi:.2f}%), exceeds the ±{margin:.1f}% margin.",
            "equiv_mean1_label": "Lot 1 Mean",
            "equiv_mean2_label": "Lot 2 Mean",
            "equiv_chart_title": "Batch Comparison",
            "tab_batch": "Batch Screening",
            "batch_title": "🔬 Batch Sample Screening (CNV Screening)",
            "batch_description": "Designed for quickly screening many samples (e.g. a cohort) against a single reference/expected ratio. Instead of building replicate groups per sample, upload a CSV and a Poisson-based confidence interval is computed for each sample individually.",
            "batch_uploader": "Upload batch screening CSV file",
            "batch_assay_title": "**Assay Assignment**",
            "batch_target_label": "Target gene assay",
            "batch_ref_label": "Reference gene assay(s) (partitions pooled if multiple selected)",
            "batch_settings_title": "**Screening Settings**",
            "batch_expected_ratio": "Expected Ratio (no-change reference)",
            "batch_ploidy": "Reference Ploidy",
            "batch_run_btn": "▶ Run Screening",
            "batch_results_title": "📋 Screening Results",
            "batch_col_sample": "Sample",
            "batch_col_lambda_t": "λ Target",
            "batch_col_lambda_r": "λ Reference",
            "batch_col_ratio": "Ratio",
            "batch_col_ci": "95% CI (Ratio)",
            "batch_col_cn": "Copy Number (CN)",
            "batch_col_conc": "Concentration (copies/µL)",
            "batch_col_class": "Classification",
            "batch_class_gain": "📈 Gain",
            "batch_class_loss": "📉 Loss",
            "batch_class_normal": "➖ Normal",
            "batch_chart_title": "Copy Number Screening Across Samples",
            "batch_download_btn": "📥 Download Screening Results (CSV)",
            "batch_no_data": "No results yet — upload a file and run the screening first.",
            "batch_n_flagged": "{n} / {total} samples are statistically significantly different from the expected ratio (95% CI does not include the no-change value).",
            "batch_multi_gene_help": "If you select multiple target genes, a heatmap across all genes × samples is automatically generated in the results section.",
            "batch_heatmap_title": "🌡️ Multi-Gene Heatmap",
            "batch_heatmap_description": "Shows Copy Number or Ratio values for all selected target genes across samples in a single grid — helps quickly spot patterns across a cohort (e.g. multi-gene gain in specific samples).",
            "batch_heatmap_metric_label": "Metric to display",
            "tab_vaf": "VAF Calculator",
            "vaf_title": "🧬 VAF / Mutation Fraction Calculator",
            "vaf_description": "Designed for liquid biopsy / ctDNA monitoring. Upload a CSV containing Mutant and Wild-type assays; the Fractional Abundance (FA%, variant allele fraction) with 95% CI is calculated for each sample.",
            "vaf_uploader": "Upload VAF CSV file",
            "vaf_assay_title": "**Assay Assignment**",
            "vaf_mutant_label": "Mutant assay",
            "vaf_wt_label": "Wild-type assay",
            "vaf_run_btn": "▶ Calculate VAF",
            "vaf_results_title": "📋 VAF Results",
            "vaf_col_sample": "Sample",
            "vaf_col_lambda_mut": "λ Mutant",
            "vaf_col_lambda_wt": "λ Wild-type",
            "vaf_col_fa": "FA (%)",
            "vaf_col_ci": "95% CI (FA%)",
            "vaf_col_conc_mut": "Mutant Concentration (copies/µL)",
            "vaf_col_detected": "Detection Status",
            "vaf_detected_yes": "✅ Detected",
            "vaf_detected_no": "❌ Not Detected",
            "vaf_chart_title": "Fractional Abundance (VAF%) Across Samples",
            "vaf_download_btn": "📥 Download VAF Results (CSV)",
            "vaf_no_data": "No results yet — upload a file and run the calculation first.",
            "vaf_method_note": "ℹ️ FA = λ(mutant) / (λ(mutant) + λ(wild-type)). 95% CI calculated via the delta method (Hindson et al. 2013, Anal Chem). This assumes mutant and wild-type assays were measured on the same partition set (multiplex) or under equivalent conditions.",
            "vaf_n_detected": "{n} / {total} samples had the mutant allele detected (≥1 positive partition).",
        }
    }
    _t = translations[language_code]

    # ═══════════════════════════════════════════════════════════════════════════════
    # PDF FONT SYSTEM
    # ═══════════════════════════════════════════════════════════════════════════════
    def _find_font(candidates):
        import glob as _glob
        for p in candidates:
            if os.path.exists(p):
                return p
        all_ttf = _glob.glob('/usr/share/fonts/**/*.ttf', recursive=True)
        return all_ttf[0] if all_ttf else None

    _NOTO_REGULAR = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSans.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ])
    _NOTO_BOLD = _find_font([
        '/usr/share/fonts/truetype/noto/NotoSans-Bold.ttf',
        '/usr/share/fonts/truetype/freefont/FreeSansBold.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    ])
    try:
        pdfmetrics.registerFont(TTFont('NotoSans', _NOTO_REGULAR))
        pdfmetrics.registerFont(TTFont('NotoSans-Bold', _NOTO_BOLD))
        PDF_FONT, PDF_FONT_BOLD = 'NotoSans', 'NotoSans-Bold'
    except Exception:
        PDF_FONT, PDF_FONT_BOLD = 'Helvetica', 'Helvetica-Bold'

    def safe_str(text):
        if not isinstance(text, str):
            text = str(text)
        return text.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # ═══════════════════════════════════════════════════════════════════════════════
    # CORE dPCR MATH — Poisson statistics
    # ═══════════════════════════════════════════════════════════════════════════════
    def poisson_lambda(positive, total, alpha=0.05):
        """
        Returns (lambda, ci_low, ci_high, status) for a single partition (droplet/well) count.
        lambda = copies per partition, derived from the Poisson distribution.
        status: 'ok', 'saturated' (all partitions positive), or 'invalid'.

        The 95% CI is computed via the EXACT (Clopper-Pearson) binomial
        confidence interval on p = positive/total, transformed to lambda-space
        via lambda = -ln(1-p). This is more rigorous than a normal/delta-method
        approximation, which becomes miscalibrated at low positive counts or
        when p is close to 0 or 1 — precisely the regime relevant near an
        assay's limit of detection. Clopper-Pearson bounds are obtained from
        the Beta distribution quantiles (the standard exact-CI construction for
        a binomial proportion; see e.g. Clopper & Pearson, Biometrika 1934).
        """
        if total is None or total <= 0 or positive is None or positive < 0 or positive > total:
            return None, None, None, "invalid"
        p = positive / total
        if p >= 1.0:
            return None, None, None, "saturated"
        lam = -np.log(1 - p)

        x, n = positive, total
        p_low = 0.0 if x == 0 else stats.beta.ppf(alpha / 2, x, n - x + 1)
        p_high = 1.0 if x == n else stats.beta.ppf(1 - alpha / 2, x + 1, n - x)
        ci_low = -np.log(1 - p_low) if p_low < 1.0 else 0.0
        ci_high = -np.log(1 - p_high) if p_high < 1.0 else np.inf
        ci_low = max(ci_low, 0.0)
        return lam, ci_low, ci_high, "ok"

    def poisson_se(positive, total):
        """
        Standard-error-like quantity for lambda, used throughout the app for
        error propagation into derived quantities (ratios, fractional
        abundance, etc.) that need an SE rather than a full confidence
        interval. Back-derived from the exact Clopper-Pearson CI half-width
        (se = (ci_high - ci_low) / (2*z)), which is better calibrated than the
        raw normal-approximation formula at low counts, while still providing
        a simple SE-like value for the standard delta-method propagation
        formulas used elsewhere in this app. Falls back to the normal
        approximation sqrt(p/(n*(1-p))) if the exact CI is degenerate
        (e.g. at very small or very large p, where the Beta-based bounds can
        become numerically unstable).
        """
        lam, ci_low, ci_high, status = poisson_lambda(positive, total)
        if status != "ok" or ci_low is None or ci_high is None or np.isinf(ci_high):
            p = positive / total if total else np.nan
            return np.sqrt(p / (total * (1 - p))) if (total and (1 - p) > 0) else np.nan
        z = stats.norm.ppf(0.975)
        return (ci_high - ci_low) / (2 * z)

    def geometric_mean(values):
        values = np.array(values, dtype=float)
        values = values[values > 0]
        if len(values) == 0:
            return np.nan
        return np.exp(np.mean(np.log(values)))

    def mean_ci(values):
        """
        Mean, 95% CI (t-distribution, based on replicate-to-replicate variability),
        and CV% for a set of replicate values (e.g. concentration or ratio).
        Returns (mean, ci_low, ci_high, cv_pct, n).
        """
        values = np.array(values, dtype=float)
        values = values[~np.isnan(values)]
        n = len(values)
        if n == 0:
            return np.nan, np.nan, np.nan, np.nan, 0
        m = float(np.mean(values))
        if n < 2:
            return m, np.nan, np.nan, np.nan, n
        sd = float(np.std(values, ddof=1))
        sem = sd / np.sqrt(n)
        t_crit = stats.t.ppf(0.975, df=n - 1)
        ci_low, ci_high = m - t_crit * sem, m + t_crit * sem
        cv_pct = (sd / m * 100) if m != 0 else np.nan
        return m, ci_low, ci_high, cv_pct, n

    def lambda_to_conc(lam_array, partition_vol_nl_local):
        """Convert lambda (copies/partition) to concentration in copies/µL."""
        lam_array = np.array(lam_array, dtype=float)
        return lam_array / partition_vol_nl_local * 1000.0

    def poisson_ratio_ci(lam_t, se_t, lam_r, se_r):
        """
        95% CI for the ratio R = lambda_t / lambda_r of two independent
        Poisson-derived rates, using the log-scale delta method:
            Var(ln R) ~= (SE_t/lambda_t)^2 + (SE_r/lambda_r)^2
        This avoids negative CI bounds (unlike a linear delta method) and is the
        standard approach for ratios of counting-statistics-derived quantities.
        Returns (ratio, ci_low, ci_high).
        """
        if lam_t is None or lam_r is None or lam_r == 0 or np.isnan(lam_t) or np.isnan(lam_r):
            return np.nan, np.nan, np.nan
        ratio = lam_t / lam_r
        if se_t is None or se_r is None or np.isnan(se_t) or np.isnan(se_r) or lam_t == 0:
            return ratio, np.nan, np.nan
        var_ln_r = (se_t / lam_t) ** 2 + (se_r / lam_r) ** 2
        se_ln_r = np.sqrt(var_ln_r)
        ci_low = ratio * np.exp(-1.96 * se_ln_r)
        ci_high = ratio * np.exp(1.96 * se_ln_r)
        return ratio, ci_low, ci_high

    def pool_and_compute_batch(std_df, target_assay, ref_assays, partition_vol_nl_local):
        """
        For each unique Sample in std_df, pools (sums) positive/total partition
        counts across replicate rows for the target assay and for the combined
        set of reference assays, then computes lambda/SE for each and the
        Ratio with its 95% CI via poisson_ratio_ci.
        Returns a list of per-sample result dicts.
        """
        results = []
        samples = sorted(std_df["Sample"].unique())
        for sample in samples:
            tgt_rows = std_df[(std_df["Sample"] == sample) & (std_df["Target"] == target_assay)]
            ref_rows = std_df[(std_df["Sample"] == sample) & (std_df["Target"].isin(ref_assays))]
            if tgt_rows.empty or ref_rows.empty:
                continue

            pos_t, tot_t = float(tgt_rows["Positives"].sum()), float(tgt_rows["Total"].sum())
            pos_r, tot_r = float(ref_rows["Positives"].sum()), float(ref_rows["Total"].sum())

            lam_t, _, _, status_t = poisson_lambda(pos_t, tot_t)
            lam_r, _, _, status_r = poisson_lambda(pos_r, tot_r)
            if status_t != "ok" or status_r != "ok":
                continue

            p_t, p_r = pos_t / tot_t, pos_r / tot_r
            se_t = poisson_se(pos_t, tot_t)
            se_r = poisson_se(pos_r, tot_r)

            ratio, ci_low, ci_high = poisson_ratio_ci(lam_t, se_t, lam_r, se_r)
            conc = lam_t / partition_vol_nl_local * 1000.0

            results.append({
                "Sample": sample, "lambda_t": lam_t, "lambda_r": lam_r,
                "ratio": ratio, "ci_low": ci_low, "ci_high": ci_high, "conc": conc,
                "n_partitions_t": int(tot_t), "n_partitions_r": int(tot_r),
            })
        return results

    def compute_vaf(lam_mut, se_mut, lam_wt, se_wt):
        """
        Fractional Abundance (FA, i.e. variant allele fraction) for a mutant vs
        wild-type ddPCR assay pair: FA = lambda_mut / (lambda_mut + lambda_wt).
        95% CI via the linear delta method on the ratio-to-sum form, which is
        the standard approach used in ddPCR liquid-biopsy literature
        (e.g. Hindson et al. 2013 Anal Chem):
            Var(FA) ~= [y/(x+y)^2]^2 * Var(x) + [x/(x+y)^2]^2 * Var(y)
        Returns (FA, ci_low, ci_high) as fractions in [0, 1].
        """
        x, y = lam_mut, lam_wt
        if x is None or y is None or np.isnan(x) or np.isnan(y) or (x + y) <= 0:
            return np.nan, np.nan, np.nan
        fa = x / (x + y)
        if se_mut is None or se_wt is None or np.isnan(se_mut) or np.isnan(se_wt):
            return fa, np.nan, np.nan
        denom_sq = (x + y) ** 4
        var_fa = ((y ** 2) * (se_mut ** 2) + (x ** 2) * (se_wt ** 2)) / denom_sq
        se_fa = np.sqrt(var_fa)
        ci_low = max(fa - 1.96 * se_fa, 0.0)
        ci_high = min(fa + 1.96 * se_fa, 1.0)
        return fa, ci_low, ci_high

    def pool_and_compute_vaf(std_df, mutant_assay, wt_assay, partition_vol_nl_local):
        """
        For each unique Sample in std_df, pools positive/total partition counts
        for the mutant assay and the wild-type assay, computes lambda/SE for
        each, and the Fractional Abundance (FA%) with 95% CI via compute_vaf.
        Returns a list of per-sample result dicts.
        """
        results = []
        samples = sorted(std_df["Sample"].unique())
        for sample in samples:
            mut_rows = std_df[(std_df["Sample"] == sample) & (std_df["Target"] == mutant_assay)]
            wt_rows = std_df[(std_df["Sample"] == sample) & (std_df["Target"] == wt_assay)]
            if mut_rows.empty or wt_rows.empty:
                continue

            pos_m, tot_m = float(mut_rows["Positives"].sum()), float(mut_rows["Total"].sum())
            pos_w, tot_w = float(wt_rows["Positives"].sum()), float(wt_rows["Total"].sum())

            lam_m, _, _, status_m = poisson_lambda(pos_m, tot_m)
            lam_w, _, _, status_w = poisson_lambda(pos_w, tot_w)
            if status_m not in ("ok", "saturated") or status_w not in ("ok", "saturated"):
                continue
            if status_m == "saturated":
                lam_m = -np.log(1.0 / tot_m)  # fallback, extreme edge case
            if status_w == "saturated":
                lam_w = -np.log(1.0 / tot_w)
            lam_m = 0.0 if pos_m == 0 else lam_m
            lam_w = 0.0 if pos_w == 0 else lam_w

            p_m = pos_m / tot_m if tot_m > 0 else np.nan
            p_w = pos_w / tot_w if tot_w > 0 else np.nan
            se_m = poisson_se(pos_m, tot_m)
            se_w = poisson_se(pos_w, tot_w)

            fa, ci_low, ci_high = compute_vaf(lam_m, se_m, lam_w, se_w)
            conc_mut = lam_m / partition_vol_nl_local * 1000.0
            detected = pos_m >= 1

            results.append({
                "Sample": sample, "lambda_mut": lam_m, "lambda_wt": lam_w,
                "fa": fa, "ci_low": ci_low, "ci_high": ci_high,
                "conc_mut": conc_mut, "detected": detected,
                "n_partitions_mut": int(tot_m), "n_partitions_wt": int(tot_w),
                "pos_mut": int(pos_m),
            })
        return results

    _PROJECT_SCALAR_KEYS = [
        "gene_count", "patient_count", "num_ref_genes", "ploidy", "partition_vol",
        "qc_min", "outlier_enabled", "outlier_method", "grubbs_alpha", "iqr_mult",
        "advanced_mode",
    ]
    _PROJECT_TEXT_PREFIXES = (
        "ctrl_tgt_pos_", "ctrl_tgt_tot_", "ctrl_ref_pos_", "ctrl_ref_tot_",
        "smp_tgt_pos_", "smp_tgt_tot_", "smp_ref_pos_", "smp_ref_tot_",
        "ntc_pos_", "ntc_tot_",
    )

    def export_project_state():
        """Serializes all data-entry-relevant session_state keys into a JSON-safe dict."""
        project = {"_absolutegene_project_version": 1, "_absolutegene_app_version": APP_VERSION}
        for k in _PROJECT_SCALAR_KEYS:
            if k in st.session_state:
                v = st.session_state[k]
                project[k] = v if isinstance(v, (int, float, str, bool)) else str(v)
        for k in list(st.session_state.keys()):
            if isinstance(k, str) and k.startswith(_PROJECT_TEXT_PREFIXES):
                v = st.session_state[k]
                if isinstance(v, str):
                    project[k] = v
        return project

    def import_project_state(project_dict):
        """
        Restores session_state from a previously exported project dict. Returns
        count of keys restored.

        Special case: "advanced_mode" controls a widget (the sidebar toggle)
        that is instantiated EARLIER in the script than this import logic runs.
        Streamlit does not allow writing to st.session_state for a key whose
        widget has already been instantiated in the current run (it raises an
        exception), so we can't set it directly here. Instead we stash it in a
        "_pending_advanced_mode" key, which is picked up and applied at the top
        of the script (before the toggle widget is instantiated) on the next
        rerun.
        """
        if not isinstance(project_dict, dict):
            return 0
        count = 0
        for k, v in project_dict.items():
            if k == "_absolutegene_project_version":
                continue
            if k == "advanced_mode":
                st.session_state["_pending_advanced_mode"] = v
                count += 1
                continue
            if k in _PROJECT_SCALAR_KEYS or (isinstance(k, str) and k.startswith(_PROJECT_TEXT_PREFIXES)):
                st.session_state[k] = v
                count += 1
        return count

    def compute_effective_dilution(reaction_volume_ul, template_volume_ul, pre_dilution_factor=1.0):
        """
        Computes the effective dilution factor between the dPCR reaction and the
        original stock sample, from volumes a user actually measures at the
        bench (reaction volume, template volume added, and any pre-dilution
        performed before adding the template to the reaction).

        When V_template µL of (possibly pre-diluted) sample is added to a total
        reaction volume V_rxn, the sample is diluted by V_rxn/V_template within
        the reaction. If the added material was itself pre-diluted by
        pre_dilution_factor from the original stock, the total dilution factor
        (to back-calculate the original stock concentration from the measured
        reaction concentration) is:
            total_dilution = (V_rxn / V_template) * pre_dilution_factor
        """
        if not reaction_volume_ul or not template_volume_ul or template_volume_ul <= 0:
            return 1.0
        return (reaction_volume_ul / template_volume_ul) * (pre_dilution_factor or 1.0)

    def render_dilution_input(key_prefix, expanded=False):
        """
        Renders a user-friendly dilution-factor input, reused across Data Entry,
        Batch Screening, and the VAF Calculator. Defaults to the more intuitive
        volume-based mode (reaction volume + template volume added, with an
        optional pre-dilution factor), since most users know these bench values
        more readily than an already-computed dilution factor. A manual-entry
        fallback is available for users who already know their exact factor.
        Returns the effective dilution factor (float, >=1.0).
        """
        with st.expander(_t['dilution_settings_expander'], expanded=expanded):
            _mode = st.radio(
                _t['dilution_mode_selector_label'],
                options=[_t['dilution_mode_volumes'], _t['dilution_mode_manual']],
                key=f"{key_prefix}_dilmode", horizontal=True
            )
            if _mode == _t['dilution_mode_volumes']:
                vc1, vc2, vc3 = st.columns(3)
                with vc1:
                    _rxn_vol = st.number_input(_t['dilution_rxn_vol_label'], min_value=0.1, value=20.0,
                                                step=0.1, key=f"{key_prefix}_rxnvol", help=_t['dilution_rxn_vol_help'])
                with vc2:
                    _tmpl_vol = st.number_input(_t['dilution_tmpl_vol_label'], min_value=0.01, value=2.0,
                                                 step=0.1, key=f"{key_prefix}_tmplvol", help=_t['dilution_tmpl_vol_help'])
                with vc3:
                    _pre_dil = st.number_input(_t['dilution_pre_dilution_label'], min_value=1.0, value=1.0,
                                                step=1.0, key=f"{key_prefix}_predil", help=_t['dilution_pre_dilution_help'])
                _factor = compute_effective_dilution(_rxn_vol, _tmpl_vol, _pre_dil)
                _ratio = _rxn_vol / _tmpl_vol if _tmpl_vol else 1.0
                _pre_txt = (_t['dilution_computed_factor_pre_txt'].format(pre=_pre_dil)
                            if _pre_dil and _pre_dil != 1.0 else _t['dilution_no_pre'])
                st.caption(_t['dilution_computed_factor'].format(factor=_factor, rxn=_rxn_vol, tmpl=_tmpl_vol,
                                                                   ratio=_ratio, pre_txt=_pre_txt))
                return _factor
            else:
                return st.number_input(_t['dilution_factor_field_label'], min_value=1.0, value=1.0, step=1.0,
                                        key=f"{key_prefix}_manualval", help=_t['dilution_factor_field_help'])

    def recommend_dilution(current_lambda, target_lambda=1.6):
        """
        Given the current observed lambda (copies/partition), recommends a
        dilution adjustment factor to bring the assay into the optimal dynamic
        range for Poisson-based quantification.

        Ideal operating range: lambda ~0.05-3.0 (acceptable), ~1.0-2.0 (optimal
        precision/dynamic-range balance; default target 1.6, near the point of
        minimal relative CI width for Poisson-derived concentration estimates).

        Returns dict: status ('too_low'/'optimal'/'too_high'/'saturated_risk'),
        dilution_factor (multiply current dilution by this value to reach
        target_lambda; >1 means dilute further, <1 means concentrate/dilute less),
        message_key (translation key hint for status).
        """
        if current_lambda is None or np.isnan(current_lambda) or current_lambda <= 0:
            return None
        dilution_factor = current_lambda / target_lambda
        if current_lambda < 0.05:
            status = "too_low"
        elif current_lambda > 4.0:
            status = "saturated_risk"
        elif current_lambda > 3.0:
            status = "too_high"
        else:
            status = "optimal"
        return {
            "current_lambda": current_lambda, "target_lambda": target_lambda,
            "dilution_factor": dilution_factor, "status": status,
        }

    # ═══════════════════════════════════════════════════════════════════════════════
    # CLINICAL TOOLS — Measurement Uncertainty, RCV, Precision Study, Method Comparison
    # ═══════════════════════════════════════════════════════════════════════════════
    def compute_mu_budget(u_poisson_pct, u_pipetting_pct, u_precision_pct, k=2.0):
        """
        Combined and expanded measurement uncertainty following the GUM
        (Guide to the Expression of Uncertainty in Measurement) approach:
        relative standard uncertainties from independent sources are combined
        in quadrature, then multiplied by a coverage factor k (k=2 for an
        approximate 95% confidence level, assuming a normal distribution).
            u_c(%) = sqrt(u_poisson^2 + u_pipetting^2 + u_precision^2)
            U(%) = k * u_c
        Returns dict with u_c_pct, U_pct, and the individual contributions
        (useful for a budget table showing each source's relative share).
        """
        components = {
            "Poisson counting statistics": u_poisson_pct or 0.0,
            "Pipetting / dilution": u_pipetting_pct or 0.0,
            "Inter-run precision": u_precision_pct or 0.0,
        }
        u_c_pct = np.sqrt(sum(v ** 2 for v in components.values()))
        U_pct = k * u_c_pct
        return {"components": components, "u_c_pct": u_c_pct, "U_pct": U_pct, "k": k}

    def compute_rcv(result1, result2, cv_analytical_pct, cv_biological_pct, z=1.96):
        """
        Reference Change Value (RCV) — determines whether the difference between
        two serial results exceeds what would be expected from analytical and
        within-subject biological variation alone (Fraser CG, Harris EK.
        Generation and application of data on biological variation in clinical
        chemistry. Crit Rev Clin Lab Sci 1989).
            RCV(%) = sqrt(2) * z * sqrt(CV_analytical^2 + CV_biological^2)
        A percent change between result1 and result2 exceeding RCV% is
        considered a statistically significant change beyond combined
        analytical + biological noise.
        Returns dict with rcv_pct, percent_change, significant (bool).
        """
        if result1 is None or result2 is None or result1 == 0:
            return None
        rcv_pct = np.sqrt(2) * z * np.sqrt(cv_analytical_pct ** 2 + cv_biological_pct ** 2)
        percent_change = (result2 - result1) / abs(result1) * 100.0
        significant = abs(percent_change) > rcv_pct
        return {"rcv_pct": rcv_pct, "percent_change": percent_change, "significant": significant}

    def compute_precision_study(day_groups):
        """
        Nested one-way random-effects ANOVA (day/run as the grouping factor)
        for a basic precision (repeatability / intermediate precision) study,
        following the general approach of CLSI EP05-A3.

        Correctly handles UNBALANCED designs (unequal replicate counts per
        day) using the standard method-of-moments variance-component
        correction term n0 for unbalanced one-way random-effects models
        (Searle, Casella & McCulloch, "Variance Components", 1992):
            n0 = (N - sum(n_i^2)/N) / (k-1)
            sigma_between^2 = max((MS_between - MS_within) / n0, 0)
        For a balanced design (all n_i equal), n0 reduces exactly to n_i and
        this is equivalent to the simple balanced-ANOVA formula — no data is
        discarded in either case.

        day_groups: list of arrays, one array of replicate values per day/run.

        Returns dict with repeatability_cv, between_day_cv, total_cv (all %),
        grand_mean, the underlying ANOVA mean squares, and is_balanced (bool)
        plus n_per_day (list) so the caller can inform the user when the
        design was unbalanced.
        """
        day_groups = [np.array(g, dtype=float) for g in day_groups if len(g) > 0]
        k = len(day_groups)
        if k < 2:
            return None
        n_per_day = np.array([len(g) for g in day_groups])
        if np.any(n_per_day < 2):
            return None
        N = int(np.sum(n_per_day))
        is_balanced = len(set(n_per_day.tolist())) == 1

        all_values = np.concatenate(day_groups)
        grand_mean = np.mean(all_values)
        day_means = np.array([np.mean(g) for g in day_groups])

        ss_within = sum(np.sum((g - np.mean(g)) ** 2) for g in day_groups)
        df_within = N - k
        ms_within = ss_within / df_within if df_within > 0 else np.nan

        ss_between = np.sum(n_per_day * (day_means - grand_mean) ** 2)
        df_between = k - 1
        ms_between = ss_between / df_between if df_between > 0 else np.nan

        n0 = (N - np.sum(n_per_day ** 2) / N) / df_between if df_between > 0 else np.nan

        s_between_day_sq = max((ms_between - ms_within) / n0, 0.0) if (not np.isnan(ms_between) and n0) else 0.0
        s_repeatability = np.sqrt(ms_within) if ms_within >= 0 else np.nan
        s_total = np.sqrt(ms_within + s_between_day_sq)

        return {
            "grand_mean": grand_mean, "k_days": k, "n_per_day": n_per_day.tolist(),
            "is_balanced": is_balanced, "n0": n0,
            "ms_within": ms_within, "ms_between": ms_between,
            "repeatability_cv": (s_repeatability / grand_mean * 100) if grand_mean != 0 else np.nan,
            "between_day_cv": (np.sqrt(s_between_day_sq) / grand_mean * 100) if grand_mean != 0 else np.nan,
            "total_cv": (s_total / grand_mean * 100) if grand_mean != 0 else np.nan,
        }

    def compute_bland_altman(method1_vals, method2_vals):
        """
        Bland-Altman agreement analysis for paired measurements from two
        methods (Bland JM, Altman DG. Statistical methods for assessing
        agreement between two methods of clinical measurement. Lancet 1986).

        In addition to the classic bias/limits-of-agreement, this also checks
        two common assumptions/pitfalls that are frequently overlooked:
          - Normality of the differences (Shapiro-Wilk, n>=3): the +-1.96 SD
            limits of agreement are only strictly valid if differences are
            approximately normally distributed.
          - Proportional bias: regresses the differences on the means; if the
            slope is significantly different from zero, the bias is not
            constant across the measurement range (a common real-world
            pattern, e.g. increasing divergence at higher concentrations),
            which the single mean-difference "bias" figure would otherwise
            hide.

        Returns dict with mean_diff (bias), sd_diff, loa_low, loa_high,
        means, diffs, shapiro_p (normality of diffs), proportional_bias_slope,
        proportional_bias_p.
        """
        m1 = np.array(method1_vals, dtype=float)
        m2 = np.array(method2_vals, dtype=float)
        n = min(len(m1), len(m2))
        m1, m2 = m1[:n], m2[:n]
        if n < 2:
            return None
        diffs = m2 - m1
        means = (m1 + m2) / 2
        mean_diff = float(np.mean(diffs))
        sd_diff = float(np.std(diffs, ddof=1))

        shapiro_p = np.nan
        if n >= 3:
            try:
                shapiro_p = float(stats.shapiro(diffs).pvalue)
            except Exception:
                shapiro_p = np.nan

        prop_slope, prop_p = np.nan, np.nan
        if n >= 3 and np.std(means) > 0:
            try:
                _res = stats.linregress(means, diffs)
                prop_slope, prop_p = float(_res.slope), float(_res.pvalue)
            except Exception:
                pass

        return {
            "means": means, "diffs": diffs, "mean_diff": mean_diff, "sd_diff": sd_diff,
            "loa_low": mean_diff - 1.96 * sd_diff, "loa_high": mean_diff + 1.96 * sd_diff,
            "n": n, "shapiro_p": shapiro_p,
            "proportional_bias_slope": prop_slope, "proportional_bias_p": prop_p,
        }

    def compute_deming_regression(method1_vals, method2_vals, variance_ratio=1.0, n_boot=1000, seed=42):
        """
        Deming regression: linear regression accounting for measurement error
        in both variables (unlike ordinary least squares, which assumes the
        x-variable is error-free). variance_ratio (lambda) = Var(error_y)/Var(error_x);
        lambda=1 assumes equal error variance in both methods (a common default
        when no independent estimate of the error ratio is available — note
        that this assumption is frequently WRONG for two genuinely different
        assay technologies, and should be set from prior precision data when
        possible).
        Reference: Linnet K. Estimation of the linear relationship between the
        measurements of two methods with proportional errors. Stat Med 1990.

        95% confidence intervals for slope and intercept are obtained via
        non-parametric bootstrap resampling (percentile method), since no
        simple closed-form CI exists for the Deming estimator under arbitrary
        variance-ratio assumptions.

        Returns dict with slope, intercept, slope_ci_low/high, intercept_ci_low/high.
        """
        x = np.array(method1_vals, dtype=float)
        y = np.array(method2_vals, dtype=float)
        n = min(len(x), len(y))
        x, y = x[:n], y[:n]
        if n < 3:
            return None

        def _fit(xx, yy, lam):
            mx, my = np.mean(xx), np.mean(yy)
            sxx = np.sum((xx - mx) ** 2) / (len(xx) - 1)
            syy = np.sum((yy - my) ** 2) / (len(yy) - 1)
            sxy = np.sum((xx - mx) * (yy - my)) / (len(xx) - 1)
            if sxy == 0:
                return np.nan, np.nan
            s = (syy - lam * sxx + np.sqrt((syy - lam * sxx) ** 2 + 4 * lam * sxy ** 2)) / (2 * sxy)
            b = my - s * mx
            return s, b

        slope, intercept = _fit(x, y, variance_ratio)

        rng = np.random.default_rng(seed)
        boot_slopes, boot_intercepts = [], []
        for _ in range(n_boot):
            idx = rng.integers(0, n, n)
            s_b, i_b = _fit(x[idx], y[idx], variance_ratio)
            if not np.isnan(s_b):
                boot_slopes.append(s_b)
                boot_intercepts.append(i_b)

        if len(boot_slopes) >= 20:
            slope_ci_low, slope_ci_high = np.percentile(boot_slopes, [2.5, 97.5])
            intercept_ci_low, intercept_ci_high = np.percentile(boot_intercepts, [2.5, 97.5])
        else:
            slope_ci_low = slope_ci_high = intercept_ci_low = intercept_ci_high = np.nan

        return {
            "slope": slope, "intercept": intercept, "n": n,
            "slope_ci_low": slope_ci_low, "slope_ci_high": slope_ci_high,
            "intercept_ci_low": intercept_ci_low, "intercept_ci_high": intercept_ci_high,
        }

    def compute_passing_bablok(method1_vals, method2_vals, n_boot=1000, seed=42):
        """
        Passing-Bablok regression: a non-parametric method-comparison
        regression that, unlike Deming regression, does not require an
        assumption about the ratio of measurement error variances between the
        two methods, and is robust to outliers and non-normally distributed
        data (Passing H, Bablok W. A new biometrical procedure for testing the
        equality of measurements from two different analytical methods.
        J Clin Chem Clin Biochem 1983;21:709-720). This is the method
        generally recommended by CLSI EP09 for method-comparison studies where
        the error-variance ratio is unknown, making it a useful complement (or
        alternative) to Deming regression in this app.

        95% confidence intervals for slope and intercept are obtained via
        non-parametric bootstrap resampling (percentile method), matching the
        approach used for the Deming regression CI elsewhere in this app —
        the original Passing-Bablok paper describes an analytical CI based on
        the sorted pairwise-slope distribution, but the bootstrap approach is
        simpler to implement correctly and gives comparable coverage in
        practice for the sample sizes typical of method-comparison studies.

        Returns dict with slope, intercept, n_pairs, slope_ci_low/high,
        intercept_ci_low/high.
        """
        x = np.asarray(method1_vals, dtype=float)
        y = np.asarray(method2_vals, dtype=float)
        n = min(len(x), len(y))
        x, y = x[:n], y[:n]
        if n < 3:
            return None

        def _fit_pb(xx, yy):
            nn = len(xx)
            slopes = []
            for i in range(nn - 1):
                for j in range(i + 1, nn):
                    dx = xx[j] - xx[i]
                    if dx != 0:
                        s = (yy[j] - yy[i]) / dx
                        if s != -1:
                            slopes.append(s)
            if not slopes:
                return np.nan, np.nan, 0
            slopes_sorted = np.sort(np.array(slopes))
            N = len(slopes_sorted)
            K = int(np.sum(slopes_sorted < -1))
            if N % 2 == 0:
                idx1 = min(max(N // 2 + K, 1), N)
                idx2 = min(max(idx1 + 1, 1), N)
                s_est = 0.5 * (slopes_sorted[idx1 - 1] + slopes_sorted[idx2 - 1])
            else:
                idx = min(max((N + 1) // 2 + K, 1), N)
                s_est = slopes_sorted[idx - 1]
            i_est = float(np.median(yy - s_est * xx))
            return float(s_est), i_est, N

        slope, intercept, n_pairs = _fit_pb(x, y)
        if np.isnan(slope):
            return None

        rng = np.random.default_rng(seed)
        boot_slopes, boot_intercepts = [], []
        for _ in range(n_boot):
            idx = rng.integers(0, n, n)
            s_b, i_b, _ = _fit_pb(x[idx], y[idx])
            if not np.isnan(s_b):
                boot_slopes.append(s_b)
                boot_intercepts.append(i_b)

        if len(boot_slopes) >= 20:
            slope_ci_low, slope_ci_high = np.percentile(boot_slopes, [2.5, 97.5])
            intercept_ci_low, intercept_ci_high = np.percentile(boot_intercepts, [2.5, 97.5])
        else:
            slope_ci_low = slope_ci_high = intercept_ci_low = intercept_ci_high = np.nan

        return {
            "slope": slope, "intercept": intercept, "n_pairs": n_pairs,
            "slope_ci_low": slope_ci_low, "slope_ci_high": slope_ci_high,
            "intercept_ci_low": intercept_ci_low, "intercept_ci_high": intercept_ci_high,
        }

    # ═══════════════════════════════════════════════════════════════════════════════
    # CRM PRODUCTION — Homogeneity, Stability, Assigned Value/Uncertainty Budget (per
    # ISO Guide 35 / ISO 17034 general principles), and Certificate generation.
    # ═══════════════════════════════════════════════════════════════════════════════
    def compute_homogeneity(unit_groups):
        """
        One-way ANOVA-based between-unit homogeneity test, following the general
        approach described in ISO Guide 35 and Linsinger TP, Pauwels J, van der
        Veen AMH, Schimmel H, Lamberty A. "Homogeneity and stability of reference
        materials." Accred Qual Assur 2001;6:20-25.

        Correctly handles UNBALANCED designs (unequal replicate counts per
        unit/vial): the F-test itself generalizes naturally to unequal n_i
        using properly weighted sums of squares, and the between-unit
        variance-component estimate (s_bb^2 / u_bb) uses the standard
        method-of-moments correction term n0 for unbalanced one-way
        random-effects models (Searle, Casella & McCulloch, 1992), so no
        replicate data is discarded.

        unit_groups: list of arrays, one array of replicate measurements per
        CRM unit/vial.

        Returns dict with ms_within, ms_between, F, F_crit, p_value,
        is_homogeneous (bool, F <= F_crit at alpha=0.05), s_bb (between-unit
        standard deviation estimate), u_bb (standard uncertainty contribution
        from potential inhomogeneity), grand_mean, is_balanced, n_per_unit.
        """
        unit_groups = [np.array(g, dtype=float) for g in unit_groups if len(g) > 0]
        p = len(unit_groups)
        if p < 2:
            return None
        n_per_unit = np.array([len(g) for g in unit_groups])
        if np.any(n_per_unit < 2):
            return None
        N = int(np.sum(n_per_unit))
        is_balanced = len(set(n_per_unit.tolist())) == 1

        all_values = np.concatenate(unit_groups)
        grand_mean = float(np.mean(all_values))
        unit_means = np.array([np.mean(g) for g in unit_groups])

        ss_within = sum(np.sum((g - np.mean(g)) ** 2) for g in unit_groups)
        df_within = N - p
        ms_within = ss_within / df_within if df_within > 0 else np.nan

        ss_between = np.sum(n_per_unit * (unit_means - grand_mean) ** 2)
        df_between = p - 1
        ms_between = ss_between / df_between if df_between > 0 else np.nan

        F = ms_between / ms_within if (ms_within and ms_within > 0) else np.nan
        F_crit = stats.f.ppf(0.95, df_between, df_within) if df_within > 0 else np.nan
        p_value = stats.f.sf(F, df_between, df_within) if not np.isnan(F) else np.nan
        is_homogeneous = (F <= F_crit) if not np.isnan(F) else True

        n0 = (N - np.sum(n_per_unit ** 2) / N) / df_between if df_between > 0 else np.nan

        if ms_between > ms_within and n0:
            s_bb_sq = (ms_between - ms_within) / n0
            u_bb = np.sqrt(s_bb_sq)
        else:
            # Conservative minimum-uncertainty estimate when between-unit
            # variance is not resolvable from within-unit noise (Linsinger et al. 2001).
            # Uses the harmonic-mean replicate count for the unbalanced case.
            _n_harmonic = p / np.sum(1.0 / n_per_unit)
            u_bb = np.sqrt(ms_within / _n_harmonic) * (2.0 / df_between) ** 0.25 if df_between > 0 else np.nan
            s_bb_sq = 0.0

        return {
            "grand_mean": grand_mean, "p_units": p, "n_per_unit": n_per_unit.tolist(),
            "is_balanced": is_balanced,
            "ms_within": ms_within, "ms_between": ms_between, "F": F, "F_crit": F_crit,
            "p_value": p_value, "is_homogeneous": is_homogeneous,
            "s_bb": np.sqrt(s_bb_sq), "u_bb": u_bb,
            "df_within": df_within, "df_between": df_between,
        }

    def compute_stability(time_points, values, study_duration=None):
        """
        Linear-regression-based stability assessment: tests whether the analyte
        concentration/ratio shows a statistically significant trend over time
        (ISO Guide 35 classical stability study approach). If a significant
        trend is found, this may indicate degradation and should inform the
        material's shelf-life claim.

        time_points, values: paired arrays (e.g. storage duration in days/months
        and the corresponding measured value).
        study_duration: the shelf-life/study period of interest, used to compute
        the uncertainty contribution from potential instability
        (u_stab = SE(slope) * study_duration). If None, uses max(time_points).

        Returns dict with slope, intercept, se_slope, t_stat, p_value,
        is_stable (bool, p >= 0.05), u_stab.
        """
        x = np.array(time_points, dtype=float)
        y = np.array(values, dtype=float)
        n = min(len(x), len(y))
        x, y = x[:n], y[:n]
        if n < 3:
            return None

        slope, intercept, r_value, p_value, se_slope = stats.linregress(x, y)
        t_stat = slope / se_slope if se_slope > 0 else np.nan
        is_stable = p_value >= 0.05

        duration = study_duration if study_duration is not None else float(np.max(x))
        u_stab = abs(se_slope) * duration

        return {
            "slope": slope, "intercept": intercept, "se_slope": se_slope,
            "t_stat": t_stat, "p_value": p_value, "r_value": r_value,
            "is_stable": is_stable, "u_stab": u_stab, "study_duration": duration, "n": n,
        }

    def compute_assigned_value_uncertainty(assigned_value, u_char, u_bb, u_stab,
                                            extra_components=None, k=2.0):
        """
        Combines characterization uncertainty (u_char), between-unit
        inhomogeneity uncertainty (u_bb), stability uncertainty (u_stab), and
        any additional user-specified components into a combined standard
        uncertainty and expanded uncertainty, following the GUM approach
        (components combined in quadrature, assuming independence):
            u_c = sqrt(u_char^2 + u_bb^2 + u_stab^2 + sum(extra_i^2))
            U = k * u_c
        All uncertainty inputs are absolute (same units as assigned_value), not
        relative percentages, consistent with ISO Guide 35 uncertainty budgets
        for certified reference materials.
        Returns dict with components, u_c, U, U_rel_pct.
        """
        components = {"Characterization": u_char or 0.0, "Homogeneity (u_bb)": u_bb or 0.0,
                      "Stability (u_stab)": u_stab or 0.0}
        if extra_components:
            components.update(extra_components)
        u_c = np.sqrt(sum(v ** 2 for v in components.values()))
        U = k * u_c
        U_rel_pct = (U / assigned_value * 100) if assigned_value else np.nan
        return {"components": components, "u_c": u_c, "U": U, "U_rel_pct": U_rel_pct, "k": k,
                "assigned_value": assigned_value}

    def compute_lot_equivalence(lot1_values, lot2_values, equivalence_margin_pct, alpha=0.05):
        """
        Two One-Sided Tests (TOST) equivalence procedure, commonly used for
        batch/lot-to-lot equivalence assessment of reference materials and in
        bioequivalence studies (Schuirmann DJ. A comparison of the two
        one-sided tests procedure and the power approach for assessing the
        equivalence of average bioavailability. J Pharmacokinet Biopharm 1987).

        Two lots are declared equivalent if the (1-2*alpha)*100% confidence
        interval (typically 90% for alpha=0.05) of the mean difference lies
        entirely within +/- equivalence_margin_pct (relative to the mean of
        lot1, taken as the reference).

        Returns dict with mean1, mean2, percent_diff, ci_low_pct, ci_high_pct,
        margin_pct, is_equivalent.
        """
        l1 = np.array(lot1_values, dtype=float)
        l2 = np.array(lot2_values, dtype=float)
        n1, n2 = len(l1), len(l2)
        if n1 < 2 or n2 < 2:
            return None

        m1, m2 = np.mean(l1), np.mean(l2)
        s1, s2 = np.std(l1, ddof=1), np.std(l2, ddof=1)
        se_diff = np.sqrt(s1 ** 2 / n1 + s2 ** 2 / n2)
        df = (s1 ** 2 / n1 + s2 ** 2 / n2) ** 2 / (
            (s1 ** 2 / n1) ** 2 / (n1 - 1) + (s2 ** 2 / n2) ** 2 / (n2 - 1)
        )
        t_crit = stats.t.ppf(1 - alpha, df)

        mean_diff = m2 - m1
        ci_low = mean_diff - t_crit * se_diff
        ci_high = mean_diff + t_crit * se_diff

        # Express in percent relative to lot 1 (reference)
        pct_diff = mean_diff / m1 * 100 if m1 != 0 else np.nan
        ci_low_pct = ci_low / m1 * 100 if m1 != 0 else np.nan
        ci_high_pct = ci_high / m1 * 100 if m1 != 0 else np.nan

        is_equivalent = (ci_low_pct >= -equivalence_margin_pct) and (ci_high_pct <= equivalence_margin_pct)

        return {
            "mean1": m1, "mean2": m2, "percent_diff": pct_diff,
            "ci_low_pct": ci_low_pct, "ci_high_pct": ci_high_pct,
            "margin_pct": equivalence_margin_pct, "is_equivalent": is_equivalent,
            "confidence_pct": (1 - 2 * alpha) * 100,
        }

    def compute_multiplex_cluster(n11, n10, n01, n00):
        """
        Converts 2-color ddPCR multiplex cluster counts (e.g. FAM=target,
        HEX=reference, both measured on the SAME physical partitions) into
        marginal positive/total counts for target and reference, and computes
        the target/reference ratio using a covariance-aware delta method that
        properly accounts for the correlation induced by sharing the same
        partition set (rather than assuming independence, which would be
        incorrect for same-well multiplex data).

        n11 = double-positive (target+, ref+) partition count
        n10 = target-positive only
        n01 = reference-positive only
        n00 = double-negative

        Multinomial covariance derivation:
            Cov(p_target, p_ref) = (p11*p00 - p10*p01) / N
        propagated to lambda via the delta method (dlambda/dp = 1/(1-p)),
        then to the log-ratio via the standard multivariate delta method.

        Returns dict with marginal positive/total counts, lambda_target,
        lambda_ref, ratio, ci_low, ci_high (95%), and the covariance term.
        """
        N = n11 + n10 + n01 + n00
        if N <= 0:
            return None
        pos_target = n11 + n10
        pos_ref = n11 + n01

        lam_t, _, _, status_t = poisson_lambda(pos_target, N)
        lam_r, _, _, status_r = poisson_lambda(pos_ref, N)
        if status_t != "ok" or status_r != "ok":
            return None

        p_t, p_r = pos_target / N, pos_ref / N
        p11_hat, p00_hat, p10_hat, p01_hat = n11 / N, n00 / N, n10 / N, n01 / N

        var_p_t = p_t * (1 - p_t) / N
        var_p_r = p_r * (1 - p_r) / N
        cov_p = (p11_hat * p00_hat - p10_hat * p01_hat) / N

        var_lam_t = var_p_t / (1 - p_t) ** 2
        var_lam_r = var_p_r / (1 - p_r) ** 2
        cov_lam = cov_p / ((1 - p_t) * (1 - p_r))

        ratio = lam_t / lam_r
        var_ln_r = var_lam_t / lam_t ** 2 + var_lam_r / lam_r ** 2 - 2 * cov_lam / (lam_t * lam_r)
        var_ln_r = max(var_ln_r, 0.0)
        se_ln_r = np.sqrt(var_ln_r)
        ci_low = ratio * np.exp(-1.96 * se_ln_r)
        ci_high = ratio * np.exp(1.96 * se_ln_r)

        # For comparison: what the (incorrect) independence assumption would give
        se_t_indep = np.sqrt(var_p_t) / (1 - p_t)
        se_r_indep = np.sqrt(var_p_r) / (1 - p_r)
        _, ci_low_indep, ci_high_indep = poisson_ratio_ci(lam_t, se_t_indep, lam_r, se_r_indep)

        return {
            "n11": n11, "n10": n10, "n01": n01, "n00": n00, "N": N,
            "pos_target": pos_target, "pos_ref": pos_ref,
            "lambda_target": lam_t, "lambda_ref": lam_r,
            "ratio": ratio, "ci_low": ci_low, "ci_high": ci_high,
            "cov_lam": cov_lam, "ci_low_indep": ci_low_indep, "ci_high_indep": ci_high_indep,
        }

    def compute_lod_loq(ntc_pos, ntc_tot, partition_vol_nl_local):
        """
        Estimate LOD (limit of detection) and LOQ (limit of quantification) from
        pooled No-Template Control (NTC) replicate partition counts.

        If zero positives are observed across all pooled NTC partitions, LOD is
        derived from the "rule of three" upper 95% confidence bound on the true
        positive rate: p_upper ≈ 3/n_total. If NTC shows background positives
        (possible contamination), LOD is set to the NTC's own λ plus a one-sided
        95% margin (background + 1.645×SEM).

        LOQ is reported as 3×LOD, a commonly used simplified heuristic; true LOQ
        should ideally be validated empirically via replicate reproducibility (CV%).

        Returns dict with keys: lod_lambda, loq_lambda, lod_conc, loq_conc,
        ntc_lambda, ntc_pos_total, ntc_tot_total, contamination (bool).
        """
        ntc_pos = np.array(ntc_pos, dtype=float)
        ntc_tot = np.array(ntc_tot, dtype=float)
        n = min(len(ntc_pos), len(ntc_tot))
        if n == 0:
            return None
        ntc_pos, ntc_tot = ntc_pos[:n], ntc_tot[:n]

        pos_total = float(np.sum(ntc_pos))
        tot_total = float(np.sum(ntc_tot))
        if tot_total <= 0:
            return None

        if pos_total == 0:
            # Rule-of-three upper bound on the true positive rate
            p_upper = 3.0 / tot_total
            lod_lambda = -np.log(1 - p_upper) if p_upper < 1 else np.nan
            ntc_lambda = 0.0
            contamination = False
        else:
            ntc_lambda, _, _, _ = poisson_lambda(pos_total, tot_total)
            se = poisson_se(pos_total, tot_total)
            lod_lambda = ntc_lambda + 1.645 * se if not np.isnan(se) else ntc_lambda
            contamination = True

        loq_lambda = lod_lambda * 3.0
        return {
            "lod_lambda": lod_lambda, "loq_lambda": loq_lambda,
            "lod_conc": lod_lambda / partition_vol_nl_local * 1000.0,
            "loq_conc": loq_lambda / partition_vol_nl_local * 1000.0,
            "ntc_lambda": ntc_lambda, "ntc_pos_total": pos_total, "ntc_tot_total": tot_total,
            "contamination": contamination,
        }

    def compute_stability_m(ref_lambda_matrix):
        """geNorm-style stability M-value, computed in log-space on lambda values."""
        n_refs, n_samples = ref_lambda_matrix.shape
        if n_refs < 2:
            return np.array([0.0])
        log_mat = np.log(np.clip(ref_lambda_matrix, 1e-9, None))
        m_values = []
        for i in range(n_refs):
            pairwise_sd = []
            for j in range(n_refs):
                if i == j:
                    continue
                ratio = log_mat[i] - log_mat[j]
                pairwise_sd.append(np.std(ratio, ddof=1) if len(ratio) > 1 else 0.0)
            m_values.append(np.mean(pairwise_sd))
        return np.array(m_values)

    def parse_input_data(input_data):
        values = [x.replace(",", ".").strip() for x in input_data.split() if x.strip()]
        out = []
        for v in values:
            try:
                out.append(float(v))
            except ValueError:
                continue
        return np.array(out)

    # ═══════════════════════════════════════════════════════════════════════════════
    # INSTRUMENT CSV IMPORT (QuantaSoft / QX Manager / QIAcuity / generic)
    # ═══════════════════════════════════════════════════════════════════════════════
    _COLUMN_ALIASES = {
        "sample": ["sample", "samplename", "sample name", "well name", "wellname",
                   "sample description 1", "sample_id", "sampleid"],
        "target": ["target", "target name", "targettype", "target type", "assay",
                   "biomarker", "assay name", "gene", "channel"],
        "positives": ["positives", "positive droplets", "positives (rain excluded)",
                      "ch1 positives", "ch1+", "positivecount", "positive count",
                      "positive partitions", "positive reactions"],
        "total": ["accepteddroplets", "accepted droplets", "total droplets",
                  "total accepted droplets", "valid partitions", "partitions (valid)",
                  "total", "total partitions", "total reactions", "total conf. droplets",
                  "droplet count", "well droplet count"],
        "dilution_factor": ["dilution factor", "dilution", "df", "dilutionfactor"],
        "reaction_volume": ["reaction volume", "rxn volume", "reaction vol", "rxn vol",
                             "total reaction volume", "reaction volume (ul)", "reaction volume (µl)"],
        "template_volume": ["template volume", "sample volume", "dna volume", "rna volume",
                             "dna/rna volume", "input volume", "template vol",
                             "template volume (ul)", "template volume (µl)"],
    }

    def _norm_col(col):
        return str(col).strip().lower()

    def _auto_detect_column(df_columns, field):
        normed = {_norm_col(c): c for c in df_columns}
        for alias in _COLUMN_ALIASES[field]:
            if alias in normed:
                return normed[alias]
        # partial/substring match fallback
        for norm_c, orig_c in normed.items():
            for alias in _COLUMN_ALIASES[field]:
                if alias in norm_c or norm_c in alias:
                    return orig_c
        return None

    def _decode_csv_bytes(file_bytes):
        """
        Decodes uploaded CSV/TSV bytes robustly. Tries UTF-8 (with BOM handling)
        first; if that produces a suspiciously high proportion of replacement
        characters (a sign the file is actually in a different encoding, e.g.
        Latin-1/Windows-1252, common from some Windows-based instrument export
        tools in non-English locales), falls back to latin-1, which can decode
        any byte sequence without raising and is a reasonable practical default
        for Western European lab software exports.
        """
        try:
            text_utf8 = file_bytes.decode("utf-8-sig", errors="replace")
            bad_ratio = text_utf8.count("\ufffd") / max(len(text_utf8), 1)
            if bad_ratio < 0.01:
                return text_utf8
        except Exception:
            pass
        try:
            return file_bytes.decode("latin-1")
        except Exception:
            return file_bytes.decode("utf-8", errors="replace")

    def _find_header_row(lines, sep, max_scan=25):
        """
        Some instrument exports include a few metadata/title lines (e.g. run
        name, date, instrument serial) before the actual column-header row.
        Scans the first `max_scan` lines and returns the index of the line
        that best matches a "column header" pattern (highest number of cells
        that match one of our known column-name aliases), so that leading
        metadata rows can be skipped automatically instead of breaking the
        parse (or silently mis-parsing metadata as data).
        Returns 0 if the first line already looks like a good header (the
        common case), avoiding any behavior change for well-formed files.
        """
        all_aliases = set()
        for aliases in _COLUMN_ALIASES.values():
            all_aliases.update(aliases)

        best_idx, best_score = 0, -1
        for idx, line in enumerate(lines[:max_scan]):
            cells = [c.strip().lower() for c in line.split(sep)]
            score = sum(1 for c in cells if c in all_aliases or any(a in c for a in all_aliases if len(a) > 3))
            if score > best_score:
                best_score, best_idx = score, idx
        return best_idx

    def parse_instrument_csv(file_bytes):
        """
        Parses a generic dPCR/ddPCR instrument export (QuantaSoft, QX Manager,
        QIAcuity, or similar CSV/TSV exports). Auto-detects Sample, Target,
        Positives, and Total (accepted partitions) columns using common column
        name aliases across instruments. Handles common real-world quirks:
        non-UTF-8 encodings, leading metadata rows before the header, and
        comma-decimal locales (handled downstream in build_standard_import_df).
        Returns (df, detected_cols, error) where df has standardized columns:
        Sample, Target, Positives, Total. detected_cols is a dict
        {field: column_name_or_None} for any fields that need manual mapping.
        """
        import io as _io
        try:
            content = _decode_csv_bytes(file_bytes)
        except Exception as e:
            return None, None, f"File decoding error: {e}"

        if not content.strip():
            return None, None, "File is empty."

        lines = content.splitlines()
        sep = "\t" if content.count("\t") > content.count(",") else ","
        header_idx = _find_header_row(lines, sep)

        try:
            df = pd.read_csv(_io.StringIO(content), sep=sep, skiprows=header_idx)
        except Exception as e:
            return None, None, f"CSV parse error: {e}"

        if df.empty or len(df.columns) < 2:
            return None, None, "No usable columns found in file."

        df.columns = [str(c).strip() for c in df.columns]
        detected = {field: _auto_detect_column(df.columns, field) for field in _COLUMN_ALIASES}
        return df, detected, None

    def _to_numeric_robust(series):
        """
        Converts a column to numeric, tolerating comma-decimal locale
        formatting (e.g. "1234,56" instead of "1234.56"), which is common in
        CSV exports from instruments configured for European locales. Falls
        back to standard pandas numeric coercion (NaN on failure) for values
        that still don't parse.
        """
        numeric = pd.to_numeric(series, errors="coerce")
        if numeric.isna().any():
            fallback = pd.to_numeric(
                series.astype(str).str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce"
            )
            numeric = numeric.fillna(fallback)
        return numeric

    def build_standard_import_df(raw_df, col_map):
        """
        col_map: {"sample": colname, "target": colname, "positives": colname, "total": colname}
        Returns a cleaned DataFrame with columns Sample, Target, Positives, Total (numeric).
        """
        out = pd.DataFrame()
        out["Sample"] = raw_df[col_map["sample"]].astype(str).str.strip()
        out["Target"] = raw_df[col_map["target"]].astype(str).str.strip()
        out["Positives"] = _to_numeric_robust(raw_df[col_map["positives"]])
        out["Total"] = _to_numeric_robust(raw_df[col_map["total"]])
        out = out.dropna(subset=["Positives", "Total"])
        out = out[out["Total"] > 0]
        return out.reset_index(drop=True)

    def apply_csv_import_to_session(std_df, target_assays, ref_assays, ctrl_label, patient_labels):
        """
        Given the standardized import DataFrame (Sample, Target, Positives, Total)
        and the user's assay/group assignments, fills st.session_state text_area
        values for each target gene / reference gene / control / patient group
        combination. Returns the number of individual partition-count values filled.
        """
        if std_df is None or std_df.empty:
            return 0

        ctrl_keywords = [s.strip() for s in ctrl_label.split(",") if s.strip()]

        def is_ctrl(name):
            return any(kw.lower() in name.lower() for kw in ctrl_keywords)

        count = 0
        for gi, target_name in enumerate(target_assays):
            tgt_df = std_df[std_df["Target"] == target_name]

            ctrl_rows = tgt_df[tgt_df["Sample"].apply(is_ctrl)]
            if len(ctrl_rows) > 0:
                st.session_state[f"ctrl_tgt_pos_{gi}"] = "\n".join(str(int(v)) for v in ctrl_rows["Positives"])
                st.session_state[f"ctrl_tgt_tot_{gi}"] = "\n".join(str(int(v)) for v in ctrl_rows["Total"])
                count += len(ctrl_rows) * 2

            for ri, ref_name in enumerate(ref_assays):
                ref_df = std_df[std_df["Target"] == ref_name]
                ref_ctrl_rows = ref_df[ref_df["Sample"].apply(is_ctrl)]
                if len(ref_ctrl_rows) > 0:
                    st.session_state[f"ctrl_ref_pos_{gi}_{ri}"] = "\n".join(str(int(v)) for v in ref_ctrl_rows["Positives"])
                    st.session_state[f"ctrl_ref_tot_{gi}_{ri}"] = "\n".join(str(int(v)) for v in ref_ctrl_rows["Total"])
                    count += len(ref_ctrl_rows) * 2

            for pj, pat_label in enumerate(patient_labels):
                pat_keywords = [s.strip() for s in pat_label.split(",") if s.strip()]

                def is_pat(name, kws=pat_keywords):
                    return any(kw.lower() in name.lower() for kw in kws)

                pat_rows = tgt_df[tgt_df["Sample"].apply(is_pat)]
                if len(pat_rows) > 0:
                    st.session_state[f"smp_tgt_pos_{gi}_{pj}"] = "\n".join(str(int(v)) for v in pat_rows["Positives"])
                    st.session_state[f"smp_tgt_tot_{gi}_{pj}"] = "\n".join(str(int(v)) for v in pat_rows["Total"])
                    count += len(pat_rows) * 2

                for ri, ref_name in enumerate(ref_assays):
                    ref_df = std_df[std_df["Target"] == ref_name]
                    ref_pat_rows = ref_df[ref_df["Sample"].apply(is_pat)]
                    if len(ref_pat_rows) > 0:
                        st.session_state[f"smp_ref_pos_{gi}_{pj}_{ri}"] = "\n".join(str(int(v)) for v in ref_pat_rows["Positives"])
                        st.session_state[f"smp_ref_tot_{gi}_{pj}_{ri}"] = "\n".join(str(int(v)) for v in ref_pat_rows["Total"])
                        count += len(ref_pat_rows) * 2

        st.session_state["gene_count"] = max(1, len(target_assays))
        st.session_state["num_ref_genes"] = max(1, len(ref_assays))
        st.session_state["patient_count"] = max(1, len(patient_labels))
        return count

    # ═══════════════════════════════════════════════════════════════════════════════
    # EXAMPLE CSV FILES (for the three upload-based workflows)
    # ═══════════════════════════════════════════════════════════════════════════════
    EXAMPLE_CSV_MAIN_IMPORT = """Sample,Target,Positives,AcceptedDroplets
    Control_1,MYCN,1890,20000
    Control_1,RPP30,1895,20000
    Control_2,MYCN,1920,20100
    Control_2,RPP30,1915,20100
    Control_3,MYCN,1875,19850
    Control_3,RPP30,1870,19850
    Control_4,MYCN,1905,20050
    Control_4,RPP30,1900,20050
    Control_5,MYCN,1898,19980
    Control_5,RPP30,1892,19980
    Patient_1,MYCN,2780,20000
    Patient_1,RPP30,1895,20000
    Patient_2,MYCN,2820,20100
    Patient_2,RPP30,1915,20100
    Patient_3,MYCN,2755,19850
    Patient_3,RPP30,1870,19850
    Patient_4,MYCN,2800,20050
    Patient_4,RPP30,1900,20050
    Patient_5,MYCN,2790,19980
    Patient_5,RPP30,1892,19980
    """

    EXAMPLE_CSV_BATCH_SCREENING = """Sample,Target,Positives,AcceptedDroplets
    Sample_01,MYCN,1902,20000
    Sample_01,RPP30,1898,20000
    Sample_02,MYCN,1875,19950
    Sample_02,RPP30,1910,19950
    Sample_03,MYCN,1940,20100
    Sample_03,RPP30,1885,20100
    Sample_04,MYCN,1888,19870
    Sample_04,RPP30,1901,19870
    Sample_05,MYCN,1915,20050
    Sample_05,RPP30,1893,20050
    Sample_06,MYCN,1867,19920
    Sample_06,RPP30,1878,19920
    Sample_07,MYCN,1933,20080
    Sample_07,RPP30,1922,20080
    Sample_08,MYCN,4980,20000
    Sample_08,RPP30,1899,20000
    Sample_09,MYCN,5120,20100
    Sample_09,RPP30,1907,20100
    Sample_10,MYCN,4870,19950
    Sample_10,RPP30,1884,19950
    """

    EXAMPLE_CSV_VAF = """Sample,Target,Positives,AcceptedDroplets
    Baseline,Mutant_KRASG12D,450,18000
    Baseline,WT_KRAS,1200,18000
    Post-Treatment,Mutant_KRASG12D,2,19500
    Post-Treatment,WT_KRAS,1350,19500
    Follow-up-3mo,Mutant_KRASG12D,0,19800
    Follow-up-3mo,WT_KRAS,1400,19800
    Relapse,Mutant_KRASG12D,180,18500
    Relapse,WT_KRAS,1250,18500
    """

    # ═══════════════════════════════════════════════════════════════════════════════
    # OUTLIER DETECTION (Grubbs / IQR) — identical logic to GeneQuantify
    # ═══════════════════════════════════════════════════════════════════════════════
    def detect_outliers_grubbs(data, alpha=0.05):
        data = np.array(data, dtype=float)
        n = len(data)
        if n < 3:
            return []
        outlier_indices = []
        working = data.copy()
        original_indices = list(range(n))
        while len(working) >= 3:
            mean_w = np.mean(working)
            std_w = np.std(working, ddof=1)
            if std_w == 0:
                break
            g_vals = np.abs(working - mean_w) / std_w
            max_idx = np.argmax(g_vals)
            G = g_vals[max_idx]
            t_crit = stats.t.ppf(1 - alpha / (2 * len(working)), df=len(working) - 2)
            G_crit = ((len(working) - 1) / np.sqrt(len(working))) * \
                     np.sqrt(t_crit**2 / (len(working) - 2 + t_crit**2))
            if G > G_crit:
                outlier_indices.append(original_indices[max_idx])
                original_indices.pop(max_idx)
                working = np.delete(working, max_idx)
            else:
                break
        return outlier_indices

    def detect_outliers_iqr(data, multiplier=1.5):
        data = np.array(data, dtype=float)
        q1, q3 = np.percentile(data, [25, 75])
        iqr = q3 - q1
        lower = q1 - multiplier * iqr
        upper = q3 + multiplier * iqr
        return [i for i, v in enumerate(data) if v < lower or v > upper]

    def render_outlier_ui(data, label, key_prefix, method, alpha=0.05, k=1.5):
        data = np.array(data, dtype=float)
        detected = detect_outliers_grubbs(data, alpha) if method == "Grubbs" else detect_outliers_iqr(data, k)
        if not detected:
            return data, []
        st.warning(
            f"⚠️ **Potential outlier(s) in {label}** ({method}): "
            f"replicate(s) **{[i+1 for i in detected]}** — λ values: **{[round(data[i], 4) for i in detected]}**"
        )
        excluded = []
        for idx in detected:
            confirm = st.checkbox(
                f"Exclude replicate {idx+1} (λ = {data[idx]:.4f}) from {label}",
                value=False, key=f"{key_prefix}_excl_{idx}"
            )
            if confirm:
                excluded.append(idx)
        if excluded:
            cleaned = np.delete(data, excluded)
            st.info(f"ℹ️ {len(excluded)} replicate(s) excluded from {label}. Remaining n = {len(cleaned)}.")
            return cleaned, excluded
        return data, []

    # ═══════════════════════════════════════════════════════════════════════════════
    # EXAMPLE / VALIDATION SCENARIOS (dPCR partition counts — Poisson-derived)
    # ═══════════════════════════════════════════════════════════════════════════════
    SCENARIOS = {
        "S1 — Basic CNV gain (1 gene, n=5)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "ploidy": 2, "partition_vol": 0.85, "qc_min": 10000,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description_tr": "1 hedef gen, 1 hasta grubu, n=5. Hasta grubunda hedef lokusta ~1.5x kopya kazanımı (CN ~3).",
            "description_en": "1 target gene, 1 patient group, n=5. Patient group shows ~1.5x copy gain at the target locus (CN ~3).",
            "ctrl_tgt_pos_0": "1890\n1920\n1875\n1905\n1898",
            "ctrl_tgt_tot_0": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_0_0": "1895\n1915\n1870\n1900\n1892",
            "ctrl_ref_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_tgt_pos_0_0": "2780\n2820\n2755\n2800\n2790",
            "smp_tgt_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_0_0": "20000\n20100\n19850\n20050\n19980",
        },
        "S2 — Multi-gene + dual reference (2 genes, 2 groups, n=5)": {
            "gene_count": 2, "patient_count": 2, "num_ref_genes": 2,
            "ploidy": 2, "partition_vol": 0.85, "qc_min": 10000,
            "outlier_method": "IQR", "outlier_enabled": True,
            "description_tr": "2 hedef gen, 2 hasta grubu, n=5. İkili referans normalizasyonu (geNorm). Gen 1: değişim yok. Gen 2: Grup 1 kayıp, Grup 2 kazanım.",
            "description_en": "2 target genes, 2 patient groups, n=5. Dual-reference normalization (geNorm-style). Gene 1: no change. Gene 2: Group 1 = loss, Group 2 = gain.",
            # Gene 1 — control
            "ctrl_tgt_pos_0": "2000\n2020\n1980\n2010\n1995",
            "ctrl_tgt_tot_0": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_0_0": "1900\n1920\n1880\n1910\n1898",
            "ctrl_ref_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_0_1": "2100\n2120\n2080\n2110\n2098",
            "ctrl_ref_tot_0_1": "20000\n20100\n19850\n20050\n19980",
            # Gene 1 — Group 1 (no change)
            "smp_tgt_pos_0_0": "2005\n2025\n1985\n2015\n2000",
            "smp_tgt_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_0": "1900\n1920\n1880\n1910\n1898",
            "smp_ref_tot_0_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_1": "2100\n2120\n2080\n2110\n2098",
            "smp_ref_tot_0_0_1": "20000\n20100\n19850\n20050\n19980",
            # Gene 1 — Group 2 (no change)
            "smp_tgt_pos_0_1": "1995\n2015\n1975\n2005\n1990",
            "smp_tgt_tot_0_1": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_1_0": "1900\n1920\n1880\n1910\n1898",
            "smp_ref_tot_0_1_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_1_1": "2100\n2120\n2080\n2110\n2098",
            "smp_ref_tot_0_1_1": "20000\n20100\n19850\n20050\n19980",
            # Gene 2 — control
            "ctrl_tgt_pos_1": "2000\n2020\n1980\n2010\n1995",
            "ctrl_tgt_tot_1": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_1_0": "1900\n1920\n1880\n1910\n1898",
            "ctrl_ref_tot_1_0": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_1_1": "2100\n2120\n2080\n2110\n2098",
            "ctrl_ref_tot_1_1": "20000\n20100\n19850\n20050\n19980",
            # Gene 2 — Group 1 (loss, ratio ~0.5)
            "smp_tgt_pos_1_0": "1030\n1045\n1015\n1038\n1022",
            "smp_tgt_tot_1_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_1_0_0": "1900\n1920\n1880\n1910\n1898",
            "smp_ref_tot_1_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_1_0_1": "2100\n2120\n2080\n2110\n2098",
            "smp_ref_tot_1_0_1": "20000\n20100\n19850\n20050\n19980",
            # Gene 2 — Group 2 (gain, ratio ~2.0)
            "smp_tgt_pos_1_1": "3790\n3830\n3760\n3810\n3795",
            "smp_tgt_tot_1_1": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_1_1_0": "1900\n1920\n1880\n1910\n1898",
            "smp_ref_tot_1_1_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_1_1_1": "2100\n2120\n2080\n2110\n2098",
            "smp_ref_tot_1_1_1": "20000\n20100\n19850\n20050\n19980",
        },
        "S3 — Outlier detection demo (n=6)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "ploidy": 2, "partition_vol": 0.85, "qc_min": 10000,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description_tr": "1 hedef gen, n=6. Kontrol grubunda replikat 5 aykırı değer (kontaminasyon benzeri yüksek λ). Grubbs testiyle tespit edilir.",
            "description_en": "1 target gene, n=6. Replicate 5 in the control group is an outlier (contamination-like elevated λ). Detected by Grubbs' test.",
            "ctrl_tgt_pos_0": "1890\n1920\n1875\n1905\n5800\n1898",
            "ctrl_tgt_tot_0": "20000\n20100\n19850\n20050\n20000\n19980",
            "ctrl_ref_pos_0_0": "1895\n1915\n1870\n1900\n1888\n1892",
            "ctrl_ref_tot_0_0": "20000\n20100\n19850\n20050\n19990\n19980",
            "smp_tgt_pos_0_0": "3780\n3820\n3755\n3800\n3790",
            "smp_tgt_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_0_0": "20000\n20100\n19850\n20050\n19980",
        },
        "S4 — Multi-group ANOVA (3 groups, n=5)": {
            "gene_count": 1, "patient_count": 3, "num_ref_genes": 1,
            "ploidy": 2, "partition_vol": 0.85, "qc_min": 10000,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description_tr": "1 hedef gen, 3 hasta grubu, n=5. Grup 1: hafif kazanım, Grup 2: güçlü kazanım, Grup 3: değişim yok. Tek yönlü ANOVA + Tukey HSD.",
            "description_en": "1 target gene, 3 patient groups, n=5. Group 1: mild gain, Group 2: strong gain, Group 3: no change. One-way ANOVA + Tukey HSD.",
            "ctrl_tgt_pos_0": "1890\n1920\n1875\n1905\n1898",
            "ctrl_tgt_tot_0": "20000\n20100\n19850\n20050\n19980",
            "ctrl_ref_pos_0_0": "1895\n1915\n1870\n1900\n1892",
            "ctrl_ref_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            # Group 1: mild gain (ratio ~1.3)
            "smp_tgt_pos_0_0": "2440\n2470\n2415\n2455\n2445",
            "smp_tgt_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_0_0": "20000\n20100\n19850\n20050\n19980",
            # Group 2: strong gain (ratio ~2.5)
            "smp_tgt_pos_0_1": "4520\n4570\n4480\n4545\n4510",
            "smp_tgt_tot_0_1": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_1_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_1_0": "20000\n20100\n19850\n20050\n19980",
            # Group 3: no change
            "smp_tgt_pos_0_2": "1885\n1915\n1870\n1900\n1892",
            "smp_tgt_tot_0_2": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_2_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_2_0": "20000\n20100\n19850\n20050\n19980",
        },
        "S5 — QC & saturation demo (n=5)": {
            "gene_count": 1, "patient_count": 1, "num_ref_genes": 1,
            "ploidy": 2, "partition_vol": 0.85, "qc_min": 10000,
            "outlier_method": "Grubbs", "outlier_enabled": True,
            "description_tr": "1 hedef gen, n=5. Kontrol grubunda 1 replikat düşük partisyon (QC hatası), 1 replikat doygun (%100 pozitif). Kalite kontrol uyarılarını gösterir.",
            "description_en": "1 target gene, n=5. Control group has 1 low-partition replicate (QC failure) and 1 saturated replicate (100% positive). Demonstrates QC warnings.",
            "ctrl_tgt_pos_0": "1890\n1920\n500\n20000\n1898",
            "ctrl_tgt_tot_0": "20000\n20100\n3000\n20000\n19980",
            "ctrl_ref_pos_0_0": "1895\n1915\n505\n1901\n1892",
            "ctrl_ref_tot_0_0": "20000\n20100\n3000\n20050\n19980",
            "smp_tgt_pos_0_0": "3780\n3820\n3755\n3800\n3790",
            "smp_tgt_tot_0_0": "20000\n20100\n19850\n20050\n19980",
            "smp_ref_pos_0_0_0": "1895\n1915\n1870\n1900\n1892",
            "smp_ref_tot_0_0_0": "20000\n20100\n19850\n20050\n19980",
        },
    }

    # ═══════════════════════════════════════════════════════════════════════════════
    # HEADER
    # ═══════════════════════════════════════════════════════════════════════════════
    _title_parts = _t['title'].split(' ', 1)
    _header_emoji = _title_parts[0]
    _header_title_text = _title_parts[1] if len(_title_parts) > 1 else _t['title']

    st.markdown(
        f"""
        <div style="background:linear-gradient(90deg,#004d40,#00796b);
                    color:white;padding:16px 18px;border-radius:8px;margin-bottom:8px;
                    display:flex;align-items:center;gap:12px;">
            <span style="font-size:28px;line-height:1;flex-shrink:0;">{_header_emoji}</span>
            <div style="display:flex;flex-direction:column;justify-content:center;">
                <span style="font-size:20px;font-weight:800;line-height:1.3;">{_header_title_text}</span>
                <span style="font-size:11px;opacity:0.8;margin-top:3px;line-height:1.3;">{_t['subtitle']}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True
    )

    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — Simple / Advanced Mode toggle
    # ═══════════════════════════════════════════════════════════════════════════════
    if "_pending_advanced_mode" in st.session_state:
        st.session_state["advanced_mode"] = st.session_state.pop("_pending_advanced_mode")

    advanced_mode = st.sidebar.toggle(
        _t['advanced_mode_label'], key="advanced_mode", help=_t['advanced_mode_help'],
        **({} if "advanced_mode" in st.session_state else {"value": False})
    )
    if not advanced_mode:
        st.sidebar.caption(_t['simple_mode_caption'])
    st.sidebar.divider()

    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — User guide, links
    # ═══════════════════════════════════════════════════════════════════════════════
    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — Example data loader
    # ═══════════════════════════════════════════════════════════════════════════════
    st.sidebar.markdown(
        f"<div style='font-size:12px;font-weight:600;color:#004d40;margin-bottom:2px;'>"
        f"{_t['sidebar_example_title']}</div>", unsafe_allow_html=True
    )
    selected_scenario = st.sidebar.selectbox(
        _t['sidebar_example_select'], options=["—"] + list(SCENARIOS.keys()),
        key="scenario_selector", label_visibility="collapsed"
    )
    if selected_scenario != "—":
        _sc = SCENARIOS[selected_scenario]
        st.sidebar.caption(_sc.get(f"description_{language_code}", _sc.get("description_en", "")))
        if st.sidebar.button(_t['sidebar_example_load_btn'], key="load_scenario_btn", use_container_width=True):
            for _key, _val in _sc.items():
                if _key.startswith("description"):
                    continue
                st.session_state[_key] = _val
            st.sidebar.success(_t['sidebar_example_loaded'].format(s=selected_scenario))

    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — Instrument CSV import
    # ═══════════════════════════════════════════════════════════════════════════════
    with st.sidebar.expander(_t['csv_import_expander'], expanded=False):
        st.caption(_t['csv_import_description'])
        st.download_button(
            _t['download_example_csv'], data=EXAMPLE_CSV_MAIN_IMPORT.encode("utf-8"),
            file_name="example_import.csv", mime="text/csv", key="dl_example_main_import",
            use_container_width=True
        )
        csv_file = st.file_uploader(_t['csv_uploader'], type=["csv", "tsv", "txt"], key="csv_import_uploader")
        if csv_file is not None:
            _raw_bytes = csv_file.read()
            _raw_df, _detected_cols, _parse_err = parse_instrument_csv(_raw_bytes)
            if _parse_err:
                st.error(_t['csv_parse_error'].format(err=_parse_err))
            else:
                st.markdown(_t['csv_col_mapping_title'])
                _col_options = list(_raw_df.columns)
                _col_map = {}
                for _field, _label_key in [("sample", "csv_col_sample"), ("target", "csv_col_target"),
                                            ("positives", "csv_col_positives"), ("total", "csv_col_total")]:
                    _detected = _detected_cols.get(_field)
                    _status = _t['csv_col_auto_detected'] if _detected else _t['csv_col_manual_needed']
                    _default_idx = _col_options.index(_detected) if _detected in _col_options else 0
                    _col_map[_field] = st.selectbox(
                        f"{_t[_label_key]} — {_status}", options=_col_options,
                        index=_default_idx, key=f"csv_colmap_{_field}"
                    )

                _std_df = build_standard_import_df(_raw_df, _col_map)
                if not _std_df.empty:
                    with st.expander(_t['csv_preview_title'], expanded=False):
                        st.dataframe(_std_df, use_container_width=True)

                    st.markdown(_t['csv_assay_assignment_title'])
                    _unique_targets = sorted(_std_df["Target"].unique())
                    _sel_target_assays = st.multiselect(
                        _t['csv_target_assays_label'], options=_unique_targets,
                        default=_unique_targets[:1] if _unique_targets else [],
                        key="csv_target_assays"
                    )
                    _remaining_targets = [t for t in _unique_targets if t not in _sel_target_assays]
                    _sel_ref_assays = st.multiselect(
                        _t['csv_ref_assays_label'], options=_remaining_targets,
                        default=_remaining_targets[:1] if _remaining_targets else [],
                        key="csv_ref_assays"
                    )

                    st.markdown(_t['csv_group_assignment_title'])
                    _unique_samples = sorted(_std_df["Sample"].unique())
                    _ctrl_label_csv = st.text_input(
                        _t['csv_ctrl_label'],
                        value=_unique_samples[0] if _unique_samples else "",
                        key="csv_ctrl_label_input"
                    )
                    _n_pat_csv = st.number_input(
                        _t['csv_n_patient_groups'], min_value=1, max_value=10, value=1, step=1, key="csv_n_pat"
                    )
                    _patient_labels_csv = []
                    for _pg in range(int(_n_pat_csv)):
                        _default_pat = _unique_samples[_pg + 1] if _pg + 1 < len(_unique_samples) else ""
                        _pat_lbl = st.text_input(
                            _t['csv_patient_label'].format(i=_pg + 1),
                            value=_default_pat, key=f"csv_pat_label_{_pg}"
                        )
                        _patient_labels_csv.append(_pat_lbl)

                    if st.button(_t['csv_apply_btn'], key="csv_apply_btn", use_container_width=True):
                        if _sel_target_assays and _sel_ref_assays:
                            _n_filled = apply_csv_import_to_session(
                                _std_df, _sel_target_assays, _sel_ref_assays, _ctrl_label_csv, _patient_labels_csv
                            )
                            if _n_filled > 0:
                                st.success(_t['csv_apply_success'].format(n=_n_filled))
                            else:
                                st.warning(_t['csv_apply_warning'])
                        else:
                            st.warning(_t['csv_apply_warning'])

    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — Project save/load
    # ═══════════════════════════════════════════════════════════════════════════════
    with st.sidebar.expander(_t['project_expander'], expanded=False):
        st.caption(_t['project_description'])
        _project_data = export_project_state()
        st.download_button(
            _t['project_export_btn'],
            data=json.dumps(_project_data, ensure_ascii=False, indent=2).encode("utf-8"),
            file_name="absolutegene_project.json", mime="application/json",
            key="project_export_btn", use_container_width=True
        )
        project_file = st.file_uploader(_t['project_import_uploader'], type=["json"], key="project_import_uploader")
        if project_file is not None:
            try:
                _imported_project = json.loads(project_file.read().decode("utf-8"))
                _n_restored = import_project_state(_imported_project)
                st.success(_t['project_import_success'].format(n=_n_restored))
            except Exception as _e:
                st.error(_t['project_import_error'].format(err=str(_e)))

    # ═══════════════════════════════════════════════════════════════════════════════
    # SIDEBAR — Session History Panel (in-session snapshots, not persisted across
    # browser reloads — Streamlit has no server-side storage without extra infra;
    # for permanent storage use the JSON Export/Import above instead).
    # ═══════════════════════════════════════════════════════════════════════════════
    with st.sidebar.expander(_t['history_expander'], expanded=False):
        st.caption(_t['history_description'])
        if "_history_snapshots" not in st.session_state:
            st.session_state["_history_snapshots"] = []

        _hist_name_default = f"Snapshot {len(st.session_state['_history_snapshots']) + 1}"
        _hist_name = st.text_input(_t['history_name_label'], value=_hist_name_default, key="history_name_input")
        if st.button(_t['history_save_btn'], key="history_save_btn", use_container_width=True):
            import datetime as _dt_hist
            st.session_state["_history_snapshots"].append({
                "name": _hist_name or _hist_name_default,
                "timestamp": _dt_hist.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "data": export_project_state(),
            })
            st.success(_t['history_saved_success'].format(name=_hist_name or _hist_name_default))

        _snapshots = st.session_state["_history_snapshots"]
        if _snapshots:
            st.markdown("---")
            _snapshot_labels = [f"{s['name']} ({s['timestamp']})" for s in _snapshots]
            _selected_snapshot_idx = st.selectbox(
                _t['history_select_label'], options=range(len(_snapshots)),
                format_func=lambda i: _snapshot_labels[i], key="history_select"
            )
            hcol1, hcol2, hcol3 = st.columns(3)
            with hcol1:
                if st.button(_t['history_restore_btn'], key="history_restore_btn", use_container_width=True):
                    _n = import_project_state(_snapshots[_selected_snapshot_idx]["data"])
                    st.success(_t['history_restore_success'].format(n=_n))
            with hcol2:
                st.download_button(
                    _t['history_download_btn'],
                    data=json.dumps(_snapshots[_selected_snapshot_idx]["data"], ensure_ascii=False, indent=2).encode("utf-8"),
                    file_name=f"{_snapshots[_selected_snapshot_idx]['name'].replace(' ', '_')}.json",
                    mime="application/json", key="history_download_btn", use_container_width=True
                )
            with hcol3:
                if st.button(_t['history_delete_btn'], key="history_delete_btn", use_container_width=True):
                    st.session_state["_history_snapshots"].pop(_selected_snapshot_idx)
                    st.rerun()
        else:
            st.caption(_t['history_no_snapshots'])

    guide_clicked = st.sidebar.button(_t['guide_btn'], use_container_width=True)
    if guide_clicked:
        @st.dialog("📘 AbsoluteGene — User Guide" if language_code == "en" else "📘 AbsoluteGene — Kullanım Kılavuzu", width="large")
        def show_guide():
            if language_code == "en":
                st.markdown("""
    ### How dPCR analysis differs from qPCR (GeneQuantify)

    Digital PCR partitions each sample into thousands of micro-reactions (droplets or
    wells). Each partition ends up **positive** (target present) or **negative**
    (target absent). Instead of a Cq value, the input here is:

    - **Positive partition count** (per replicate)
    - **Total accepted partition count** (per replicate, after excluding failed/rain partitions in your instrument software)

    From these two numbers, the app calculates the average copies per partition (λ)
    using the Poisson distribution:

    ```
    p = positive / total
    λ = -ln(1 - p)
    ```

    A 95% confidence interval is calculated using the standard delta-method normal
    approximation for Poisson-derived λ.

    **Normalization & Copy Number:**
    ```
    NF (normalization factor) = geometric mean of reference-locus λ values
    Ratio = λ(target) / NF
    CN (copy number) = ploidy(reference) × Ratio
    Fold Change = Ratio(sample) / Ratio(control)
    ```

    Unlike qPCR, dPCR is an **absolute, endpoint** measurement — no amplification
    efficiency correction (Pfaffl) or standard curve is required, since efficiency
    does not affect the binary positive/negative call at reaction endpoint.

    **Quality control:** Replicates with total accepted partitions below your
    configured threshold are flagged (dMIQE recommends ≥10,000 for ddPCR
    instruments as a general guideline — check your instrument's own
    specifications).

    **Statistical testing:** Because λ and Ratio are already on a linear scale (no
    log transform needed, unlike ΔCt), Shapiro-Wilk → Levene → t-test/Welch/Mann-Whitney
    (2 groups) or ANOVA/Kruskal-Wallis + post-hoc (≥3 groups) are applied directly
    to the Ratio values.

    **Reference:** Huggett JF et al. *The digital MIQE guidelines: Minimum Information
    for Publication of Quantitative Digital PCR Experiments.* Clin Chem 2013;
    updated Huggett et al. 2020.

    **Disclaimer:** For research and educational use only. Not validated for
    clinical diagnostic decision-making.

    **Contact:** mailtoburhanettin@gmail.com
    """)
            else:
                st.markdown("""
    ### dPCR analizi qPCR'den (GeneQuantify) nasıl farklıdır?

    Dijital PCR, her örneği binlerce mikro-reaksiyona (damlacık veya kuyucuk) böler.
    Her partisyon **pozitif** (hedef var) veya **negatif** (hedef yok) olarak sonuçlanır.
    Cq değeri yerine buraya girilen veri:

    - **Pozitif partisyon sayısı** (replikat başına)
    - **Toplam kabul edilen partisyon sayısı** (replikat başına; cihaz yazılımınızda
      başarısız/rain partisyonlar hariç tutulduktan sonra)

    Bu iki sayıdan, Poisson dağılımı kullanılarak partisyon başına ortalama kopya
    sayısı (λ) hesaplanır:

    ```
    p = pozitif / toplam
    λ = -ln(1 - p)
    ```

    %95 güven aralığı, Poisson-türevli λ için standart delta-method normal
    yaklaşımı kullanılarak hesaplanır.

    **Normalizasyon & Kopya Sayısı:**
    ```
    NF (normalizasyon faktörü) = referans lokus λ değerlerinin geometrik ortalaması
    Oran = λ(hedef) / NF
    CN (kopya sayısı) = ploidi(referans) × Oran
    Kat Değişimi = Oran(örnek) / Oran(kontrol)
    ```

    qPCR'nin aksine dPCR, **mutlak ve endpoint** bir ölçümdür — amplifikasyon
    verimliliği düzeltmesi (Pfaffl) veya standart eğri gerekmez, çünkü verimlilik
    reaksiyon sonundaki ikili pozitif/negatif sonucu etkilemez.

    **Kalite kontrol:** Toplam kabul edilen partisyonu ayarladığınız eşiğin altında
    kalan replikatlar işaretlenir (dMIQE genel öneri olarak ddPCR cihazları için
    tipik olarak ≥10.000 önerir — kendi cihazınızın spesifikasyonlarını kontrol edin).

    **İstatistiksel test:** λ ve Oran zaten doğrusal ölçekte olduğundan (ΔCt'nin
    aksine log dönüşümüne gerek yoktur), Shapiro-Wilk → Levene → t-testi/Welch/
    Mann-Whitney (2 grup) veya ANOVA/Kruskal-Wallis + post-hoc (≥3 grup) doğrudan
    Oran değerlerine uygulanır.

    **Referans:** Huggett JF ve ark. *The digital MIQE guidelines.* Clin Chem 2013;
    güncelleme Huggett ve ark. 2020.

    **Sorumluluk Reddi:** Yalnızca araştırma ve eğitim amaçlıdır. Klinik tanı
    kararları için doğrulanmamıştır.

    **İletişim:** mailtoburhanettin@gmail.com
    """)
        show_guide()

    st.sidebar.divider()
    st.sidebar.link_button(
        _t['sidebar_github_btn'],
        "https://github.com/burhanettiny/GeneQuantify",
        use_container_width=True
    )
    st.sidebar.link_button(
        _t['sidebar_sister_tool'],
        "https://GeneQuantify.streamlit.app/",
        use_container_width=True
    )
    st.sidebar.caption(f"AbsoluteGene v{APP_VERSION} — GPL-3.0 | mailtoburhanettin@gmail.com")

    tab_data, tab_results, tab_batch, tab_vaf, tab_clinical, tab_crm, tab_report = st.tabs([
        f"📥 {_t['tab_data']}", f"📊 {_t['tab_results']}", f"🔬 {_t['tab_batch']}",
        f"🧬 {_t['tab_vaf']}", f"🩺 {_t['tab_clinical']}", f"🏭 {_t['tab_crm']}", f"📄 {_t['tab_report']}"
    ])

    # ═══════════════════════════════════════════════════════════════════════════════
    # DATA CONTAINERS
    # ═══════════════════════════════════════════════════════════════════════════════
    input_values_table = []   # per-replicate rows (for display + PDF)
    data = []                 # per gene/group summary rows (ratio, CN, FC)
    stats_data = []            # pairwise stats rows

    with tab_data:
        # ── Study Design ──────────────────────────────────────────────────────────
        with st.container(border=True):
            st.markdown(f"**{_t['study_design']}**")
            sd_c1, sd_c2, sd_c3 = st.columns(3)
            with sd_c1:
                num_target_genes = st.number_input(_t['num_target_genes'], min_value=1, step=1, key="gene_count")
            with sd_c2:
                num_patient_groups = st.number_input(_t['num_patient_groups'], min_value=1, step=1, key="patient_count")
            with sd_c3:
                num_ref_genes = st.number_input(_t['num_ref_genes'], min_value=1, max_value=10, step=1,
                                                 key="num_ref_genes", help=_t['ref_gene_help'])
            sd_c4, sd_c5, sd_c6 = st.columns(3)
            with sd_c4:
                ploidy = st.number_input(_t['ploidy_label'], min_value=1, max_value=10,
                                          **({} if "ploidy" in st.session_state else {"value": 2}),
                                          step=1, key="ploidy", help=_t['ploidy_help'])
            with sd_c5:
                partition_vol_nl = st.number_input(
                    _t['partition_vol_label'], min_value=0.01, max_value=100.0,
                    **({} if "partition_vol" in st.session_state else {"value": 0.85}),
                    step=0.01, format="%.2f", key="partition_vol", help=_t['partition_vol_help']
                )
            with sd_c6:
                qc_min_partitions = st.number_input(
                    _t['qc_min_partitions'], min_value=100, max_value=100000,
                    **({} if "qc_min" in st.session_state else {"value": 10000}),
                    step=500, key="qc_min", help=_t['qc_min_partitions_help']
                )

        # ── Outlier Detection Settings ────────────────────────────────────────────
        with st.container(border=True):
            st.markdown(_t['outlier_section_title'])
            out_c1, out_c2 = st.columns([1, 2])
            with out_c1:
                outlier_enabled = st.checkbox(
                    _t['outlier_enable'], key="outlier_enabled", help=_t['outlier_enable_help'],
                    **({} if "outlier_enabled" in st.session_state else {"value": True})
                )
                outlier_method = st.radio(_t['outlier_method_label'], options=["Grubbs", "IQR"],
                                           key="outlier_method", help=_t['outlier_method_help'])
                if outlier_method == "Grubbs":
                    st.caption(_t['grubbs_power_warning'])
            with out_c2:
                if outlier_method == "Grubbs":
                    grubbs_alpha = st.number_input(
                        _t['outlier_alpha_label'], min_value=0.01, max_value=0.10,
                        **({} if "grubbs_alpha" in st.session_state else {"value": 0.05}),
                        step=0.01, format="%.2f", key="grubbs_alpha"
                    )
                    iqr_multiplier = 1.5
                else:
                    iqr_multiplier = st.number_input(
                        _t['outlier_iqr_label'], min_value=1.0, max_value=3.0,
                        **({} if "iqr_mult" in st.session_state else {"value": 1.5}),
                        step=0.25, format="%.2f", key="iqr_mult"
                    )
                    grubbs_alpha = 0.05

        st.divider()

        # ── Dilution / Dynamic Range Calculator ───────────────────────────────────
        with st.expander(_t['dilution_expander'], expanded=False):
            st.caption(_t['dilution_description'])
            dil_mode = st.radio(
                _t['dilution_mode_label'],
                options=[_t['dilution_mode_counts'], _t['dilution_mode_lambda']],
                key="dilution_mode", horizontal=True
            )
            dcol1, dcol2 = st.columns(2)
            if dil_mode == _t['dilution_mode_counts']:
                with dcol1:
                    dil_pos = st.number_input(_t['dilution_positive_label'], min_value=0, value=1000, step=10, key="dil_pos")
                with dcol2:
                    dil_tot = st.number_input(_t['dilution_total_label'], min_value=1, value=20000, step=100, key="dil_tot")
                _dil_lambda_input, _, _, _ = poisson_lambda(dil_pos, dil_tot)
            else:
                with dcol1:
                    _dil_lambda_input = st.number_input(_t['dilution_lambda_label'], min_value=0.001, value=1.6,
                                                         step=0.01, format="%.3f", key="dil_lambda_direct")

            dil_target = st.number_input(_t['dilution_target_label'], min_value=0.1, max_value=3.0,
                                          value=1.6, step=0.1, format="%.2f", key="dil_target_lambda")

            if st.button(_t['dilution_calc_btn'], key="dil_calc_btn"):
                _dil_result = recommend_dilution(_dil_lambda_input, dil_target)
                if _dil_result is None:
                    st.error("Invalid input.")
                else:
                    _lam = _dil_result["current_lambda"]
                    _factor = _dil_result["dilution_factor"]
                    if _dil_result["status"] == "optimal":
                        st.success(_t['dilution_result_optimal'].format(lam=_lam))
                    elif _dil_result["status"] == "too_low":
                        st.warning(_t['dilution_result_too_low'].format(lam=_lam, factor=_factor))
                    elif _dil_result["status"] == "too_high":
                        st.warning(_t['dilution_result_too_high'].format(lam=_lam, factor=_factor))
                    else:
                        st.error(_t['dilution_result_saturated'].format(lam=_lam, factor=_factor))
                    st.caption(_t['dilution_factor_note'])

        # ── 2-Color Multiplex Cluster Converter (advanced) ────────────────────────
        if advanced_mode:
            with st.expander(_t['multiplex_expander'], expanded=False):
                st.caption(_t['multiplex_description'])
                mx_c1, mx_c2, mx_c3, mx_c4 = st.columns(4)
                with mx_c1:
                    _mx_n11 = st.number_input(_t['multiplex_n11_label'], min_value=0, value=1500, step=10, key="mx_n11")
                with mx_c2:
                    _mx_n10 = st.number_input(_t['multiplex_n10_label'], min_value=0, value=400, step=10, key="mx_n10")
                with mx_c3:
                    _mx_n01 = st.number_input(_t['multiplex_n01_label'], min_value=0, value=300, step=10, key="mx_n01")
                with mx_c4:
                    _mx_n00 = st.number_input(_t['multiplex_n00_label'], min_value=0, value=17800, step=10, key="mx_n00")

                if st.button(_t['multiplex_calc_btn'], key="mx_calc_btn"):
                    _mx_result = compute_multiplex_cluster(_mx_n11, _mx_n10, _mx_n01, _mx_n00)
                    if _mx_result is None:
                        st.error("Invalid input — check that counts are non-negative and total > 0.")
                    else:
                        st.markdown(_t['multiplex_result_title'])
                        mxc1, mxc2, mxc3 = st.columns(3)
                        mxc1.metric(_t['multiplex_pos_target_label'], _mx_result["pos_target"])
                        mxc2.metric(_t['multiplex_pos_ref_label'], _mx_result["pos_ref"])
                        mxc3.metric(_t['multiplex_total_label'], _mx_result["N"])

                        mxr1, mxr2 = st.columns(2)
                        mxr1.metric(
                            _t['multiplex_ratio_covaware'],
                            f"{_mx_result['ratio']:.4f}",
                            delta=f"{_mx_result['ci_low']:.4f}–{_mx_result['ci_high']:.4f}",
                            delta_color="off"
                        )
                        mxr2.metric(
                            _t['multiplex_ratio_indep'],
                            f"{_mx_result['ratio']:.4f}",
                            delta=f"{_mx_result['ci_low_indep']:.4f}–{_mx_result['ci_high_indep']:.4f}",
                            delta_color="off"
                        )
                        st.info(_t['multiplex_paste_hint'])

        st.markdown(
            f"<div style='font-size:15px;font-weight:700;color:#004d40;margin-bottom:6px;'>"
            f"{_t['patient_data_header']}</div>", unsafe_allow_html=True
        )
        st.caption(_t['input_format_info'])

        # ── Helper: sync target + all reference loci, compute lambda/NF/ratio ──────
        def sync_and_compute(target_pos_txt, target_tot_txt, ref_pos_txts, ref_tot_txts, label, key_prefix):
            """
            Parses target & reference-locus positive/total partition text areas,
            aligns them to a common replicate count, computes Poisson lambda for
            each locus/replicate, applies QC (min partitions) and outlier
            detection (cascading across target + all reference loci), then
            computes the normalization factor (geometric mean of reference
            lambdas) and the normalized Ratio = lambda(target) / NF per kept
            replicate.
            Returns None if any required field is empty.
            """
            pos_t = parse_input_data(target_pos_txt)
            tot_t = parse_input_data(target_tot_txt)
            ref_pos = [parse_input_data(x) for x in ref_pos_txts]
            ref_tot = [parse_input_data(x) for x in ref_tot_txts]

            lengths = [len(pos_t), len(tot_t)] + [len(a) for a in ref_pos] + [len(a) for a in ref_tot]
            if any(l == 0 for l in lengths):
                return None

            n_common = min(lengths)
            if len(set(lengths)) > 1:
                st.warning(f"⚠️ {label}: " + (
                    f"Farklı replikat sayıları tespit edildi — analiz n={n_common} ile devam edecek."
                    if language_code == "tr" else
                    f"Unequal replicate counts detected — analysis will proceed with n={n_common}."
                ))

            pos_t, tot_t = pos_t[:n_common], tot_t[:n_common]
            ref_pos = [a[:n_common] for a in ref_pos]
            ref_tot = [a[:n_common] for a in ref_tot]

            # --- target lambda/status ---
            lam_t, status_t = [], []
            for p, t in zip(pos_t, tot_t):
                lam, cl, ch, s = poisson_lambda(p, t)
                lam_t.append(lam if lam is not None else np.nan)
                status_t.append(s)
            lam_t = np.array(lam_t)
            qc_t = tot_t < qc_min_partitions

            # --- reference loci lambda/status ---
            lam_refs, status_refs, qc_refs = [], [], []
            for r in range(len(ref_pos)):
                lam_r, status_r = [], []
                for p, t in zip(ref_pos[r], ref_tot[r]):
                    lam, cl, ch, s = poisson_lambda(p, t)
                    lam_r.append(lam if lam is not None else np.nan)
                    status_r.append(s)
                lam_refs.append(np.array(lam_r))
                status_refs.append(status_r)
                qc_refs.append(ref_tot[r] < qc_min_partitions)

            # --- combined validity mask (all loci must be ok + QC-pass at a replicate) ---
            valid_mask = np.array([s == "ok" for s in status_t]) & (~qc_t)
            for r in range(len(ref_pos)):
                valid_mask &= np.array([s == "ok" for s in status_refs[r]]) & (~qc_refs[r])

            n_qc_fail = int(np.sum(qc_t) + sum(int(np.sum(q)) for q in qc_refs))
            n_saturated = int(sum(1 for s in status_t if s == "saturated") +
                              sum(sum(1 for s in sr if s == "saturated") for sr in status_refs))
            if n_qc_fail > 0:
                st.warning(_t['qc_fail_warning'].format(n=n_qc_fail, thr=int(qc_min_partitions)))
            if n_saturated > 0:
                st.warning(_t['saturation_warning'].format(n=n_saturated))

            # --- cascading outlier detection: target first, then each ref locus ---
            excluded_idx = set()
            valid_positions = [i for i in range(n_common) if valid_mask[i]]

            if outlier_enabled and len(valid_positions) >= 3:
                target_vals = lam_t[valid_positions]
                _, excl_local = render_outlier_ui(
                    target_vals, f"{label} — {_t['target_gene']}", f"{key_prefix}_tgt",
                    outlier_method, grubbs_alpha, iqr_multiplier
                )
                excluded_idx.update(valid_positions[i] for i in excl_local)

            valid_positions = [i for i in valid_positions if i not in excluded_idx]
            for r in range(len(ref_pos)):
                if outlier_enabled and len(valid_positions) >= 3:
                    ref_vals = lam_refs[r][valid_positions]
                    _, excl_local = render_outlier_ui(
                        ref_vals, f"{label} — {_t['reference_gene']} {r+1}", f"{key_prefix}_ref{r}",
                        outlier_method, grubbs_alpha, iqr_multiplier
                    )
                    new_excl = [valid_positions[i] for i in excl_local]
                    excluded_idx.update(new_excl)
                    valid_positions = [i for i in valid_positions if i not in new_excl]

            kept = sorted(valid_positions)

            if kept:
                ref_matrix_kept = np.vstack([lam_refs[r][kept] for r in range(len(ref_pos))])
                NF_kept = np.array([geometric_mean(ref_matrix_kept[:, k]) for k in range(len(kept))])
                lam_t_kept = lam_t[kept]
                ratio_kept = lam_t_kept / NF_kept
            else:
                ref_matrix_kept = np.zeros((len(ref_pos), 0))
                NF_kept = np.array([])
                lam_t_kept = np.array([])
                ratio_kept = np.array([])

            detail_rows = []
            for i in range(n_common):
                detail_rows.append({
                    "replicate": i + 1,
                    "target_pos": pos_t[i], "target_tot": tot_t[i],
                    "target_lambda": lam_t[i], "target_status": status_t[i],
                    "ref_lambdas": [lam_refs[r][i] for r in range(len(ref_pos))],
                    "qc_fail": bool(qc_t[i]) or any(bool(qc_refs[r][i]) for r in range(len(ref_pos))),
                    "excluded": i in excluded_idx,
                    "used": i in kept,
                })

            return {
                "n_common": n_common, "kept": kept,
                "lam_target_kept": lam_t_kept, "NF_kept": NF_kept, "ratio_kept": ratio_kept,
                "ref_matrix_kept": ref_matrix_kept,
                "detail_rows": detail_rows,
                "pos_target_kept": pos_t[kept] if kept else np.array([]),
                "tot_target_kept": tot_t[kept] if kept else np.array([]),
            }

        def show_exclusion_diagnosis(result_dict):
            """
            When sync_and_compute returns a valid (non-None) result but zero
            replicates survived processing, this shows a diagnostic message
            that actually explains WHY (all QC-failed / all saturated / all
            outlier-excluded / a mix) instead of a generic, potentially
            misleading message about mismatched line counts (which is not
            actually what happened in this case — the input parsed fine).
            """
            rows = result_dict["detail_rows"] if result_dict else []
            if not rows:
                st.error(_t['warning_field_empty'])
                return
            n = len(rows)
            n_qc = sum(1 for r in rows if r["qc_fail"])
            n_sat = sum(1 for r in rows if r["target_status"] == "saturated")
            n_excl = sum(1 for r in rows if r["excluded"])
            if n_qc == n:
                st.error(_t['warning_all_excluded_qc'].format(thr=int(qc_min_partitions)))
            elif n_sat == n:
                st.error(_t['warning_all_excluded_saturated'])
            elif n_excl == n:
                st.error(_t['warning_all_excluded_outlier'])
            else:
                st.error(_t['warning_all_excluded_mixed'])

        def _ta(label, key):
            """text_area wrapper that avoids the Streamlit session-state/value conflict warning."""
            kwargs = {} if key in st.session_state else {"value": ""}
            return st.text_area(label, key=key, **kwargs)

        # ── Main per-gene loop ────────────────────────────────────────────────────
        for i in range(num_target_genes):
            st.markdown(
                f"<h4 style='margin-top:14px;margin-bottom:4px;color:#004d40;'>"
                f"🧬 {_t['target_gene']} {i+1}</h4>", unsafe_allow_html=True
            )

            # ── Side-by-side input row: Control | Group 1 | Group 2 | ... ────────
            # (Laid out this way so users can paste a Control column and Patient
            # column(s) directly next to each other, matching how the data is
            # usually laid out in a spreadsheet.)
            _n_entry_cols = 1 + num_patient_groups
            _entry_cols = st.columns(_n_entry_cols)

            with _entry_cols[0]:
                st.markdown(f"**{_t['control_group']}**")
                ctrl_target_pos_txt = _ta(
                    f"Control {i+1} — {_t['positive_partitions']} ({_t['target_gene']})", f"ctrl_tgt_pos_{i}"
                )
                ctrl_target_tot_txt = _ta(
                    f"Control {i+1} — {_t['total_partitions']} ({_t['target_gene']})", f"ctrl_tgt_tot_{i}"
                )
                ctrl_dilution_factor = render_dilution_input(f"ctrl_{i}")
                ctrl_ref_pos_txts, ctrl_ref_tot_txts = [], []
                for r in range(num_ref_genes):
                    ref_lbl = f"{_t['reference_gene']} {r+1}" if num_ref_genes > 1 else _t['reference_gene']
                    rp = _ta(f"Control {i+1} — {_t['positive_partitions']} ({ref_lbl})", f"ctrl_ref_pos_{i}_{r}")
                    rt = _ta(f"Control {i+1} — {_t['total_partitions']} ({ref_lbl})", f"ctrl_ref_tot_{i}_{r}")
                    ctrl_ref_pos_txts.append(rp)
                    ctrl_ref_tot_txts.append(rt)

            smp_target_pos_txts, smp_target_tot_txts = [], []
            smp_dilution_factors = []
            smp_ref_pos_txts_all, smp_ref_tot_txts_all = [], []
            for j in range(num_patient_groups):
                with _entry_cols[j + 1]:
                    st.markdown(f"**{_t['patient_group']} {j+1}**")
                    _pos_txt = _ta(
                        f"Group {j+1} — {_t['positive_partitions']} ({_t['target_gene']} {i+1})", f"smp_tgt_pos_{i}_{j}"
                    )
                    _tot_txt = _ta(
                        f"Group {j+1} — {_t['total_partitions']} ({_t['target_gene']} {i+1})", f"smp_tgt_tot_{i}_{j}"
                    )
                    _dil = render_dilution_input(f"smp_{i}_{j}")
                    _ref_pos_list, _ref_tot_list = [], []
                    for r in range(num_ref_genes):
                        ref_lbl = f"{_t['reference_gene']} {r+1}" if num_ref_genes > 1 else _t['reference_gene']
                        rp = _ta(f"Group {j+1} — {_t['positive_partitions']} ({ref_lbl})", f"smp_ref_pos_{i}_{j}_{r}")
                        rt = _ta(f"Group {j+1} — {_t['total_partitions']} ({ref_lbl})", f"smp_ref_tot_{i}_{j}_{r}")
                        _ref_pos_list.append(rp)
                        _ref_tot_list.append(rt)
                    smp_target_pos_txts.append(_pos_txt)
                    smp_target_tot_txts.append(_tot_txt)
                    smp_dilution_factors.append(_dil)
                    smp_ref_pos_txts_all.append(_ref_pos_list)
                    smp_ref_tot_txts_all.append(_ref_tot_list)

            st.divider()

            # ── Control group processing (computation, full width) ───────────────
            ctrl_result = sync_and_compute(
                ctrl_target_pos_txt, ctrl_target_tot_txt, ctrl_ref_pos_txts, ctrl_ref_tot_txts,
                f"{_t['control_group']} {i+1}", f"ctrl_{i}"
            )
            if ctrl_result is None:
                st.error(_t['warning_field_empty'])
                continue
            if len(ctrl_result["kept"]) == 0:
                show_exclusion_diagnosis(ctrl_result)
                continue

            # ── geNorm-style reference stability (control) ──────────────────────
            if num_ref_genes >= 2 and ctrl_result["ref_matrix_kept"].shape[1] >= 2:
                m_vals = compute_stability_m(ctrl_result["ref_matrix_kept"])
                st.markdown(f"##### 📊 {_t['genorm_title']} — {_t['control_group']} {i+1}")
                stab_cols = st.columns(num_ref_genes)
                for r, col in enumerate(stab_cols):
                    with col:
                        st.metric(label=f"{_t['reference_gene']} {r+1}", value=f"M = {m_vals[r]:.3f}")
                        if m_vals[r] < 0.5:
                            st.caption(f"✅ {_t['stable']}")
                        elif m_vals[r] < 1.0:
                            st.caption(f"⚠️ {_t['borderline']}")
                        else:
                            st.caption(f"❌ {_t['unstable']}")

            # ── Store control replicate rows in input_values_table ───────────────
            for row in ctrl_result["detail_rows"]:
                status_label = (
                    f"{_t['outlier_excluded_yes']} ({outlier_method})" if row["excluded"]
                    else (_t['qc_fail'] if row["qc_fail"] else
                          (_t['qc_saturated'] if row["target_status"] == "saturated" else _t['outlier_excluded_no']))
                )
                input_values_table.append({
                    "__gene__": f"Gene {i+1}", "Grup": "Control",
                    "__replicate__": row["replicate"],
                    "__positive__": row["target_pos"], "__total__": row["target_tot"],
                    "__lambda__": round(row["target_lambda"], 5) if not np.isnan(row["target_lambda"]) else None,
                    "__conc__": round(row["target_lambda"] / partition_vol_nl * 1000.0, 2) if not np.isnan(row["target_lambda"]) else None,
                    "Outlier Excluded": status_label,
                    "__used__": row["used"],
                })

            avg_ctrl_ratio = float(np.mean(ctrl_result["ratio_kept"])) if len(ctrl_result["ratio_kept"]) > 0 else None

            # ── NTC / LOD-LOQ (optional, advanced) ────────────────────────────────
            lod_loq_result = None
            if advanced_mode:
                with st.expander(_t['ntc_expander'], expanded=False):
                    st.caption(_t['ntc_description'])
                    ntc_c1, ntc_c2 = st.columns(2)
                    with ntc_c1:
                        ntc_pos_txt = _ta(f"{_t['ntc_positive_label']} — {_t['target_gene']} {i+1}", f"ntc_pos_{i}")
                    with ntc_c2:
                        ntc_tot_txt = _ta(f"{_t['ntc_total_label']} — {_t['target_gene']} {i+1}", f"ntc_tot_{i}")

                    ntc_pos_arr = parse_input_data(ntc_pos_txt)
                    ntc_tot_arr = parse_input_data(ntc_tot_txt)
                    if len(ntc_pos_arr) > 0 and len(ntc_tot_arr) > 0:
                        lod_loq_result = compute_lod_loq(ntc_pos_arr, ntc_tot_arr, partition_vol_nl)
                        if lod_loq_result is not None:
                            _gene_label_lod = f"{_t['target_gene']} {i+1}"
                            st.markdown(f"**{_t['lod_result_title'].format(gene=_gene_label_lod)}**")
                            lod_c1, lod_c2 = st.columns(2)
                            lod_c1.metric(_t['lod_label'], f"{lod_loq_result['lod_conc']:.2f}")
                            lod_c2.metric(_t['loq_label'], f"{lod_loq_result['loq_conc']:.2f}")
                            if lod_loq_result["contamination"]:
                                st.warning(_t['ntc_contamination_warning'].format(
                                    lam=lod_loq_result["ntc_lambda"],
                                    pos=int(lod_loq_result["ntc_pos_total"]), tot=int(lod_loq_result["ntc_tot_total"])
                                ))
                            else:
                                st.info(_t['ntc_zero_note'].format(n=int(lod_loq_result["ntc_tot_total"])))
                            st.caption(_t['loq_heuristic_note'])

            # ── Patient groups (computation + results, full width) ───────────────
            for j in range(num_patient_groups):
                smp_target_pos_txt = smp_target_pos_txts[j]
                smp_target_tot_txt = smp_target_tot_txts[j]
                smp_dilution_factor = smp_dilution_factors[j]
                smp_ref_pos_txts = smp_ref_pos_txts_all[j]
                smp_ref_tot_txts = smp_ref_tot_txts_all[j]

                st.markdown(f"**{_t['patient_group']} {j+1} — {_t['target_gene']} {i+1}**")

                smp_result = sync_and_compute(
                    smp_target_pos_txt, smp_target_tot_txt, smp_ref_pos_txts, smp_ref_tot_txts,
                    f"{_t['patient_group']} {j+1} — Gene {i+1}", f"smp_{i}_{j}"
                )
                if smp_result is None:
                    st.error(_t['warning_field_empty'])
                    continue
                if len(smp_result["kept"]) == 0:
                    show_exclusion_diagnosis(smp_result)
                    continue

                if num_ref_genes >= 2 and smp_result["ref_matrix_kept"].shape[1] >= 2:
                    m_vals_s = compute_stability_m(smp_result["ref_matrix_kept"])
                    st.markdown(f"##### 📊 {_t['genorm_title']} — {_t['patient_group']} {j+1}")
                    stab_cols_s = st.columns(num_ref_genes)
                    for r, col in enumerate(stab_cols_s):
                        with col:
                            st.metric(label=f"{_t['reference_gene']} {r+1}", value=f"M = {m_vals_s[r]:.3f}")
                            if m_vals_s[r] < 0.5:
                                st.caption(f"✅ {_t['stable']}")
                            elif m_vals_s[r] < 1.0:
                                st.caption(f"⚠️ {_t['borderline']}")
                            else:
                                st.caption(f"❌ {_t['unstable']}")

                for row in smp_result["detail_rows"]:
                    status_label = (
                        f"{_t['outlier_excluded_yes']} ({outlier_method})" if row["excluded"]
                        else (_t['qc_fail'] if row["qc_fail"] else
                              (_t['qc_saturated'] if row["target_status"] == "saturated" else _t['outlier_excluded_no']))
                    )
                    input_values_table.append({
                        "__gene__": f"Gene {i+1}", "Grup": f"Group {j+1}",
                        "__replicate__": row["replicate"],
                        "__positive__": row["target_pos"], "__total__": row["target_tot"],
                        "__lambda__": round(row["target_lambda"], 5) if not np.isnan(row["target_lambda"]) else None,
                        "__conc__": round(row["target_lambda"] / partition_vol_nl * 1000.0, 2) if not np.isnan(row["target_lambda"]) else None,
                        "Outlier Excluded": status_label,
                        "__used__": row["used"],
                    })

                avg_smp_ratio = float(np.mean(smp_result["ratio_kept"]))
                fold_change = avg_smp_ratio / avg_ctrl_ratio if avg_ctrl_ratio else float("nan")
                cn_ctrl = ploidy * avg_ctrl_ratio if avg_ctrl_ratio else float("nan")
                cn_smp = ploidy * avg_smp_ratio

                # ── Concentration (copies/µL) with 95% CI (replicate-based) ─────────
                conc_ctrl_arr = lambda_to_conc(ctrl_result["lam_target_kept"], partition_vol_nl)
                conc_smp_arr = lambda_to_conc(smp_result["lam_target_kept"], partition_vol_nl)
                conc_ctrl_mean, conc_ctrl_lo, conc_ctrl_hi, conc_ctrl_cv, _ = mean_ci(conc_ctrl_arr)
                conc_smp_mean, conc_smp_lo, conc_smp_hi, conc_smp_cv, _ = mean_ci(conc_smp_arr)

                # ── Stock/original concentration (back-calculated using dilution factor) ──
                stock_conc_ctrl_arr = conc_ctrl_arr * ctrl_dilution_factor
                stock_conc_smp_arr = conc_smp_arr * smp_dilution_factor
                stock_conc_ctrl_mean, stock_conc_ctrl_lo, stock_conc_ctrl_hi, _, _ = mean_ci(stock_conc_ctrl_arr)
                stock_conc_smp_mean, stock_conc_smp_lo, stock_conc_smp_hi, _, _ = mean_ci(stock_conc_smp_arr)

                # ── Dynamic range QC (based on mean target lambda, independent of dilution) ──
                _mean_lam_smp = float(np.mean(smp_result["lam_target_kept"]))
                if _mean_lam_smp < 0.05:
                    dynamic_range_flag = _t['dynamic_range_low']
                elif _mean_lam_smp > 4.0:
                    dynamic_range_flag = _t['dynamic_range_saturated']
                elif _mean_lam_smp > 3.0:
                    dynamic_range_flag = _t['dynamic_range_high']
                else:
                    dynamic_range_flag = _t['dynamic_range_ok']

                # ── Theoretical Poisson-only relative uncertainty (for MU budget) ────
                # Inverse-variance pooling of each kept replicate's Poisson SE (as a
                # relative %), giving a pure counting-statistics uncertainty estimate
                # that is free of pipetting/biological replicate-to-replicate scatter
                # (unlike __conc_smp_cv__, which reflects total observed variability).
                _pos_kept = smp_result.get("pos_target_kept", np.array([]))
                _tot_kept = smp_result.get("tot_target_kept", np.array([]))
                if len(_pos_kept) > 0:
                    _rel_vars = []
                    for _pp, _tt, _ll in zip(_pos_kept, _tot_kept, smp_result["lam_target_kept"]):
                        if _ll > 0:
                            _se_i = poisson_se(_pp, _tt)
                            if not np.isnan(_se_i):
                                _rel_vars.append((_se_i / _ll) ** 2)
                    if _rel_vars:
                        _poisson_rel_se_pct = np.sqrt(1.0 / np.sum(1.0 / np.array(_rel_vars))) * 100
                    else:
                        _poisson_rel_se_pct = np.nan
                else:
                    _poisson_rel_se_pct = np.nan

                # ── LOD/LOQ QC flag (if NTC data provided for this gene) ─────────────
                lod_loq_flag = "—"
                if lod_loq_result is not None:
                    if conc_smp_mean < lod_loq_result["lod_conc"]:
                        lod_loq_flag = _t['below_lod_flag']
                    elif conc_smp_mean < lod_loq_result["loq_conc"]:
                        lod_loq_flag = _t['between_lod_loq_flag']
                    else:
                        lod_loq_flag = _t['above_loq_flag']

                # ── Empirical LOQ cross-check (replicate-reproducibility-based) ──────
                # The LOD/LOQ above rests on the 3x-LOD heuristic, which is a common
                # analytical-chemistry convention but not independently validated for
                # Poisson-based dPCR. As a cross-check, flag results whose *observed*
                # replicate-to-replicate CV% exceeds a reproducibility threshold
                # (commonly ~20-25% in bioanalytical method validation guidance,
                # e.g. FDA/EMA), independent of the LOD/LOQ calculation entirely.
                empirical_loq_flag = "—"
                if not np.isnan(conc_smp_cv):
                    empirical_loq_flag = (_t['empirical_loq_fail'] if conc_smp_cv > 25
                                           else _t['empirical_loq_pass'])

                if fold_change >= 1.5:
                    regulation = _t['upregulated']
                elif fold_change <= 0.67:
                    regulation = _t['downregulated']
                else:
                    regulation = _t['no_change']

                st.markdown(f"#### {_t['method_comparison']} — {_t['target_gene']} {i+1} / {_t['patient_group']} {j+1}")
                rcol1, rcol2, rcol3, rcol4 = st.columns(4)
                rcol1.metric(_t['ratio_col'], f"{avg_smp_ratio:.4f}")
                rcol2.metric(_t['cn_col'], f"{cn_smp:.3f}")
                rcol3.metric(_t['fc_col'], f"{fold_change:.4f}", delta=regulation)
                _conc_ci_txt = (f"95% CI: {conc_smp_lo:.1f}–{conc_smp_hi:.1f}"
                                 if not np.isnan(conc_smp_lo) else "n<2, no CI")
                rcol4.metric(_t['conc_col'], f"{conc_smp_mean:.1f}", delta=_conc_ci_txt, delta_color="off")
                if smp_dilution_factor != 1.0:
                    st.caption(f"{_t['stock_conc_col']}: {stock_conc_smp_mean:.1f} "
                               f"(95% CI: {stock_conc_smp_lo:.1f}–{stock_conc_smp_hi:.1f}, "
                               f"×{smp_dilution_factor:.0f} {_t['dilution_factor_field_label']})")
                if lod_loq_result is not None:
                    st.caption(f"{_t['lod_qc_col']}: {lod_loq_flag}")
                if not np.isnan(conc_smp_cv):
                    st.caption(f"{_t['empirical_loq_label']}: {empirical_loq_flag} (CV={conc_smp_cv:.1f}%)")
                st.caption(f"{_t['dynamic_range_warning_label']}: {dynamic_range_flag}")

                # ── Statistics (Control vs this patient group), directly on Ratio ──
                control_ratios = ctrl_result["ratio_kept"]
                sample_ratios = smp_result["ratio_kept"]
                n_ctrl, n_smp = len(control_ratios), len(sample_ratios)

                if n_ctrl < 2 or n_smp < 2:
                    test_pvalue, test_method, test_type, significance = float('nan'), "N/A (n < 2)", "—", "—"
                    control_normal = sample_normal = True
                    equal_variance = True
                    shapiro_control = shapiro_sample = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                    levene_test = type('LV', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                else:
                    _MIN_N_SHAPIRO = 8
                    if n_ctrl >= _MIN_N_SHAPIRO and n_smp >= _MIN_N_SHAPIRO:
                        shapiro_control = stats.shapiro(control_ratios)
                        shapiro_sample = stats.shapiro(sample_ratios)
                        control_normal = shapiro_control.pvalue > 0.05
                        sample_normal = shapiro_sample.pvalue > 0.05
                    else:
                        shapiro_control = shapiro_sample = type('SW', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                        control_normal = sample_normal = True

                    try:
                        levene_test = stats.levene(control_ratios, sample_ratios)
                        equal_variance = (levene_test.pvalue > 0.05) if not np.isnan(levene_test.pvalue) else True
                    except Exception:
                        levene_test = type('LV', (), {'statistic': float('nan'), 'pvalue': float('nan')})()
                        equal_variance = True

                    try:
                        if control_normal and sample_normal:
                            if equal_variance:
                                test_pvalue = stats.ttest_ind(control_ratios, sample_ratios).pvalue
                                test_method = _t['t_test']
                            else:
                                test_pvalue = stats.ttest_ind(control_ratios, sample_ratios, equal_var=False).pvalue
                                test_method = _t['welch_t_test']
                            test_type = _t['parametric']
                        else:
                            test_pvalue = stats.mannwhitneyu(control_ratios, sample_ratios, alternative='two-sided').pvalue
                            test_method = _t['mann_whitney_u_test']
                            test_type = _t['non_parametric']
                    except Exception:
                        test_pvalue, test_method, test_type = float('nan'), "Error", "—"

                    significance = (_t['significant'] if (not np.isnan(test_pvalue) and test_pvalue < 0.05)
                                     else (_t['insignificant'] if not np.isnan(test_pvalue) else "—"))

                with st.expander(f"{_t['stat_decision_title']} — {_t['target_gene']} {i+1} / Group {j+1}", expanded=False):
                    st.markdown(_t['stat_decision_steps'])
                    if n_ctrl >= 8 and n_smp >= 8:
                        sw_ctrl_sym = "✅" if control_normal else "❌"
                        sw_smp_sym = "✅" if sample_normal else "❌"
                        st.markdown(
                            f"{_t['stat_shapiro_title']}  \n"
                            f"- Control: W={shapiro_control.statistic:.4f}, p={shapiro_control.pvalue:.4f} {sw_ctrl_sym} "
                            f"{_t['stat_normal'] if control_normal else _t['stat_nonnormal']}  \n"
                            f"- Group {j+1}: W={shapiro_sample.statistic:.4f}, p={shapiro_sample.pvalue:.4f} {sw_smp_sym} "
                            f"{_t['stat_normal'] if sample_normal else _t['stat_nonnormal']}"
                        )
                    else:
                        st.info(f"ℹ️ Shapiro-Wilk skipped (n={min(n_ctrl, n_smp)} < 8). Normality assumed.")

                    if control_normal and sample_normal:
                        lev_sym = "✅" if equal_variance else "⚠️"
                        st.markdown(
                            f"{_t['stat_levene_title']}  \n"
                            f"- F={levene_test.statistic:.4f}, p={levene_test.pvalue:.4f} {lev_sym} "
                            f"{_t['stat_equal_var'] if equal_variance else _t['stat_unequal_var']}"
                        )
                    else:
                        st.markdown(_t['stat_levene_skipped'])

                    if not control_normal or not sample_normal:
                        reason = _t['stat_reason_nonnormal']
                    elif equal_variance:
                        reason = _t['stat_reason_normal_equal']
                    else:
                        reason = _t['stat_reason_normal_unequal']

                    st.success(
                        f"{_t['stat_selected_test']} {test_method}  \n"
                        f"{_t['stat_reason']} {reason}  \n"
                        f"{_t['stat_result']} p = {test_pvalue:.4f} → **{significance}**"
                    )
                    if num_patient_groups >= 2:
                        st.caption(_t['stat_multigroup_note'])

                stats_data.append({
                    "__gene__": f"Gene {i+1}", "__group__": f"Group {j+1}",
                    "__test_type__": test_type, "__test_method__": test_method,
                    "__pvalue__": test_pvalue, "__significance__": significance,
                    "Comparison": f"Control vs Group {j+1}",
                    "__ratio_ctrl__": control_ratios, "__ratio_smp__": sample_ratios,
                })

                data.append({
                    "__gene__": f"Gene {i+1}", "__group__": f"Group {j+1}",
                    "__ratio_ctrl__": avg_ctrl_ratio, "__ratio_smp__": avg_smp_ratio,
                    "__cn_ctrl__": cn_ctrl, "__cn_smp__": cn_smp,
                    "__fc__": fold_change, "__regulation__": regulation,
                    "__n_ctrl__": n_ctrl, "__n_smp__": n_smp,
                    "__conc_ctrl__": conc_ctrl_mean, "__conc_ctrl_lo__": conc_ctrl_lo, "__conc_ctrl_hi__": conc_ctrl_hi,
                    "__conc_smp__": conc_smp_mean, "__conc_smp_lo__": conc_smp_lo, "__conc_smp_hi__": conc_smp_hi,
                    "__conc_smp_cv__": conc_smp_cv, "__lod_loq_flag__": lod_loq_flag,
                    "__dilution_factor__": smp_dilution_factor,
                    "__stock_conc_smp__": stock_conc_smp_mean, "__stock_conc_smp_lo__": stock_conc_smp_lo,
                    "__stock_conc_smp_hi__": stock_conc_smp_hi, "__dynamic_range_flag__": dynamic_range_flag,
                    "__lambda_smp_mean__": _mean_lam_smp, "__poisson_rel_se_pct__": _poisson_rel_se_pct,
                    "__empirical_loq_flag__": empirical_loq_flag,
                })


    # ═══════════════════════════════════════════════════════════════════════════════
    # MULTI-GROUP ANALYSIS (≥3 groups per gene) — omnibus + post-hoc + correction
    # ═══════════════════════════════════════════════════════════════════════════════
    multigroup_results = []

    for i in range(num_target_genes):
        gene_label = f"Gene {i+1}"
        gene_rows = [r for r in stats_data if r["__gene__"] == gene_label]
        if not gene_rows:
            continue

        ctrl_ratios = gene_rows[0]["__ratio_ctrl__"]
        group_ratio_map = {r["__group__"]: r["__ratio_smp__"] for r in gene_rows}

        all_groups = [list(ctrl_ratios)] + [list(v) for v in group_ratio_map.values()]
        all_group_names = ["Control"] + list(group_ratio_map.keys())
        n_groups = len(all_groups)

        if n_groups < 3:
            multigroup_results.append({
                "gene": gene_label, "n_groups": n_groups, "note": "2-group",
                "omnibus_test": "—", "omnibus_p": None, "posthoc_rows": []
            })
            continue

        normality_ok = all((len(g) < 8 or stats.shapiro(g).pvalue > 0.05) for g in all_groups if len(g) >= 3)
        levene_p = stats.levene(*all_groups).pvalue if n_groups >= 2 else 1.0
        variance_ok = levene_p > 0.05

        if normality_ok and variance_ok:
            omnibus_stat, omnibus_p = stats.f_oneway(*all_groups)
            omnibus_test, omnibus_type, posthoc_method = "One-way ANOVA", "parametric", "Tukey HSD"
        elif normality_ok and not variance_ok:
            try:
                from scipy.stats import alexandergovern
                result = alexandergovern(*all_groups)
                omnibus_p, omnibus_stat = result.pvalue, result.statistic
            except Exception:
                omnibus_stat, omnibus_p = stats.f_oneway(*all_groups)
            omnibus_test, omnibus_type, posthoc_method = "Welch ANOVA", "parametric", "Games-Howell (approx.)"
        else:
            omnibus_stat, omnibus_p = stats.kruskal(*all_groups)
            omnibus_test, omnibus_type, posthoc_method = "Kruskal-Wallis", "non-parametric", "Dunn (Mann-Whitney U)"

        omnibus_sig = _t['multigroup_significant'] if omnibus_p < 0.05 else _t['multigroup_not_significant']

        pairs, raw_pvals = [], []
        for a in range(n_groups):
            for b in range(a + 1, n_groups):
                g_a, g_b = all_groups[a], all_groups[b]
                if omnibus_type == "parametric" and variance_ok:
                    p = stats.ttest_ind(g_a, g_b).pvalue
                elif omnibus_type == "parametric" and not variance_ok:
                    p = stats.ttest_ind(g_a, g_b, equal_var=False).pvalue
                else:
                    p = stats.mannwhitneyu(g_a, g_b, alternative="two-sided").pvalue
                pairs.append((all_group_names[a], all_group_names[b]))
                raw_pvals.append(p)

        n_tests = len(raw_pvals)
        bonf_pvals = [min(p * n_tests, 1.0) for p in raw_pvals]
        ranked = sorted(range(n_tests), key=lambda k: raw_pvals[k])
        fdr_pvals = [1.0] * n_tests
        for rank, idx in enumerate(ranked):
            fdr_pvals[idx] = min(raw_pvals[idx] * n_tests / (rank + 1), 1.0)
        for k in range(n_tests - 2, -1, -1):
            fdr_pvals[ranked[k]] = min(fdr_pvals[ranked[k]], fdr_pvals[ranked[k + 1]])

        posthoc_rows = []
        for idx, (pa, pb) in enumerate(pairs):
            posthoc_rows.append({
                "Comparison": f"{pa} vs {pb}", "Raw p": round(raw_pvals[idx], 4),
                "Bonferroni p": round(bonf_pvals[idx], 4), "FDR p (B-H)": round(fdr_pvals[idx], 4),
                "Sig (raw)": "✅" if raw_pvals[idx] < 0.05 else "—",
                "Sig (Bonferroni)": "✅" if bonf_pvals[idx] < 0.05 else "—",
                "Sig (FDR)": "✅" if fdr_pvals[idx] < 0.05 else "—",
            })

        multigroup_results.append({
            "gene": gene_label, "n_groups": n_groups,
            "omnibus_test": omnibus_test, "omnibus_type": omnibus_type, "omnibus_p": omnibus_p,
            "omnibus_sig": omnibus_sig, "posthoc_method": posthoc_method, "posthoc_rows": posthoc_rows,
            "normality_ok": normality_ok, "variance_ok": variance_ok, "note": None
        })

    # ═══════════════════════════════════════════════════════════════════════════════
    # RESULTS TAB
    # ═══════════════════════════════════════════════════════════════════════════════
    with tab_results:

        # ── QC Summary Panel ──────────────────────────────────────────────────────
        if input_values_table:
            _n_total_rep = len(input_values_table)
            _n_qc_fail = sum(1 for r in input_values_table if r["Outlier Excluded"] == _t['qc_fail'])
            _n_saturated = sum(1 for r in input_values_table if r["Outlier Excluded"] == _t['qc_saturated'])
            _n_outlier = sum(1 for r in input_values_table if str(r["Outlier Excluded"]).startswith(_t['outlier_excluded_yes']))
            _n_below_lod = sum(1 for r in data if r.get("__lod_loq_flag__") in (_t['below_lod_flag'], _t['between_lod_loq_flag']))
            _n_high_cv = sum(1 for r in data if r.get("__conc_smp_cv__") is not None
                              and not np.isnan(r["__conc_smp_cv__"]) and r["__conc_smp_cv__"] > 25)
            _n_dynamic_range = sum(1 for r in data if r.get("__dynamic_range_flag__", _t['dynamic_range_ok']) != _t['dynamic_range_ok'])

            with st.container(border=True):
                st.markdown(f"#### {_t['qc_panel_title']}")
                qc_c1, qc_c2, qc_c3, qc_c4, qc_c5, qc_c6, qc_c7 = st.columns(7)
                qc_c1.metric(_t['qc_panel_total'], _n_total_rep)
                qc_c2.metric(_t['qc_panel_qc_fail'], _n_qc_fail)
                qc_c3.metric(_t['qc_panel_saturated'], _n_saturated)
                qc_c4.metric(_t['qc_panel_outlier'], _n_outlier)
                qc_c5.metric(_t['qc_panel_below_lod'], _n_below_lod)
                qc_c6.metric(_t['qc_panel_high_cv'], _n_high_cv)
                qc_c7.metric(_t['dynamic_range_qc_panel_label'], _n_dynamic_range)

                _n_flags_total = _n_qc_fail + _n_saturated + _n_outlier + _n_below_lod + _n_high_cv + _n_dynamic_range
                _flag_ratio = _n_flags_total / max(_n_total_rep, 1)
                if _flag_ratio == 0:
                    st.success(_t['qc_panel_verdict_good'])
                elif _flag_ratio < 0.15:
                    st.warning(_t['qc_panel_verdict_caution'])
                else:
                    st.error(_t['qc_panel_verdict_poor'])
            st.markdown("---")
        else:
            st.info(_t['qc_panel_no_data'])

        # ── Multi-group display ───────────────────────────────────────────────────
        if any(r["n_groups"] >= 3 for r in multigroup_results):
            st.markdown("---")
            st.markdown(_t['multigroup_title'])
            for res in multigroup_results:
                if res["n_groups"] < 3:
                    continue
                st.markdown(f"### 🧬 {res['gene']} — {res['n_groups']} {_t['patient_group'].replace('🩸 ', '')}")
                if res["normality_ok"] and res["variance_ok"]:
                    st.success(_t['multigroup_decision_normal_equal'])
                elif res["normality_ok"] and not res["variance_ok"]:
                    st.warning(_t['multigroup_decision_normal_unequal'])
                else:
                    st.warning(_t['multigroup_decision_nonnormal'])

                omni_col1, omni_col2, omni_col3 = st.columns(3)
                omni_col1.metric(_t['multigroup_omnibus_test'], res["omnibus_test"])
                omni_col2.metric(_t['multigroup_pvalue'], f"{res['omnibus_p']:.4f}")
                omni_col3.metric(_t['multigroup_result'], res["omnibus_sig"])
                if res["omnibus_p"] >= 0.05:
                    st.info(_t['multigroup_omnibus_ns'])

                st.markdown(f"{_t['multigroup_posthoc_label']} {res['posthoc_method']}")
                ph_df = pd.DataFrame(res["posthoc_rows"])
                st.dataframe(ph_df, use_container_width=True)

                fig_ph = go.Figure()
                comparisons = [r["Comparison"] for r in res["posthoc_rows"]]
                fig_ph.add_trace(go.Bar(name="Raw p", x=comparisons, y=[r["Raw p"] for r in res["posthoc_rows"]], marker_color="#00796b"))
                fig_ph.add_trace(go.Bar(name="Bonferroni p", x=comparisons, y=[r["Bonferroni p"] for r in res["posthoc_rows"]], marker_color="#f57c00"))
                fig_ph.add_trace(go.Bar(name="FDR p (B-H)", x=comparisons, y=[r["FDR p (B-H)"] for r in res["posthoc_rows"]], marker_color="#455a64"))
                fig_ph.add_hline(y=0.05, line_dash="dash", line_color="red", annotation_text="α = 0.05")
                fig_ph.update_layout(barmode="group", title=f"{res['gene']} — Post-hoc p-values", yaxis_title="p-value", height=350)
                st.plotly_chart(fig_ph, use_container_width=True, key=f"posthoc_{res['gene']}")

                ph_csv = ph_df.to_csv(index=False).encode("utf-8")
                st.download_button(f"{_t['multigroup_dl_button']} {res['gene']}", data=ph_csv,
                                    file_name=f"posthoc_{res['gene'].replace(' ', '_')}.csv", mime="text/csv",
                                    key=f"ph_dl_{res['gene']}")
        elif num_patient_groups >= 2 and multigroup_results:
            st.markdown("---")
            st.info(_t['multigroup_2group_note'])

        # ── Input data table ──────────────────────────────────────────────────────
        if input_values_table:
            st.subheader(_t['gr_tbl'])
            _rename = {
                "__gene__": _t['target_gene'], "Grup": "Group", "__replicate__": _t['sample_number'],
                "__positive__": _t['positive_partitions'], "__total__": _t['total_partitions'],
                "__lambda__": _t['lambda_col'], "__conc__": _t['conc_col'], "Outlier Excluded": "Status",
            }
            display_df = pd.DataFrame(input_values_table).drop(columns=["__used__"]).rename(columns=_rename)
            st.dataframe(display_df, use_container_width=True)
            csv = display_df.to_csv(index=False).encode("utf-8")
            st.download_button(_t['download_csv'], data=csv, file_name="dpcr_input_data.csv", mime="text/csv", key="dl_input_csv")

        # ── Results summary table ─────────────────────────────────────────────────
        if data:
            st.subheader(_t['nil_mine'])
            _res_display_rows = []
            for r in data:
                _res_display_rows.append({
                    _t['target_gene']: r["__gene__"], "Group": r["__group__"],
                    f"{_t['ratio_col']} (Control)": r["__ratio_ctrl__"], f"{_t['ratio_col']} (Sample)": r["__ratio_smp__"],
                    f"{_t['cn_col']} (Control)": r["__cn_ctrl__"], f"{_t['cn_col']} (Sample)": r["__cn_smp__"],
                    f"{_t['conc_col']} (Control)": (f"{r['__conc_ctrl__']:.1f}" if not np.isnan(r['__conc_ctrl__']) else "—"),
                    f"{_t['conc_col']} (Sample)": (f"{r['__conc_smp__']:.1f}" if not np.isnan(r['__conc_smp__']) else "—"),
                    "95% CI (Sample, copies/µL)": (
                        f"{r['__conc_smp_lo__']:.1f}–{r['__conc_smp_hi__']:.1f}"
                        if not np.isnan(r['__conc_smp_lo__']) else "n<2"
                    ),
                    "CV% (Sample)": (f"{r['__conc_smp_cv__']:.1f}" if not np.isnan(r['__conc_smp_cv__']) else "—"),
                    _t['lod_qc_col']: r.get("__lod_loq_flag__", "—"),
                    "Dilution Factor": r.get("__dilution_factor__", 1.0),
                    f"{_t['stock_conc_col']} (Sample)": (
                        f"{r['__stock_conc_smp__']:.1f}" if r.get("__stock_conc_smp__") is not None
                        and not np.isnan(r["__stock_conc_smp__"]) else "—"
                    ),
                    _t['dynamic_range_warning_label']: r.get("__dynamic_range_flag__", "—"),
                    _t['fc_col']: r["__fc__"], _t['regulation_status']: r["__regulation__"],
                    "n (Control)": r["__n_ctrl__"], "n (Sample)": r["__n_smp__"],
                })
            res_df = pd.DataFrame(_res_display_rows)
            st.dataframe(res_df, use_container_width=True)
            csv2 = res_df.to_csv(index=False).encode("utf-8")
            st.download_button(_t['download_csv'], data=csv2, file_name="dpcr_results.csv", mime="text/csv", key="dl_res_csv")

        # ── Statistics table ───────────────────────────────────────────────────────
        if stats_data:
            st.subheader(_t['statistical_results'])
            _rename3 = {
                "__gene__": _t['target_gene'], "__group__": "Group", "__test_type__": _t['test_type'],
                "__test_method__": _t['test_method'], "__pvalue__": _t['test_pvalue'], "__significance__": _t['significance'],
            }
            stats_df = pd.DataFrame(stats_data).drop(columns=["__ratio_ctrl__", "__ratio_smp__"]).rename(columns=_rename3)
            st.dataframe(stats_df, use_container_width=True)
            csv3 = stats_df.to_csv(index=False).encode("utf-8")
            st.download_button(_t['download_csv'], data=csv3, file_name="dpcr_statistics.csv", mime="text/csv", key="dl_stats_csv")

        # ── Multi-gene p-value correction ─────────────────────────────────────────
        if stats_data and num_target_genes >= 2:
            st.markdown("---")
            st.markdown(_t['multigene_title'])
            correction_rows = [
                {"Gene": r["__gene__"], "Group": r["__group__"], "Raw p": r["__pvalue__"], "Test": r["__test_method__"]}
                for r in stats_data if r.get("__pvalue__") is not None and not np.isnan(r["__pvalue__"])
            ]
            if correction_rows:
                n_tests = len(correction_rows)
                raw_pvals = [r["Raw p"] for r in correction_rows]
                bonf = [min(p * n_tests, 1.0) for p in raw_pvals]
                ranked = sorted(range(n_tests), key=lambda k: raw_pvals[k])
                fdr = [1.0] * n_tests
                for rank, idx in enumerate(ranked):
                    fdr[idx] = min(raw_pvals[idx] * n_tests / (rank + 1), 1.0)
                for k in range(n_tests - 2, -1, -1):
                    fdr[ranked[k]] = min(fdr[ranked[k]], fdr[ranked[k + 1]])
                for idx, row in enumerate(correction_rows):
                    row["Bonferroni p"] = round(bonf[idx], 4)
                    row["FDR p (B-H)"] = round(fdr[idx], 4)
                    row["Sig (raw)"] = "✅" if raw_pvals[idx] < 0.05 else "—"
                    row["Sig (Bonferroni)"] = "✅" if bonf[idx] < 0.05 else "—"
                    row["Sig (FDR)"] = "✅" if fdr[idx] < 0.05 else "—"

                corr_df = pd.DataFrame(correction_rows)
                st.dataframe(corr_df, use_container_width=True)

                n_raw_sig = sum(1 for p in raw_pvals if p < 0.05)
                n_fdr_sig = sum(1 for p in fdr if p < 0.05)
                sc1, sc2, sc3 = st.columns(3)
                sc1.metric(_t['multigene_sig_raw'], f"{n_raw_sig} / {n_tests}")
                sc2.metric(_t['multigene_sig_bonf'], f"{sum(1 for p in bonf if p < 0.05)} / {n_tests}")
                sc3.metric(_t['multigene_sig_fdr'], f"{n_fdr_sig} / {n_tests}")

                if n_raw_sig > n_fdr_sig:
                    st.warning(_t['multigene_warning'].format(lost=n_raw_sig - n_fdr_sig))
                elif n_raw_sig == n_fdr_sig and n_raw_sig > 0:
                    st.success(_t['multigene_success'].format(n=n_raw_sig))
                else:
                    st.info(_t['multigene_no_sig'])

                corr_csv = corr_df.to_csv(index=False).encode("utf-8")
                st.download_button(_t['multigene_dl_button'], data=corr_csv, file_name="multi_gene_correction.csv",
                                    mime="text/csv", key="multigene_corr_dl")
        elif stats_data and num_target_genes == 1:
            st.markdown("---")
            st.info(_t['multigene_1gene_note'])

        # ── Distribution plots ────────────────────────────────────────────────────
        st.markdown("---")
        plot_mode = st.radio(
            _t['dist_plot_mode_label'],
            options=[_t['dist_plot_ratio'], _t['dist_plot_lambda'], _t['dist_plot_fc']],
            index=0, horizontal=True, key="dist_plot_mode"
        )
        _mode_id = "RATIO" if plot_mode == _t['dist_plot_ratio'] else ("LAMBDA" if plot_mode == _t['dist_plot_lambda'] else "FC")

        for i in range(num_target_genes):
            gene_label = f"Gene {i+1}"
            gene_rows_st = [r for r in stats_data if r["__gene__"] == gene_label]
            if not gene_rows_st:
                continue
            st.subheader(f"{_t['target_gene']} {i+1} - {_t['distribution_graph']}")

            ctrl_ratios = gene_rows_st[0]["__ratio_ctrl__"]
            ctrl_lambdas = np.array([
                row["__lambda__"] for row in input_values_table
                if row["__gene__"] == gene_label and row["Grup"] == "Control" and row["__used__"] and row["__lambda__"] is not None
            ])

            if _mode_id == "RATIO":
                ctrl_vals = np.array(ctrl_ratios)
                y_label = _t['ratio_col']
            elif _mode_id == "LAMBDA":
                ctrl_vals = ctrl_lambdas
                y_label = _t['lambda_col']
            else:
                ctrl_vals = np.ones(len(ctrl_ratios))  # FC of control vs itself = 1
                y_label = _t['fc_col']

            fig = go.Figure()
            avg_ctrl = float(np.mean(ctrl_vals)) if len(ctrl_vals) else np.nan
            fig.add_trace(go.Scatter(x=[0.8, 1.2], y=[avg_ctrl, avg_ctrl], mode='lines',
                                      line=dict(color='black', width=4), name="Control avg"))
            fig.add_trace(go.Scatter(
                x=np.ones(len(ctrl_vals)) + np.random.uniform(-0.05, 0.05, len(ctrl_vals)),
                y=ctrl_vals, mode='markers', name="Control", marker=dict(color='#00796b')
            ))

            for j in range(num_patient_groups):
                row_match = [r for r in gene_rows_st if r["__group__"] == f"Group {j+1}"]
                if not row_match:
                    continue
                smp_ratios = np.array(row_match[0]["__ratio_smp__"])
                if _mode_id == "RATIO":
                    smp_vals = smp_ratios
                elif _mode_id == "LAMBDA":
                    smp_vals = np.array([
                        row["__lambda__"] for row in input_values_table
                        if row["__gene__"] == gene_label and row["Grup"] == f"Group {j+1}" and row["__used__"] and row["__lambda__"] is not None
                    ])
                else:
                    smp_vals = smp_ratios / avg_ctrl if avg_ctrl else smp_ratios

                avg_smp = float(np.mean(smp_vals)) if len(smp_vals) else np.nan
                fig.add_trace(go.Scatter(x=[j + 1.8, j + 2.2], y=[avg_smp, avg_smp], mode='lines',
                                          line=dict(color='black', width=4), name=f"Group {j+1} avg"))
                fig.add_trace(go.Scatter(
                    x=np.ones(len(smp_vals)) * (j + 2) + np.random.uniform(-0.05, 0.05, len(smp_vals)),
                    y=smp_vals, mode='markers', name=f"Group {j+1}", marker=dict(color='#d84315')
                ))

            fig.update_layout(
                title=f"{_t['target_gene']} {i+1} — {y_label}",
                xaxis=dict(tickvals=[1] + [j + 2 for j in range(num_patient_groups)],
                           ticktext=["Control"] + [f"Group {j+1}" for j in range(num_patient_groups)],
                           title=_t['x_axis_title']),
                yaxis=dict(title=y_label), showlegend=True
            )
            st.plotly_chart(fig, use_container_width=True, key=f"dist_chart_{i}")

    # ═══════════════════════════════════════════════════════════════════════════════
    # PDF REPORT GENERATION
    # ═══════════════════════════════════════════════════════════════════════════════
    def create_excel_report(data, stats_data, input_values_table, lang):
        """
        Builds a formatted multi-sheet Excel workbook (openpyxl) with:
          - Summary sheet (study parameters)
          - Input Data sheet (raw replicate values, outlier rows highlighted red)
          - Results sheet (Ratio/CN/Fold Change, conditionally colored by
            regulation status: red=upregulated, blue=downregulated)
          - Statistics sheet (p-values, significant results bolded/highlighted)
        Column widths are auto-sized and headers are frozen/styled for a more
        usable deliverable than a plain CSV.
        """
        wb = Workbook()

        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="00695C", end_color="00695C", fill_type="solid")
        up_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
        down_fill = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")
        excluded_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        sig_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        thin_border = Border(*(Side(style="thin", color="CCCCCC"),) * 4)

        def _write_sheet(ws, headers, rows, row_fill_fn=None):
            for c, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
            for i, row in enumerate(rows):
                r = i + 2
                fill = row_fill_fn(row, i) if row_fill_fn else None
                for c, h in enumerate(headers, start=1):
                    val = row.get(h, "")
                    if isinstance(val, float) and np.isnan(val):
                        val = ""
                    cell = ws.cell(row=r, column=c, value=val)
                    cell.border = thin_border
                    if fill:
                        cell.fill = fill
            for c, h in enumerate(headers, start=1):
                max_len = max([len(str(h))] + [len(str(row.get(h, ""))) for row in rows]) if rows else len(str(h))
                ws.column_dimensions[get_column_letter(c)].width = min(max(max_len + 2, 10), 40)
            ws.freeze_panes = "A2"

        # ── Summary sheet ─────────────────────────────────────────────────────────
        ws_summary = wb.active
        ws_summary.title = "Summary"
        n_genes = len(set(r["__gene__"] for r in data)) if data else 0
        n_groups = len(set(r["__group__"] for r in data)) if data else 0
        summary_rows = [
            {"Parameter": "AbsoluteGene version", "Value": APP_VERSION},
            {"Parameter": "Generated", "Value": pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")},
            {"Parameter": "Target genes analyzed", "Value": n_genes},
            {"Parameter": "Patient groups", "Value": n_groups},
            {"Parameter": "Total replicates", "Value": len(input_values_table)},
            {"Parameter": "Calculation method", "Value": "Poisson statistics (dPCR, absolute quantification)"},
        ]
        _write_sheet(ws_summary, ["Parameter", "Value"], summary_rows)

        # ── Input Data sheet ─────────────────────────────────────────────────────
        if input_values_table:
            ws_input = wb.create_sheet("Input Data")
            headers = ["__gene__", "Grup", "__replicate__", "__positive__", "__total__",
                       "__lambda__", "__conc__", "Outlier Excluded"]
            rename = {"__gene__": "Gene", "Grup": "Group", "__replicate__": "Replicate #",
                       "__positive__": "Positive", "__total__": "Total", "__lambda__": "Lambda",
                       "__conc__": "Concentration (copies/uL)", "Outlier Excluded": "Status"}
            clean_rows = [{rename[h]: row.get(h, "") for h in headers} for row in input_values_table]

            def _input_fill(row, idx):
                status = str(row.get("Status", ""))
                return excluded_fill if status.startswith(("Yes", "Evet")) else None

            _write_sheet(ws_input, list(rename.values()), clean_rows, row_fill_fn=_input_fill)

        # ── Results sheet ────────────────────────────────────────────────────────
        if data:
            ws_res = wb.create_sheet("Results")
            headers = ["Gene", "Group", "Ratio (Control)", "Ratio (Sample)", "CN (Control)", "CN (Sample)",
                        "Concentration (copies/uL)", "Fold Change", "Regulation"]
            clean_rows = []
            for r in data:
                clean_rows.append({
                    "Gene": r["__gene__"], "Group": r["__group__"],
                    "Ratio (Control)": round(r["__ratio_ctrl__"], 4) if r["__ratio_ctrl__"] is not None else "",
                    "Ratio (Sample)": round(r["__ratio_smp__"], 4),
                    "CN (Control)": round(r["__cn_ctrl__"], 3) if r["__cn_ctrl__"] is not None else "",
                    "CN (Sample)": round(r["__cn_smp__"], 3),
                    "Concentration (copies/uL)": (round(r["__conc_smp__"], 1)
                                                   if r.get("__conc_smp__") is not None and not np.isnan(r["__conc_smp__"]) else ""),
                    "Fold Change": round(r["__fc__"], 4),
                    "Regulation": r["__regulation__"],
                })

            def _res_fill(row, idx):
                reg = str(row.get("Regulation", "")).lower()
                if "up" in reg or "yukarı" in reg or "kazan" in reg or "gain" in reg:
                    return up_fill
                if "down" in reg or "aşağı" in reg or "kayıp" in reg or "loss" in reg:
                    return down_fill
                return None

            _write_sheet(ws_res, headers, clean_rows, row_fill_fn=_res_fill)

        # ── Statistics sheet ─────────────────────────────────────────────────────
        if stats_data:
            ws_stat = wb.create_sheet("Statistics")
            headers = ["Gene", "Comparison", "Test Type", "Test Method", "p-value", "Significance"]
            clean_rows = []
            sig_flags = []
            for r in stats_data:
                pval = r.get("__pvalue__", float("nan"))
                clean_rows.append({
                    "Gene": r["__gene__"], "Comparison": r.get("Comparison", ""),
                    "Test Type": r["__test_type__"], "Test Method": r["__test_method__"],
                    "p-value": round(pval, 4) if not np.isnan(pval) else "",
                    "Significance": r["__significance__"],
                })
                # Determined directly from the numeric p-value (language-independent),
                # rather than string-matching the localized "Significant"/"Anlamlı"
                # label, which would silently fail to highlight anything in non-English
                # reports.
                sig_flags.append(bool(not np.isnan(pval) and pval < 0.05))

            def _stat_fill(row, idx):
                return sig_fill if sig_flags[idx] else None

            _write_sheet(ws_stat, headers, clean_rows, row_fill_fn=_stat_fill)

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def create_simple_pdf(report_title, subtitle, description, summary_rows, table_header,
                           table_rows, chart_png_bytes, chart_caption, footer_note, references=None):
        """
        Generic lightweight PDF builder shared by the Batch Screening and VAF
        Calculator tabs. Produces: title/subtitle, description paragraph,
        a summary key-value table, a results table, an optional chart image,
        a footer note, and an optional references list.
        """
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50, topMargin=60, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('RT', parent=styles['Title'], fontName=PDF_FONT_BOLD, fontSize=18,
                                      textColor=colors.HexColor('#004d40'), spaceAfter=6, alignment=1)
        sub_style = ParagraphStyle('RS', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10,
                                    textColor=colors.HexColor('#555555'), spaceAfter=4, alignment=1)
        body_style = ParagraphStyle('BD', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9, leading=13, spaceAfter=8)
        small_style = ParagraphStyle('SM', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8, leading=11,
                                      textColor=colors.HexColor('#444444'))
        caption_style = ParagraphStyle('CA', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8,
                                        textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=6)

        def hr():
            return HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=8, spaceBefore=4)

        def make_table(rows, col_widths=None, header=True):
            if not rows:
                return Spacer(1, 1)
            styled_rows = []
            for ri, row in enumerate(rows):
                styled_row = []
                for cell in row:
                    cell_str = safe_str(cell)
                    if ri == 0 and header:
                        p = Paragraph(cell_str, ParagraphStyle('TH', fontName=PDF_FONT_BOLD, fontSize=7,
                                                                textColor=colors.white, alignment=1))
                    else:
                        p = Paragraph(cell_str, ParagraphStyle('TD', fontName=PDF_FONT, fontSize=7, alignment=1))
                    styled_row.append(p)
                styled_rows.append(styled_row)
            tbl = Table(styled_rows, colWidths=col_widths, repeatRows=1 if header else 0)
            tbl_style = [
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e0f2f1')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            if header:
                tbl_style.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#004d40')))
            tbl.setStyle(TableStyle(tbl_style))
            return tbl

        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        elements.append(Spacer(1, 30))
        elements.append(Paragraph(safe_str(report_title), title_style))
        elements.append(Paragraph(safe_str(subtitle), sub_style))
        elements.append(Paragraph(safe_str(f"Generated: {now}"), sub_style))
        elements.append(Spacer(1, 14))
        elements.append(hr())
        elements.append(Paragraph(safe_str(description), body_style))

        if summary_rows:
            elements.append(make_table(summary_rows, col_widths=[260, 200]))
            elements.append(Spacer(1, 14))

        if table_header and table_rows:
            cw = (letter[0] - 100) / len(table_header)
            elements.append(make_table([table_header] + table_rows, col_widths=[cw] * len(table_header)))
            elements.append(Spacer(1, 10))

        if chart_png_bytes:
            elements.append(RLImage(BytesIO(chart_png_bytes), width=460, height=250))
            if chart_caption:
                elements.append(Paragraph(safe_str(chart_caption), caption_style))

        if references:
            elements.append(Spacer(1, 12))
            elements.append(hr())
            for ref in references:
                elements.append(Paragraph(safe_str(f"• {ref}"), small_style))
                elements.append(Spacer(1, 3))

        elements.append(Spacer(1, 14))
        elements.append(hr())
        elements.append(Paragraph(safe_str(footer_note), small_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer


    def create_pdf(results, stat_rows, input_df, lang, multigroup_results=None):
        T = translations[lang]
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter, leftMargin=50, rightMargin=50, topMargin=60, bottomMargin=50)
        elements = []
        styles = getSampleStyleSheet()

        title_style = ParagraphStyle('RT', parent=styles['Title'], fontName=PDF_FONT_BOLD, fontSize=20,
                                      textColor=colors.HexColor('#004d40'), spaceAfter=6, alignment=1)
        sub_style = ParagraphStyle('RS', parent=styles['Normal'], fontName=PDF_FONT, fontSize=10,
                                    textColor=colors.HexColor('#555555'), spaceAfter=4, alignment=1)
        h1_style = ParagraphStyle('H1', parent=styles['Heading1'], fontName=PDF_FONT_BOLD, fontSize=13,
                                   textColor=colors.HexColor('#004d40'), spaceBefore=16, spaceAfter=5)
        h2_style = ParagraphStyle('H2', parent=styles['Heading2'], fontName=PDF_FONT_BOLD, fontSize=11,
                                   textColor=colors.HexColor('#00695c'), spaceBefore=10, spaceAfter=4)
        body_style = ParagraphStyle('BD', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9, leading=13, spaceAfter=4)
        small_style = ParagraphStyle('SM', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8, leading=11,
                                      textColor=colors.HexColor('#444444'))
        caption_style = ParagraphStyle('CA', parent=styles['Normal'], fontName=PDF_FONT, fontSize=8,
                                        textColor=colors.HexColor('#666666'), alignment=1, spaceAfter=6)
        warn_style = ParagraphStyle('WN', parent=styles['Normal'], fontName=PDF_FONT, fontSize=9, leading=13,
                                     backColor=colors.HexColor('#fff8e1'), borderPad=6, leftIndent=8, rightIndent=8, spaceAfter=6)

        def hr():
            return HRFlowable(width="100%", thickness=0.5, color=colors.HexColor('#cccccc'), spaceAfter=8, spaceBefore=4)

        def make_table(rows, col_widths=None, header=True):
            if not rows:
                return Spacer(1, 1)
            styled_rows = []
            for ri, row in enumerate(rows):
                styled_row = []
                for cell in row:
                    if isinstance(cell, Flowable):
                        styled_row.append(cell)
                        continue
                    cell_str = safe_str(cell)
                    if ri == 0 and header:
                        p = Paragraph(cell_str, ParagraphStyle('TH', fontName=PDF_FONT_BOLD, fontSize=7,
                                                                textColor=colors.white, alignment=1))
                    else:
                        p = Paragraph(cell_str, ParagraphStyle('TD', fontName=PDF_FONT, fontSize=7, alignment=1))
                    styled_row.append(p)
                styled_rows.append(styled_row)
            tbl = Table(styled_rows, colWidths=col_widths, repeatRows=1 if header else 0)
            tbl_style = [
                ('FONTNAME', (0, 0), (-1, -1), PDF_FONT), ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'), ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cccccc')),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#e0f2f1')]),
                ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ]
            if header:
                tbl_style.append(('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#004d40')))
            tbl.setStyle(TableStyle(tbl_style))
            return tbl

        import datetime
        now = datetime.datetime.now().strftime("%Y-%m-%d %H:%M")

        # ── COVER ──────────────────────────────────────────────────────────────────
        elements.append(Spacer(1, 40))
        elements.append(Paragraph(safe_str("AbsoluteGene"), title_style))
        elements.append(Paragraph(safe_str(T['pdf_report']), sub_style))
        elements.append(Spacer(1, 6))
        elements.append(Paragraph(safe_str(f"Generated: {now}"), sub_style))
        elements.append(Paragraph(safe_str("Related tool for qPCR: GeneQuantify — Yalçınkaya B (2026), Mol Cell Biochem, "
                                            "https://doi.org/10.1007/s11010-026-05621-y"), sub_style))
        elements.append(Spacer(1, 20))
        elements.append(hr())

        n_genes = len(set(r["__gene__"] for r in results)) if results else 0
        n_groups = len(set(r["__group__"] for r in results)) if results else 0
        n_replicates = len(input_df)
        n_excluded = sum(1 for _, row in input_df.iterrows() if str(row.get('Outlier Excluded', '')).startswith(('Yes', 'Evet')))
        summary_rows = [
            ["Parameter", "Value"],
            ["Target genes analyzed", str(n_genes)],
            ["Patient groups", str(n_groups)],
            ["Total replicates", str(n_replicates)],
            ["Excluded / flagged replicates", str(n_excluded)],
            ["Reference loci per gene", str(num_ref_genes)],
            ["Reference ploidy", str(ploidy)],
            ["Partition volume (nL)", str(partition_vol_nl)],
            ["QC minimum partitions", str(int(qc_min_partitions))],
            ["Calculation method", "Poisson statistics (dPCR, absolute quantification)"],
        ]
        elements.append(make_table(summary_rows, col_widths=[260, 200]))
        elements.append(Spacer(1, 14))
        elements.append(Paragraph(safe_str(
            "This report was generated automatically by AbsoluteGene. All calculations follow the "
            "digital MIQE (dMIQE) guidelines (Huggett et al., Clin Chem 2013; updated 2020). "
            "For research and educational use only — not validated for clinical diagnostic decision-making."
        ), small_style))
        elements.append(PageBreak())

        # ── SECTION 1: METHODS ────────────────────────────────────────────────────
        elements.append(Paragraph(safe_str("1. Methods and Analysis Settings"), h1_style))
        elements.append(hr())
        elements.append(Paragraph(safe_str("1.1 Poisson Quantification"), h2_style))
        elements.append(Paragraph(safe_str(
            "For each replicate, the fraction of positive partitions p = positive/total was converted to "
            "copies per partition (λ) using the Poisson distribution: λ = -ln(1-p). A 95% confidence "
            "interval was calculated using the delta-method normal approximation: "
            "Var(λ) ≈ p / (n·(1-p))."
        ), body_style))
        elements.append(Paragraph(safe_str("1.2 Normalization and Copy Number"), h2_style))
        elements.append(Paragraph(safe_str(
            f"Normalization factor (NF) = geometric mean of reference-locus λ values "
            f"({num_ref_genes} reference locus/loci used). Ratio = λ(target) / NF. "
            f"Copy number (CN) = ploidy(reference) × Ratio, using reference ploidy = {ploidy}. "
            f"Fold Change = Ratio(sample) / Ratio(control). Because dPCR is an absolute endpoint "
            f"measurement, no amplification efficiency correction is required."
        ), body_style))
        elements.append(Paragraph(safe_str("1.3 Quality Control"), h2_style))
        elements.append(Paragraph(safe_str(
            f"Replicates with fewer than {int(qc_min_partitions)} total accepted partitions were flagged "
            f"as low quality and excluded from λ-based calculations. Replicates with 100% positive "
            f"partitions (saturation) were also excluded, as λ is undefined at p=1; further dilution "
            f"is recommended for such samples."
        ), body_style))
        elements.append(Paragraph(safe_str("1.4 Outlier Detection"), h2_style))
        if outlier_enabled:
            method_txt = (f"Grubbs' test (alpha = {grubbs_alpha})" if outlier_method == "Grubbs"
                           else f"IQR method (multiplier k = {iqr_multiplier})")
            elements.append(Paragraph(safe_str(
                f"{method_txt} was applied to λ values (copies/partition) for the target locus and each "
                f"reference locus, cascading across all loci within a replicate. Flagged replicates were "
                f"confirmed for exclusion by the user; {n_excluded} replicate(s) were excluded/flagged in "
                f"this analysis."
            ), body_style))
        else:
            elements.append(Paragraph(safe_str("Outlier detection was disabled for this analysis."), body_style))
        elements.append(PageBreak())

        # ── SECTION 2: INPUT DATA ─────────────────────────────────────────────────
        elements.append(Paragraph(safe_str("2. Input Data"), h1_style))
        elements.append(hr())
        elements.append(Paragraph(safe_str(
            "Positive and total accepted partition counts entered by the user, with derived λ "
            "(copies/partition) values. Rows flagged as excluded were removed from ratio calculations."
        ), body_style))
        elements.append(Spacer(1, 6))
        if not input_df.empty:
            _pdf_rename = {
                "__gene__": "Gene", "Grup": "Group", "__replicate__": "Replicate #",
                "__positive__": "Positive", "__total__": "Total", "__lambda__": "Lambda (copies/partition)",
                "Outlier Excluded": "Status",
            }
            input_df_disp = input_df.drop(columns=["__used__"], errors="ignore").rename(columns=_pdf_rename)
            cols = input_df_disp.columns.tolist()
            page_w = letter[0] - 100
            cw = page_w / max(len(cols), 1)
            tbl_rows = [cols]
            for _, row in input_df_disp.iterrows():
                is_excl = str(row.get('Status', '')).startswith(('Yes', 'Evet'))
                row_cells = []
                for c in cols:
                    v = row.get(c, '')
                    cell_str = safe_str(str(v) if v is not None else '')
                    style = ParagraphStyle('EX' if is_excl else 'TD', fontName=PDF_FONT, fontSize=7, alignment=1,
                                            textColor=colors.HexColor('#cc0000') if is_excl else colors.black)
                    row_cells.append(Paragraph(cell_str, style))
                tbl_rows.append(row_cells)
            elements.append(make_table(tbl_rows, col_widths=[cw] * len(cols)))
        elements.append(PageBreak())

        # ── SECTION 3: RESULTS ────────────────────────────────────────────────────
        elements.append(Paragraph(safe_str("3. Results"), h1_style))
        elements.append(hr())
        res_cols = ["Gene", "Group", "Ratio (Ctrl)", "Ratio (Sample)", "CN (Ctrl)", "CN (Sample)",
                    "Conc. Sample (copies/µL)", "95% CI (copies/µL)", "Fold Change", "Regulation", "LOD/LOQ Status"]
        res_rows = [res_cols]
        for r in results:
            conc_ci = (f"{r['__conc_smp_lo__']:.1f}\u2013{r['__conc_smp_hi__']:.1f}"
                       if r.get('__conc_smp_lo__') is not None and not np.isnan(r['__conc_smp_lo__']) else "n<2")
            res_rows.append([
                r["__gene__"], r["__group__"],
                f"{r['__ratio_ctrl__']:.4f}" if r['__ratio_ctrl__'] is not None else "—",
                f"{r['__ratio_smp__']:.4f}",
                f"{r['__cn_ctrl__']:.3f}" if r['__cn_ctrl__'] is not None else "—",
                f"{r['__cn_smp__']:.3f}",
                f"{r.get('__conc_smp__', float('nan')):.1f}" if r.get('__conc_smp__') is not None and not np.isnan(r.get('__conc_smp__', np.nan)) else "—",
                conc_ci,
                f"{r['__fc__']:.4f}", r["__regulation__"],
                r.get("__lod_loq_flag__", "—"),
            ])
        cw11r = (letter[0] - 100) / 11
        elements.append(make_table(res_rows, col_widths=[cw11r] * 11))
        elements.append(Spacer(1, 8))

        if results:
            try:
                fig_fc, ax_fc = plt.subplots(figsize=(7, 3.5))
                labels_fc = [f"{r['__gene__']} /\n{r['__group__']}" for r in results]
                vals_fc = [r["__fc__"] for r in results]
                bars = ax_fc.bar(range(len(labels_fc)), vals_fc, color='#00796b', alpha=0.85)
                ax_fc.axhline(y=1, color='black', linestyle='--', linewidth=0.8, alpha=0.6)
                ax_fc.set_xticks(range(len(labels_fc)))
                ax_fc.set_xticklabels(labels_fc, fontsize=7)
                ax_fc.set_ylabel('Fold Change', fontsize=9)
                ax_fc.set_title('Fold Change by Gene/Group', fontsize=10, fontweight='bold')
                ax_fc.spines['top'].set_visible(False)
                ax_fc.spines['right'].set_visible(False)
                for bar in bars:
                    ax_fc.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 0.02,
                               f'{bar.get_height():.2f}', ha='center', va='bottom', fontsize=6)
                plt.tight_layout()
                ib = BytesIO()
                plt.savefig(ib, format='png', dpi=150, bbox_inches='tight')
                plt.close()
                ib.seek(0)
                elements.append(RLImage(ib, width=460, height=230))
                elements.append(Paragraph(safe_str("Figure 1. Fold change by gene/group. Dashed line y=1 = no change vs control."), caption_style))
            except Exception as _chart_err:
                elements.append(Paragraph(safe_str(
                    f"[Chart could not be generated: {_chart_err}]"
                ), small_style))
        elements.append(PageBreak())

        # ── SECTION 4: STATISTICS ─────────────────────────────────────────────────
        elements.append(Paragraph(safe_str("4. Statistical Analysis"), h1_style))
        elements.append(hr())
        elements.append(Paragraph(safe_str(
            "Statistical tests were performed directly on the normalized Ratio values (linear scale, "
            "no log transform required). Test selection is automatic based on normality (Shapiro-Wilk, "
            "n≥8) and variance homogeneity (Levene). Significance threshold: p < 0.05."
        ), body_style))
        elements.append(Spacer(1, 6))

        _has_multigroup = multigroup_results is not None and any(r.get("n_groups", 0) >= 3 for r in multigroup_results)
        stat_cols = ["Gene", "Comparison", "Type", "Method", "p-value", "Significance"]
        stat_table_rows = [stat_cols]
        for sr in stat_rows:
            stat_table_rows.append([
                sr["__gene__"], sr["Comparison"], sr["__test_type__"], sr["__test_method__"],
                f"{sr['__pvalue__']:.4f}" if not np.isnan(sr['__pvalue__']) else "—", sr["__significance__"],
            ])
        cw6 = (letter[0] - 100) / 6
        elements.append(make_table(stat_table_rows, col_widths=[cw6] * 6))
        elements.append(Spacer(1, 10))

        if _has_multigroup:
            elements.append(Paragraph(safe_str("4.1 Multi-Group Comparison (≥3 Groups)"), h2_style))
            for mg in (multigroup_results or []):
                if mg.get("n_groups", 0) < 3:
                    continue
                elements.append(Paragraph(safe_str(f"▶ {mg['gene']} — {mg['omnibus_test']} (p = {mg['omnibus_p']:.4f})"), body_style))
                ph_rows = mg.get("posthoc_rows", [])
                if ph_rows:
                    ph_header = ["Comparison", "Raw p", "Bonferroni p", "FDR p (B-H)"]
                    ph_table = [ph_header]
                    for ph in ph_rows:
                        ph_table.append([ph["Comparison"], f"{ph['Raw p']:.4f}", f"{ph['Bonferroni p']:.4f}", f"{ph['FDR p (B-H)']:.4f}"])
                    elements.append(make_table(ph_table, col_widths=[(letter[0] - 100) / 4] * 4))
                    elements.append(Spacer(1, 8))
        elements.append(PageBreak())

        # ── SECTION 5: DISTRIBUTION PLOTS ─────────────────────────────────────────
        elements.append(Paragraph(safe_str("5. Normalized Ratio Distribution Plots"), h1_style))
        elements.append(hr())
        elements.append(Paragraph(safe_str(
            "Distribution of normalized Ratio (Target λ / reference NF) values per group. "
            "Each point = one biological replicate; horizontal bars = group means."
        ), body_style))
        elements.append(Spacer(1, 8))

        for i in range(num_target_genes):
            gene_label = f"Gene {i+1}"
            gene_rows_pdf = [r for r in stat_rows if r["__gene__"] == gene_label]
            if not gene_rows_pdf:
                continue
            try:
                fig_d, ax_d = plt.subplots(figsize=(6, 3.2))
                all_vals, all_labels = [list(gene_rows_pdf[0]["__ratio_ctrl__"])], ["Control"]
                for gr in gene_rows_pdf:
                    all_vals.append(list(gr["__ratio_smp__"]))
                    all_labels.append(gr["__group__"])
                palette = ['#00796b', '#d84315', '#455a64', '#f9a825', '#5e35b1']
                for k, (vals, lbl) in enumerate(zip(all_vals, all_labels)):
                    col = palette[k % len(palette)]
                    jitter = np.random.uniform(-0.08, 0.08, len(vals))
                    ax_d.scatter([k + 1 + j for j in jitter], vals, color=col, alpha=0.75, s=28, zorder=3)
                    ax_d.hlines(np.mean(vals), k + 0.75, k + 1.25, colors='black', linewidths=2, zorder=4)
                ax_d.set_xticks(range(1, len(all_labels) + 1))
                ax_d.set_xticklabels(all_labels, fontsize=8)
                ax_d.set_ylabel('Normalized Ratio', fontsize=9)
                ax_d.set_title(f'{gene_label} — Ratio Distribution', fontsize=10, fontweight='bold')
                ax_d.spines['top'].set_visible(False)
                ax_d.spines['right'].set_visible(False)
                plt.tight_layout()
                ib3 = BytesIO()
                plt.savefig(ib3, format='png', dpi=150, bbox_inches='tight')
                plt.close()
                ib3.seek(0)
                elements.append(RLImage(ib3, width=420, height=210))
                elements.append(Paragraph(safe_str(f"Figure. Ratio distribution for {gene_label}."), caption_style))
                elements.append(Spacer(1, 10))
            except Exception as _chart_err:
                elements.append(Paragraph(safe_str(
                    f"[Chart for {gene_label} could not be generated: {_chart_err}]"
                ), small_style))
        elements.append(PageBreak())

        # ── SECTION 6: INTERPRETATION ─────────────────────────────────────────────
        elements.append(Paragraph(safe_str("6. How to Interpret Your Results"), h1_style))
        elements.append(hr())
        fc_hdr = ["Fold Change", "Interpretation", "Biological Significance"]
        fc_rows = [
            [">2.0", "Strong upregulation / gain", "Consider biologically relevant"],
            ["1.5-2.0", "Moderate upregulation / gain", "May be relevant; verify"],
            ["1.0-1.5", "Weak upregulation", "Likely not significant alone"],
            ["1.0", "No change", "No differential expression / CNV"],
            ["0.67-1.0", "Weak downregulation", "Likely not significant alone"],
            ["0.5-0.67", "Moderate downregulation / loss", "May be relevant; verify"],
            ["<0.5", "Strong downregulation / loss", "Consider biologically relevant"],
        ]
        elements.append(make_table([fc_hdr] + fc_rows, col_widths=[(letter[0] - 100) / 3] * 3))
        elements.append(Spacer(1, 8))
        elements.append(Paragraph(safe_str(
            "Note: statistical significance (p < 0.05) and biological significance (fold change "
            "magnitude, and the width of the Poisson 95% CI relative to the observed effect) "
            "should be considered together."
        ), warn_style))
        elements.append(PageBreak())

        # ── SECTION 7: REFERENCES ─────────────────────────────────────────────────
        elements.append(Paragraph(safe_str("7. References"), h1_style))
        elements.append(hr())
        refs = [
            "Yalçınkaya B (2026). GeneQuantify: a web-based tool for qPCR gene expression and copy number variation analysis. Molecular and Cellular Biochemistry. https://doi.org/10.1007/s11010-026-05621-y",
            "Huggett JF et al. (2013). The digital MIQE guidelines. Clinical Chemistry, 59(6), 892-902.",
            "Huggett JF (2020). The digital MIQE guidelines update. Clinical Chemistry, 66(8), 1012-1029.",
            "Vandesompele J et al. (2002). Genome Biology, 3(7). (geNorm normalization concept, adapted here to λ values)",
            "Grubbs FE (1969). Technometrics, 11(1), 1-21. (Outlier detection)",
            "Tukey JW (1977). Exploratory Data Analysis. Addison-Wesley. (IQR outlier method)",
            "Benjamini Y & Hochberg Y (1995). J Royal Stat Soc B, 57(1), 289-300. (FDR correction)",
            "Dube S, Qin J, Ramakrishnan R (2008). PLoS ONE, 3(8), e2876. (Poisson statistics for digital PCR)",
        ]
        for ref in refs:
            elements.append(Paragraph(safe_str(f"• {ref}"), small_style))
            elements.append(Spacer(1, 3))

        elements.append(Spacer(1, 16))
        elements.append(hr())
        elements.append(Paragraph(safe_str(
            f"AbsoluteGene v{APP_VERSION} — For research and educational use only. Not validated for clinical "
            f"diagnostic purposes. | Generated: {now} | Contact: mailtoburhanettin@gmail.com"
        ), small_style))

        doc.build(elements)
        buffer.seek(0)
        return buffer


    # ═══════════════════════════════════════════════════════════════════════════════
    # BATCH SCREENING TAB
    # ═══════════════════════════════════════════════════════════════════════════════
    with tab_batch:
        st.markdown(f"### {_t['batch_title']}")
        st.caption(_t['batch_description'])
        st.download_button(
            _t['download_example_csv'], data=EXAMPLE_CSV_BATCH_SCREENING.encode("utf-8"),
            file_name="example_batch_screening.csv", mime="text/csv", key="dl_example_batch"
        )
        st.markdown("---")

        batch_file = st.file_uploader(_t['batch_uploader'], type=["csv", "tsv", "txt"], key="batch_uploader")

        if batch_file is not None:
            _batch_raw_bytes = batch_file.read()
            _batch_raw_df, _batch_detected, _batch_err = parse_instrument_csv(_batch_raw_bytes)
            if _batch_err:
                st.error(_t['csv_parse_error'].format(err=_batch_err))
            else:
                _batch_col_options = list(_batch_raw_df.columns)
                _batch_col_map = {}
                bcol1, bcol2, bcol3, bcol4 = st.columns(4)
                for _field, _label_key, _col_widget in [
                    ("sample", "csv_col_sample", bcol1), ("target", "csv_col_target", bcol2),
                    ("positives", "csv_col_positives", bcol3), ("total", "csv_col_total", bcol4)
                ]:
                    _detected = _batch_detected.get(_field)
                    _default_idx = _batch_col_options.index(_detected) if _detected in _batch_col_options else 0
                    with _col_widget:
                        _batch_col_map[_field] = st.selectbox(
                            _t[_label_key], options=_batch_col_options,
                            index=_default_idx, key=f"batch_colmap_{_field}"
                        )

                _batch_std_df = build_standard_import_df(_batch_raw_df, _batch_col_map)

                if not _batch_std_df.empty:
                    st.markdown(_t['batch_assay_title'])
                    _batch_unique_targets = sorted(_batch_std_df["Target"].unique())
                    bacol1, bacol2 = st.columns(2)
                    with bacol1:
                        _batch_target_assays = st.multiselect(
                            _t['batch_target_label'], options=_batch_unique_targets,
                            default=_batch_unique_targets[:1] if _batch_unique_targets else [],
                            key="batch_target_assays", help=_t['batch_multi_gene_help']
                        )
                    with bacol2:
                        _batch_ref_options = [t for t in _batch_unique_targets if t not in _batch_target_assays]
                        _batch_ref_assays = st.multiselect(
                            _t['batch_ref_label'], options=_batch_ref_options,
                            default=_batch_ref_options[:1] if _batch_ref_options else [],
                            key="batch_ref_assays"
                        )

                    st.markdown(_t['batch_settings_title'])
                    bscol1, bscol2 = st.columns(2)
                    with bscol1:
                        _batch_expected_ratio = st.number_input(
                            _t['batch_expected_ratio'], min_value=0.01, max_value=100.0,
                            value=1.0, step=0.1, key="batch_expected_ratio"
                        )
                    with bscol2:
                        _batch_ploidy = st.number_input(
                            _t['batch_ploidy'], min_value=1, max_value=10, value=2, step=1, key="batch_ploidy"
                        )
                    _batch_dilution = render_dilution_input("batch")

                    if st.button(_t['batch_run_btn'], key="batch_run_btn", use_container_width=True) and _batch_ref_assays and _batch_target_assays:
                        _batch_results_by_gene = {}
                        for _assay in _batch_target_assays:
                            _batch_results_by_gene[_assay] = pool_and_compute_batch(
                                _batch_std_df, _assay, _batch_ref_assays, partition_vol_nl
                            )
                        st.session_state["_batch_results_by_gene_cache"] = _batch_results_by_gene
                        st.session_state["_batch_expected_ratio_cache"] = _batch_expected_ratio
                        st.session_state["_batch_ploidy_cache"] = _batch_ploidy
                        st.session_state["_batch_dilution_cache"] = _batch_dilution

        # ── Display cached results (persists across reruns/tab switches) ───────────
        _batch_results_by_gene = st.session_state.get("_batch_results_by_gene_cache")
        _batch_results = None
        if _batch_results_by_gene:
            _first_gene = list(_batch_results_by_gene.keys())[0]
            _batch_results = _batch_results_by_gene[_first_gene]
            _exp_ratio = st.session_state.get("_batch_expected_ratio_cache", 1.0)
            _b_ploidy = st.session_state.get("_batch_ploidy_cache", 2)
            _b_target_assay = _first_gene
            _b_dilution = st.session_state.get("_batch_dilution_cache", 1.0)

            st.markdown(f"#### {_t['batch_results_title']}")
            _batch_rows = []
            _n_flagged = 0
            for r in _batch_results:
                if np.isnan(r["ci_low"]) or np.isnan(r["ci_high"]):
                    classification = "—"
                elif _exp_ratio < r["ci_low"]:
                    classification = _t['batch_class_gain']
                    _n_flagged += 1
                elif _exp_ratio > r["ci_high"]:
                    classification = _t['batch_class_loss']
                    _n_flagged += 1
                else:
                    classification = _t['batch_class_normal']

                _batch_lam_t = r["lambda_t"]
                if _batch_lam_t < 0.05:
                    _batch_dr_flag = _t['dynamic_range_low']
                elif _batch_lam_t > 4.0:
                    _batch_dr_flag = _t['dynamic_range_saturated']
                elif _batch_lam_t > 3.0:
                    _batch_dr_flag = _t['dynamic_range_high']
                else:
                    _batch_dr_flag = _t['dynamic_range_ok']

                _batch_rows.append({
                    _t['batch_col_sample']: r["Sample"],
                    _t['batch_col_lambda_t']: round(r["lambda_t"], 5),
                    _t['batch_col_lambda_r']: round(r["lambda_r"], 5),
                    _t['batch_col_ratio']: round(r["ratio"], 4),
                    _t['batch_col_ci']: (f"{r['ci_low']:.4f}\u2013{r['ci_high']:.4f}"
                                           if not np.isnan(r["ci_low"]) else "—"),
                    _t['batch_col_cn']: round(_b_ploidy * r["ratio"], 3),
                    _t['batch_col_conc']: round(r["conc"], 1),
                    _t['stock_conc_col']: round(r["conc"] * _b_dilution, 1),
                    _t['dynamic_range_warning_label']: _batch_dr_flag,
                    _t['batch_col_class']: classification,
                })

            st.info(_t['batch_n_flagged'].format(n=_n_flagged, total=len(_batch_results)))
            batch_df = pd.DataFrame(_batch_rows)
            st.dataframe(batch_df, use_container_width=True)

            batch_csv = batch_df.to_csv(index=False).encode("utf-8")
            st.download_button(_t['batch_download_btn'], data=batch_csv,
                                file_name="batch_screening_results.csv", mime="text/csv", key="batch_dl_csv")

            # ── Screening plot ────────────────────────────────────────────────────
            fig_batch = go.Figure()
            _colors_map = {
                _t['batch_class_gain']: "#e53935", _t['batch_class_loss']: "#1e88e5",
                _t['batch_class_normal']: "#43a047", "—": "#9e9e9e"
            }
            _sample_names = [r["Sample"] for r in _batch_results]
            _ratios = [r["ratio"] for r in _batch_results]
            _err_low = [max(r["ratio"] - r["ci_low"], 0) if not np.isnan(r["ci_low"]) else 0 for r in _batch_results]
            _err_high = [r["ci_high"] - r["ratio"] if not np.isnan(r["ci_high"]) else 0 for r in _batch_results]
            _point_colors = [_colors_map.get(row[_t['batch_col_class']], "#9e9e9e") for row in _batch_rows]

            fig_batch.add_trace(go.Scatter(
                x=_sample_names, y=_ratios, mode="markers",
                marker=dict(color=_point_colors, size=9),
                error_y=dict(type="data", symmetric=False, array=_err_high, arrayminus=_err_low),
                name="Ratio (95% CI)"
            ))
            fig_batch.add_hline(y=_exp_ratio, line_dash="dash", line_color="black",
                                annotation_text=f"Expected = {_exp_ratio}")
            fig_batch.update_layout(
                title=_t['batch_chart_title'], xaxis_title=_t['batch_col_sample'],
                yaxis_title=_t['batch_col_ratio'], height=420, showlegend=False
            )
            st.plotly_chart(fig_batch, use_container_width=True, key="batch_screening_chart")

            # ── Multi-gene heatmap (only when ≥2 target genes were analyzed) ───────
            if len(_batch_results_by_gene) >= 2:
                st.markdown("---")
                st.markdown(f"#### {_t['batch_heatmap_title']}")
                st.caption(_t['batch_heatmap_description'])

                _heatmap_metric = st.radio(
                    _t['batch_heatmap_metric_label'],
                    options=[_t['batch_col_cn'], _t['batch_col_ratio']],
                    key="batch_heatmap_metric", horizontal=True
                )

                _all_genes = list(_batch_results_by_gene.keys())
                _all_samples = sorted({r["Sample"] for res in _batch_results_by_gene.values() for r in res})
                _matrix = np.full((len(_all_genes), len(_all_samples)), np.nan)
                for _gi, _gene in enumerate(_all_genes):
                    _res_by_sample = {r["Sample"]: r for r in _batch_results_by_gene[_gene]}
                    for _si, _sample in enumerate(_all_samples):
                        _r = _res_by_sample.get(_sample)
                        if _r is not None:
                            _val = _b_ploidy * _r["ratio"] if _heatmap_metric == _t['batch_col_cn'] else _r["ratio"]
                            _matrix[_gi, _si] = _val

                fig_heatmap = go.Figure(go.Heatmap(
                    z=_matrix, x=_all_samples, y=_all_genes,
                    colorscale="RdBu_r", zmid=_exp_ratio * (_b_ploidy if _heatmap_metric == _t['batch_col_cn'] else 1),
                    colorbar=dict(title=_heatmap_metric),
                    text=np.round(_matrix, 3), texttemplate="%{text}", textfont=dict(size=9),
                ))
                fig_heatmap.update_layout(
                    title=f"{_heatmap_metric} — {_t['batch_heatmap_title']}",
                    xaxis_title=_t['batch_col_sample'], yaxis_title=_t['batch_target_label'],
                    height=max(300, 60 * len(_all_genes) + 150)
                )
                st.plotly_chart(fig_heatmap, use_container_width=True, key="batch_heatmap_chart")

                # Per-gene results in expanders (since the main table above only shows the first gene)
                for _gene in _all_genes[1:]:
                    with st.expander(f"{_t['batch_results_title']} — {_gene}", expanded=False):
                        _gene_rows = []
                        for r in _batch_results_by_gene[_gene]:
                            _gene_rows.append({
                                _t['batch_col_sample']: r["Sample"],
                                _t['batch_col_ratio']: round(r["ratio"], 4),
                                _t['batch_col_cn']: round(_b_ploidy * r["ratio"], 3),
                                _t['batch_col_conc']: round(r["conc"], 1),
                            })
                        st.dataframe(pd.DataFrame(_gene_rows), use_container_width=True)

            # ── PDF report ────────────────────────────────────────────────────────
            if st.button(_t['batch_pdf_btn'], key="batch_pdf_btn"):
                try:
                    fig_mpl, ax_mpl = plt.subplots(figsize=(7, 3.5))
                    _mpl_colors = {
                        _t['batch_class_gain']: "#e53935", _t['batch_class_loss']: "#1e88e5",
                        _t['batch_class_normal']: "#43a047", "—": "#9e9e9e"
                    }
                    _mpl_point_colors = [_mpl_colors.get(row[_t['batch_col_class']], "#9e9e9e") for row in _batch_rows]
                    ax_mpl.errorbar(range(len(_sample_names)), _ratios, yerr=[_err_low, _err_high],
                                     fmt='none', ecolor='gray', capsize=3, zorder=2)
                    ax_mpl.scatter(range(len(_sample_names)), _ratios, c=_mpl_point_colors, s=50, zorder=3)
                    ax_mpl.axhline(y=_exp_ratio, color='black', linestyle='--', linewidth=1)
                    ax_mpl.set_xticks(range(len(_sample_names)))
                    ax_mpl.set_xticklabels(_sample_names, rotation=45, ha='right', fontsize=7)
                    ax_mpl.set_ylabel(_t['batch_col_ratio'], fontsize=9)
                    ax_mpl.set_title(_t['batch_chart_title'], fontsize=10, fontweight='bold')
                    ax_mpl.spines['top'].set_visible(False)
                    ax_mpl.spines['right'].set_visible(False)
                    plt.tight_layout()
                    _img_buf = BytesIO()
                    plt.savefig(_img_buf, format='png', dpi=150, bbox_inches='tight')
                    plt.close()
                    _chart_bytes = _img_buf.getvalue()
                except Exception as _chart_err:
                    _chart_bytes = None
                    st.warning(f"⚠️ Chart could not be generated for the PDF ({_chart_err}). "
                               f"The report will still be created without this figure.")

                _batch_summary_rows = [
                    ["Parameter", "Value"],
                    ["Target assay", _b_target_assay],
                    ["Samples screened", str(len(_batch_results))],
                    ["Expected ratio", f"{_exp_ratio}"],
                    ["Reference ploidy", str(_b_ploidy)],
                    ["Samples flagged (significant deviation)", str(_n_flagged)],
                ]
                _batch_pdf_table_header = [_t['batch_col_sample'], _t['batch_col_lambda_t'], _t['batch_col_lambda_r'],
                                            _t['batch_col_ratio'], _t['batch_col_ci'], _t['batch_col_cn'],
                                            _t['batch_col_conc'], _t['batch_col_class']]
                _batch_pdf_table_rows = [[str(v) for v in row.values()] for row in _batch_rows]

                batch_pdf_buffer = create_simple_pdf(
                    report_title="AbsoluteGene",
                    subtitle=_t['batch_pdf_report'],
                    description=_t['batch_pdf_description'],
                    summary_rows=_batch_summary_rows,
                    table_header=_batch_pdf_table_header,
                    table_rows=_batch_pdf_table_rows,
                    chart_png_bytes=_chart_bytes,
                    chart_caption="Figure. Ratio (95% CI) per sample. Dashed line = expected (no-change) ratio.",
                    footer_note=f"AbsoluteGene v{APP_VERSION} — For research and educational use only. Not validated for clinical diagnostic purposes.",
                    references=[
                        "Poisson-based dPCR quantification: Dube S, Qin J, Ramakrishnan R (2008). PLoS ONE, 3(8), e2876.",
                        "digital MIQE guidelines: Huggett JF et al. (2013). Clin Chem, 59(6), 892-902.",
                    ]
                )
                st.download_button(
                    "⬇️ " + _t['batch_pdf_report'], data=batch_pdf_buffer,
                    file_name="batch_screening_report.pdf", mime="application/pdf", key="batch_pdf_dl"
                )
        else:
            st.info(_t['batch_no_data'])


    # ═══════════════════════════════════════════════════════════════════════════════
    # VAF / MUTATION FRACTION CALCULATOR TAB
    # ═══════════════════════════════════════════════════════════════════════════════
    with tab_vaf:
        st.markdown(f"### {_t['vaf_title']}")
        st.caption(_t['vaf_description'])
        st.info(_t['vaf_method_note'])
        st.download_button(
            _t['download_example_csv'], data=EXAMPLE_CSV_VAF.encode("utf-8"),
            file_name="example_vaf.csv", mime="text/csv", key="dl_example_vaf"
        )
        st.markdown("---")

        vaf_file = st.file_uploader(_t['vaf_uploader'], type=["csv", "tsv", "txt"], key="vaf_uploader")

        if vaf_file is not None:
            _vaf_raw_bytes = vaf_file.read()
            _vaf_raw_df, _vaf_detected, _vaf_err = parse_instrument_csv(_vaf_raw_bytes)
            if _vaf_err:
                st.error(_t['csv_parse_error'].format(err=_vaf_err))
            else:
                _vaf_col_options = list(_vaf_raw_df.columns)
                _vaf_col_map = {}
                vcol1, vcol2, vcol3, vcol4 = st.columns(4)
                for _field, _label_key, _col_widget in [
                    ("sample", "csv_col_sample", vcol1), ("target", "csv_col_target", vcol2),
                    ("positives", "csv_col_positives", vcol3), ("total", "csv_col_total", vcol4)
                ]:
                    _detected = _vaf_detected.get(_field)
                    _default_idx = _vaf_col_options.index(_detected) if _detected in _vaf_col_options else 0
                    with _col_widget:
                        _vaf_col_map[_field] = st.selectbox(
                            _t[_label_key], options=_vaf_col_options,
                            index=_default_idx, key=f"vaf_colmap_{_field}"
                        )

                _vaf_std_df = build_standard_import_df(_vaf_raw_df, _vaf_col_map)

                if not _vaf_std_df.empty:
                    st.markdown(_t['vaf_assay_title'])
                    _vaf_unique_targets = sorted(_vaf_std_df["Target"].unique())
                    vacol1, vacol2 = st.columns(2)
                    with vacol1:
                        _vaf_mutant_assay = st.selectbox(
                            _t['vaf_mutant_label'], options=_vaf_unique_targets, key="vaf_mutant_assay"
                        )
                    with vacol2:
                        _vaf_wt_options = [t for t in _vaf_unique_targets if t != _vaf_mutant_assay]
                        _vaf_wt_assay = st.selectbox(
                            _t['vaf_wt_label'], options=_vaf_wt_options, key="vaf_wt_assay"
                        ) if _vaf_wt_options else None
                    _vaf_dilution = render_dilution_input("vaf")

                    if st.button(_t['vaf_run_btn'], key="vaf_run_btn", use_container_width=True) and _vaf_wt_assay:
                        _vaf_results = pool_and_compute_vaf(
                            _vaf_std_df, _vaf_mutant_assay, _vaf_wt_assay, partition_vol_nl
                        )
                        st.session_state["_vaf_results_cache"] = _vaf_results
                        st.session_state["_vaf_mutant_assay_cache"] = _vaf_mutant_assay
                        st.session_state["_vaf_wt_assay_cache"] = _vaf_wt_assay
                        st.session_state["_vaf_dilution_cache"] = _vaf_dilution

        # ── Display cached results ──────────────────────────────────────────────────
        _vaf_results = st.session_state.get("_vaf_results_cache")
        if _vaf_results:
            st.markdown(f"#### {_t['vaf_results_title']}")
            _v_dilution = st.session_state.get("_vaf_dilution_cache", 1.0)
            _vaf_rows = []
            _n_detected = 0
            for r in _vaf_results:
                if r["detected"]:
                    _n_detected += 1
                _vaf_rows.append({
                    _t['vaf_col_sample']: r["Sample"],
                    _t['vaf_col_lambda_mut']: round(r["lambda_mut"], 5),
                    _t['vaf_col_lambda_wt']: round(r["lambda_wt"], 5),
                    _t['vaf_col_fa']: round(r["fa"] * 100, 3),
                    _t['vaf_col_ci']: (f"{r['ci_low']*100:.3f}\u2013{r['ci_high']*100:.3f}"
                                         if not np.isnan(r["ci_low"]) else "—"),
                    _t['vaf_col_conc_mut']: round(r["conc_mut"], 2),
                    _t['stock_conc_col']: round(r["conc_mut"] * _v_dilution, 2),
                    _t['vaf_col_detected']: _t['vaf_detected_yes'] if r["detected"] else _t['vaf_detected_no'],
                })

            st.info(_t['vaf_n_detected'].format(n=_n_detected, total=len(_vaf_results)))
            vaf_df = pd.DataFrame(_vaf_rows)
            st.dataframe(vaf_df, use_container_width=True)

            vaf_csv = vaf_df.to_csv(index=False).encode("utf-8")
            st.download_button(_t['vaf_download_btn'], data=vaf_csv,
                                file_name="vaf_results.csv", mime="text/csv", key="vaf_dl_csv")

            # ── VAF plot ───────────────────────────────────────────────────────────
            fig_vaf = go.Figure()
            _vaf_samples = [r["Sample"] for r in _vaf_results]
            _vaf_pct = [r["fa"] * 100 for r in _vaf_results]
            _vaf_err_low = [max((r["fa"] - r["ci_low"]) * 100, 0) if not np.isnan(r["ci_low"]) else 0 for r in _vaf_results]
            _vaf_err_high = [(r["ci_high"] - r["fa"]) * 100 if not np.isnan(r["ci_high"]) else 0 for r in _vaf_results]
            _vaf_colors = ["#e53935" if r["detected"] else "#9e9e9e" for r in _vaf_results]

            fig_vaf.add_trace(go.Bar(
                x=_vaf_samples, y=_vaf_pct, marker_color=_vaf_colors,
                error_y=dict(type="data", symmetric=False, array=_vaf_err_high, arrayminus=_vaf_err_low),
                name="FA%"
            ))
            fig_vaf.update_layout(
                title=_t['vaf_chart_title'], xaxis_title=_t['vaf_col_sample'],
                yaxis_title=_t['vaf_col_fa'], height=420, showlegend=False
            )
            st.plotly_chart(fig_vaf, use_container_width=True, key="vaf_chart")

            # ── PDF report ────────────────────────────────────────────────────────
            if st.button(_t['vaf_pdf_btn'], key="vaf_pdf_btn"):
                try:
                    fig_mpl, ax_mpl = plt.subplots(figsize=(7, 3.5))
                    _mpl_vaf_colors = ["#e53935" if r["detected"] else "#9e9e9e" for r in _vaf_results]
                    ax_mpl.bar(range(len(_vaf_samples)), _vaf_pct, color=_mpl_vaf_colors,
                               yerr=[_vaf_err_low, _vaf_err_high], capsize=3)
                    ax_mpl.set_xticks(range(len(_vaf_samples)))
                    ax_mpl.set_xticklabels(_vaf_samples, rotation=45, ha='right', fontsize=7)
                    ax_mpl.set_ylabel(_t['vaf_col_fa'], fontsize=9)
                    ax_mpl.set_title(_t['vaf_chart_title'], fontsize=10, fontweight='bold')
                    ax_mpl.spines['top'].set_visible(False)
                    ax_mpl.spines['right'].set_visible(False)
                    plt.tight_layout()
                    _img_buf = BytesIO()
                    plt.savefig(_img_buf, format='png', dpi=150, bbox_inches='tight')
                    plt.close()
                    _vaf_chart_bytes = _img_buf.getvalue()
                except Exception as _chart_err:
                    _vaf_chart_bytes = None
                    st.warning(f"⚠️ Chart could not be generated for the PDF ({_chart_err}). "
                               f"The report will still be created without this figure.")

                _vaf_mutant_cache = st.session_state.get("_vaf_mutant_assay_cache", "—")
                _vaf_wt_cache = st.session_state.get("_vaf_wt_assay_cache", "—")
                _vaf_summary_rows = [
                    ["Parameter", "Value"],
                    ["Mutant assay", _vaf_mutant_cache],
                    ["Wild-type assay", _vaf_wt_cache],
                    ["Samples analyzed", str(len(_vaf_results))],
                    ["Samples with mutant detected", f"{_n_detected} / {len(_vaf_results)}"],
                ]
                _vaf_pdf_table_header = [_t['vaf_col_sample'], _t['vaf_col_lambda_mut'], _t['vaf_col_lambda_wt'],
                                          _t['vaf_col_fa'], _t['vaf_col_ci'], _t['vaf_col_conc_mut'], _t['vaf_col_detected']]
                _vaf_pdf_table_rows = [[str(v) for v in row.values()] for row in _vaf_rows]

                vaf_pdf_buffer = create_simple_pdf(
                    report_title="AbsoluteGene",
                    subtitle=_t['vaf_pdf_report'],
                    description=_t['vaf_pdf_description'],
                    summary_rows=_vaf_summary_rows,
                    table_header=_vaf_pdf_table_header,
                    table_rows=_vaf_pdf_table_rows,
                    chart_png_bytes=_vaf_chart_bytes,
                    chart_caption="Figure. Fractional Abundance (FA%, 95% CI) per sample. Red = mutant detected.",
                    footer_note=f"AbsoluteGene v{APP_VERSION} — For research and educational use only. Not validated for clinical diagnostic purposes.",
                    references=[
                        "Delta-method CI for ddPCR fractional abundance: Hindson CM et al. (2013). Nat Methods, 10(10), 1003-1005.",
                        "digital MIQE guidelines: Huggett JF et al. (2013). Clin Chem, 59(6), 892-902.",
                    ]
                )
                st.download_button(
                    "⬇️ " + _t['vaf_pdf_report'], data=vaf_pdf_buffer,
                    file_name="vaf_report.pdf", mime="application/pdf", key="vaf_pdf_dl"
                )
        else:
            st.info(_t['vaf_no_data'])


    # ═══════════════════════════════════════════════════════════════════════════════
    # CLINICAL TOOLS TAB
    # ═══════════════════════════════════════════════════════════════════════════════
    with tab_clinical:
        st.markdown(f"### {_t['clinical_title']}")
        st.caption(_t['clinical_description'])
        st.markdown("---")

        if advanced_mode:

            clinical_mode = st.radio(
                _t['clinical_mode_label'],
                options=[_t['clinical_mode_mu'], _t['clinical_mode_rcv'],
                         _t['clinical_mode_precision'], _t['clinical_mode_comparison']],
                key="clinical_mode", horizontal=True
            )
            st.markdown("---")

            # ── MU Budget ──────────────────────────────────────────────────────────────
            if clinical_mode == _t['clinical_mode_mu']:
                st.caption(_t['mu_description'])

                _mu_gene_options = [f"{r['__gene__']} / {r['__group__']}" for r in data] if data else []
                mu_c1, mu_c2 = st.columns(2)
                with mu_c1:
                    _mu_source_mode = st.radio(_t['mu_poisson_source'],
                                                options=[_t['mu_poisson_auto_replicate'], _t['mu_poisson_auto_theoretical'],
                                                         _t['mu_poisson_manual']],
                                                key="mu_source_mode")
                with mu_c2:
                    if _mu_source_mode == _t['mu_poisson_auto_replicate'] and _mu_gene_options:
                        _mu_selected = st.selectbox(_t['mu_gene_select'], options=_mu_gene_options, key="mu_gene_select")
                        _mu_match = data[_mu_gene_options.index(_mu_selected)]
                        _u_poisson_val = _mu_match.get("__conc_smp_cv__", 0.0)
                        _u_poisson_val = 0.0 if (_u_poisson_val is None or np.isnan(_u_poisson_val)) else _u_poisson_val
                        st.metric(_t['mu_poisson_source'], f"{_u_poisson_val:.2f}%")
                        st.warning(_t['mu_double_count_warning'])
                    elif _mu_source_mode == _t['mu_poisson_auto_theoretical'] and _mu_gene_options:
                        _mu_selected = st.selectbox(_t['mu_gene_select'], options=_mu_gene_options, key="mu_gene_select_theo")
                        _mu_match = data[_mu_gene_options.index(_mu_selected)]
                        _u_poisson_val = _mu_match.get("__poisson_rel_se_pct__", np.nan)
                        if _u_poisson_val is not None and not np.isnan(_u_poisson_val):
                            st.metric(_t['mu_poisson_source'], f"{_u_poisson_val:.2f}%")
                            st.caption(_t['mu_theoretical_note'])
                        else:
                            _u_poisson_val = 0.0
                            st.info(_t['mu_no_lambda_cache'])
                    else:
                        _u_poisson_val = st.number_input(_t['mu_poisson_source'], min_value=0.0, max_value=100.0,
                                                          value=5.0, step=0.1, key="mu_poisson_manual_val",
                                                          help=_t['mu_poisson_help'])

                mu_c3, mu_c4, mu_c5 = st.columns(3)
                with mu_c3:
                    _u_pipetting_val = st.number_input(_t['mu_pipetting_label'], min_value=0.0, max_value=50.0,
                                                        value=1.5, step=0.1, key="mu_pipetting_val",
                                                        help=_t['mu_pipetting_help'])
                with mu_c4:
                    _u_precision_val = st.number_input(_t['mu_precision_label'], min_value=0.0, max_value=50.0,
                                                        value=0.0, step=0.1, key="mu_precision_val",
                                                        help=_t['mu_precision_help'])
                with mu_c5:
                    _mu_k = st.number_input(_t['mu_k_label'], min_value=1.0, max_value=3.0, value=2.0,
                                             step=0.5, key="mu_k_val")

                if (_mu_source_mode == _t['mu_poisson_auto_replicate']) and (_u_pipetting_val > 0 or _u_precision_val > 0):
                    st.error(_t['mu_double_count_error'])

                if st.button(_t['mu_calc_btn'], key="mu_calc_btn"):
                    _mu_result = compute_mu_budget(_u_poisson_val, _u_pipetting_val, _u_precision_val, _mu_k)
                    _val_label = _mu_selected if (_mu_source_mode == _t['mu_poisson_auto'] and _mu_gene_options) else "—"
                    st.success(_t['mu_result_title'].format(value=_val_label, U=_mu_result["U_pct"], k=_mu_k))

                    st.markdown(_t['mu_budget_table_title'])
                    _mu_table_rows = [{_t['mu_col_source']: k, _t['mu_col_contribution']: f"{v:.3f}%"}
                                       for k, v in _mu_result["components"].items()]
                    _mu_table_rows.append({_t['mu_col_source']: _t['mu_col_combined'], _t['mu_col_contribution']: f"{_mu_result['u_c_pct']:.3f}%"})
                    _mu_table_rows.append({_t['mu_col_source']: f"{_t['mu_col_expanded']} (k={_mu_k})", _t['mu_col_contribution']: f"{_mu_result['U_pct']:.3f}%"})
                    st.dataframe(pd.DataFrame(_mu_table_rows), use_container_width=True)

                    fig_mu = go.Figure(go.Bar(
                        x=list(_mu_result["components"].keys()), y=list(_mu_result["components"].values()),
                        marker_color="#00796b"
                    ))
                    fig_mu.update_layout(title="Uncertainty Component Contributions", yaxis_title="Relative uncertainty (%)", height=350)
                    st.plotly_chart(fig_mu, use_container_width=True, key="mu_chart")

            # ── RCV ────────────────────────────────────────────────────────────────────
            elif clinical_mode == _t['clinical_mode_rcv']:
                st.caption(_t['rcv_description'])
                rcv_c1, rcv_c2 = st.columns(2)
                with rcv_c1:
                    _rcv_result1 = st.number_input(_t['rcv_result1_label'], value=10.0, step=0.1, key="rcv_result1")
                with rcv_c2:
                    _rcv_result2 = st.number_input(_t['rcv_result2_label'], value=12.0, step=0.1, key="rcv_result2")
                rcv_c3, rcv_c4, rcv_c5 = st.columns(3)
                with rcv_c3:
                    _rcv_cv_a = st.number_input(_t['rcv_cv_analytical_label'], min_value=0.0, max_value=100.0,
                                                 value=5.0, step=0.1, key="rcv_cv_a", help=_t['rcv_cv_analytical_help'])
                with rcv_c4:
                    _rcv_cv_b = st.number_input(_t['rcv_cv_biological_label'], min_value=0.0, max_value=200.0,
                                                 value=15.0, step=0.5, key="rcv_cv_b", help=_t['rcv_cv_biological_help'])
                with rcv_c5:
                    _rcv_z = st.number_input(_t['rcv_z_label'], min_value=1.0, max_value=3.0, value=1.96,
                                              step=0.01, key="rcv_z")

                if st.button(_t['rcv_calc_btn'], key="rcv_calc_btn"):
                    _rcv_res = compute_rcv(_rcv_result1, _rcv_result2, _rcv_cv_a, _rcv_cv_b, _rcv_z)
                    if _rcv_res is None:
                        st.error("Invalid input.")
                    else:
                        if _rcv_res["significant"]:
                            st.error(_t['rcv_result_significant'].format(change=_rcv_res["percent_change"], rcv=_rcv_res["rcv_pct"]))
                        else:
                            st.success(_t['rcv_result_not_significant'].format(change=_rcv_res["percent_change"], rcv=_rcv_res["rcv_pct"]))
                        rc1, rc2 = st.columns(2)
                        rc1.metric("RCV (%)", f"±{_rcv_res['rcv_pct']:.2f}%")
                        rc2.metric("Observed Change (%)", f"{_rcv_res['percent_change']:+.2f}%")

            # ── Precision Study ───────────────────────────────────────────────────────
            elif clinical_mode == _t['clinical_mode_precision']:
                st.caption(_t['precision_description'])
                _prec_n_days = st.number_input(_t['precision_n_days'], min_value=2, max_value=20, value=3, step=1, key="prec_n_days")
                _prec_day_arrays = []
                _prec_cols = st.columns(min(int(_prec_n_days), 4))
                for _d in range(int(_prec_n_days)):
                    with _prec_cols[_d % len(_prec_cols)]:
                        _txt = st.text_area(_t['precision_day_label'].format(i=_d + 1), key=f"prec_day_{_d}", height=120)
                        _prec_day_arrays.append(parse_input_data(_txt))

                if st.button(_t['precision_calc_btn'], key="prec_calc_btn"):
                    _prec_result = compute_precision_study(_prec_day_arrays)
                    if _prec_result is None:
                        st.error("Insufficient data — need at least 2 days with ≥2 replicates each.")
                    else:
                        if not _prec_result["is_balanced"]:
                            st.info(_t['unbalanced_design_note'].format(n_list=_prec_result["n_per_day"]))
                        pc1, pc2, pc3, pc4 = st.columns(4)
                        pc1.metric(_t['precision_grand_mean'], f"{_prec_result['grand_mean']:.4f}")
                        pc2.metric(_t['precision_repeatability_cv'], f"{_prec_result['repeatability_cv']:.2f}%")
                        pc3.metric(_t['precision_between_day_cv'], f"{_prec_result['between_day_cv']:.2f}%")
                        pc4.metric(_t['precision_total_cv'], f"{_prec_result['total_cv']:.2f}%")

                        fig_prec = go.Figure()
                        for _d, arr in enumerate(_prec_day_arrays):
                            if len(arr) > 0:
                                fig_prec.add_trace(go.Box(y=arr, name=f"Day {_d+1}", boxpoints="all"))
                        fig_prec.update_layout(title="Precision Study — Per-Day Distribution", height=380)
                        st.plotly_chart(fig_prec, use_container_width=True, key="prec_chart")

            # ── Method Comparison ─────────────────────────────────────────────────────
            else:
                st.caption(_t['comparison_description'])
                comp_c1, comp_c2 = st.columns(2)
                with comp_c1:
                    _comp_m1_txt = st.text_area(_t['comparison_method1_label'], key="comp_m1", height=150)
                with comp_c2:
                    _comp_m2_txt = st.text_area(_t['comparison_method2_label'], key="comp_m2", height=150)
                _comp_variance_ratio = st.number_input(_t['comparison_variance_ratio_label'], min_value=0.01, max_value=100.0,
                                                        value=1.0, step=0.1, key="comp_var_ratio")

                if st.button(_t['comparison_calc_btn'], key="comp_calc_btn"):
                    _m1_arr = parse_input_data(_comp_m1_txt)
                    _m2_arr = parse_input_data(_comp_m2_txt)
                    _ba_result = compute_bland_altman(_m1_arr, _m2_arr)
                    _deming_result = compute_deming_regression(_m1_arr, _m2_arr, _comp_variance_ratio)
                    _pb_result = compute_passing_bablok(_m1_arr, _m2_arr)

                    if _ba_result is None:
                        st.error("Need at least 2 paired values.")
                    else:
                        bc1, bc2, bc3 = st.columns(3)
                        bc1.metric(_t['comparison_bias_label'], f"{_ba_result['mean_diff']:.4f}")
                        bc2.metric(_t['comparison_loa_label'], f"{_ba_result['loa_low']:.3f} to {_ba_result['loa_high']:.3f}")
                        if _deming_result:
                            bc3.metric(_t['comparison_deming_label'], f"y={_deming_result['slope']:.3f}x+{_deming_result['intercept']:.3f}")

                        # ── Assumption checks (normality, proportional bias) ──────────────
                        if not np.isnan(_ba_result["shapiro_p"]):
                            if _ba_result["shapiro_p"] >= 0.05:
                                st.success(_t['comparison_normality_ok'].format(p=_ba_result["shapiro_p"]))
                            else:
                                st.warning(_t['comparison_normality_warn'].format(p=_ba_result["shapiro_p"]))
                        if not np.isnan(_ba_result["proportional_bias_p"]):
                            if _ba_result["proportional_bias_p"] >= 0.05:
                                st.success(_t['comparison_prop_bias_ok'].format(p=_ba_result["proportional_bias_p"]))
                            else:
                                st.warning(_t['comparison_prop_bias_warn'].format(
                                    slope=_ba_result["proportional_bias_slope"], p=_ba_result["proportional_bias_p"]))

                        # ── Regression method comparison table (Deming vs Passing-Bablok) ──
                        if _deming_result or _pb_result:
                            st.markdown(f"**{_t['comparison_deming_label']} vs {_t['comparison_pb_label']}**")
                            _reg_rows = []
                            if _deming_result:
                                _slope_ci = (f"{_deming_result['slope_ci_low']:.4f}–{_deming_result['slope_ci_high']:.4f}"
                                              if not np.isnan(_deming_result['slope_ci_low']) else "—")
                                _int_ci = (f"{_deming_result['intercept_ci_low']:.4f}–{_deming_result['intercept_ci_high']:.4f}"
                                           if not np.isnan(_deming_result['intercept_ci_low']) else "—")
                                _reg_rows.append({
                                    "Method": _t['comparison_deming_label'],
                                    "Slope": round(_deming_result["slope"], 4),
                                    f"Slope {_t['comparison_deming_ci_label']}": _slope_ci,
                                    "Intercept": round(_deming_result["intercept"], 4),
                                    "Intercept 95% CI": _int_ci,
                                })
                            if _pb_result:
                                _pb_slope_ci = (f"{_pb_result['slope_ci_low']:.4f}–{_pb_result['slope_ci_high']:.4f}"
                                                 if not np.isnan(_pb_result['slope_ci_low']) else "—")
                                _pb_int_ci = (f"{_pb_result['intercept_ci_low']:.4f}–{_pb_result['intercept_ci_high']:.4f}"
                                              if not np.isnan(_pb_result['intercept_ci_low']) else "—")
                                _reg_rows.append({
                                    "Method": _t['comparison_pb_label'],
                                    "Slope": round(_pb_result["slope"], 4),
                                    f"Slope {_t['comparison_deming_ci_label']}": _pb_slope_ci,
                                    "Intercept": round(_pb_result["intercept"], 4),
                                    "Intercept 95% CI": _pb_int_ci,
                                })
                            st.dataframe(pd.DataFrame(_reg_rows), use_container_width=True)
                            st.caption(_t['comparison_pb_note'])

                        fig_ba = go.Figure()
                        fig_ba.add_trace(go.Scatter(x=_ba_result["means"], y=_ba_result["diffs"], mode="markers",
                                                     marker=dict(color="#00796b")))
                        fig_ba.add_hline(y=_ba_result["mean_diff"], line_color="black", line_dash="solid")
                        fig_ba.add_hline(y=_ba_result["loa_low"], line_color="red", line_dash="dash")
                        fig_ba.add_hline(y=_ba_result["loa_high"], line_color="red", line_dash="dash")
                        fig_ba.update_layout(title=_t['comparison_ba_chart_title'], xaxis_title="Mean of methods",
                                              yaxis_title="Difference (M2-M1)", height=380)
                        st.plotly_chart(fig_ba, use_container_width=True, key="ba_chart")

                        if _deming_result or _pb_result:
                            fig_dem = go.Figure()
                            fig_dem.add_trace(go.Scatter(x=_m1_arr, y=_m2_arr, mode="markers", marker=dict(color="#00796b"), name="Data"))
                            _x_line = np.array([min(_m1_arr), max(_m1_arr)])
                            if _deming_result:
                                _y_line = _deming_result["slope"] * _x_line + _deming_result["intercept"]
                                fig_dem.add_trace(go.Scatter(x=_x_line, y=_y_line, mode="lines", line=dict(color="red"),
                                                              name=_t['comparison_deming_label']))
                            if _pb_result:
                                _y_line_pb = _pb_result["slope"] * _x_line + _pb_result["intercept"]
                                fig_dem.add_trace(go.Scatter(x=_x_line, y=_y_line_pb, mode="lines", line=dict(color="blue", dash="dot"),
                                                              name=_t['comparison_pb_label']))
                            fig_dem.update_layout(title=_t['comparison_deming_chart_title'], xaxis_title="Method 1",
                                                   yaxis_title="Method 2", height=380)
                            st.plotly_chart(fig_dem, use_container_width=True, key="deming_chart")


        # ═══════════════════════════════════════════════════════════════════════════════
        # CRM PRODUCTION TAB
        # ═══════════════════════════════════════════════════════════════════════════════
        else:
            st.info(f"{_t['advanced_gate_title']}\n\n{_t['advanced_gate_message']}")


    with tab_crm:
        st.markdown(f"### {_t['crm_title']}")
        st.caption(_t['crm_description'])
        st.markdown("---")

        if advanced_mode:

            crm_mode = st.radio(
                _t['crm_mode_label'],
                options=[_t['crm_mode_homogeneity'], _t['crm_mode_stability'],
                         _t['crm_mode_uncertainty'], _t['crm_mode_equivalence'], _t['crm_mode_coa']],
                key="crm_mode", horizontal=True
            )
            st.markdown("---")

            # ── Homogeneity Testing ───────────────────────────────────────────────────
            if crm_mode == _t['crm_mode_homogeneity']:
                st.caption(_t['homog_description'])
                _homog_n_units = st.number_input(_t['homog_n_units'], min_value=2, max_value=30, value=10, step=1, key="homog_n_units")
                _homog_unit_arrays = []
                _homog_cols = st.columns(min(int(_homog_n_units), 5))
                for _u in range(int(_homog_n_units)):
                    with _homog_cols[_u % len(_homog_cols)]:
                        _txt = st.text_area(_t['homog_unit_label'].format(i=_u + 1), key=f"homog_unit_{_u}", height=100)
                        _homog_unit_arrays.append(parse_input_data(_txt))

                if st.button(_t['homog_calc_btn'], key="homog_calc_btn"):
                    _homog_result = compute_homogeneity(_homog_unit_arrays)
                    if _homog_result is None:
                        st.error("Insufficient data — need at least 2 units with ≥2 replicates each.")
                    else:
                        st.session_state["_homog_result_cache"] = _homog_result
                        if _homog_result["is_homogeneous"]:
                            st.success(_t['homog_result_homogeneous'].format(
                                F=_homog_result["F"], fcrit=_homog_result["F_crit"], p=_homog_result["p_value"]))
                        else:
                            st.error(_t['homog_result_inhomogeneous'].format(
                                F=_homog_result["F"], fcrit=_homog_result["F_crit"], p=_homog_result["p_value"]))
                        if not _homog_result["is_balanced"]:
                            st.info(_t['unbalanced_design_note'].format(n_list=_homog_result["n_per_unit"]))

                        hc1, hc2, hc3 = st.columns(3)
                        hc1.metric(_t['homog_grand_mean_label'], f"{_homog_result['grand_mean']:.5f}")
                        hc2.metric(_t['homog_ubb_label'], f"{_homog_result['u_bb']:.5f}")
                        hc3.metric("F / F_crit", f"{_homog_result['F']:.3f} / {_homog_result['F_crit']:.3f}")

                        fig_homog = go.Figure()
                        for _u, arr in enumerate(_homog_unit_arrays):
                            if len(arr) > 0:
                                fig_homog.add_trace(go.Box(y=arr, name=f"Unit {_u+1}", boxpoints="all"))
                        fig_homog.add_hline(y=_homog_result["grand_mean"], line_dash="dash", line_color="black")
                        fig_homog.update_layout(title="Homogeneity — Per-Unit Distribution", height=400)
                        st.plotly_chart(fig_homog, use_container_width=True, key="homog_chart")

            # ── Stability Testing ─────────────────────────────────────────────────────
            elif crm_mode == _t['crm_mode_stability']:
                st.caption(_t['stab_description'])
                stab_c1, stab_c2 = st.columns(2)
                with stab_c1:
                    _stab_time_txt = st.text_area(_t['stab_time_label'], key="stab_time", height=150)
                with stab_c2:
                    _stab_value_txt = st.text_area(_t['stab_value_label'], key="stab_value", height=150)
                _stab_duration = st.number_input(_t['stab_duration_label'], min_value=0.1, value=365.0, step=1.0, key="stab_duration")

                if st.button(_t['stab_calc_btn'], key="stab_calc_btn"):
                    _stab_time_arr = parse_input_data(_stab_time_txt)
                    _stab_value_arr = parse_input_data(_stab_value_txt)
                    _stab_result = compute_stability(_stab_time_arr, _stab_value_arr, _stab_duration)
                    if _stab_result is None:
                        st.error("Need at least 3 paired time/value points.")
                    else:
                        st.session_state["_stab_result_cache"] = _stab_result
                        if _stab_result["is_stable"]:
                            st.success(_t['stab_result_stable'].format(p=_stab_result["p_value"]))
                        else:
                            st.warning(_t['stab_result_unstable'].format(p=_stab_result["p_value"]))

                        sc1, sc2, sc3 = st.columns(3)
                        sc1.metric(_t['stab_slope_label'], f"{_stab_result['slope']:.6f}")
                        sc2.metric(_t['stab_ustab_label'], f"{_stab_result['u_stab']:.6f}")
                        sc3.metric("R²", f"{_stab_result['r_value']**2:.4f}")

                        fig_stab = go.Figure()
                        fig_stab.add_trace(go.Scatter(x=_stab_time_arr, y=_stab_value_arr, mode="markers",
                                                       marker=dict(color="#00796b", size=9), name="Data"))
                        _x_line = np.array([min(_stab_time_arr), max(_stab_time_arr)])
                        _y_line = _stab_result["slope"] * _x_line + _stab_result["intercept"]
                        fig_stab.add_trace(go.Scatter(x=_x_line, y=_y_line, mode="lines", line=dict(color="red"), name="Trend"))
                        fig_stab.update_layout(title=_t['stab_chart_title'], xaxis_title="Time", yaxis_title="Value", height=400)
                        st.plotly_chart(fig_stab, use_container_width=True, key="stab_chart")

            # ── Assigned Value & Uncertainty Budget ───────────────────────────────────
            elif crm_mode == _t['crm_mode_uncertainty']:
                st.caption(_t['unc_description'])

                _cached_homog = st.session_state.get("_homog_result_cache")
                _cached_stab = st.session_state.get("_stab_result_cache")
                _use_cached = st.checkbox(_t['unc_use_cached'], value=True, key="unc_use_cached")
                if _use_cached and not (_cached_homog or _cached_stab):
                    st.info(_t['unc_no_cache'])

                unc_c1, unc_c2 = st.columns(2)
                with unc_c1:
                    _unc_assigned_value = st.number_input(_t['unc_assigned_value_label'], value=100.0, step=0.1, key="unc_assigned_value")
                    _unc_u_char = st.number_input(_t['unc_u_char_label'], min_value=0.0, value=1.0, step=0.1,
                                                   key="unc_u_char", help=_t['unc_u_char_help'])
                with unc_c2:
                    _default_u_bb = _cached_homog["u_bb"] if (_use_cached and _cached_homog) else 0.0
                    _unc_u_bb = st.number_input(_t['unc_u_bb_label'], min_value=0.0, value=float(_default_u_bb), step=0.01, key="unc_u_bb")
                    _default_u_stab = _cached_stab["u_stab"] if (_use_cached and _cached_stab) else 0.0
                    _unc_u_stab = st.number_input(_t['unc_u_stab_label'], min_value=0.0, value=float(_default_u_stab), step=0.01, key="unc_u_stab")

                _unc_k = st.number_input(_t['unc_k_label'], min_value=1.0, max_value=3.0, value=2.0, step=0.5, key="unc_k")

                if st.button(_t['unc_calc_btn'], key="unc_calc_btn"):
                    _unc_result = compute_assigned_value_uncertainty(_unc_assigned_value, _unc_u_char, _unc_u_bb, _unc_u_stab, k=_unc_k)
                    st.session_state["_unc_result_cache"] = _unc_result
                    st.success(_t['unc_result_title'].format(
                        value=_unc_result["assigned_value"], U=_unc_result["U"], k=_unc_k, urel=_unc_result["U_rel_pct"]))

                    _unc_table_rows = [{_t['mu_col_source']: k, _t['mu_col_contribution']: f"{v:.5f}"}
                                        for k, v in _unc_result["components"].items()]
                    _unc_table_rows.append({_t['mu_col_source']: _t['mu_col_combined'], _t['mu_col_contribution']: f"{_unc_result['u_c']:.5f}"})
                    _unc_table_rows.append({_t['mu_col_source']: f"{_t['mu_col_expanded']} (k={_unc_k})", _t['mu_col_contribution']: f"{_unc_result['U']:.5f}"})
                    st.dataframe(pd.DataFrame(_unc_table_rows), use_container_width=True)

                    fig_unc = go.Figure(go.Bar(
                        x=list(_unc_result["components"].keys()), y=list(_unc_result["components"].values()),
                        marker_color="#00695c"
                    ))
                    fig_unc.update_layout(title="Uncertainty Component Contributions (absolute units)", height=350)
                    st.plotly_chart(fig_unc, use_container_width=True, key="unc_chart")

            # ── Batch-to-Batch (Lot) Equivalence ──────────────────────────────────────
            elif crm_mode == _t['crm_mode_equivalence']:
                st.caption(_t['equiv_description'])
                eq_c1, eq_c2 = st.columns(2)
                with eq_c1:
                    _eq_lot1_txt = st.text_area(_t['equiv_lot1_label'], key="eq_lot1", height=150)
                with eq_c2:
                    _eq_lot2_txt = st.text_area(_t['equiv_lot2_label'], key="eq_lot2", height=150)
                _eq_margin = st.number_input(_t['equiv_margin_label'], min_value=0.1, max_value=100.0,
                                              value=10.0, step=0.5, key="eq_margin")

                if st.button(_t['equiv_calc_btn'], key="eq_calc_btn"):
                    _eq_lot1_arr = parse_input_data(_eq_lot1_txt)
                    _eq_lot2_arr = parse_input_data(_eq_lot2_txt)
                    _eq_result = compute_lot_equivalence(_eq_lot1_arr, _eq_lot2_arr, _eq_margin)
                    if _eq_result is None:
                        st.error("Need at least 2 values per lot.")
                    else:
                        if _eq_result["is_equivalent"]:
                            st.success(_t['equiv_result_equivalent'].format(
                                diff=_eq_result["percent_diff"], lo=_eq_result["ci_low_pct"],
                                hi=_eq_result["ci_high_pct"], margin=_eq_margin))
                        else:
                            st.error(_t['equiv_result_not_equivalent'].format(
                                diff=_eq_result["percent_diff"], lo=_eq_result["ci_low_pct"],
                                hi=_eq_result["ci_high_pct"], margin=_eq_margin))

                        eqc1, eqc2, eqc3 = st.columns(3)
                        eqc1.metric(_t['equiv_mean1_label'], f"{_eq_result['mean1']:.4f}")
                        eqc2.metric(_t['equiv_mean2_label'], f"{_eq_result['mean2']:.4f}")
                        eqc3.metric("% Difference", f"{_eq_result['percent_diff']:+.2f}%")

                        fig_eq = go.Figure()
                        fig_eq.add_trace(go.Bar(x=["% Difference"], y=[_eq_result["percent_diff"]],
                                                 marker_color="#00796b",
                                                 error_y=dict(type="data", symmetric=False,
                                                               array=[_eq_result["ci_high_pct"] - _eq_result["percent_diff"]],
                                                               arrayminus=[_eq_result["percent_diff"] - _eq_result["ci_low_pct"]])))
                        fig_eq.add_hline(y=_eq_margin, line_dash="dash", line_color="red")
                        fig_eq.add_hline(y=-_eq_margin, line_dash="dash", line_color="red")
                        fig_eq.update_layout(title=_t['equiv_chart_title'], yaxis_title="% Difference (90% CI)", height=400)
                        st.plotly_chart(fig_eq, use_container_width=True, key="eq_chart")

            # ── CoA Generator ──────────────────────────────────────────────────────────
            else:
                st.caption(_t['coa_description'])
                _cached_unc = st.session_state.get("_unc_result_cache")

                coa_c1, coa_c2 = st.columns(2)
                with coa_c1:
                    _coa_material = st.text_input(_t['coa_material_name'], value="", key="coa_material")
                    _coa_lot = st.text_input(_t['coa_lot_number'], value="", key="coa_lot")
                    _coa_producer = st.text_input(_t['coa_producer'], value="", key="coa_producer")
                with coa_c2:
                    _default_val = _cached_unc["assigned_value"] if _cached_unc else 100.0
                    _default_U = _cached_unc["U"] if _cached_unc else 2.0
                    _default_k = _cached_unc["k"] if _cached_unc else 2.0
                    _coa_value = st.number_input(_t['coa_assigned_value_label'], value=float(_default_val), step=0.01, key="coa_value")
                    _coa_unit = st.text_input(_t['coa_unit_label'], value="copies/µL", key="coa_unit")
                    _coa_U = st.number_input(_t['coa_expanded_unc_label'], value=float(_default_U), step=0.01, key="coa_U")
                    _coa_k = st.number_input(_t['coa_k_label'], value=float(_default_k), step=0.5, key="coa_k")

                _coa_traceability = st.text_area(_t['coa_traceability_label'], value=_t['coa_traceability_default'], key="coa_traceability")
                _coa_validity = st.text_input(_t['coa_validity_label'], value="", key="coa_validity")

                if st.button(_t['coa_generate_btn'], key="coa_generate_btn"):
                    _coa_summary_rows = [
                        ["Parameter", "Value"],
                        ["Material", _coa_material or "—"],
                        ["Batch/Lot", _coa_lot or "—"],
                        ["Producer/Laboratory", _coa_producer or "—"],
                        ["Assigned Value", f"{_coa_value} {_coa_unit}"],
                        ["Expanded Uncertainty (U)", f"± {_coa_U} {_coa_unit} (k={_coa_k})"],
                        ["Validity / Shelf Life", _coa_validity or "—"],
                    ]
                    coa_pdf_buffer = create_simple_pdf(
                        report_title="Certificate of Analysis",
                        subtitle=_coa_material or "AbsoluteGene CRM Certificate",
                        description=_coa_traceability,
                        summary_rows=_coa_summary_rows,
                        table_header=None, table_rows=None, chart_png_bytes=None, chart_caption=None,
                        footer_note=(f"This certificate was generated by AbsoluteGene v{APP_VERSION} for research and internal "
                                      "documentation purposes. It does not constitute a formally accredited "
                                      "Certificate of Analysis under ISO 17034 unless issued by an accredited "
                                      "reference material producer following full validation."),
                        references=[
                            "ISO Guide 35:2017. Reference materials — Guidance for characterization and assessment of homogeneity and stability.",
                            "ISO 17034:2016. General requirements for the competence of reference material producers.",
                            "Linsinger TP et al. (2001). Homogeneity and stability of reference materials. Accred Qual Assur, 6, 20-25.",
                        ]
                    )
                    st.download_button(
                        _t['coa_download_btn'], data=coa_pdf_buffer,
                        file_name="certificate_of_analysis.pdf", mime="application/pdf", key="coa_pdf_dl"
                    )


        else:
            st.info(f"{_t['advanced_gate_title']}\n\n{_t['advanced_gate_message']}")


    with tab_report:
        st.markdown(f"### 📄 {_t['pdf_report']}")
        st.markdown("---")
        if not input_values_table:
            st.info(_t['error_no_data'])
        else:
            st.success('✅ ' + _t['pdf_ready'].format(n=len(input_values_table)))
            if st.button(f"📥 {_t['generate_pdf']}", key="pdf_btn"):
                pdf_buffer = create_pdf(data, stats_data, pd.DataFrame(input_values_table), language_code,
                                         multigroup_results=multigroup_results)
                st.download_button(
                    label=f"⬇️ {_t['pdf_report']}", data=pdf_buffer,
                    file_name="absolutegene_report.pdf", mime="application/pdf", key="pdf_dl"
                )

            st.markdown("---")
            st.markdown(f"#### {_t['excel_export_title']}")
            st.caption(_t['excel_export_description'])
            if st.button(_t['excel_export_btn'], key="excel_export_btn"):
                excel_buffer = create_excel_report(data, stats_data, input_values_table, language_code)
                st.download_button(
                    label=f"⬇️ {_t['excel_export_btn']}", data=excel_buffer,
                    file_name="absolutegene_report.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="excel_dl"
                )

            st.markdown("---")
            st.markdown(f"#### {_t['lims_export_title']}")
            st.caption(_t['lims_export_description'])
            lims_c1, lims_c2, lims_c3 = st.columns(3)
            with lims_c1:
                _lims_operator = st.text_input(_t['lims_operator_label'], value="", key="lims_operator")
            with lims_c2:
                _lims_instrument = st.text_input(_t['lims_instrument_label'], value="", key="lims_instrument")
            with lims_c3:
                import datetime as _dt
                _lims_run_date = st.text_input(_t['lims_run_date_label'],
                                                value=_dt.date.today().isoformat(), key="lims_run_date")

            if data:
                _lims_rows = []
                for r in data:
                    _lims_rows.append({
                        _t['lims_col_sample_id']: f"{r['__gene__']}_{r['__group__']}",
                        _t['lims_col_test_code']: "dPCR_RATIO",
                        _t['lims_col_analyte']: r["__gene__"],
                        _t['lims_col_result_value']: f"{r['__ratio_smp__']:.4f}",
                        _t['lims_col_result_unit']: "ratio",
                        _t['lims_col_reference']: (f"{r['__ratio_ctrl__']:.4f}" if r["__ratio_ctrl__"] is not None else ""),
                        _t['lims_col_flag']: r["__regulation__"],
                        _t['lims_col_result_date']: _lims_run_date,
                        _t['lims_col_instrument']: _lims_instrument,
                        _t['lims_col_operator']: _lims_operator,
                        _t['lims_col_comments']: r.get("__dynamic_range_flag__", ""),
                    })
                    _lims_rows.append({
                        _t['lims_col_sample_id']: f"{r['__gene__']}_{r['__group__']}",
                        _t['lims_col_test_code']: "dPCR_CONC",
                        _t['lims_col_analyte']: r["__gene__"],
                        _t['lims_col_result_value']: (f"{r['__conc_smp__']:.2f}"
                                                        if r.get("__conc_smp__") is not None and not np.isnan(r["__conc_smp__"]) else ""),
                        _t['lims_col_result_unit']: "copies/uL",
                        _t['lims_col_reference']: "",
                        _t['lims_col_flag']: r.get("__lod_loq_flag__", ""),
                        _t['lims_col_result_date']: _lims_run_date,
                        _t['lims_col_instrument']: _lims_instrument,
                        _t['lims_col_operator']: _lims_operator,
                        _t['lims_col_comments']: "",
                    })
                lims_df = pd.DataFrame(_lims_rows)
                lims_csv = lims_df.to_csv(index=False).encode("utf-8")
                st.download_button(_t['lims_export_btn'], data=lims_csv,
                                    file_name="absolutegene_lims_export.csv", mime="text/csv", key="lims_export_dl")
                with st.expander("Preview", expanded=False):
                    st.dataframe(lims_df, use_container_width=True)

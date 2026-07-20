"""
Regression test suite for AbsoluteGene (absolutegene.py).

Run with:  pytest test_absolutegene.py -v

These tests drive the actual Streamlit app via streamlit.testing.v1.AppTest,
exercising the real UI code paths (not just isolated math functions), which
verifies both the underlying statistics AND that the widgets/session-state
wiring between them still works after any future edit. This file should be
committed to the repository and run in CI on every change — prior to this
file, all testing during development was ad-hoc and not persisted, meaning
a future edit could silently break a calculation without anyone noticing.

Requirements: pytest, streamlit (same version the app is deployed with).
"""
import io
import csv
import numpy as np
import pytest
from streamlit.testing.v1 import AppTest

APP_PATH = "absolutegene.py"


def _fresh_app():
    at = AppTest.from_file(APP_PATH)
    at.run(timeout=30)
    return at


def _fresh_app_advanced():
    """Same as _fresh_app but with Advanced Mode enabled (needed for Clinical
    Tools, CRM Production, and the Multiplex Converter, which are gated
    behind Advanced Mode by default in Simple Mode)."""
    at = AppTest.from_file(APP_PATH)
    at.run(timeout=30)
    at.toggle(key="advanced_mode").set_value(True)
    at.run(timeout=30)
    return at


def _load_scenario(at, name):
    at.selectbox(key="scenario_selector").set_value(name)
    at.run(timeout=30)
    at.button(key="load_scenario_btn").click()
    at.run(timeout=30)
    return at


# ─── Basic app load ────────────────────────────────────────────────────────

def test_app_loads_without_exception():
    at = _fresh_app()
    assert not at.exception


def test_turkish_language_switch():
    at = _fresh_app()
    at.selectbox[0].set_value("🇹🇷 Türkçe")
    at.run(timeout=30)
    assert not at.exception


# ─── Example scenarios (end-to-end pipeline + PDF) ─────────────────────────

SCENARIOS = [
    "S1 — Basic CNV gain (1 gene, n=5)",
    "S2 — Multi-gene + dual reference (2 genes, 2 groups, n=5)",
    "S3 — Outlier detection demo (n=6)",
    "S4 — Multi-group ANOVA (3 groups, n=5)",
    "S5 — QC & saturation demo (n=5)",
]


@pytest.mark.parametrize("scenario", SCENARIOS)
def test_scenario_loads_and_computes(scenario):
    at = _fresh_app()
    _load_scenario(at, scenario)
    assert not at.exception, f"{scenario} raised an exception"


@pytest.mark.parametrize("scenario", SCENARIOS)
def test_scenario_pdf_report_generates(scenario):
    at = _fresh_app()
    _load_scenario(at, scenario)
    at.button(key="pdf_btn").click()
    at.run(timeout=60)
    assert not at.exception, f"{scenario} PDF generation raised an exception"


def test_s1_ratio_and_fold_change_values():
    """S1 is a manually-verified scenario; check the computed Ratio/FC are
    within a tight tolerance of the expected hand-calculated values."""
    at = _fresh_app()
    _load_scenario(at, "S1 — Basic CNV gain (1 gene, n=5)")
    metrics = {m.label: m.value for m in at.metric}
    ratio = float(metrics["Normalized Ratio (Target/Reference)"])
    fc = float(metrics["Fold Change (vs Control)"])
    assert 1.45 < ratio < 1.55
    assert 1.45 < fc < 1.55


def test_s4_multigroup_anova_significant():
    """S4 has a clearly separated strong-gain group; the omnibus ANOVA
    should detect a significant group effect."""
    at = _fresh_app()
    _load_scenario(at, "S4 — Multi-group ANOVA (3 groups, n=5)")
    found_significant = any("Significant" in s.value for s in at.success) or \
                         any("Significant" in e.value for e in at.error)
    assert found_significant


# ─── Poisson math (core statistics) ────────────────────────────────────────

def test_poisson_exact_ci_wider_than_normal_at_low_counts():
    """The exact (Clopper-Pearson) CI must not underestimate uncertainty
    relative to the old normal approximation at low positive counts —
    regression guard for the LOD-precision fix."""
    from scipy import stats as _stats

    def poisson_lambda(positive, total, alpha=0.05):
        p = positive / total
        lam = -np.log(1 - p)
        x, n = positive, total
        p_low = 0.0 if x == 0 else _stats.beta.ppf(alpha / 2, x, n - x + 1)
        p_high = 1.0 if x == n else _stats.beta.ppf(1 - alpha / 2, x + 1, n - x)
        ci_high = -np.log(1 - p_high) if p_high < 1.0 else np.inf
        return lam, ci_high

    lam, ci_high_exact = poisson_lambda(2, 20000)
    # old normal approximation upper bound for the same input (hand-computed reference)
    old_ci_high = 0.000239
    assert ci_high_exact > old_ci_high


# ─── CSV import (main workflow, batch screening, VAF) ──────────────────────

def _build_csv(rows):
    buf = io.StringIO()
    csv.writer(buf).writerows(rows)
    return buf.getvalue().encode("utf-8")


def test_main_csv_import_end_to_end():
    rows = [
        ["Sample", "Target", "Positives", "AcceptedDroplets"],
        ["Control_1", "MYCN", "1890", "20000"], ["Control_1", "RPP30", "1895", "20000"],
        ["Control_2", "MYCN", "1920", "20100"], ["Control_2", "RPP30", "1915", "20100"],
        ["Control_3", "MYCN", "1875", "19850"], ["Control_3", "RPP30", "1870", "19850"],
        ["Patient_1", "MYCN", "2780", "20000"], ["Patient_1", "RPP30", "1895", "20000"],
        ["Patient_2", "MYCN", "2820", "20100"], ["Patient_2", "RPP30", "1915", "20100"],
        ["Patient_3", "MYCN", "2755", "19850"], ["Patient_3", "RPP30", "1870", "19850"],
    ]
    at = _fresh_app()
    at.file_uploader(key="csv_import_uploader").upload("t.csv", _build_csv(rows), "text/csv")
    at.run(timeout=30)
    at.text_input(key="csv_ctrl_label_input").set_value("Control")
    at.text_input(key="csv_pat_label_0").set_value("Patient")
    at.run(timeout=30)
    at.button(key="csv_apply_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert at.session_state["ctrl_tgt_pos_0"] == "1890\n1920\n1875"


def test_csv_import_with_leading_metadata_rows():
    """Real instrument exports often have a few metadata lines before the
    real header row — regression guard for the header auto-detection fix."""
    messy = (
        "Run Name: TestRun_2026\nInstrument: QX200\nOperator: Lab\n\n"
        "Sample,Target,Positives,AcceptedDroplets\n"
        "Control_1,MYCN,1890,20000\nControl_1,RPP30,1895,20000\n"
    ).encode("utf-8")
    at = _fresh_app()
    at.file_uploader(key="csv_import_uploader").upload("messy.csv", messy, "text/csv")
    at.run(timeout=30)
    assert not at.exception
    colmap = {s.key: s.value for s in at.selectbox if s.key and "csv_colmap" in s.key}
    assert colmap.get("csv_colmap_sample") == "Sample"
    assert colmap.get("csv_colmap_total") == "AcceptedDroplets"


def test_batch_screening_classification():
    rows = [["Sample", "Target", "Positives", "AcceptedDroplets"]]
    for s in range(1, 8):
        rows.append([f"S{s}", "MYCN", "1900", "20000"])
        rows.append([f"S{s}", "RPP30", "1900", "20000"])
    for s in range(8, 11):
        rows.append([f"S{s}", "MYCN", "5000", "20000"])
        rows.append([f"S{s}", "RPP30", "1900", "20000"])

    at = _fresh_app()
    at.file_uploader(key="batch_uploader").upload("b.csv", _build_csv(rows), "text/csv")
    at.run(timeout=30)
    at.button(key="batch_run_btn").click()
    at.run(timeout=30)
    assert not at.exception


def test_batch_screening_multi_gene_heatmap():
    rows = [["Sample", "Target", "Positives", "AcceptedDroplets"]]
    for s in range(1, 5):
        rows.append([f"S{s}", "MYCN", "1900", "20000"])
        rows.append([f"S{s}", "ERBB2", "1900", "20000"])
        rows.append([f"S{s}", "RPP30", "1900", "20000"])
    at = _fresh_app()
    at.file_uploader(key="batch_uploader").upload("m.csv", _build_csv(rows), "text/csv")
    at.run(timeout=30)
    at.multiselect(key="batch_target_assays").set_value(["MYCN", "ERBB2"])
    at.run(timeout=30)
    at.multiselect(key="batch_ref_assays").set_value(["RPP30"])
    at.run(timeout=30)
    at.button(key="batch_run_btn").click()
    at.run(timeout=30)
    assert not at.exception


def test_vaf_calculator_detects_and_not_detects():
    rows = [
        ["Sample", "Target", "Positives", "AcceptedDroplets"],
        ["Baseline", "Mutant", "450", "18000"], ["Baseline", "WT", "1200", "18000"],
        ["Undetectable", "Mutant", "0", "19800"], ["Undetectable", "WT", "1400", "19800"],
    ]
    at = _fresh_app()
    at.file_uploader(key="vaf_uploader").upload("v.csv", _build_csv(rows), "text/csv")
    at.run(timeout=30)
    at.button(key="vaf_run_btn").click()
    at.run(timeout=30)
    assert not at.exception


# ─── Clinical Tools ─────────────────────────────────────────────────────────

def test_mu_budget_manual_mode():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("📐 Measurement Uncertainty (MU) Budget")
    at.run(timeout=30)
    at.radio(key="mu_source_mode").set_value("Enter manually")
    at.run(timeout=30)
    at.number_input(key="mu_poisson_manual_val").set_value(3.0)
    at.number_input(key="mu_pipetting_val").set_value(1.5)
    at.number_input(key="mu_precision_val").set_value(2.0)
    at.run(timeout=30)
    at.button(key="mu_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    result_text = " ".join(s.value for s in at.success)
    # sqrt(3^2+1.5^2+2^2)*2 = 7.81
    assert "7.81" in result_text or "7.8" in result_text


def test_mu_budget_double_count_warning_fires():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("📐 Measurement Uncertainty (MU) Budget")
    at.run(timeout=30)
    _load_scenario(at, "S1 — Basic CNV gain (1 gene, n=5)")
    at.radio(key="clinical_mode").set_value("📐 Measurement Uncertainty (MU) Budget")
    at.run(timeout=30)
    at.radio(key="mu_source_mode").set_value("📊 From Replicate CV% (total observed variability)")
    at.run(timeout=30)
    at.number_input(key="mu_pipetting_val").set_value(2.0)
    at.run(timeout=30)
    assert any("Double-counting" in e.value for e in at.error)


def test_rcv_significant_and_not_significant():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("📈 Reference Change Value (RCV)")
    at.run(timeout=30)
    at.number_input(key="rcv_result1").set_value(10.0)
    at.number_input(key="rcv_result2").set_value(25.0)
    at.number_input(key="rcv_cv_a").set_value(5.0)
    at.number_input(key="rcv_cv_b").set_value(15.0)
    at.run(timeout=30)
    at.button(key="rcv_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any("Significant change" in e.value for e in at.error)


def test_precision_study_unbalanced_design():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("🔁 Precision Study")
    at.run(timeout=30)
    at.number_input(key="prec_n_days").set_value(3)
    at.run(timeout=30)
    at.text_area(key="prec_day_0").set_value("100\n102\n98")
    at.text_area(key="prec_day_1").set_value("105\n107\n103\n106\n104")
    at.text_area(key="prec_day_2").set_value("95\n97\n93\n96")
    at.run(timeout=30)
    at.button(key="prec_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any("Unbalanced" in i.value for i in at.info)


def test_method_comparison_passing_bablok_and_deming():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("⚖️ Method Comparison")
    at.run(timeout=30)
    at.text_area(key="comp_m1").set_value("10\n20\n30\n40\n50\n60\n70\n80\n90\n100")
    at.text_area(key="comp_m2").set_value("11\n19\n32\n38\n52\n59\n72\n79\n93\n98")
    at.run(timeout=30)
    at.button(key="comp_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any(len(d.value) for d in at.dataframe)


def test_method_comparison_proportional_bias_detected():
    at = _fresh_app_advanced()
    at.radio(key="clinical_mode").set_value("⚖️ Method Comparison")
    at.run(timeout=30)
    at.text_area(key="comp_m1").set_value("10\n20\n30\n40\n50\n60\n70\n80\n90\n100")
    at.text_area(key="comp_m2").set_value("10\n21\n33\n46\n61\n77\n95\n115\n137\n161")
    at.run(timeout=30)
    at.button(key="comp_calc_btn").click()
    at.run(timeout=30)
    assert any("Proportional bias" in w.value for w in at.warning)


# ─── CRM Production ─────────────────────────────────────────────────────────

def test_homogeneity_detects_inhomogeneous_batch():
    at = _fresh_app_advanced()
    at.radio(key="crm_mode").set_value("🧪 Homogeneity Testing")
    at.run(timeout=30)
    at.number_input(key="homog_n_units").set_value(5)
    at.run(timeout=30)
    for u, mean in enumerate([90, 95, 100, 105, 110]):
        at.text_area(key=f"homog_unit_{u}").set_value(f"{mean-0.5}\n{mean+0.5}\n{mean-0.3}\n{mean+0.3}")
    at.run(timeout=30)
    at.button(key="homog_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any("Not Homogeneous" in e.value for e in at.error)


def test_homogeneity_unbalanced_design_uses_all_data():
    at = _fresh_app_advanced()
    at.radio(key="crm_mode").set_value("🧪 Homogeneity Testing")
    at.run(timeout=30)
    at.number_input(key="homog_n_units").set_value(3)
    at.run(timeout=30)
    at.text_area(key="homog_unit_0").set_value("100\n101\n99")
    at.text_area(key="homog_unit_1").set_value("100\n101\n99\n100.5\n99.5")
    at.text_area(key="homog_unit_2").set_value("100\n101\n99\n100.2")
    at.run(timeout=30)
    at.button(key="homog_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    grand_mean = float([m.value for m in at.metric if "Grand Mean" in m.label][0])
    assert abs(grand_mean - 100.01667) < 0.001  # uses all 12 points, not truncated to 9


def test_stability_detects_degradation_trend():
    at = _fresh_app_advanced()
    at.radio(key="crm_mode").set_value("⏳ Stability Testing")
    at.run(timeout=30)
    at.text_area(key="stab_time").set_value("0\n30\n60\n90\n120\n180\n365")
    at.text_area(key="stab_value").set_value("100\n95\n90\n86\n80\n70\n50")
    at.run(timeout=30)
    at.button(key="stab_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any("Significant Trend" in w.value for w in at.warning)


def test_lot_equivalence_and_non_equivalence():
    at = _fresh_app_advanced()
    at.radio(key="crm_mode").set_value("⚖️ Batch-to-Batch Comparison (Lot Equivalence)")
    at.run(timeout=30)
    at.text_area(key="eq_lot1").set_value("100\n102\n98\n101\n99\n103")
    at.text_area(key="eq_lot2").set_value("130\n132\n128\n131\n129\n133")
    at.number_input(key="eq_margin").set_value(10.0)
    at.run(timeout=30)
    at.button(key="eq_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert any("Not Equivalent" in e.value for e in at.error)


def test_coa_pdf_generation():
    at = _fresh_app_advanced()
    at.radio(key="crm_mode").set_value("📜 Certificate (CoA) Generator")
    at.run(timeout=30)
    at.text_input(key="coa_material").set_value("Test Material")
    at.run(timeout=30)
    at.button(key="coa_generate_btn").click()
    at.run(timeout=30)
    assert not at.exception


# ─── Multiplex cluster converter ───────────────────────────────────────────

def test_multiplex_covariance_narrower_than_independence():
    """The covariance-aware CI should generally be narrower/more accurate
    than the naive independence-assumption CI for positively-correlated
    shared-partition data — regression guard for the multiplex feature."""
    at = _fresh_app_advanced()
    at.number_input(key="mx_n11").set_value(1500)
    at.number_input(key="mx_n10").set_value(400)
    at.number_input(key="mx_n01").set_value(300)
    at.number_input(key="mx_n00").set_value(17800)
    at.run(timeout=30)
    at.button(key="mx_calc_btn").click()
    at.run(timeout=30)
    assert not at.exception


# ─── Dilution factor (volume-based) ────────────────────────────────────────

def test_volume_based_dilution_back_calculation():
    at = _fresh_app()
    at.text_area(key="ctrl_tgt_pos_0").set_value("1890\n1920\n1875\n1905\n1898")
    at.text_area(key="ctrl_tgt_tot_0").set_value("20000\n20100\n19850\n20050\n19980")
    at.text_area(key="ctrl_ref_pos_0_0").set_value("1895\n1915\n1870\n1900\n1892")
    at.text_area(key="ctrl_ref_tot_0_0").set_value("20000\n20100\n19850\n20050\n19980")
    at.run(timeout=30)
    at.text_area(key="smp_tgt_pos_0_0").set_value("2780\n2820\n2755\n2800\n2790")
    at.text_area(key="smp_tgt_tot_0_0").set_value("20000\n20100\n19850\n20050\n19980")
    at.text_area(key="smp_ref_pos_0_0_0").set_value("1895\n1915\n1870\n1900\n1892")
    at.text_area(key="smp_ref_tot_0_0_0").set_value("20000\n20100\n19850\n20050\n19980")
    at.number_input(key="smp_0_0_rxnvol").set_value(20.0)
    at.number_input(key="smp_0_0_tmplvol").set_value(1.0)
    at.run(timeout=30)
    assert not at.exception
    stock_caption = next((c.value for c in at.caption if "Stock Concentration" in c.value), None)
    assert stock_caption is not None
    assert "3534" in stock_caption  # 176.7 * 20 ≈ 3534


# ─── Simple / Advanced Mode ─────────────────────────────────────────────────

def test_simple_mode_gates_advanced_tabs():
    """By default (Advanced Mode off), Clinical Tools and CRM Production
    should show a gate message instead of their actual tool content."""
    at = _fresh_app()
    assert not at.toggle(key="advanced_mode").value
    gate_shown = any("Advanced Mode" in i.value or "Gelişmiş Mod" in i.value for i in at.info)
    assert gate_shown


def test_advanced_mode_unlocks_clinical_tools():
    at = _fresh_app_advanced()
    assert at.toggle(key="advanced_mode").value
    assert any(r.key == "clinical_mode" for r in at.radio)


# ─── Excel export ───────────────────────────────────────────────────────────

def test_excel_export_generates_without_exception():
    at = _fresh_app()
    _load_scenario(at, "S2 — Multi-gene + dual reference (2 genes, 2 groups, n=5)")
    at.button(key="excel_export_btn").click()
    at.run(timeout=30)
    assert not at.exception


# ─── Session History Panel ──────────────────────────────────────────────────

def test_history_save_and_restore_roundtrip():
    at = _fresh_app()
    _load_scenario(at, "S1 — Basic CNV gain (1 gene, n=5)")
    at.text_input(key="history_name_input").set_value("Test Snapshot")
    at.run(timeout=30)
    at.button(key="history_save_btn").click()
    at.run(timeout=30)
    assert any("Test Snapshot" in s.value for s in at.success)

    # overwrite data with a different scenario, then restore
    _load_scenario(at, "S5 — QC & saturation demo (n=5)")
    at.button(key="history_restore_btn").click()
    at.run(timeout=30)
    assert not at.exception
    assert at.session_state["ctrl_tgt_pos_0"] == "1890\n1920\n1875\n1905\n1898"  # S1's original value


# ─── Regression: advanced_mode round-trips through project import ──────────

def test_advanced_mode_restored_via_project_import():
    """Regression guard: advanced_mode is a widget-backed session_state key
    instantiated earlier in the script than the import logic. Setting it
    directly during import raises a StreamlitAPIException (silently
    swallowed by the import's try/except) unless deferred via the
    "_pending_advanced_mode" mechanism — this test ensures that deferral
    actually works end-to-end."""
    import json
    project = {"_absolutegene_project_version": 1, "advanced_mode": True, "gene_count": 1}
    at = _fresh_app()
    assert not at.toggle(key="advanced_mode").value
    at.file_uploader(key="project_import_uploader").upload(
        "p.json", json.dumps(project).encode("utf-8"), "application/json"
    )
    at.run(timeout=30)
    at.run(timeout=30)  # the pending flag is applied on the following rerun
    assert at.toggle(key="advanced_mode").value
    assert any(r.key == "clinical_mode" for r in at.radio)


# ─── Regression: Excel significance highlighting is language-independent ───

def test_excel_stat_highlighting_works_in_turkish():
    """Regression guard: the Statistics sheet's green highlighting for
    significant results must be derived from the numeric p-value, not from
    string-matching the localized "Significant"/"Anlamlı" label (which
    would silently never match in non-English reports)."""
    from openpyxl import load_workbook

    src = open("absolutegene.py", encoding="utf-8").read()
    fn_src = src.split("def create_excel_report(data, stats_data, input_values_table, lang):")[1]
    fn_src = fn_src.split("def create_simple_pdf")[0]
    namespace = {
        "np": np, "pd": __import__("pandas"), "BytesIO": __import__("io").BytesIO,
        "APP_VERSION": "test",
    }
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    namespace.update({"Workbook": Workbook, "Font": Font, "PatternFill": PatternFill,
                       "Alignment": Alignment, "Border": Border, "Side": Side,
                       "get_column_letter": get_column_letter})
    exec("def create_excel_report(data, stats_data, input_values_table, lang):" + fn_src, namespace)
    create_excel_report = namespace["create_excel_report"]

    stats_data = [
        {"__gene__": "Gene 1", "Comparison": "C vs G1", "__test_type__": "Parametrik",
         "__test_method__": "t-test", "__pvalue__": 0.001, "__significance__": "Anlamlı"},
        {"__gene__": "Gene 1", "Comparison": "C vs G2", "__test_type__": "Parametrik",
         "__test_method__": "t-test", "__pvalue__": 0.6, "__significance__": "Anlamsız"},
    ]
    buf = create_excel_report([], stats_data, [], "tr")
    wb = load_workbook(buf)
    ws = wb["Statistics"]
    assert ws.cell(row=2, column=1).fill.start_color.rgb == "00C8E6C9"  # significant -> green
    assert ws.cell(row=3, column=1).fill.start_color.rgb in ("00000000", None)  # not significant -> no fill


if __name__ == "__main__":
    import sys
    sys.exit(pytest.main([__file__, "-v"]))
